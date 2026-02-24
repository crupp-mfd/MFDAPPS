from __future__ import annotations

from dataclasses import dataclass
from datetime import date, datetime, time
from decimal import Decimal
from io import BytesIO
from pathlib import Path
import re
from typing import Any, Sequence

from openpyxl import load_workbook
from openpyxl.workbook.workbook import Workbook


@dataclass(frozen=True)
class SheetContent:
    name: str
    headers: list[str]
    rows: list[dict[str, Any]]
    source_row_count: int


@dataclass(frozen=True)
class WorkbookContent:
    source_name: str
    sheets: list[SheetContent]


_NULL_MARKERS = {"#n/a", "n/a", "na", "00:00:00", "0:00:00"}
_CM_ALLOWED_ROW_TYPES = {"header", "batch", "wagon"}
_CM_EXCESS_COLUMN = "excess_mileage_eur_per_km"
_CM_EXCESS_AMOUNT_COLUMN = "excess_mileage_amount_eur"
_CM_EXCESS_BASIS_COLUMN = "excess_mileage_basis_km"
_CM_SUMMARY_KEYWORDS = ("total", "subtotal", "sum", "summe", "gesamt")
_CM_EXCESS_10T_PATTERN = re.compile(r"\b(\d+)\s*t\b", re.IGNORECASE)
_CM_EXCESS_NUMBER_PATTERN = re.compile(r"[-+]?\d+(?:[.,]\d+)?")
_CM_EXCESS_KEYWORD_BASIS_PATTERN = re.compile(
    r"(?:per|pro|je|all|alle)\s*([0-9][0-9 .']*)\s*(?:km|kilometer|kilometern)?",
    re.IGNORECASE,
)
_CM_EXCESS_PER_KM_PATTERN = re.compile(r"(?:per|pro|je|all|alle)\s*km\b", re.IGNORECASE)
_CM_EXCESS_BASIS_PATTERN = re.compile(r"(?<![A-Za-z0-9])(\d{1,3}(?:[ .']\d{3})+|\d+)\s*(km)?\b", re.IGNORECASE)


def _normalize_text_marker(value: str) -> str | None:
    stripped = value.strip()
    if not stripped:
        return ""
    if stripped.lower() in _NULL_MARKERS:
        return None
    return value


def _normalize_cell(value: Any) -> Any:
    if value is None:
        return None
    if isinstance(value, str):
        return _normalize_text_marker(value)
    if isinstance(value, time):
        # CM nutzt 00:00:00 als Pseudo-NULL in Datumsfeldern.
        if value == time(0, 0, 0):
            return None
        return value.isoformat()
    if isinstance(value, (datetime, date, time)):
        return value.isoformat()
    if isinstance(value, Decimal):
        return str(value)
    if isinstance(value, bytes):
        return value.decode("utf-8", errors="replace")
    if isinstance(value, (str, int, float, bool)):
        return value
    return str(value)


def _sanitize_identifier(raw: Any, fallback: str) -> str:
    text = str(raw).strip().lower() if raw is not None else ""
    text = re.sub(r"\s+", "_", text)
    text = re.sub(r"[^a-z0-9_]", "_", text)
    text = re.sub(r"_+", "_", text).strip("_")
    if not text:
        text = fallback
    if text[0].isdigit():
        text = f"col_{text}"
    return text


def _deduplicate(names: list[str]) -> list[str]:
    used: dict[str, int] = {}
    unique: list[str] = []
    for name in names:
        base = name
        count = used.get(base, 0)
        if count == 0:
            unique.append(base)
        else:
            unique.append(f"{base}_{count + 1}")
        used[base] = count + 1
    return unique


def _normalize_number_token(raw: str) -> float | None:
    token = raw.strip()
    if not token:
        return None
    token = token.replace("'", "").replace(" ", "")
    if "," in token and "." in token:
        if token.rfind(",") > token.rfind("."):
            token = token.replace(".", "").replace(",", ".")
        else:
            token = token.replace(",", "")
    else:
        token = token.replace(",", ".")
    try:
        return float(token)
    except ValueError:
        return None


def _normalize_basis_token(raw: str) -> int | None:
    token = raw.strip().replace("'", "").replace(" ", "").replace(".", "")
    if not token:
        return None
    if not token.isdigit():
        return None
    value = int(token)
    return value if value > 0 else None


def _format_numeric(value: float) -> str:
    return format(value, "g")


def _is_empty_row(row: dict[str, Any]) -> bool:
    for value in row.values():
        if value is None:
            continue
        if isinstance(value, str) and value.strip() == "":
            continue
        return False
    return True


def _row_contains_keyword(row: dict[str, Any], keywords: tuple[str, ...]) -> bool:
    for value in row.values():
        if value is None:
            continue
        text = str(value).strip().lower()
        if not text:
            continue
        if any(keyword in text for keyword in keywords):
            return True
    return False


def _is_cm_summary_row(row: dict[str, Any]) -> bool:
    row_type = str(row.get("row_type") or "").strip().lower()
    if row_type in _CM_ALLOWED_ROW_TYPES:
        return False

    contract_number = str(row.get("contract_number_m3") or "").strip().lower()
    if contract_number in _CM_SUMMARY_KEYWORDS:
        return True

    # Fallback für Summenzeilen ohne sauberen Row Type.
    return row_type == "" and _row_contains_keyword(row=row, keywords=_CM_SUMMARY_KEYWORDS)


def _parse_cm_excess_mileage(raw_value: Any) -> tuple[str | None, int | None]:
    if raw_value is None:
        return None, None

    text = str(raw_value).strip()
    if not text:
        return None, None

    numeric_only = _normalize_number_token(text)
    if numeric_only is not None:
        return _format_numeric(numeric_only), 1

    amount_match = _CM_EXCESS_NUMBER_PATTERN.search(text)
    amount_value = _normalize_number_token(amount_match.group(0)) if amount_match else None

    lower = text.lower()
    ten_t_match = _CM_EXCESS_10T_PATTERN.search(lower)
    if ten_t_match:
        ten_t_value = int(ten_t_match.group(1))
        basis = ten_t_value * 1000
        return (_format_numeric(amount_value) if amount_value is not None else None), basis

    keyword_basis = _CM_EXCESS_KEYWORD_BASIS_PATTERN.search(text)
    if keyword_basis:
        basis_from_keyword = _normalize_basis_token(keyword_basis.group(1))
        if basis_from_keyword is not None:
            return (_format_numeric(amount_value) if amount_value is not None else None), basis_from_keyword
    if _CM_EXCESS_PER_KM_PATTERN.search(lower):
        return (_format_numeric(amount_value) if amount_value is not None else None), 1

    basis_candidates: list[int] = []
    for basis_match in _CM_EXCESS_BASIS_PATTERN.finditer(text):
        token = basis_match.group(1)
        normalized = _normalize_basis_token(token)
        if normalized is not None:
            basis_candidates.append(normalized)

    basis_value: int | None = None
    if basis_candidates:
        basis_value = max(basis_candidates)
    elif any(marker in lower for marker in ("per km", "je km", "pro km")):
        basis_value = 1

    return (_format_numeric(amount_value) if amount_value is not None else None), basis_value


class ExcelWorkbookReader:
    def read(
        self,
        file_path: Path,
        sheet_names: Sequence[str] | None,
        header_row: int,
        skip_rows: int,
        max_rows_per_sheet: int | None,
    ) -> WorkbookContent:
        workbook = load_workbook(filename=file_path, read_only=True, data_only=True, keep_links=False)
        return self._read_loaded_workbook(
            workbook=workbook,
            source_name=str(file_path),
            sheet_names=sheet_names,
            header_row=header_row,
            skip_rows=skip_rows,
            max_rows_per_sheet=max_rows_per_sheet,
        )

    def read_bytes(
        self,
        content: bytes,
        source_name: str,
        sheet_names: Sequence[str] | None,
        header_row: int,
        skip_rows: int,
        max_rows_per_sheet: int | None,
    ) -> WorkbookContent:
        workbook = load_workbook(filename=BytesIO(content), read_only=True, data_only=True, keep_links=False)
        return self._read_loaded_workbook(
            workbook=workbook,
            source_name=source_name,
            sheet_names=sheet_names,
            header_row=header_row,
            skip_rows=skip_rows,
            max_rows_per_sheet=max_rows_per_sheet,
        )

    def _read_loaded_workbook(
        self,
        workbook: Workbook,
        source_name: str,
        sheet_names: Sequence[str] | None,
        header_row: int,
        skip_rows: int,
        max_rows_per_sheet: int | None,
    ) -> WorkbookContent:
        try:
            selected_sheets = list(sheet_names) if sheet_names else list(workbook.sheetnames)
            missing = [sheet_name for sheet_name in selected_sheets if sheet_name not in workbook.sheetnames]
            if missing:
                raise ValueError(f"Unknown sheet(s): {', '.join(missing)}")

            sheets: list[SheetContent] = []
            for sheet_name in selected_sheets:
                worksheet = workbook[sheet_name]
                raw_rows = [tuple(_normalize_cell(value) for value in row) for row in worksheet.iter_rows(values_only=True)]

                header_index = header_row - 1
                start_index = header_index + 1 + skip_rows
                if start_index < 0:
                    start_index = 0

                row_values = raw_rows[start_index:] if start_index < len(raw_rows) else []
                if max_rows_per_sheet is not None:
                    row_values = row_values[:max_rows_per_sheet]

                header_source = raw_rows[header_index] if 0 <= header_index < len(raw_rows) else ()
                max_width = max(
                    len(header_source),
                    max((len(row) for row in row_values), default=0),
                )

                headers = self._build_headers(header_source=header_source, width=max_width)
                rows = [self._map_row(headers=headers, row_values=row) for row in row_values]
                rows = [row for row in rows if not _is_empty_row(row)]
                headers, rows = self._apply_sheet_rules(sheet_name=sheet_name, headers=headers, rows=rows)

                sheets.append(
                    SheetContent(
                        name=sheet_name,
                        headers=headers,
                        rows=rows,
                        source_row_count=len(rows),
                    )
                )

            return WorkbookContent(source_name=source_name, sheets=sheets)
        finally:
            workbook.close()

    def _build_headers(self, header_source: Sequence[Any], width: int) -> list[str]:
        if width <= 0:
            return []
        source_values = list(header_source[:width]) + [None] * max(0, width - len(header_source))
        raw_headers = [
            _sanitize_identifier(raw=value, fallback=f"column_{index}")
            for index, value in enumerate(source_values, start=1)
        ]
        return _deduplicate(raw_headers)

    def _map_row(self, headers: list[str], row_values: Sequence[Any]) -> dict[str, Any]:
        if not headers:
            return {}
        padded = list(row_values[: len(headers)]) + [None] * max(0, len(headers) - len(row_values))
        return {header: padded[index] for index, header in enumerate(headers)}

    def _apply_sheet_rules(
        self,
        sheet_name: str,
        headers: list[str],
        rows: list[dict[str, Any]],
    ) -> tuple[list[str], list[dict[str, Any]]]:
        if sheet_name.strip().lower() != "cm":
            return headers, rows

        if _CM_EXCESS_AMOUNT_COLUMN not in headers:
            headers.append(_CM_EXCESS_AMOUNT_COLUMN)
        if _CM_EXCESS_BASIS_COLUMN not in headers:
            headers.append(_CM_EXCESS_BASIS_COLUMN)

        transformed: list[dict[str, Any]] = []
        for row in rows:
            if _is_cm_summary_row(row):
                continue
            amount, basis = _parse_cm_excess_mileage(row.get(_CM_EXCESS_COLUMN))
            row[_CM_EXCESS_AMOUNT_COLUMN] = amount
            row[_CM_EXCESS_BASIS_COLUMN] = basis
            transformed.append(row)

        return headers, transformed
