from __future__ import annotations

from copy import copy
from datetime import datetime, time as dt_time, timezone
from pathlib import Path
import re
import sqlite3
from threading import Lock, Thread
from typing import Callable, Literal
from uuid import uuid4

from openpyxl import Workbook, load_workbook

from app_excel_import.config import ExcelImportSettings
from app_excel_import.models import (
    ExportSnapshotJobStartResponse,
    ExportSnapshotJobStatusResponse,
    ExportSnapshotResponse,
    GoFabricResponse,
    ImportJobResponse,
    ImportRequest,
    ImportSheetResult,
    ImportSummary,
    PreviewRequest,
    SharePointImportRequest,
    SharePointPreviewRequest,
    SheetPreview,
    TransferTestResponse,
    WorkbookPreviewResponse,
)
from app_excel_import.services.database_writer import SQLiteImportWriter
from app_excel_import.services.excel_reader import ExcelWorkbookReader, WorkbookContent
from app_excel_import.services.fabric_writer import FabricSqlWriter
from app_excel_import.services.sharepoint_client import SharePointGraphClient


class ExcelImportService:
    def __init__(self, settings: ExcelImportSettings) -> None:
        self.settings = settings
        self.reader = ExcelWorkbookReader()
        self.writer = SQLiteImportWriter(settings.sqlite_path)
        self.fabric_writer = FabricSqlWriter(settings)
        self.sharepoint_client = SharePointGraphClient(settings)
        self._export_jobs_lock = Lock()
        self._export_jobs: dict[str, dict[str, object | None]] = {}

    def configuration_health(self) -> dict:
        has_client_credentials = bool(
            self.settings.sharepoint_tenant_id and self.settings.sharepoint_client_id and self.settings.sharepoint_client_secret
        )
        return {
            "runtime_root": str(self.settings.runtime_root),
            "input_root": str(self.settings.input_root),
            "database_path": str(self.settings.sqlite_path),
            "database_exists": self.settings.sqlite_path.exists(),
            "sharepoint": {
                "configured": self.sharepoint_client.is_configured(),
                "site_hostname": self.settings.sharepoint_site_hostname,
                "site_path": self.settings.sharepoint_site_path,
                "default_workbook_path": self.settings.sharepoint_default_workbook_path,
                "auth_mode": (
                    "managed_identity+client_credentials"
                    if self.settings.sharepoint_use_managed_identity and has_client_credentials
                    else "managed_identity"
                    if self.settings.sharepoint_use_managed_identity
                    else "client_credentials"
                    if has_client_credentials
                    else "unconfigured"
                ),
            },
            "fabric": {
                "configured": self.fabric_writer.is_configured(),
                "missing": self.fabric_writer.missing_configuration(),
                "sql_server": self.settings.fabric_sql_server,
                "sql_database": self.settings.fabric_sql_database,
                "target_schema": self.settings.go_target_schema,
                "raw_table": self.settings.go_raw_table,
                "target_table": self.settings.go_target_table,
            },
            "go_job": {
                "workbook_path": self.settings.go_workbook_path,
                "sheet_name": self.settings.go_sheet_name,
                "header_row": self.settings.go_header_row,
                "skip_rows": self.settings.go_skip_rows,
                "export_dir": str(self.settings.go_export_dir),
            },
        }

    def run_transfer_test(self) -> TransferTestResponse:
        self.settings.runtime_root.mkdir(parents=True, exist_ok=True)
        self.settings.sqlite_path.parent.mkdir(parents=True, exist_ok=True)

        tested_at_utc = datetime.now(timezone.utc).isoformat()
        with sqlite3.connect(self.settings.sqlite_path) as connection:
            connection.execute(
                """
                CREATE TABLE IF NOT EXISTS transfer_test_events (
                    id INTEGER PRIMARY KEY AUTOINCREMENT,
                    tested_at_utc TEXT NOT NULL,
                    source TEXT NOT NULL
                )
                """
            )
            connection.execute(
                """
                INSERT INTO transfer_test_events (tested_at_utc, source)
                VALUES (?, ?)
                """,
                (tested_at_utc, "go_button"),
            )
            connection.commit()

        return TransferTestResponse(
            status="ok",
            message="Transfer test successful (frontend -> API -> database).",
            tested_at_utc=tested_at_utc,
            database_path=str(self.settings.sqlite_path),
        )

    def run_go_fabric_import(self) -> GoFabricResponse:
        started_at_utc = datetime.now(timezone.utc).isoformat()
        workbook_path = self._resolve_input_path(self.settings.go_workbook_path)
        workbook = self.reader.read(
            file_path=workbook_path,
            sheet_names=[self.settings.go_sheet_name],
            header_row=self.settings.go_header_row,
            skip_rows=self.settings.go_skip_rows,
            max_rows_per_sheet=None,
        )
        if not workbook.sheets:
            raise ValueError("Keine Sheets fuer GO-Import gefunden.")
        sheet = workbook.sheets[0]
        if sheet.source_row_count == 0:
            raise ValueError(f"Sheet '{sheet.name}' enthaelt keine Datenzeilen.")

        result = self.fabric_writer.import_contract_management_daily(
            workbook=workbook,
            target_schema=self.settings.go_target_schema,
            raw_table=self.settings.go_raw_table,
            clean_table=self.settings.go_target_table,
            source_file=str(workbook_path),
        )
        finished_at_utc = datetime.now(timezone.utc).isoformat()
        target_tables = (
            f"{self.settings.go_target_schema}.{result['header_table']}, "
            f"{self.settings.go_target_schema}.{result['batch_table']}, "
            f"{self.settings.go_target_schema}.{result['wagon_table']}, "
            f"{self.settings.go_target_schema}.{result['structure_table']}"
        )
        return GoFabricResponse(
            status="ok",
            message="GO Delta-Import nach Fabric erfolgreich abgeschlossen (Header/Batch/Wagon + Structure).",
            started_at_utc=started_at_utc,
            finished_at_utc=finished_at_utc,
            workbook_path=str(workbook_path),
            sheet_name=result["sheet_name"],
            raw_table=f"{self.settings.go_target_schema}.{self.settings.go_raw_table}",
            target_table=target_tables,
            batch_id=result["batch_id"],
            import_timestamp_utc=result["import_timestamp_utc"],
            input_rows=result["input_rows"],
            deduplicated_rows=result["deduplicated_rows"],
            raw_inserted_rows=result["raw_inserted_rows"],
            clean_new_rows=result["clean_new_rows"],
            clean_unchanged_rows=result["clean_unchanged_rows"],
            clean_closed_rows=result["clean_closed_rows"],
            inserted_rows=result["raw_inserted_rows"],
        )

    def export_contract_management_snapshot(self, snapshot_iso: str | None) -> ExportSnapshotResponse:
        requested_snapshot = self._parse_snapshot(snapshot_iso) if snapshot_iso else None
        snapshot_payload = self.fabric_writer.export_contract_management_snapshot(
            target_schema=self.settings.go_target_schema,
            clean_table=self.settings.go_target_table,
            snapshot_at_utc=requested_snapshot,
        )
        headers = [str(header) for header in snapshot_payload["headers"]]
        rows = [dict(row) for row in snapshot_payload["rows"]]
        effective_snapshot = snapshot_payload["effective_snapshot"]
        source_file = str(snapshot_payload["source_file"] or "").strip() or None

        self.settings.go_export_dir.mkdir(parents=True, exist_ok=True)
        file_suffix = effective_snapshot.strftime("%Y%m%dT%H%M%SZ")
        export_path = self.settings.go_export_dir / f"ContractManagement_snapshot_{file_suffix}.xlsx"
        self._write_snapshot_workbook(
            export_path=export_path,
            headers=headers,
            rows=rows,
            source_template=source_file,
            progress_callback=None,
        )

        return ExportSnapshotResponse(
            status="ok",
            message="Snapshot als Excel exportiert (Originalformat aus Vorlage).",
            requested_snapshot_utc=requested_snapshot.isoformat() if requested_snapshot else None,
            effective_snapshot_utc=effective_snapshot.isoformat(),
            target_table=f"{self.settings.go_target_schema}.{self.settings.go_target_table}",
            rows_exported=len(rows),
            workbook_path=str(export_path),
        )

    def start_export_snapshot_job(self, snapshot_iso: str | None) -> ExportSnapshotJobStartResponse:
        requested_snapshot = self._parse_snapshot(snapshot_iso) if snapshot_iso else None
        now_utc = datetime.now(timezone.utc)
        job_id = f"export-{uuid4()}"
        self._set_export_job(
            job_id=job_id,
            updates={
                "job_id": job_id,
                "status": "running",
                "message": "Export-Job gestartet.",
                "target_table": f"{self.settings.go_target_schema}.{self.settings.go_target_table}",
                "requested_snapshot_utc": requested_snapshot.isoformat() if requested_snapshot else None,
                "effective_snapshot_utc": None,
                "started_at_utc": now_utc.isoformat(),
                "finished_at_utc": None,
                "total_rows": None,
                "exported_rows": 0,
                "workbook_path": None,
                "download_url": None,
                "error": None,
            },
        )

        worker = Thread(
            target=self._run_export_snapshot_job,
            args=(job_id, requested_snapshot),
            daemon=True,
        )
        worker.start()
        return ExportSnapshotJobStartResponse(
            status="accepted",
            message="Export-Job wurde gestartet.",
            job_id=job_id,
            created_at_utc=now_utc.isoformat(),
        )

    def get_export_snapshot_job(self, job_id: str) -> ExportSnapshotJobStatusResponse:
        job = self._get_export_job(job_id)
        if job is None:
            raise ValueError(f"Export-Job nicht gefunden: {job_id}")

        total_rows = int(job["total_rows"]) if job.get("total_rows") is not None else None
        exported_rows = int(job.get("exported_rows") or 0)
        progress_pct = None
        if total_rows and total_rows > 0:
            progress_pct = round(min(100.0, max(0.0, (exported_rows / total_rows) * 100.0)), 1)

        return ExportSnapshotJobStatusResponse(
            job_id=job_id,
            status=str(job.get("status") or "failed"),
            message=str(job.get("message") or ""),
            target_table=str(job.get("target_table") or ""),
            requested_snapshot_utc=str(job["requested_snapshot_utc"]) if job.get("requested_snapshot_utc") else None,
            effective_snapshot_utc=str(job["effective_snapshot_utc"]) if job.get("effective_snapshot_utc") else None,
            started_at_utc=str(job.get("started_at_utc") or ""),
            finished_at_utc=str(job["finished_at_utc"]) if job.get("finished_at_utc") else None,
            total_rows=total_rows,
            exported_rows=exported_rows,
            progress_pct=progress_pct,
            workbook_path=str(job["workbook_path"]) if job.get("workbook_path") else None,
            download_url=str(job["download_url"]) if job.get("download_url") else None,
            error=str(job["error"]) if job.get("error") else None,
        )

    def get_export_snapshot_file(self, job_id: str) -> Path:
        job = self._get_export_job(job_id)
        if job is None:
            # Fallback: Nach Server-Neustart ist der In-Memory-Job weg,
            # die exportierte Datei kann aber weiterhin vorhanden sein.
            fallback = self._find_export_file_by_job_id(job_id)
            if fallback is not None:
                return fallback
            raise ValueError(f"Export-Job nicht gefunden: {job_id}")
        if str(job.get("status")) != "completed":
            raise ValueError(f"Export-Job ist noch nicht abgeschlossen: {job_id}")
        workbook_path_raw = str(job.get("workbook_path") or "").strip()
        if not workbook_path_raw:
            raise RuntimeError(f"Export-Datei nicht gesetzt fuer Job: {job_id}")
        workbook_path = Path(workbook_path_raw).resolve()
        if not workbook_path.exists() or not workbook_path.is_file():
            raise RuntimeError(f"Export-Datei nicht gefunden: {workbook_path}")
        return workbook_path

    def _find_export_file_by_job_id(self, job_id: str) -> Path | None:
        export_dir = self.settings.go_export_dir
        if not export_dir.exists() or not export_dir.is_dir():
            return None
        pattern = f"ContractManagement_snapshot_*_{job_id}.xlsx"
        matches = sorted(export_dir.glob(pattern))
        if not matches:
            return None
        latest = matches[-1].resolve()
        if latest.exists() and latest.is_file():
            return latest
        return None

    def _run_export_snapshot_job(self, job_id: str, snapshot: datetime | None) -> None:
        try:
            self._set_export_job(
                job_id=job_id,
                updates={"message": "Lese Snapshot aus Fabric SQL ..."},
            )
            snapshot_payload = self.fabric_writer.export_contract_management_snapshot(
                target_schema=self.settings.go_target_schema,
                clean_table=self.settings.go_target_table,
                snapshot_at_utc=snapshot,
            )
            headers = [str(header) for header in snapshot_payload["headers"]]
            rows = [dict(row) for row in snapshot_payload["rows"]]
            effective_snapshot = snapshot_payload["effective_snapshot"]
            source_file = str(snapshot_payload["source_file"] or "").strip() or None

            self._set_export_job(
                job_id=job_id,
                updates={
                    "effective_snapshot_utc": effective_snapshot.isoformat(),
                    "total_rows": len(rows),
                    "message": "Schreibe Excel-Datei aus Vorlage ...",
                },
            )

            self.settings.go_export_dir.mkdir(parents=True, exist_ok=True)
            file_suffix = effective_snapshot.strftime("%Y%m%dT%H%M%SZ")
            export_path = self.settings.go_export_dir / f"ContractManagement_snapshot_{file_suffix}_{job_id}.xlsx"
            self._write_snapshot_workbook(
                export_path=export_path,
                headers=headers,
                rows=rows,
                source_template=source_file,
                progress_callback=lambda current, total: self._set_export_job(
                    job_id=job_id,
                    updates={
                        "exported_rows": current,
                        "message": f"Schreibe Excel-Zeilen: {current}/{total}",
                    },
                ),
            )
            self._set_export_job(
                job_id=job_id,
                updates={
                    "status": "completed",
                    "message": "Export abgeschlossen.",
                    "exported_rows": len(rows),
                    "workbook_path": str(export_path),
                    "download_url": f"/api/excel-import/export/cm/jobs/{job_id}/download",
                    "finished_at_utc": datetime.now(timezone.utc).isoformat(),
                },
            )
        except Exception as exc:
            self._set_export_job(
                job_id=job_id,
                updates={
                    "status": "failed",
                    "message": "Export fehlgeschlagen.",
                    "error": str(exc),
                    "finished_at_utc": datetime.now(timezone.utc).isoformat(),
                },
            )

    def _set_export_job(self, job_id: str, updates: dict[str, object | None]) -> None:
        with self._export_jobs_lock:
            current = self._export_jobs.get(job_id, {}).copy()
            current.update(updates)
            self._export_jobs[job_id] = current

    def _get_export_job(self, job_id: str) -> dict[str, object | None] | None:
        with self._export_jobs_lock:
            value = self._export_jobs.get(job_id)
            return value.copy() if value else None

    def preview_workbook(self, payload: PreviewRequest) -> WorkbookPreviewResponse:
        workbook_path = self._resolve_input_path(payload.file_path)
        workbook = self.reader.read(
            file_path=workbook_path,
            sheet_names=payload.sheet_names,
            header_row=payload.header_row,
            skip_rows=payload.skip_rows,
            max_rows_per_sheet=payload.max_rows_per_sheet,
        )
        return self._build_preview_response(workbook)

    def preview_sharepoint_workbook(self, payload: SharePointPreviewRequest) -> WorkbookPreviewResponse:
        download = self.sharepoint_client.download_workbook(
            workbook_path=payload.workbook_path,
            site_hostname=payload.site_hostname,
            site_path=payload.site_path,
        )
        workbook = self.reader.read_bytes(
            content=download.content,
            source_name=download.source_name,
            sheet_names=payload.sheet_names,
            header_row=payload.header_row,
            skip_rows=payload.skip_rows,
            max_rows_per_sheet=payload.max_rows_per_sheet,
        )
        return self._build_preview_response(workbook)

    def submit_import(self, payload: ImportRequest) -> ImportJobResponse:
        workbook_path = self._resolve_input_path(payload.file_path)
        workbook = self.reader.read(
            file_path=workbook_path,
            sheet_names=payload.sheet_names,
            header_row=payload.header_row,
            skip_rows=payload.skip_rows,
            max_rows_per_sheet=payload.max_rows_per_sheet,
        )
        return self._submit_prepared_import(
            workbook=workbook,
            source_name=str(workbook_path),
            target_table_prefix=payload.target_table_prefix,
            if_exists=payload.if_exists,
            dry_run=payload.dry_run,
        )

    def submit_sharepoint_import(self, payload: SharePointImportRequest) -> ImportJobResponse:
        download = self.sharepoint_client.download_workbook(
            workbook_path=payload.workbook_path,
            site_hostname=payload.site_hostname,
            site_path=payload.site_path,
        )
        workbook = self.reader.read_bytes(
            content=download.content,
            source_name=download.source_name,
            sheet_names=payload.sheet_names,
            header_row=payload.header_row,
            skip_rows=payload.skip_rows,
            max_rows_per_sheet=payload.max_rows_per_sheet,
        )
        return self._submit_prepared_import(
            workbook=workbook,
            source_name=download.source_name,
            target_table_prefix=payload.target_table_prefix,
            if_exists=payload.if_exists,
            dry_run=payload.dry_run,
        )

    def submit_default_sharepoint_import(self, dry_run: bool) -> ImportJobResponse:
        default_workbook_path = (self.settings.sharepoint_default_workbook_path or "").strip()
        if not default_workbook_path:
            raise ValueError("SHAREPOINT_DEFAULT_WORKBOOK_PATH is not configured.")

        payload = SharePointImportRequest(
            workbook_path=default_workbook_path,
            site_hostname=self.settings.sharepoint_site_hostname,
            site_path=self.settings.sharepoint_site_path,
            target_table_prefix=self.settings.sharepoint_default_table_prefix,
            if_exists="append",
            dry_run=dry_run,
        )
        return self.submit_sharepoint_import(payload)

    def _resolve_input_path(self, raw_path: str) -> Path:
        candidate = Path(raw_path).expanduser()
        if not candidate.is_absolute():
            candidate = (self.settings.input_root / candidate).resolve()
        else:
            candidate = candidate.resolve()
        if not candidate.exists():
            raise FileNotFoundError(f"Workbook file not found: {candidate}")
        if not candidate.is_file():
            raise ValueError(f"Path is not a file: {candidate}")
        return candidate

    def _build_preview_response(self, workbook: WorkbookContent) -> WorkbookPreviewResponse:
        return WorkbookPreviewResponse(
            file_path=workbook.source_name,
            sheet_count=len(workbook.sheets),
            sheets=[
                SheetPreview(
                    sheet_name=sheet.name,
                    columns=sheet.headers,
                    total_rows=sheet.source_row_count,
                    sample_rows=sheet.rows[: self.settings.preview_limit],
                )
                for sheet in workbook.sheets
            ],
        )

    def _parse_snapshot(self, raw: str) -> datetime:
        text = raw.strip()
        if not text:
            raise ValueError("Snapshot darf nicht leer sein.")
        normalized = text.replace("Z", "+00:00")
        try:
            parsed = datetime.fromisoformat(normalized)
        except ValueError as exc:
            raise ValueError(
                "Snapshot-Format ungueltig. Erwartet: YYYY-MM-DD oder ISO-8601 (z.B. 2026-02-23T18:00:00Z)."
            ) from exc

        if parsed.tzinfo is None:
            if len(text) == 10:
                parsed = datetime.combine(parsed.date(), dt_time.max)
            parsed = parsed.replace(tzinfo=timezone.utc)
        else:
            parsed = parsed.astimezone(timezone.utc)

        if len(text) == 10:
            parsed = parsed.replace(hour=23, minute=59, second=59, microsecond=0)
        return parsed

    def _sanitize_excel_header(self, raw: object | None, fallback: str) -> str:
        text = str(raw).strip().lower() if raw is not None else ""
        text = re.sub(r"\s+", "_", text)
        text = re.sub(r"[^a-z0-9_]", "_", text)
        text = re.sub(r"_+", "_", text).strip("_")
        if not text:
            text = fallback
        if text[0].isdigit():
            text = f"col_{text}"
        return text

    def _resolve_template_path(self, source_template: str | None) -> Path | None:
        raw = str(source_template or "").strip()
        if not raw:
            return None
        candidate = Path(raw).expanduser()
        if not candidate.is_absolute():
            candidate = (self.settings.input_root / candidate).resolve()
        else:
            candidate = candidate.resolve()
        if not candidate.exists() or not candidate.is_file():
            return None
        return candidate

    def _build_plain_export_rows(
        self,
        headers: list[str],
        rows: list[dict[str, object | None]],
    ) -> list[list[object | None]]:
        materialized_rows: list[list[object | None]] = []
        for row in rows:
            materialized_rows.append([row.get(header) for header in headers])
        return materialized_rows

    def _write_snapshot_workbook_plain(
        self,
        export_path: Path,
        headers: list[str],
        rows: list[dict[str, object | None]],
        progress_callback: Callable[[int, int], None] | None,
    ) -> None:
        workbook = Workbook()
        worksheet = workbook.active
        worksheet.title = self.settings.go_sheet_name
        worksheet.append(headers)
        data_rows = self._build_plain_export_rows(headers=headers, rows=rows)
        total_rows = len(data_rows)
        for index, row in enumerate(data_rows, start=1):
            worksheet.append(row)
            if progress_callback is not None and (index == 1 or index % 50 == 0 or index == total_rows):
                progress_callback(index, total_rows)
        workbook.save(export_path)

    def _capture_template_style_blueprints(
        self,
        worksheet: object,
        header_row: int,
        row_type_col: int | None,
        max_col: int,
    ) -> tuple[dict[str, dict[int, object]], dict[str, float | None], dict[int, object], float | None]:
        style_by_type: dict[str, dict[int, object]] = {}
        height_by_type: dict[str, float | None] = {}
        default_style: dict[int, object] = {}
        default_height: float | None = None
        for source_row in range(header_row + 1, worksheet.max_row + 1):
            current_style = {col: copy(worksheet.cell(row=source_row, column=col)._style) for col in range(1, max_col + 1)}
            current_height = worksheet.row_dimensions[source_row].height
            if not default_style:
                default_style = current_style
                default_height = current_height

            row_type = ""
            if row_type_col is not None:
                row_type_value = worksheet.cell(row=source_row, column=row_type_col).value
                row_type = str(row_type_value or "").strip().lower()
            if row_type in {"header", "batch", "wagon"} and row_type not in style_by_type:
                style_by_type[row_type] = current_style
                height_by_type[row_type] = current_height
            if len(style_by_type) >= 3 and default_style:
                break

        if not default_style:
            default_style = {col: copy(worksheet.cell(row=header_row, column=col)._style) for col in range(1, max_col + 1)}
        return style_by_type, height_by_type, default_style, default_height

    def _write_snapshot_workbook_from_template(
        self,
        export_path: Path,
        template_path: Path,
        headers: list[str],
        rows: list[dict[str, object | None]],
        progress_callback: Callable[[int, int], None] | None,
    ) -> None:
        template_workbook = load_workbook(filename=template_path, keep_links=False)
        try:
            sheet_name = self.settings.go_sheet_name
            template_sheet = template_workbook[sheet_name] if sheet_name in template_workbook.sheetnames else template_workbook.active
            header_row = self.settings.go_header_row
            max_col = template_sheet.max_column

            header_to_col: dict[str, int] = {}
            for col in range(1, max_col + 1):
                key = self._sanitize_excel_header(
                    raw=template_sheet.cell(row=header_row, column=col).value,
                    fallback=f"column_{col}",
                )
                if key not in header_to_col:
                    header_to_col[key] = col

            next_col = max_col + 1
            for header in headers:
                if header in header_to_col:
                    continue
                header_to_col[header] = next_col
                next_col += 1
            max_col = max(max_col, next_col - 1)

            workbook = Workbook()
            worksheet = workbook.active
            worksheet.title = str(template_sheet.title or sheet_name)
            self._copy_template_header_layout(
                template_sheet=template_sheet,
                worksheet=worksheet,
                header_row=header_row,
                max_col=max_col,
            )
            for header, col in header_to_col.items():
                if col > template_sheet.max_column:
                    worksheet.cell(row=header_row, column=col).value = header

            row_type_col = header_to_col.get("row_type")
            style_by_type, height_by_type, default_style, default_height = self._capture_template_style_blueprints(
                worksheet=template_sheet,
                header_row=header_row,
                row_type_col=row_type_col,
                max_col=max_col,
            )

            total_rows = len(rows)
            for index, row_payload in enumerate(rows, start=1):
                target_row = header_row + index
                row_type = str(row_payload.get("row_type") or "").strip().lower()
                style_row = style_by_type.get(row_type, default_style)
                for header, col in header_to_col.items():
                    value = row_payload.get(header)
                    if isinstance(value, str) and value.strip() == "":
                        value = None
                    cell = worksheet.cell(row=target_row, column=col, value=value)
                    style = style_row.get(col)
                    if style is not None:
                        cell._style = copy(style)

                row_height = height_by_type.get(row_type, default_height)
                if row_height is not None:
                    worksheet.row_dimensions[target_row].height = row_height

                if progress_callback is not None and (index == 1 or index % 50 == 0 or index == total_rows):
                    progress_callback(index, total_rows)

            workbook.save(export_path)
            workbook.close()
        finally:
            template_workbook.close()

    def _copy_template_header_layout(
        self,
        template_sheet: object,
        worksheet: object,
        header_row: int,
        max_col: int,
    ) -> None:
        for row in range(1, header_row + 1):
            worksheet.row_dimensions[row].height = template_sheet.row_dimensions[row].height
            for col in range(1, max_col + 1):
                src = template_sheet.cell(row=row, column=col)
                dst = worksheet.cell(row=row, column=col, value=src.value)
                dst._style = copy(src._style)

        for column_key, source_dim in template_sheet.column_dimensions.items():
            target_dim = worksheet.column_dimensions[column_key]
            target_dim.width = source_dim.width
            target_dim.hidden = source_dim.hidden
            target_dim.outlineLevel = source_dim.outlineLevel
            target_dim.bestFit = source_dim.bestFit

        worksheet.freeze_panes = template_sheet.freeze_panes
        worksheet.sheet_format.defaultRowHeight = template_sheet.sheet_format.defaultRowHeight

        for merged_range in template_sheet.merged_cells.ranges:
            if merged_range.max_row <= header_row:
                worksheet.merge_cells(str(merged_range))

    def _write_snapshot_workbook(
        self,
        export_path: Path,
        headers: list[str],
        rows: list[dict[str, object | None]],
        source_template: str | None,
        progress_callback: Callable[[int, int], None] | None,
    ) -> None:
        template_path = self._resolve_template_path(source_template)
        if template_path is not None:
            try:
                self._write_snapshot_workbook_from_template(
                    export_path=export_path,
                    template_path=template_path,
                    headers=headers,
                    rows=rows,
                    progress_callback=progress_callback,
                )
                return
            except Exception:
                # Fallback auf generische Export-Datei ohne Vorlage.
                pass

        self._write_snapshot_workbook_plain(
            export_path=export_path,
            headers=headers,
            rows=rows,
            progress_callback=progress_callback,
        )

    def _submit_prepared_import(
        self,
        workbook: WorkbookContent,
        source_name: str,
        target_table_prefix: str,
        if_exists: Literal["append", "replace"],
        dry_run: bool,
    ) -> ImportJobResponse:
        if dry_run:
            planned_sheets = [
                ImportSheetResult(
                    sheet_name=sheet.name,
                    target_table=self.writer.build_table_name(target_table_prefix, sheet.name),
                    inserted_rows=sheet.source_row_count,
                )
                for sheet in workbook.sheets
            ]
            summary = ImportSummary(
                workbook_path=source_name,
                database_path=str(self.settings.sqlite_path),
                processed_sheets=len(planned_sheets),
                processed_rows=sum(sheet.inserted_rows for sheet in planned_sheets),
                dry_run=True,
                sheets=planned_sheets,
            )
            return ImportJobResponse.dry_run(summary)

        write_result = self.writer.write(
            workbook=workbook,
            source_file=source_name,
            table_prefix=target_table_prefix,
            if_exists=if_exists,
        )
        summary = ImportSummary(
            workbook_path=source_name,
            database_path=str(write_result.database_path),
            processed_sheets=len(write_result.sheets),
            processed_rows=write_result.total_rows,
            dry_run=False,
            sheets=[
                ImportSheetResult(
                    sheet_name=sheet.sheet_name,
                    target_table=sheet.table_name,
                    inserted_rows=sheet.inserted_rows,
                )
                for sheet in write_result.sheets
            ],
        )
        return ImportJobResponse.completed(summary)
