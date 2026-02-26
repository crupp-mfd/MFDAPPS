"""FastAPI server serving the loader UI and paginated wagon data."""
from __future__ import annotations

import os
import time
import io
import mimetypes
import importlib
from contextlib import contextmanager
from xml.sax.saxutils import escape as xml_escape
import sqlite3
import subprocess
import shutil
import sys
import json
import re
from urllib.parse import urlsplit, urlunsplit, urlencode, unquote, quote
from pathlib import Path
from typing import List, Dict, Any, Mapping, Optional
import threading
import uuid

from datetime import datetime, date, timedelta

from fastapi import BackgroundTasks, FastAPI, HTTPException, Query, Body, Response, Request
import logging
from fastapi.responses import PlainTextResponse, JSONResponse, FileResponse, RedirectResponse, HTMLResponse
from fastapi.middleware.cors import CORSMiddleware
from fastapi.staticfiles import StaticFiles
import psycopg
import httpx
from openai import OpenAI
from openpyxl import Workbook, load_workbook
from openpyxl.cell import WriteOnlyCell
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter

try:
    from sparepart_shared.auth import is_basic_auth_valid
    from sparepart_shared.db import create_sqlite_connection
except Exception:
    import base64

    def is_basic_auth_valid(auth_header: str, expected_user: str, expected_pass: str) -> bool:
        if not expected_user or not expected_pass:
            return False
        if not auth_header.startswith("Basic "):
            return False
        encoded = auth_header.split(" ", 1)[1].strip()
        try:
            decoded = base64.b64decode(encoded).decode("utf-8")
        except Exception:
            return False
        if ":" not in decoded:
            return False
        user, password = decoded.split(":", 1)
        return user == expected_user and password == expected_pass

    def create_sqlite_connection(path: Path | str) -> sqlite3.Connection:
        db_path = Path(path)
        db_path.parent.mkdir(parents=True, exist_ok=True)
        last_error: sqlite3.OperationalError | None = None
        for attempt in range(6):
            try:
                conn = sqlite3.connect(str(db_path), timeout=30, isolation_level=None)
                conn.row_factory = sqlite3.Row
                conn.execute("PRAGMA busy_timeout = 30000")
                return conn
            except sqlite3.OperationalError as exc:
                last_error = exc
                message = str(exc).lower()
                is_transient = (
                    "unable to open database file" in message
                    or "database is locked" in message
                )
                if not is_transient or attempt == 5:
                    raise
                time.sleep(0.05 * (attempt + 1))
        if last_error is not None:
            raise last_error
        raise sqlite3.OperationalError(f"SQLite Verbindung fehlgeschlagen: {db_path}")

from .env_loader import (
    get_credentials_root,
    get_frontend_root,
    get_runtime_root,
    load_project_dotenv,
)
from .rsrd2_sync import (
    RSRDTables,
    BASE_DETAIL_TABLE,
    BASE_JSON_TABLE,
    BASE_SNAPSHOTS_TABLE,
    BASE_WAGONS_TABLE,
    init_db as rsrd_init_db,
    resolve_env_value as rsrd_resolve_env_value,
    sync_wagons as rsrd_sync_wagons,
    tables_for_env as rsrd_tables_for_env,
)
from .rsrd_compare import (
    build_erp_payload,
    compare_erp_to_rsrd,
    serialize_diffs,
    serialize_payload,
)
from .m3_api_call import (
    load_ionapi_config,
    get_access_token_service_account,
    build_base_url,
    call_m3_mi_get,
)

load_project_dotenv()
logging.basicConfig(level=logging.INFO)
_auth_logger = logging.getLogger("auth")

_msy_text_cache: Dict[str, str] = {}
_wg_tsi_txid_cache: Dict[str, str] = {}
_wg_tsi_text_cache: Dict[str, str] = {}

PROJECT_ROOT = Path(__file__).resolve().parents[1]
RUNTIME_ROOT = get_runtime_root()
CREDENTIALS_ROOT = get_credentials_root()
FRONTEND_DIR = get_frontend_root()


def _resolve_runtime_path(value: str | None, default_name: str) -> Path:
    raw = (value or "").strip()
    if not raw:
        return RUNTIME_ROOT / default_name
    candidate = Path(raw).expanduser()
    if candidate.is_absolute():
        return candidate
    return RUNTIME_ROOT / candidate


DB_PATH = _resolve_runtime_path(os.getenv("SQLITE_PATH"), "cache.db")
API_LOG_PATH = _resolve_runtime_path(os.getenv("API_LOG_PATH"), "API.log")
DRYRUN_TRACE_PATH = _resolve_runtime_path(os.getenv("DRYRUN_TRACE_PATH"), "dryrun_api_calls.txt")
_SQLITE_JOURNAL_MODES = {"DELETE", "WAL", "TRUNCATE", "PERSIST", "MEMORY", "OFF"}
SQLITE_JOURNAL_MODE = (os.getenv("SQLITE_JOURNAL_MODE", "DELETE") or "DELETE").strip().upper()
if SQLITE_JOURNAL_MODE not in _SQLITE_JOURNAL_MODES:
    SQLITE_JOURNAL_MODE = "DELETE"

# Guard against legacy root-level DB path leaking in from older shells.
if DB_PATH.resolve() == (PROJECT_ROOT / "cache.db").resolve():
    DB_PATH = RUNTIME_ROOT / "cache.db"

IONAPI_DIR = CREDENTIALS_ROOT / "ionapi"
TST_ENV_DIR = CREDENTIALS_ROOT / "TSTEnv"
TST_COMPASS_IONAPI = TST_ENV_DIR / "Infor Compass JDBC Driver.ionapi"
TST_COMPASS_JDBC = TST_ENV_DIR / "infor-compass-jdbc-2020-09.jar"
DEFAULT_TABLE = "wagons"
SPAREPARTS_TABLE = "spareparts"
SPAREPARTS_SWAP_TABLE = "sparepart_swaps"
WAGENUMBAU_TABLE = "Wagenumbau_Wagons"
RENUMBER_WAGON_TABLE = "RENUMBER_WAGON"
RENUMBER_EXTRA_COLUMNS = [
    "SEQ",
    "WAGEN_ITNO",
    "WAGEN_SERN",
    "NEW_SERN",
    "NEW_BAUREIHE",
    "NEW_PART_ITNO",
    "NEW_PART_SER2",
    "UMBAU_DATUM",
    "UMBAU_ART",
    "PLPN",
    "MWNO",
    "MOS100_STATUS",
    "MOS180_STATUS",
    "MOS050_STATUS",
    "CRS335_STATUS",
    "STS046_STATUS",
    "STS046_ADD_STATUS",
    "MMS240_STATUS",
    "CUSEXT_STATUS",
    "OUT",
    "UPDATED_AT",
    "IN",
    "TIMESTAMP_IN",
    "ROLLBACK",
    "TIMESTAMP_ROLLBACK",
]
RSRD_ERP_TABLE = "RSRD_ERP_WAGONNO"
RSRD_ERP_FULL_TABLE = "RSRD_ERP_DATA"
RSRD_UPLOAD_TABLE = "RSRD_WAGON_UPLOAD"
RSRD_SYNC_TABLE = "RSRD_SYNC_WAGONS"
RSRD_SYNC_SELECTION_TABLE = "RSRD_SYNC_SELECTIONS"
RSRD_DOCUMENTS_TABLE = "RSRD_DOCUMENTS"
RSRD_DOCUMENT_ASSIGNMENTS_TABLE = "RSRD_DOCUMENT_ASSIGNMENTS"
TEILENUMMER_TABLE = "TEILENUMMER"
WAGENSUCHE_TABLE = "WAGENSUCHE"
TEILENUMMER_TAUSCH_TABLE = "TEILENUMMER_TAUSCH"
CACHE_STATUS_TABLE = "CACHE_TABLE_STATUS"
DATALAKE_CACHE_SNAPSHOT_TABLE = "DATALAKE_SYNC_SNAPSHOT"
DATALAKE_CACHE_TABLES_TABLE = "DATALAKE_SYNC_TABLES"
DATALAKE_SYNC_SELECTION_TABLE = "DATALAKE_SYNC_SELECTIONS"
TEILENUMMER_TAUSCH_EXTRA_COLUMNS = [
    "NITNO",
    "NSERN",
    "OUT_DELAY_MIN",
    "IN_DELAY_MIN",
    "PLAN_OUT_DATE",
    "PLAN_OUT_TIME",
    "PLAN_IN_DATE",
    "PLAN_IN_TIME",
    "UMGEBAUT",
    "TIMESTAMP",
    "OUT_STATUS",
    "MOS170_STATUS",
    "PLPN",
    "CMS100_STATUS",
    "MWNO",
    "MOS100_STATUS",
    "MOS180_STATUS",
    "MOS050_STATUS",
    "IN_STATUS",
]
DB_PATH.parent.mkdir(parents=True, exist_ok=True)
DB_PATH.touch(exist_ok=True)
API_LOG_PATH.parent.mkdir(parents=True, exist_ok=True)
DRYRUN_TRACE_PATH.parent.mkdir(parents=True, exist_ok=True)


def _init_sqlite_runtime() -> None:
    try:
        with sqlite3.connect(str(DB_PATH), timeout=10) as conn:
            conn.execute(f"PRAGMA journal_mode = {SQLITE_JOURNAL_MODE}")
            conn.execute("PRAGMA synchronous = NORMAL")
            conn.execute("PRAGMA busy_timeout = 30000")
    except sqlite3.OperationalError:
        pass


_init_sqlite_runtime()


def _ensure_sqlite_wal_ready(max_attempts: int = 20) -> str:
    desired_mode = SQLITE_JOURNAL_MODE.lower()
    last_error: Exception | None = None
    for attempt in range(max_attempts):
        try:
            with sqlite3.connect(str(DB_PATH), timeout=1, isolation_level=None) as conn:
                conn.execute("PRAGMA busy_timeout = 1000")
                current_row = conn.execute("PRAGMA journal_mode").fetchone()
                current_mode = str(current_row[0] if current_row else "").lower()
                if current_mode == desired_mode:
                    return desired_mode
                row = conn.execute(f"PRAGMA journal_mode = {SQLITE_JOURNAL_MODE}").fetchone()
                mode = str(row[0] if row else "").lower()
                return mode or (current_mode or "unknown")
        except sqlite3.OperationalError as exc:
            last_error = exc
            if not _is_sqlite_locked_error(exc) and "busy" not in str(exc).lower():
                raise
            time.sleep(0.2 + (0.1 * attempt))
    if last_error is not None:
        return f"locked ({last_error})"
    return "locked"

DEFAULT_SCHEME = os.getenv("SPAREPART_SCHEME", "datalake")
DEFAULT_CATALOG = os.getenv("SPAREPART_CATALOG")
DEFAULT_COLLECTION = os.getenv("SPAREPART_DEFAULT_COLLECTION")
SQL_FILE = PROJECT_ROOT / "sql" / "wagons_base.sql"
WAGONS_SQL_FILES = {
    "prd": PROJECT_ROOT / "sql" / "wagons_base_prd.sql",
    "tst": PROJECT_ROOT / "sql" / "wagons_base_tst.sql",
}
WAGENSUCHE_SQL_FILES = {
    "prd": PROJECT_ROOT / "sql" / "wagensuche_prd.sql",
    "tst": PROJECT_ROOT / "sql" / "wagensuche_tst.sql",
}
SPAREPARTS_SQL_FILE = PROJECT_ROOT / "sql" / "spareparts_base.sql"
RSRD_ERP_SQL_FILE = PROJECT_ROOT / "sql" / "rsrd_erp_full.sql"
TEILENUMMER_SQL_FILE = PROJECT_ROOT / "sql" / "teilenummer_base.sql"
TEILENUMMER_SQL_FILES = {
    "prd": PROJECT_ROOT / "sql" / "teilenummer_base_prd.sql",
    "tst": PROJECT_ROOT / "sql" / "teilenummer_base_tst.sql",
}
DEFAULT_ENV = os.getenv("SPAREPART_ENV", "prd").lower()
WAGENSUCHE_PG_URL = os.getenv("WAGENSUCHE_PG_URL", "").strip()
WAGENSUCHE_PG_HOST = os.getenv("WAGENSUCHE_PG_HOST", "").strip()
WAGENSUCHE_PG_PORT = os.getenv("WAGENSUCHE_PG_PORT", "").strip()
WAGENSUCHE_PG_DB = os.getenv("WAGENSUCHE_PG_DB", "").strip()
WAGENSUCHE_PG_USER = os.getenv("WAGENSUCHE_PG_USER", "").strip()
WAGENSUCHE_PG_PASS = os.getenv("WAGENSUCHE_PG_PASS", "").strip()
ENV_ALIASES = {
    "live": "prd",
    "prod": "prd",
    "prd": "prd",
    "test": "tst",
    "tst": "tst",
}
ENV_SUFFIXES = {"prd": "_PRD", "tst": "_TST"}
ENV_IONAPI = {
    "prd": {
        "compass": IONAPI_DIR / "Infor Compass JDBC Driver.ionapi",
        "mi": IONAPI_DIR / "MFD_Backend_Python_vNEW.ionapi"
        if (IONAPI_DIR / "MFD_Backend_Python_vNEW.ionapi").exists()
        else IONAPI_DIR / "MFD_Backend_Python.ionapi",
    },
    "tst": {
        "compass": TST_COMPASS_IONAPI if TST_COMPASS_IONAPI.exists() else IONAPI_DIR / "Infor Compass JDBC Driver_TST.ionapi",
        "mi": IONAPI_DIR / "TST_MFD_Backend_Python_new.ionapi",
    },
}
MOS125_DRY_RUN = os.getenv("SPAREPART_MOS125_DRY_RUN", "1").strip().lower() in {"1", "true", "yes", "y"}
MI_CONO = os.getenv("SPAREPART_M3_CONO", "").strip()
MI_CONO_PRD = os.getenv("SPAREPART_M3_CONO_PRD", "").strip()
MI_CONO_TST = os.getenv("SPAREPART_M3_CONO_TST", "881").strip()
CMS100_RETRY_DELAY_SEC = float(os.getenv("SPAREPART_CMS100_RETRY_DELAY", "3").strip() or "3")
CMS100_RETRY_MAX = int(os.getenv("SPAREPART_CMS100_RETRY_MAX", "20").strip() or "20")
WAGON_CMS100_RETRY_DELAY_SEC = float(os.getenv("SPAREPART_WAGON_CMS100_RETRY_DELAY", "5").strip() or "5")
WAGON_CMS100_RETRY_MAX = int(os.getenv("SPAREPART_WAGON_CMS100_RETRY_MAX", "8").strip() or "8")
MOS170_RETRY_DELAY_SEC = float(os.getenv("SPAREPART_MOS170_RETRY_DELAY", "3").strip() or "3")
MOS170_RETRY_MAX = int(os.getenv("SPAREPART_MOS170_RETRY_MAX", "5").strip() or "5")
MOS100_RETRY_DELAY_SEC = float(os.getenv("SPAREPART_MOS100_RETRY_DELAY", "3").strip() or "3")
MOS100_RETRY_MAX = int(os.getenv("SPAREPART_MOS100_RETRY_MAX", "10").strip() or "10")
MOS180_RETRY_DELAY_SEC = float(os.getenv("SPAREPART_MOS180_RETRY_DELAY", "3").strip() or "3")
MOS180_RETRY_MAX = int(os.getenv("SPAREPART_MOS180_RETRY_MAX", "10").strip() or "10")
MOS050_RETRY_DELAY_SEC = float(os.getenv("SPAREPART_MOS050_RETRY_DELAY", "3").strip() or "3")
MOS050_RETRY_MAX = int(os.getenv("SPAREPART_MOS050_RETRY_MAX", "10").strip() or "10")
TEILENUMMER_PHASE_SPLIT_WAIT_SEC = 5
TEILENUMMER_CMS100_RETRY_DELAY_SEC = 3.0
TEILENUMMER_CMS100_RETRY_MAX = 10
WAGON_MOS100_RETRY_MAX = int(os.getenv("SPAREPART_WAGON_MOS100_RETRY_MAX", "8").strip() or "8")
WAGON_RENUMBER_SKIP_MOS170 = os.getenv("SPAREPART_WAGON_RENUMBER_SKIP_MOS170", "").strip().lower() in {"1", "true", "yes", "y"}
WAGON_RENUMBER_FIXED_PLPN = os.getenv("SPAREPART_WAGON_RENUMBER_FIXED_PLPN", "").strip()
API_LOG_ONLY = [
    value.strip()
    for value in os.getenv("SPAREPART_API_LOG_ONLY", "").split(",")
    if value.strip()
]
IPS_COMPANY = os.getenv("SPAREPART_IPS_COMPANY", "").strip()
IPS_DIVISION = os.getenv("SPAREPART_IPS_DIVISION", "").strip()
IPS_COMPANY_TST = os.getenv("SPAREPART_IPS_COMPANY_TST", "").strip()
IPS_DIVISION_TST = os.getenv("SPAREPART_IPS_DIVISION_TST", "").strip()
MOS180_FACI = os.getenv("SPAREPART_MOS180_FACI", "100").strip()
MOS180_RESP = os.getenv("SPAREPART_MOS180_RESP", "CHRUPP").strip()
MOS180_APRB = os.getenv("SPAREPART_MOS180_APRB", "CHRUPP").strip()
MOS050_LOCATION = os.getenv("SPAREPART_MOS050_LOCATION", "EINBAU").strip()
MOS050_SERVICE = os.getenv("SPAREPART_MOS050_SERVICE", "MOS050_MONTAGE").strip()
MOS050_OPERATION = os.getenv("SPAREPART_MOS050_OPERATION", "MOS050").strip()
MOS050_NAMESPACE = os.getenv("SPAREPART_MOS050_NAMESPACE", "").strip()
MOS050_BODY_TAG = os.getenv("SPAREPART_MOS050_BODY_TAG", "MOS050").strip()
CRS335_ACRF = os.getenv("SPAREPART_CRS335_ACRF", "").strip()

JOB_LOG_LIMIT = 2000
PROGRESS_LINE = re.compile(r"^\d+/\d+\s+Datensätze gespeichert \.\.\.$")
TEILENUMMER_RELOAD_TIMEOUT_SEC = max(
    60,
    int(os.getenv("TEILENUMMER_RELOAD_TIMEOUT_SEC", "600") or 600),
)
COMPASS_PRECHECK_TIMEOUT_SEC = max(
    5,
    int(os.getenv("COMPASS_PRECHECK_TIMEOUT_SEC", "20") or 20),
)
JOB_STALE_MARGIN_SEC = max(
    30,
    int(os.getenv("MFDAPPS_JOB_STALE_MARGIN_SEC", "120") or 120),
)
_jobs_lock = threading.Lock()
_jobs: Dict[str, Dict[str, Any]] = {}
_dryrun_trace_lock = threading.Lock()
_dryrun_trace_seq = 0
_reload_merge_state_lock = threading.Lock()
_reload_merge_thread_id: int | None = None
DATALAKE_SAFE_IDENTIFIER = re.compile(r"^[A-Za-z0-9_]+$")
_DATALAKE_TABLE_NAME_PATTERN = re.compile(r"^[A-Za-z]{6}$")
_DATALAKE_TIMESTAMP_PREFERRED_COLUMNS = (
    "updated_at",
    "updatedat",
    "last_updated_at",
    "lastupdatedat",
    "last_update",
    "lastupdate",
    "modified_at",
    "modifiedat",
    "last_modified_at",
    "lastmodifiedat",
    "timestamp",
    "event_time",
    "created_at",
)
_DATALAKE_TIMESTAMP_TYPE_HINTS = ("timestamp", "datetime", "date", "time")
_DATALAKE_NIGHTLY_PREPARED_TIME = "00:30"
_DATALAKE_TIMEOUT_DEFAULT_SECONDS = 60
_DATALAKE_TIMEOUT_RETRY_SECONDS = 300
_DATALAKE_TIMEOUT_SUPERHEAVY_SECONDS = 600
_DATALAKE_TIMEOUT_TIERS = ("normal", "timeout", "superheavy")
_datalake_tables_lock = threading.Lock()


def _new_datalake_state() -> Dict[str, Any]:
    return {
        "running": False,
        "current_table": None,
        "phase": None,
        "phase_detail": None,
        "started_at_utc": None,
        "finished_at_utc": None,
        "last_error": None,
        "total_tables": 0,
        "completed_tables": 0,
        "error_tables": 0,
        "tables": {},
    }


_datalake_tables_state: Dict[str, Dict[str, Any]] = {
    "prd": _new_datalake_state(),
    "tst": _new_datalake_state(),
}

app = FastAPI(title="SPAREPART Loader API")

app.add_middleware(
    CORSMiddleware,
    allow_origins=["*"],
    allow_methods=["*"],
    allow_headers=["*"],
)


@app.exception_handler(Exception)
async def _unhandled_exception_handler(request: Request, exc: Exception) -> JSONResponse:
    logging.getLogger("web-server").exception("Unhandled exception on %s", request.url.path)
    return JSONResponse(status_code=500, content={"detail": f"Unhandled backend error: {exc}"})


def _try_include_excel_import_router() -> None:
    excel_src = PROJECT_ROOT / "apps" / "christian" / "AppExcelImport" / "src"
    if not excel_src.exists():
        return
    excel_src_text = str(excel_src)
    if excel_src_text not in sys.path:
        sys.path.insert(0, excel_src_text)
    try:
        from app_excel_import.main import router as excel_import_router  # type: ignore
    except Exception as exc:  # pragma: no cover - optional integration
        logging.getLogger("excel-import").warning(
            "AppExcelImport Router konnte nicht eingebunden werden: %s", exc
        )
        return
    app.include_router(excel_import_router)


_try_include_excel_import_router()

BASIC_AUTH_USER = os.getenv("BASIC_AUTH_USER", "").strip()
BASIC_AUTH_PASS = os.getenv("BASIC_AUTH_PASS", "").strip()
BASIC_AUTH_ENABLED = bool(BASIC_AUTH_USER and BASIC_AUTH_PASS)
M3BRIDGE_API_KEY = os.getenv("M3BRIDGE_API_KEY", "").strip()

GOLDENVIEW_QUERIES_TABLE = "GOLDENVIEW_QUERIES"
GOLDENVIEW_FIELDS_TABLE = "GOLDENVIEW_FIELDS"
GOLDENVIEW_EXPORT_DIR = _resolve_runtime_path(
    os.getenv("GOLDENVIEW_REPO_PATH"),
    "goldenview_exports",
)
GITHUB_SYNC_REPO = os.getenv("GITHUB_SYNC_REPO", "crupp-mfd/M3ChatbotExcels").strip()
GITHUB_SYNC_TOKEN = os.getenv("GITHUB_SYNC_TOKEN", "").strip()
GITHUB_SYNC_WORKFLOW = os.getenv("GITHUB_SYNC_WORKFLOW", "sync-knowledge.yml").strip()
OPENAI_API_KEY = os.getenv("OPENAI_API_KEY", "").strip()
OPENAI_VECTOR_STORE_ID = os.getenv("OPENAI_VECTOR_STORE_ID", "").strip()
GPT_ACTION_API_KEY = os.getenv("GPT_ACTION_API_KEY", "").strip()


def _basic_auth_valid(auth_header: str) -> bool:
    if not BASIC_AUTH_ENABLED:
        return True
    return is_basic_auth_valid(auth_header, BASIC_AUTH_USER, BASIC_AUTH_PASS)


class AuthStaticFiles(StaticFiles):
    async def __call__(self, scope, receive, send) -> None:  # type: ignore[override]
        if scope.get("type") != "http":
            return await super().__call__(scope, receive, send)
        headers = {k.decode().lower(): v.decode() for k, v in scope.get("headers", [])}
        auth_header = headers.get("authorization", "")
        if not _basic_auth_valid(auth_header):
            resp = Response(status_code=401, headers={"WWW-Authenticate": "Basic"})
            await resp(scope, receive, send)
            return
        await super().__call__(scope, receive, send)


@app.middleware("http")
async def _auth_middleware(request: Request, call_next):
    path = request.url.path
    if path == "/":
        portal_file = PROJECT_ROOT / "index.html"
        if portal_file.exists():
            return FileResponse(
                portal_file,
                media_type="text/html",
                headers={
                    "Cache-Control": "no-store, no-cache, must-revalidate, max-age=0",
                    "Pragma": "no-cache",
                },
            )
    raw_path = ""
    try:
        raw_path = request.scope.get("raw_path", b"").decode("utf-8", "ignore")
    except Exception:
        raw_path = ""
    if "ask_m3_knowledge" in path or "ask_m3_knowledge" in raw_path:
        _auth_logger.info(
            "auth_debug path=%s raw_path=%s has_api_key=%s method=%s",
            path,
            raw_path,
            bool(request.headers.get("x-api-key")),
            request.method,
        )
    if path.startswith("/query") or "/api/ask_m3_knowledge" in path or "ask_m3_knowledge" in raw_path:
        api_key = request.headers.get("x-api-key", "").strip()
        expected = M3BRIDGE_API_KEY if path.startswith("/query") else GPT_ACTION_API_KEY
        if "ask_m3_knowledge" in path or "ask_m3_knowledge" in raw_path:
            _auth_logger.info(
                "auth_debug expected_set=%s expected_len=%s",
                bool(expected),
                len(expected) if expected else 0,
            )
            import hashlib

            def _h(val: str) -> str:
                return hashlib.sha256(val.encode("utf-8")).hexdigest()[:8] if val else "empty"

            _auth_logger.info(
                "auth_debug api_len=%s api_hash=%s expected_hash=%s",
                len(api_key),
                _h(api_key),
                _h(expected),
            )
        auth_header = request.headers.get("authorization", "")
        if expected and api_key == expected:
            return await call_next(request)
        if _basic_auth_valid(auth_header):
            return await call_next(request)
        if expected:
            return JSONResponse({"detail": "Unauthorized"}, status_code=401)
        return Response(status_code=401, headers={"WWW-Authenticate": "Basic"})
    if request.method == "OPTIONS":
        return await call_next(request)
    auth_header = request.headers.get("authorization", "")
    if not _basic_auth_valid(auth_header):
        return Response(status_code=401, headers={"WWW-Authenticate": "Basic"})
    return await call_next(request)


@app.on_event("startup")
def _prepare_env_tables() -> None:
    if not DB_PATH.exists():
        return
    with _connect() as conn:
        _ensure_env_tables(conn)
        _ensure_datalake_cache_tables(conn)
        _init_goldenview_db(conn)
        conn.commit()


def _init_goldenview_db(conn: sqlite3.Connection) -> None:
    conn.execute(
        f"""
        CREATE TABLE IF NOT EXISTS {GOLDENVIEW_QUERIES_TABLE} (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            name TEXT,
            sql_text TEXT NOT NULL,
            description TEXT,
            excel_path TEXT,
            md_path TEXT,
            generated_at TEXT,
            commit_at TEXT,
            created_at TEXT DEFAULT (datetime('now'))
        )
        """
    )
    conn.execute(
        f"""
        CREATE TABLE IF NOT EXISTS {GOLDENVIEW_FIELDS_TABLE} (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            query_id INTEGER NOT NULL,
            field_name TEXT NOT NULL,
            field_description TEXT,
            connected_fields TEXT,
            FOREIGN KEY(query_id) REFERENCES {GOLDENVIEW_QUERIES_TABLE}(id) ON DELETE CASCADE
        )
        """
    )

    columns = {row["name"] for row in conn.execute(f"PRAGMA table_info({GOLDENVIEW_QUERIES_TABLE})").fetchall()}
    for column in ("excel_path", "md_path", "generated_at", "commit_at"):
        if column not in columns:
            conn.execute(f"ALTER TABLE {GOLDENVIEW_QUERIES_TABLE} ADD COLUMN {column} TEXT")


@contextmanager
def _connect(timeout: float = 30.0, busy_timeout_ms: int = 30000):
    current_thread_id = threading.get_ident()
    for _ in range(120):
        with _reload_merge_state_lock:
            merge_thread = _reload_merge_thread_id
        if merge_thread is None or merge_thread == current_thread_id:
            break
        time.sleep(0.05)
    else:
        raise sqlite3.OperationalError("database is locked")

    db_path = Path(DB_PATH)
    db_path.parent.mkdir(parents=True, exist_ok=True)
    db_path.touch(exist_ok=True)
    last_error: sqlite3.OperationalError | None = None
    conn: sqlite3.Connection | None = None
    for attempt in range(6):
        try:
            conn = sqlite3.connect(str(db_path), timeout=timeout, isolation_level=None)
            conn.row_factory = sqlite3.Row
            conn.execute(f"PRAGMA busy_timeout = {int(busy_timeout_ms)}")
            break
        except sqlite3.OperationalError as exc:
            last_error = exc
            if not _is_sqlite_locked_error(exc) or attempt == 5:
                raise
            time.sleep(0.05 * (attempt + 1))
    if conn is None and last_error is not None:
        raise last_error
    if conn is None:
        raise sqlite3.OperationalError(f"SQLite Verbindung fehlgeschlagen: {db_path}")
    try:
        yield conn
    finally:
        try:
            conn.close()
        except Exception:
            pass


def _ensure_swap_table(conn: sqlite3.Connection, table_name: str) -> None:
    conn.execute(
        f"""
        CREATE TABLE IF NOT EXISTS {table_name} (
            WAGEN_ITNO TEXT NOT NULL,
            WAGEN_SERN TEXT NOT NULL,
            ORIGINAL_ITNO TEXT NOT NULL,
            ORIGINAL_SERN TEXT NOT NULL,
            ERSATZ_ITNO TEXT NOT NULL,
            ERSATZ_SERN TEXT NOT NULL,
            USER TEXT,
            UPLOAD TEXT DEFAULT 'N',
            TIMESTAMP TEXT,
            PRIMARY KEY (WAGEN_ITNO, WAGEN_SERN, ORIGINAL_ITNO, ORIGINAL_SERN)
        )
        """
    )


def _validate_table(table: str) -> str:
    if not table.replace("_", "").isalnum():
        raise HTTPException(status_code=400, detail="Ungültiger Tabellenname.")
    return table


def _normalize_env(env: str | None) -> str:
    value = (env or DEFAULT_ENV).lower()
    normalized = ENV_ALIASES.get(value)
    if not normalized:
        raise HTTPException(status_code=400, detail="Ungültige Umgebung.")
    return normalized


def _normalize_rsrd_env(rsrd_env: str | None, env: str | None) -> str:
    return _normalize_env(rsrd_env or env)


def _effective_dry_run(env: str | None) -> bool:
    def _override(name: str) -> bool | None:
        raw = os.getenv(name)
        if raw is None or raw.strip() == "":
            return None
        return raw.strip().lower() in {"1", "true", "yes", "y"}

    normalized = _normalize_env(env)
    if normalized == "prd":
        override = _override("SPAREPART_PRD_DRY_RUN")
        if override is not None:
            return override
        return MOS125_DRY_RUN
    if normalized == "tst":
        override = _override("SPAREPART_TST_DRY_RUN")
        if override is not None:
            return override
        return MOS125_DRY_RUN
    return MOS125_DRY_RUN


def _table_for(base: str, env: str | None) -> str:
    normalized = _normalize_env(env)
    return f"{base}{ENV_SUFFIXES[normalized]}"


def _ionapi_path(env: str, kind: str) -> Path:
    normalized = _normalize_env(env)
    env_config = ENV_IONAPI.get(normalized)
    if not env_config or kind not in env_config:
        raise HTTPException(status_code=400, detail=f"Ionapi-Konfiguration fehlt für {normalized}/{kind}")
    path = env_config[kind]
    if not path.exists():
        raise HTTPException(status_code=500, detail=f"Ionapi-Datei nicht gefunden: {path}")
    return path


def _sanitize_url(value: str) -> str:
    if not value:
        return ""
    try:
        parts = urlsplit(value)
    except ValueError:
        return value
    hostname = parts.hostname or ""
    netloc = hostname
    if parts.port:
        netloc = f"{hostname}:{parts.port}"
    return urlunsplit((parts.scheme, netloc, parts.path, parts.query, parts.fragment))


def _safe_ionapi_url(env: str, kind: str) -> str:
    try:
        ionapi = _ionapi_path(env, kind)
    except HTTPException:
        return ""
    try:
        data = json.loads(ionapi.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError):
        return ""
    if isinstance(data, dict):
        return _sanitize_url(str(data.get("iu") or ""))
    return ""


def _resolve_rsrd_wsdl(env: str | None) -> str:
    normalized = _normalize_env(env)
    suffix = "PRD" if normalized == "prd" else "TST"
    value = os.getenv(f"RSRD_WSDL_URL_{suffix}") or os.getenv("RSRD_WSDL_URL") or ""
    return _sanitize_url(value)


def _resolve_rsrd_upload_url(env: str | None) -> str:
    wsdl = _resolve_rsrd_wsdl(env)
    if not wsdl:
        return ""
    if wsdl.lower().endswith("?wsdl"):
        return wsdl[: -len("?wsdl")]
    return wsdl


def _rsrd_upload_credentials(env: str | None) -> tuple[str, str]:
    user = rsrd_resolve_env_value("RSRD_SOAP_USER", env)
    password = rsrd_resolve_env_value("RSRD_SOAP_PASS", env)
    return user, password


def _rsrd_xml_value(value: Any) -> str:
    if isinstance(value, bool):
        return "true" if value else "false"
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, float):
        if value.is_integer():
            return str(int(value))
        return str(value)
    return str(value)


def _rsrd_payload_to_xml(tag: str, value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, list):
        return "".join(_rsrd_payload_to_xml(tag, item) for item in value if item is not None)
    if isinstance(value, dict):
        order_map = {
            "DesignDataSet": [
                "LetterMarking",
                "CombinedTransportWagonType",
                "TankCode",
                "WagonNumberOfAxles",
                "WheelSetType",
                "WheelDiameter",
                "WheelsetGauge",
                "WheelSetTransformationMethod",
                "NumberOfBogies",
                "BogiePitch",
                "BogiePivotPitch",
                "InnerWheelbase",
                "CouplingType",
                "BufferType",
                "NormalLoadingGauge",
                "MinCurveRadius",
                "MinVerticalRadiusYardHump",
                "WagonWeightEmpty",
                "LengthOverBuffers",
                "MaxAxleWeight",
                "LoadTable",
                "MaxDesignSpeed",
                "AirBrake",
                "HandBrake",
                "DerailmentDetectionDevice",
                "BrakeBlock",
                "MaxLengthOfLoad",
                "LoadArea",
                "HeightOfLoadingPlaneUnladen",
                "RemovableAccessories",
                "LoadingCapacity",
                "MaxGrossWeight",
                "VapourReturnSystem",
                "FerryPermittedFlag",
                "FerryRampAngle",
                "TemperatureRange",
                "TechnicalForwardingRestrictions",
                "MaintenancePlanRef",
                "DateLastOverhaul",
                "PlannedDateNextOverhaul",
                "OverhaulValidityPeriod",
                "PermittedTolerance",
                "DateOfNextTankInspection",
            ],
            "HandBrake": ["HandBrakeType", "HandBrakedWeight", "ParkingBrakeForce"],
            "AirBrake": [
                "NumberOfBrakes",
                "BrakeSystem",
                "AirBrakeType",
                "BrakingPowerVariationDevice",
                "AirBrakedMass",
                "LoadChangeDevice",
                "BrakeSpecialCharacteristics",
            ],
            "BrakeBlock": [
                "BrakeBlockName",
                "CompositeBrakeBlockRetrofitted",
                "CompositeBrakeBlockInstallationDate",
            ],
            "LoadTable": [
                "LoadTableProduct",
                "LoadTableCountry",
                "SpeedCategory",
                "LoadTableStars",
                "RouteClassPayloads",
            ],
        }
        order = order_map.get(tag)
        if order:
            seen = set()
            parts = []
            for key in order:
                if key in value and value[key] is not None:
                    parts.append(_rsrd_payload_to_xml(key, value[key]))
                    seen.add(key)
            for key, inner_value in value.items():
                if key in seen or inner_value is None:
                    continue
                parts.append(_rsrd_payload_to_xml(key, inner_value))
            inner = "".join(parts)
        else:
            inner = "".join(
                _rsrd_payload_to_xml(key, inner_value)
                for key, inner_value in value.items()
                if inner_value is not None
            )
        if not inner:
            return ""
        return f"<xsd:{tag}>{inner}</xsd:{tag}>"
    safe = xml_escape(_rsrd_xml_value(value))
    return f"<xsd:{tag}>{safe}</xsd:{tag}>"


def _rsrd_build_upload_xml(payload: Dict[str, Any]) -> str:
    dataset = "".join(
        _rsrd_payload_to_xml(key, value) for key, value in payload.items() if value is not None
    )
    return (
        "<?xml version=\"1.0\" encoding=\"UTF-8\" standalone=\"yes\"?>"
        "<soapenv:Envelope xmlns:soapenv=\"http://schemas.xmlsoap.org/soap/envelope/\" "
        "xmlns:xsd=\"http://www.rsrd.com/xsd\">"
        "<soapenv:Header/>"
        "<soapenv:Body>"
        "<xsd:UploadWagonDataRequest>"
        "<xsd:RollingStockDataset>"
        f"{dataset}"
        "</xsd:RollingStockDataset>"
        "</xsd:UploadWagonDataRequest>"
        "</soapenv:Body>"
        "</soapenv:Envelope>"
    )


def _rsrd_tables(env: str | None) -> RSRDTables:
    return rsrd_tables_for_env(_normalize_env(env))


def _ensure_rsrd_tables(conn: sqlite3.Connection, env: str | None) -> RSRDTables:
    tables = _rsrd_tables(env)
    rsrd_init_db(conn, tables=tables)
    return tables


def _ensure_rsrd_upload_table(conn: sqlite3.Connection, env: str | None) -> str:
    table_name = _table_for(RSRD_UPLOAD_TABLE, env)
    conn.execute(
        f"""
        CREATE TABLE IF NOT EXISTS {table_name} (
            wagon_number_freight TEXT PRIMARY KEY,
            rsrd_wagon_id TEXT,
            payload_json TEXT NOT NULL,
            diff_json TEXT NOT NULL,
            rsrd_json TEXT,
            created_at TEXT NOT NULL,
            updated_at TEXT NOT NULL
        )
        """
    )
    return table_name


def _ensure_rsrd_sync_table(conn: sqlite3.Connection, env: str | None) -> str:
    table_name = _table_for(RSRD_SYNC_TABLE, env)
    conn.execute(
        f"""
        CREATE TABLE IF NOT EXISTS {table_name} (
            wagon_number_freight TEXT PRIMARY KEY,
            enabled TEXT,
            sync_data_env TEXT,
            sync_km_env TEXT,
            sync_docs_env TEXT,
            updated_at TEXT
        )
        """
    )
    existing = {row[1].lower() for row in conn.execute(f"PRAGMA table_info({table_name})")}
    for column in ("sync_data_env", "sync_km_env", "sync_docs_env"):
        if column in existing:
            continue
        conn.execute(f'ALTER TABLE {table_name} ADD COLUMN "{column}" TEXT')
    return table_name


def _ensure_rsrd_sync_selection_table(conn: sqlite3.Connection, env: str | None) -> str:
    table_name = _table_for(RSRD_SYNC_SELECTION_TABLE, env)
    conn.execute(
        f"""
        CREATE TABLE IF NOT EXISTS {table_name} (
            wagon_number_freight TEXT PRIMARY KEY,
            sync_data_env TEXT,
            sync_km_env TEXT,
            sync_docs_env TEXT,
            one_time_transfer TEXT,
            updated_at TEXT
        )
        """
    )
    existing = {row[1].lower() for row in conn.execute(f"PRAGMA table_info({table_name})")}
    for column in ("sync_data_env", "sync_km_env", "sync_docs_env", "one_time_transfer"):
        if column in existing:
            continue
        conn.execute(f'ALTER TABLE {table_name} ADD COLUMN "{column}" TEXT')
    return table_name


def _rsrd_documents_dir() -> Path:
    directory = RUNTIME_ROOT / "rsrd_documents"
    directory.mkdir(parents=True, exist_ok=True)
    return directory


def _ensure_rsrd_document_tables(conn: sqlite3.Connection) -> None:
    conn.execute(
        f"""
        CREATE TABLE IF NOT EXISTS {RSRD_DOCUMENTS_TABLE} (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            display_name TEXT NOT NULL,
            original_name TEXT NOT NULL,
            stored_name TEXT NOT NULL UNIQUE,
            mime_type TEXT,
            size_bytes INTEGER,
            uploaded_at TEXT NOT NULL,
            updated_at TEXT NOT NULL
        )
        """
    )
    conn.execute(
        f"""
        CREATE TABLE IF NOT EXISTS {RSRD_DOCUMENT_ASSIGNMENTS_TABLE} (
            id INTEGER PRIMARY KEY AUTOINCREMENT,
            document_id INTEGER NOT NULL,
            assign_type TEXT NOT NULL,
            assign_value TEXT NOT NULL,
            created_at TEXT NOT NULL,
            UNIQUE(document_id, assign_type, assign_value)
        )
        """
    )
    conn.execute(
        f"""
        CREATE INDEX IF NOT EXISTS idx_rsrd_doc_assign_lookup
        ON {RSRD_DOCUMENT_ASSIGNMENTS_TABLE}(assign_type, assign_value)
        """
    )
    conn.execute(
        f"""
        CREATE INDEX IF NOT EXISTS idx_rsrd_doc_assign_doc
        ON {RSRD_DOCUMENT_ASSIGNMENTS_TABLE}(document_id)
        """
    )


def _normalize_doc_assign_values(raw_values: Any) -> List[str]:
    values = raw_values if isinstance(raw_values, list) else []
    seen = set()
    cleaned: List[str] = []
    for item in values:
        value = str(item or "").strip()
        if not value:
            continue
        key = value.upper()
        if key in seen:
            continue
        seen.add(key)
        cleaned.append(value)
    return cleaned


def _load_matching_documents(
    conn: sqlite3.Connection,
    rows: List[sqlite3.Row],
) -> Dict[str, List[str]]:
    _ensure_rsrd_document_tables(conn)
    baureihen = {
        str(row["baureihe"]).strip().upper()
        for row in rows
        if row["baureihe"] is not None and str(row["baureihe"]).strip()
    }
    wagen_typen = {
        str(row["wagen_typ"]).strip().upper()
        for row in rows
        if row["wagen_typ"] is not None and str(row["wagen_typ"]).strip()
    }
    where: List[str] = []
    params: List[Any] = []
    if baureihen:
        where.append(
            "(a.assign_type = 'baureihe' AND UPPER(TRIM(a.assign_value)) IN ("
            + ", ".join("?" for _ in baureihen)
            + "))"
        )
        params.extend(sorted(baureihen))
    if wagen_typen:
        where.append(
            "(a.assign_type = 'wagen_typ' AND UPPER(TRIM(a.assign_value)) IN ("
            + ", ".join("?" for _ in wagen_typen)
            + "))"
        )
        params.extend(sorted(wagen_typen))
    if not where:
        return {}

    matched = conn.execute(
        f"""
        SELECT d.display_name, a.assign_type, UPPER(TRIM(a.assign_value)) AS assign_key
        FROM {RSRD_DOCUMENT_ASSIGNMENTS_TABLE} a
        JOIN {RSRD_DOCUMENTS_TABLE} d
          ON d.id = a.document_id
        WHERE {" OR ".join(where)}
        ORDER BY d.display_name
        """,
        params,
    ).fetchall()

    by_baureihe: Dict[str, set[str]] = {}
    by_wagen_typ: Dict[str, set[str]] = {}
    for row in matched:
        name = str(row["display_name"] or "").strip()
        key = str(row["assign_key"] or "").strip()
        if not name or not key:
            continue
        if row["assign_type"] == "baureihe":
            by_baureihe.setdefault(key, set()).add(name)
        elif row["assign_type"] == "wagen_typ":
            by_wagen_typ.setdefault(key, set()).add(name)

    result: Dict[str, List[str]] = {}
    for row in rows:
        wagon = str(row["wagon_number"] or "").strip()
        if not wagon:
            continue
        names: set[str] = set()
        baureihe_key = str(row["baureihe"] or "").strip().upper()
        wagen_typ_key = str(row["wagen_typ"] or "").strip().upper()
        if baureihe_key and baureihe_key in by_baureihe:
            names.update(by_baureihe[baureihe_key])
        if wagen_typ_key and wagen_typ_key in by_wagen_typ:
            names.update(by_wagen_typ[wagen_typ_key])
        result[wagon] = sorted(names)
    return result


def _fetch_erp_wagon_numbers(
    conn: sqlite3.Connection,
    env: str | None,
    limit: int | None = None,
) -> List[str]:
    table_name = _table_for(RSRD_ERP_TABLE, env)
    try:
        query = f"SELECT wagon_sern, wagon_sern_numeric FROM {table_name} ORDER BY wagon_sern"
        params: List[int] = []
        if limit:
            query += " LIMIT ?"
            params.append(limit)
        rows = conn.execute(query, params).fetchall()
    except sqlite3.OperationalError as exc:  # table missing
        raise HTTPException(status_code=404, detail=f"Tabelle {table_name} nicht gefunden.") from exc
    wagon_numbers = []
    for row in rows:
        sern_numeric = (row["wagon_sern_numeric"] or "").strip()
        sern = (row["wagon_sern"] or "").strip()
        number = sern_numeric or sern
        if number:
            wagon_numbers.append(number)
    return wagon_numbers


def _like_pattern(value: str) -> str:
    raw = value.strip()
    if not raw:
        return "%"
    escaped = raw.replace("\\", "\\\\").replace("%", "\\%").replace("_", "\\_")
    if "*" in escaped:
        return escaped.replace("*", "%")
    return f"%{escaped}%"


def _sern_filter_pattern(value: str) -> str:
    raw = value.strip()
    filtered = "".join(ch for ch in raw if ch.isdigit() or ch == "*")
    return _like_pattern(filtered or raw)


def _ensure_table(
    conn: sqlite3.Connection,
    table: str,
    template: str | None = None,
) -> str:
    table = _validate_table(table)
    cursor = conn.execute(
        "SELECT name FROM sqlite_master WHERE type='table' AND name = ?",
        (table,),
    )
    if cursor.fetchone() is None:
        if template:
            template = _validate_table(template)
            conn.execute(
                f'CREATE TABLE IF NOT EXISTS "{table}" AS SELECT * FROM "{template}" WHERE 0=1'
            )
        else:
            raise HTTPException(status_code=404, detail=f"Tabelle '{table}' nicht gefunden.")
    return table


def _table_exists(conn: sqlite3.Connection, table: str) -> bool:
    cursor = conn.execute(
        "SELECT name FROM sqlite_master WHERE type='table' AND name = ?",
        (table,),
    )
    return cursor.fetchone() is not None


def _ensure_columns(conn: sqlite3.Connection, table: str, columns: List[str]) -> None:
    existing = {
        row[1]
        for row in conn.execute(f'PRAGMA table_info("{table}")').fetchall()
        if row and len(row) > 1
    }
    for col in columns:
        if col not in existing:
            conn.execute(f'ALTER TABLE "{table}" ADD COLUMN "{col}" TEXT')


def _table_columns(conn: sqlite3.Connection, table: str) -> List[str]:
    return [
        row[1]
        for row in conn.execute(f'PRAGMA table_info("{table}")').fetchall()
        if row and len(row) > 1
    ]


def _table_row_count(conn: sqlite3.Connection, table: str) -> int:
    return int(conn.execute(f'SELECT COUNT(*) FROM "{table}"').fetchone()[0])


def _is_sqlite_locked_error(exc: sqlite3.OperationalError) -> bool:
    msg = str(exc or "").lower()
    return "database is locked" in msg or "database schema is locked" in msg


def _is_sqlite_missing_table_error(exc: sqlite3.OperationalError) -> bool:
    msg = str(exc or "").lower()
    return "no such table" in msg


def _ensure_cache_status_table(conn: sqlite3.Connection) -> None:
    conn.execute(
        f"""
        CREATE TABLE IF NOT EXISTS {CACHE_STATUS_TABLE} (
            base_table TEXT NOT NULL,
            env TEXT NOT NULL,
            env_table TEXT NOT NULL,
            row_count INTEGER,
            loaded_at TEXT NOT NULL,
            source TEXT,
            PRIMARY KEY (base_table, env)
        )
        """
    )


def _record_cache_status(
    base_table: str,
    env: str,
    source: str,
    row_count: int | None = None,
    connect_timeout: float = 30.0,
    busy_timeout_ms: int = 30000,
    max_attempts: int = 6,
    retry_sleep_base: float = 0.25,
) -> None:
    base_table = _validate_table(base_table)
    normalized = _normalize_env(env)
    env_table = _table_for(base_table, normalized)
    loaded_at = datetime.utcnow().replace(microsecond=0).isoformat() + "Z"

    attempts = max(1, int(max_attempts))
    for attempt in range(attempts):
        try:
            with _connect(timeout=connect_timeout, busy_timeout_ms=busy_timeout_ms) as conn:
                _ensure_cache_status_table(conn)
                if row_count is None:
                    row_count = _table_row_count(conn, env_table) if _table_exists(conn, env_table) else 0
                conn.execute(
                    f"""
                    INSERT INTO {CACHE_STATUS_TABLE} (base_table, env, env_table, row_count, loaded_at, source)
                    VALUES (?, ?, ?, ?, ?, ?)
                    ON CONFLICT(base_table, env) DO UPDATE SET
                        env_table = excluded.env_table,
                        row_count = excluded.row_count,
                        loaded_at = excluded.loaded_at,
                        source = excluded.source
                    """,
                    (base_table, normalized, env_table, int(row_count), loaded_at, source),
                )
                conn.commit()
                return
        except sqlite3.OperationalError as exc:
            if not _is_sqlite_locked_error(exc) or attempt >= (attempts - 1):
                raise
            time.sleep(max(0.0, retry_sleep_base) * (attempt + 1))


def _cache_status_for(base_table: str, env: str) -> dict:
    base_table = _validate_table(base_table)
    normalized = _normalize_env(env)
    env_table = _table_for(base_table, normalized)

    for attempt in range(2):
        try:
            with _connect(timeout=0.5, busy_timeout_ms=500) as conn:
                _ensure_cache_status_table(conn)
                row = conn.execute(
                    f"SELECT base_table, env, env_table, row_count, loaded_at, source FROM {CACHE_STATUS_TABLE} WHERE base_table = ? AND env = ?",
                    (base_table, normalized),
                ).fetchone()
                if row is not None:
                    return {
                        "table": row["base_table"],
                        "env": row["env"],
                        "env_table": row["env_table"],
                        "row_count": int(row["row_count"] or 0),
                        "loaded_at": row["loaded_at"],
                        "source": row["source"] or "reload",
                    }

                if _table_exists(conn, env_table):
                    inferred_loaded_at = None
                    try:
                        inferred_loaded_at = datetime.utcfromtimestamp(DB_PATH.stat().st_mtime).replace(microsecond=0).isoformat() + "Z"
                    except OSError:
                        pass
                    return {
                        "table": base_table,
                        "env": normalized,
                        "env_table": env_table,
                        "row_count": _table_row_count(conn, env_table),
                        "loaded_at": inferred_loaded_at,
                        "source": "inferred",
                    }
                break
        except sqlite3.OperationalError as exc:
            if not _is_sqlite_locked_error(exc):
                raise
            if attempt >= 1:
                return {
                    "table": base_table,
                    "env": normalized,
                    "env_table": env_table,
                    "row_count": 0,
                    "loaded_at": None,
                    "source": "locked",
                }
            time.sleep(0.1 * (attempt + 1))

    return {
        "table": base_table,
        "env": normalized,
        "env_table": env_table,
        "row_count": 0,
        "loaded_at": None,
        "source": "missing",
    }


def _columns_from_sql_file(sql_path: Path) -> List[str]:
    if not sql_path.exists():
        return []
    try:
        sql_text = sql_path.read_text(encoding="utf-8")
    except OSError:
        return []
    matches = re.findall(r"\bas\s+(['\"])([^'\"]+)\1", sql_text, flags=re.IGNORECASE)
    columns: List[str] = []
    seen = set()
    for _, name in matches:
        if name not in seen:
            columns.append(name)
            seen.add(name)
    return columns


def _wagons_sql_file(env: str | None) -> Path:
    normalized = _normalize_env(env)
    preferred = WAGONS_SQL_FILES.get(normalized)
    if preferred and preferred.exists():
        return preferred
    if SQL_FILE.exists():
        return SQL_FILE
    return preferred or SQL_FILE


def _wagensuche_sql_file(env: str | None) -> Path:
    normalized = _normalize_env(env)
    preferred = WAGENSUCHE_SQL_FILES.get(normalized)
    if preferred and preferred.exists():
        return preferred
    fallback = WAGENSUCHE_SQL_FILES.get("prd")
    return preferred or fallback or SQL_FILE


def _teilenummer_sql_file(env: str | None) -> Path:
    normalized = _normalize_env(env)
    preferred = TEILENUMMER_SQL_FILES.get(normalized)
    if preferred and preferred.exists():
        return preferred
    if TEILENUMMER_SQL_FILE.exists():
        return TEILENUMMER_SQL_FILE
    return preferred or TEILENUMMER_SQL_FILE


def _wagensuche_pg_dsn() -> Optional[str]:
    if WAGENSUCHE_PG_URL:
        return WAGENSUCHE_PG_URL
    if not (WAGENSUCHE_PG_HOST and WAGENSUCHE_PG_DB and WAGENSUCHE_PG_USER):
        return None
    port = WAGENSUCHE_PG_PORT or "5432"
    password = WAGENSUCHE_PG_PASS
    return (
        f"postgresql://{WAGENSUCHE_PG_USER}:{password}"
        f"@{WAGENSUCHE_PG_HOST}:{port}/{WAGENSUCHE_PG_DB}"
    )


def _normalize_sern_variants(sern: str) -> List[str]:
    raw = sern.strip()
    digits = re.sub(r"\D", "", raw)
    variants = []
    if raw:
        variants.append(raw)
    if digits and digits != raw:
        variants.append(digits)
    # Deduplicate while preserving order
    seen = set()
    unique: List[str] = []
    for item in variants:
        if item not in seen:
            unique.append(item)
            seen.add(item)
    return unique


def _wagensuche_latest_position(sern: str) -> Optional[Dict[str, Any]]:
    dsn = _wagensuche_pg_dsn()
    if not dsn:
        raise HTTPException(
            status_code=500,
            detail="Postgres-Konfiguration fehlt. Bitte WAGENSUCHE_PG_* setzen.",
        )
    variants = _normalize_sern_variants(sern)
    if not variants:
        return None
    placeholders = ", ".join(["%s"] * len(variants))
    query = f"""
        SELECT
            "ITSS_TransportDeviceID",
            "GNSS_UTCtimestamp",
            "GNSS_Longitude",
            "GNSS_Latitude",
            mileage
        FROM streaming.notification
        WHERE "ITSS_TransportDeviceID" IN ({placeholders})
          AND mileage > 0
          AND "GNSS_Longitude" IS NOT NULL
          AND "GNSS_Latitude" IS NOT NULL
        ORDER BY "GNSS_UTCtimestamp" DESC
        LIMIT 1
    """
    try:
        with psycopg.connect(dsn, connect_timeout=5, options="-c statement_timeout=10000") as conn:
            with conn.cursor() as cur:
                cur.execute(query, tuple(variants))
                row = cur.fetchone()
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=500, detail=f"Postgres-Query fehlgeschlagen: {exc}") from exc
    if not row:
        return None
    return {
        "sern": row[0],
        "timestamp": row[1],
        "longitude": row[2],
        "latitude": row[3],
        "mileage": row[4],
    }


def _create_table_from_columns(conn: sqlite3.Connection, table: str, columns: List[str]) -> None:
    if _table_exists(conn, table) or not columns:
        return
    column_defs = ", ".join(f'"{col}" TEXT' for col in columns)
    conn.execute(f'CREATE TABLE IF NOT EXISTS "{table}" ({column_defs})')


def _clone_table_schema(conn: sqlite3.Connection, source: str, target: str) -> None:
    if _table_exists(conn, target) or not _table_exists(conn, source):
        return
    conn.execute(f'CREATE TABLE IF NOT EXISTS "{target}" AS SELECT * FROM "{source}" WHERE 0=1')


def _ordered_columns(columns: List[str], order_hint: List[str] | None) -> List[str]:
    if not order_hint:
        return columns
    ordered = [col for col in order_hint if col in columns]
    remaining = [col for col in columns if col not in ordered]
    ordered.extend(sorted(remaining, key=lambda name: name.upper()))
    return ordered


def _rebuild_table_with_order(
    conn: sqlite3.Connection,
    table: str,
    ordered_columns: List[str],
) -> None:
    if not ordered_columns:
        return
    existing_columns = _table_columns(conn, table)
    if existing_columns == ordered_columns:
        return
    temp_name = f"{table}__tmp"
    column_defs = ", ".join(f'"{col}" TEXT' for col in ordered_columns)
    copy_columns = [col for col in ordered_columns if col in existing_columns]
    column_list = ", ".join(f'"{col}"' for col in copy_columns)
    try:
        conn.execute("BEGIN")
        conn.execute(f'DROP TABLE IF EXISTS "{temp_name}"')
        conn.execute(f'CREATE TABLE "{temp_name}" ({column_defs})')
        if copy_columns:
            conn.execute(
                f'INSERT INTO "{temp_name}" ({column_list}) SELECT {column_list} FROM "{table}"'
            )
        conn.execute(f'DROP TABLE "{table}"')
        conn.execute(f'ALTER TABLE "{temp_name}" RENAME TO "{table}"')
        conn.commit()
    except Exception:
        conn.rollback()
        raise


def _ensure_env_table_pair(
    conn: sqlite3.Connection,
    base: str,
    columns_hint: List[str] | None = None,
    extra_columns: List[str] | None = None,
    enforce_order: bool = False,
) -> None:
    prd_table = _table_for(base, "prd")
    tst_table = _table_for(base, "tst")
    prd_exists = _table_exists(conn, prd_table)
    tst_exists = _table_exists(conn, tst_table)

    if not prd_exists and not tst_exists and columns_hint:
        _create_table_from_columns(conn, prd_table, columns_hint)
        _create_table_from_columns(conn, tst_table, columns_hint)
    elif not prd_exists and tst_exists:
        _clone_table_schema(conn, tst_table, prd_table)
    elif not tst_exists and prd_exists:
        _clone_table_schema(conn, prd_table, tst_table)

    prd_exists = _table_exists(conn, prd_table)
    tst_exists = _table_exists(conn, tst_table)
    if not prd_exists and not tst_exists:
        return

    prd_columns = _table_columns(conn, prd_table) if prd_exists else []
    tst_columns = _table_columns(conn, tst_table) if tst_exists else []
    merged = list(
        dict.fromkeys(prd_columns + tst_columns + (columns_hint or []) + (extra_columns or []))
    )

    if prd_exists:
        _ensure_columns(conn, prd_table, merged)
    if tst_exists:
        _ensure_columns(conn, tst_table, merged)

    if enforce_order:
        ordered = _ordered_columns(merged, columns_hint)
        if prd_exists:
            _rebuild_table_with_order(conn, prd_table, ordered)
        if tst_exists:
            _rebuild_table_with_order(conn, tst_table, ordered)


def _ensure_renumber_schema(conn: sqlite3.Connection, table_name: str) -> None:
    if not _table_exists(conn, table_name):
        return
    existing_info = conn.execute(f'PRAGMA table_info("{table_name}")').fetchall()
    existing_columns = [row[1] for row in existing_info if row and len(row) > 1]
    missing = [col for col in RENUMBER_EXTRA_COLUMNS if col not in existing_columns]
    if not missing:
        return
    base_columns = [col for col in existing_columns if col not in RENUMBER_EXTRA_COLUMNS]
    ordered_columns = base_columns + [
        col for col in RENUMBER_EXTRA_COLUMNS if col in existing_columns or col in missing
    ]
    temp_name = f"{table_name}__tmp"
    column_defs = ", ".join(f'"{col}" TEXT' for col in ordered_columns)
    copy_columns = [col for col in ordered_columns if col in existing_columns]
    column_list = ", ".join(f'"{col}"' for col in copy_columns)
    try:
        conn.execute("BEGIN")
        conn.execute(f'DROP TABLE IF EXISTS "{temp_name}"')
        conn.execute(f'CREATE TABLE "{temp_name}" ({column_defs})')
        if copy_columns:
            conn.execute(
                f'INSERT INTO "{temp_name}" ({column_list}) SELECT {column_list} FROM "{table_name}"'
            )
        conn.execute(f'DROP TABLE "{table_name}"')
        conn.execute(f'ALTER TABLE "{temp_name}" RENAME TO "{table_name}"')
        conn.commit()
    except Exception:
        conn.rollback()
        raise


def _ensure_env_tables(conn: sqlite3.Connection) -> None:
    wagons_columns = _columns_from_sql_file(_wagons_sql_file("prd"))
    wagensuche_columns = _columns_from_sql_file(_wagensuche_sql_file("prd"))
    spareparts_columns = _columns_from_sql_file(SPAREPARTS_SQL_FILE)
    rsrd_full_columns = _columns_from_sql_file(RSRD_ERP_SQL_FILE)

    _ensure_env_table_pair(conn, DEFAULT_TABLE, columns_hint=wagons_columns, enforce_order=True)
    _ensure_env_table_pair(conn, WAGENUMBAU_TABLE, columns_hint=wagons_columns, enforce_order=True)
    _ensure_env_table_pair(conn, WAGENSUCHE_TABLE, columns_hint=wagensuche_columns, enforce_order=True)
    _ensure_env_table_pair(conn, SPAREPARTS_TABLE, columns_hint=spareparts_columns, enforce_order=True)
    _ensure_env_table_pair(conn, RSRD_ERP_FULL_TABLE, columns_hint=rsrd_full_columns, enforce_order=True)
    _ensure_rsrd_sync_table(conn, "prd")
    _ensure_rsrd_sync_table(conn, "tst")
    _ensure_env_table_pair(
        conn,
        RSRD_ERP_TABLE,
        columns_hint=["wagon_sern", "wagon_sern_numeric", "updated_at"],
        enforce_order=True,
    )
    _ensure_env_table_pair(
        conn,
        RENUMBER_WAGON_TABLE,
        columns_hint=RENUMBER_EXTRA_COLUMNS,
        extra_columns=RENUMBER_EXTRA_COLUMNS,
    )
    _ensure_env_table_pair(conn, SPAREPARTS_SWAP_TABLE)

    _ensure_swap_table(conn, _table_for(SPAREPARTS_SWAP_TABLE, "prd"))
    _ensure_swap_table(conn, _table_for(SPAREPARTS_SWAP_TABLE, "tst"))
    _ensure_rsrd_upload_table(conn, "prd")
    _ensure_rsrd_upload_table(conn, "tst")
    _ensure_rsrd_tables(conn, "prd")
    _ensure_rsrd_tables(conn, "tst")
    _ensure_env_table_pair(conn, RSRD_UPLOAD_TABLE)
    _ensure_env_table_pair(conn, BASE_WAGONS_TABLE)
    _ensure_env_table_pair(conn, BASE_SNAPSHOTS_TABLE)
    _ensure_env_table_pair(conn, BASE_JSON_TABLE)
    _ensure_env_table_pair(conn, BASE_DETAIL_TABLE)

    _ensure_renumber_schema(conn, _table_for(RENUMBER_WAGON_TABLE, "prd"))
    _ensure_renumber_schema(conn, _table_for(RENUMBER_WAGON_TABLE, "tst"))


def _clear_table_rows(table_base: str, env: str) -> None:
    table_name = _table_for(table_base, env)
    with _connect() as conn:
        if _table_exists(conn, table_name):
            conn.execute(f'DELETE FROM "{table_name}"')
            conn.commit()


def _update_job(job_id: str, **updates: Any) -> None:
    with _jobs_lock:
        job = _jobs.get(job_id)
        if not job:
            return
        job.update(updates)


def _append_job_result(job_id: str, result: Dict[str, Any]) -> None:
    with _jobs_lock:
        job = _jobs.get(job_id)
        if not job:
            return
        job.setdefault("results", []).append(result)


def _format_yyyymmdd(value: str) -> str:
    if not value:
        return ""
    digits = "".join(ch for ch in value if ch.isdigit())
    if len(digits) >= 8:
        return digits[:8]
    return digits


def _format_ddmmyy(value: str) -> str:
    if not value:
        return ""
    digits = "".join(ch for ch in value if ch.isdigit())
    if len(digits) >= 8:
        if digits.startswith(("19", "20")):
            yyyymmdd = digits[:8]
            return f"{yyyymmdd[6:8]}{yyyymmdd[4:6]}{yyyymmdd[2:4]}"
        ddmmyyyy = digits[:8]
        return f"{ddmmyyyy[0:4]}{ddmmyyyy[6:8]}"
    if len(digits) >= 6:
        return digits[:6]
    return digits


def _hierarchy_level(value: str) -> int:
    if not value:
        return 0
    return str(value).count("-")


def _row_value(row: sqlite3.Row, *keys: str) -> str:
    for key in keys:
        if key in row.keys():
            value = row[key]
            if value is not None and value != "":
                return str(value)
    return ""


def _model_suffix(value: str) -> str:
    cleaned = (value or "").strip()
    if not cleaned:
        return ""
    if "_" in cleaned:
        return cleaned.split("_")[-1]
    return cleaned


def _wagon_serial_suffix(value: str) -> str:
    cleaned = re.sub(r"\s+", "", value or "")
    if not cleaned:
        return ""
    suffix = cleaned[-10:] if len(cleaned) > 10 else cleaned
    trimmed = suffix.lstrip("0")
    return trimmed or suffix


def _compute_part_updates(
    row: sqlite3.Row,
    new_sern: str,
    new_baureihe: str,
) -> tuple[str, str]:
    itno = _row_value(row, "ITNO")
    ser2 = _row_value(row, "SER2")
    old_model = _model_suffix(_row_value(row, "WAGEN_ITNO"))
    new_model = _model_suffix(new_baureihe)
    new_itno = ""
    if itno and old_model and new_model:
        if itno.endswith(f"_{old_model}") or itno.endswith(old_model):
            new_itno = itno[: -len(old_model)] + new_model
            if new_itno == itno:
                new_itno = ""
    old_suffix = _wagon_serial_suffix(_row_value(row, "WAGEN_SERN"))
    new_suffix = _wagon_serial_suffix(new_sern)
    new_ser2 = ""
    if ser2 and "-" in ser2 and old_suffix and new_suffix:
        if ser2.endswith(old_suffix):
            new_ser2 = ser2[: -len(old_suffix)] + new_suffix
            if new_ser2 == ser2:
                new_ser2 = ""
    return new_itno, new_ser2


def _needs_renumber(row: sqlite3.Row) -> bool:
    if not _row_value(row, "SER2"):
        return False
    return bool(_row_value(row, "NEW_PART_ITNO") or _row_value(row, "NEW_PART_SER2"))


def _is_ok_status(value: str | None) -> bool:
    normalized = str(value or "").strip().lower()
    return normalized.startswith("ok") or "success" in normalized or "erfolg" in normalized


def _renumber_row_key(row: Mapping[str, Any]) -> str:
    cfgl = row.get("CFGL") or row.get("MFGL") or ""
    parts = [
        str(cfgl),
        str(row.get("ITNO") or ""),
        str(row.get("SER2") or ""),
        str(row.get("MTRL") or ""),
        str(row.get("SERN") or ""),
    ]
    return "||".join(parts)


def _renumber_sort_key(row: Dict[str, Any]) -> str:
    return str(row.get("CFGL") or row.get("MFGL") or "")


def _wagon_log_context(row: sqlite3.Row) -> Dict[str, str]:
    return {
        "itno": _row_value(row, "WAGEN_ITNO", "NEW_BAUREIHE"),
        "sern": _row_value(row, "WAGEN_SERN", "NEW_SERN"),
        "new_itno": _row_value(row, "NEW_BAUREIHE"),
        "new_sern": _row_value(row, "NEW_SERN"),
    }


def _mi_error_message(payload: Any) -> str:
    if payload is None:
        return "Leere Antwort"
    if isinstance(payload, dict):
        status_code = payload.get("status_code") or payload.get("statusCode")
        if status_code and int(status_code) != 200:
            return f"HTTP {status_code}"
        if "text" in payload and payload.get("text"):
            text = str(payload.get("text"))
            if "error" in text.lower() or "fehler" in text.lower():
                return text
        response = payload.get("MIResponse") or payload.get("response") or payload
        if isinstance(response, dict):
            response_type = str(response.get("@type") or response.get("type") or "").strip()
            response_code = str(response.get("@code") or response.get("code") or "").strip()
            if "NOK" in response_type or "NOK" in response_code:
                messages = response.get("Messages") or response.get("messages")
                if isinstance(messages, dict):
                    message_entries = messages.get("Message") or messages.get("message") or []
                    if not isinstance(message_entries, list):
                        message_entries = [message_entries]
                    for entry in message_entries:
                        if not isinstance(entry, dict):
                            continue
                        msg_text = str(entry.get("MessageText") or entry.get("messageText") or "").strip()
                        if msg_text:
                            return msg_text
                return response_code or "ServerReturnedNOK"
            for key in ("ErrorNumber", "errorNumber", "ErrorCode", "errorCode", "Error"):
                value = response.get(key)
                if value is None or value == "":
                    continue
                if str(value) not in {"0", "00"}:
                    return f"{key}={value}"
            messages = response.get("Messages") or response.get("messages")
            if isinstance(messages, dict):
                message_entries = messages.get("Message") or messages.get("message") or []
                if not isinstance(message_entries, list):
                    message_entries = [message_entries]
                for entry in message_entries:
                    if not isinstance(entry, dict):
                        continue
                    msg_type = str(entry.get("MessageType") or entry.get("messageType") or "").strip()
                    msg_text = str(entry.get("MessageText") or entry.get("messageText") or "").strip()
                if msg_type in {"2", "3", "4", "E", "ERROR"}:
                    if msg_text and "mo96202" in msg_text.lower():
                        return ""
                    return msg_text or f"MessageType={msg_type}"
                if msg_text and ("error" in msg_text.lower() or "fehler" in msg_text.lower()):
                    return msg_text
            for key in (
                "ErrorMessage",
                "errorMessage",
                "ErrorMsg",
                "errorMsg",
                "ErrorText",
                "errorText",
            ):
                value = response.get(key)
                if value:
                    return str(value)
            message = response.get("Message") or response.get("message")
            if message:
                text = str(message)
                if "error" in text.lower() or "fehler" in text.lower():
                    return text
        for key, value in payload.items():
            lower = key.lower()
            if lower in {"error", "errormessage", "errormsg", "errortext"} and value:
                return str(value)
            if isinstance(value, (dict, list)):
                nested = _mi_error_message(value)
                if nested:
                    return nested
    if isinstance(payload, list):
        for item in payload:
            nested = _mi_error_message(item)
            if nested:
                return nested
    return ""


def _mi_extract_code_message(payload: Any) -> tuple[str, str]:
    if not isinstance(payload, dict):
        return "", ""
    response = payload.get("MIResponse") or payload.get("response") or payload
    if not isinstance(response, dict):
        return "", ""
    code = str(response.get("@code") or response.get("code") or "").strip()
    message = ""
    messages = response.get("Messages") or response.get("messages")
    if isinstance(messages, dict):
        message_entries = messages.get("Message") or messages.get("message") or []
        if not isinstance(message_entries, list):
            message_entries = [message_entries]
        for entry in message_entries:
            if not isinstance(entry, dict):
                continue
            msg_text = str(entry.get("MessageText") or entry.get("messageText") or "").strip()
            if msg_text:
                message = msg_text
                break
    if not message:
        message = str(response.get("Message") or response.get("message") or "").strip()
    return code, message


def _mi_status(payload: Any) -> tuple[bool, str, str]:
    code, message = _mi_extract_code_message(payload)
    message_lower = message.lower()
    if code == "MO96202" or "asynchronous removal" in message_lower:
        return True, "OK_ASYNC", ""
    if code == "MO12524" or "is installed in this position" in message_lower:
        return True, "OK_IDEMPOTENT", ""
    if code == "MO12527" or "status is 80" in message_lower:
        return False, "BLOCKING_WARNING", message or code
    error_message = _mi_error_message(payload)
    if error_message:
        return False, "ERROR", error_message
    return True, "OK", ""


def _objstrk_has_item(rows: List[Dict[str, Any]], cfgl: str, itno: str, ser2: str) -> bool:
    cfgl_key = str(cfgl or "").strip()
    itno_key = str(itno or "").strip()
    ser2_key = str(ser2 or "").strip()
    for row in rows:
        row_cfgl = str(row.get("CFGL") or row.get("MFGL") or "").strip()
        row_itno = str(row.get("ITNO") or "").strip()
        row_ser2 = str(row.get("SER2") or "").strip()
        if row_cfgl == cfgl_key and row_itno == itno_key and row_ser2 == ser2_key:
            return True
    return False


def _build_m3_request_url(base_url: str, program: str, transaction: str, params: Dict[str, Any]) -> str:
    base = base_url.rstrip("/") if base_url else ""
    path = f"/M3/m3api-rest/execute/{program}/{transaction}"
    url = f"{base}{path}" if base else path
    if params:
        return f"{url}?{urlencode(params)}"
    return url


def _build_ips_request_url(base_url: str, service_name: str) -> str:
    base = base_url.rstrip("/") if base_url else ""
    path = f"/M3/ips/service/{service_name}"
    return f"{base}{path}" if base else path


def _ips_company_division(env: str | None) -> tuple[str, str]:
    normalized = _normalize_env(env or DEFAULT_ENV)
    if normalized == "tst" and (IPS_COMPANY_TST or IPS_DIVISION_TST):
        return IPS_COMPANY_TST, IPS_DIVISION_TST
    return IPS_COMPANY, IPS_DIVISION


def _mi_cono(env: str | None) -> str:
    normalized = _normalize_env(env or DEFAULT_ENV)
    if normalized == "tst":
        return MI_CONO_TST or MI_CONO
    if normalized == "prd":
        return MI_CONO_PRD or MI_CONO
    return MI_CONO


def _build_ips_envelope(
    service_name: str,
    operation: str,
    params: Dict[str, str],
    namespace_override: str | None = None,
    body_tag_override: str | None = None,
    env: str | None = None,
) -> str:
    namespace = namespace_override or f"http://schemas.infor.com/ips/{service_name}/{operation}"
    parts = []
    for key, value in params.items():
        safe = xml_escape(value or "")
        parts.append(f"<chg:{key}>{safe}</chg:{key}>")
    body = "".join(parts)
    header = "<soapenv:Header/>"
    company, division = _ips_company_division(env)
    if company or division:
        cred_parts = []
        if company:
            cred_parts.append(f"<cred:company>{xml_escape(company)}</cred:company>")
        if division:
            cred_parts.append(f"<cred:division>{xml_escape(division)}</cred:division>")
        cred_body = "".join(cred_parts)
        header = f"<soapenv:Header><cred:lws>{cred_body}</cred:lws></soapenv:Header>"
    body_tag = body_tag_override or service_name
    return (
        '<soapenv:Envelope xmlns:soapenv="http://schemas.xmlsoap.org/soap/envelope/" '
        f'xmlns:chg="{namespace}" xmlns:cred="http://lawson.com/ws/credentials">'
        f"{header}"
        "<soapenv:Body>"
        f"<chg:{operation}><chg:{body_tag}>{body}</chg:{body_tag}></chg:{operation}>"
        "</soapenv:Body>"
        "</soapenv:Envelope>"
    )


def _call_ips_service(
    base_url: str,
    access_token: str,
    service_name: str,
    operation: str,
    params: Dict[str, str],
    namespace_override: str | None = None,
    body_tag_override: str | None = None,
    env: str | None = None,
) -> Dict[str, Any]:
    url = _build_ips_request_url(base_url, service_name)
    body = _build_ips_envelope(
        service_name,
        operation,
        params,
        namespace_override=namespace_override,
        body_tag_override=body_tag_override,
        env=env,
    )
    headers = {
        "Authorization": f"Bearer {access_token}",
        "Accept": "text/xml",
        "Content-Type": "text/xml; charset=utf-8",
    }
    import requests

    resp = requests.post(url, headers=headers, data=body.encode("utf-8"), timeout=60)
    return {
        "status_code": resp.status_code,
        "text": resp.text,
        "request_body": body,
    }


def _ips_mos100_already_exists_fault(response: Any) -> bool:
    if not isinstance(response, dict):
        return False
    text = str(response.get("text") or "").lower()
    return (
        "origin identity" in text
        and "already exists" in text
        and "reclassification not possible" in text
    )


def _append_api_log(
    action: str,
    params: Dict[str, Any],
    response: Any,
    ok: bool,
    error: str | None = None,
    env: str | None = None,
    wagon: Dict[str, str] | None = None,
    dry_run: bool | None = None,
    request_url: str | None = None,
    program: str = "MOS125MI",
    transaction: str = "RemoveInstall",
    request_method: str = "GET",
    status: str | None = None,
) -> None:
    if API_LOG_ONLY and action not in API_LOG_ONLY:
        return
    if action == "rollback":
        entry = {
            "ts": datetime.utcnow().isoformat(sep=" ", timespec="seconds"),
            "env": env or "",
            "action": action,
            "itno": params.get("ITNO", ""),
            "sern": params.get("SERN", ""),
            "parent_itno": params.get("PARENT_ITNO", ""),
            "parent_sern": params.get("PARENT_SERN", ""),
            "response": response,
        }
    else:
        entry = {
            "ts": datetime.utcnow().isoformat(sep=" ", timespec="seconds"),
            "env": env or "",
            "action": action,
            "wagon": wagon or {},
            "program": program,
            "transaction": transaction,
            "dry_run": bool(dry_run),
            "request": {"method": request_method, "url": request_url or ""},
            "ok": ok,
            "params": params,
            "error": error or "",
            "response": response,
        }
        if status is not None:
            entry["status"] = status
    try:
        API_LOG_PATH.parent.mkdir(parents=True, exist_ok=True)
        with API_LOG_PATH.open("a", encoding="utf-8") as handle:
            handle.write(json.dumps(entry, ensure_ascii=True, default=str))
            handle.write("\n")
    except Exception:  # noqa: BLE001
        pass
    if bool(dry_run):
        _append_dryrun_trace_entry(entry)


def _clear_api_log() -> None:
    try:
        API_LOG_PATH.parent.mkdir(parents=True, exist_ok=True)
        API_LOG_PATH.write_text("", encoding="utf-8")
    except Exception:  # noqa: BLE001
        pass


def _reset_dryrun_trace(job_name: str, env: str) -> None:
    global _dryrun_trace_seq
    with _dryrun_trace_lock:
        _dryrun_trace_seq = 0
        started = datetime.utcnow().isoformat(sep=" ", timespec="seconds")
        lines = [
            "# MFDAPPS DryRun API Trace",
            f"# Job: {job_name}",
            f"# Env: {env}",
            f"# Started (UTC): {started}",
            "# Format: seq | ts | env | action | api | method url | ok | status | wagon | params",
            "",
        ]
        DRYRUN_TRACE_PATH.parent.mkdir(parents=True, exist_ok=True)
        DRYRUN_TRACE_PATH.write_text("\n".join(lines), encoding="utf-8")


def _append_dryrun_trace_entry(entry: Dict[str, Any]) -> None:
    global _dryrun_trace_seq
    try:
        with _dryrun_trace_lock:
            _dryrun_trace_seq += 1
            seq = _dryrun_trace_seq
            ts = str(entry.get("ts") or "")
            env = str(entry.get("env") or "")
            action = str(entry.get("action") or "")
            program = str(entry.get("program") or "")
            transaction = str(entry.get("transaction") or "")
            api_name = f"{program}.{transaction}" if (program or transaction) else "-"
            request = entry.get("request") if isinstance(entry.get("request"), dict) else {}
            method = str(request.get("method") or "")
            url = str(request.get("url") or "")
            ok = "OK" if bool(entry.get("ok")) else "ERR"
            status = str(entry.get("status") or "")
            wagon_obj = entry.get("wagon") if isinstance(entry.get("wagon"), dict) else {}
            wagon_old = f"{wagon_obj.get('itno', '')}/{wagon_obj.get('sern', '')}".strip("/")
            wagon_new = f"{wagon_obj.get('new_itno', '')}/{wagon_obj.get('new_sern', '')}".strip("/")
            wagon = f"{wagon_old} -> {wagon_new}" if wagon_new else wagon_old
            params = entry.get("params") if isinstance(entry.get("params"), dict) else {}
            params_json = json.dumps(params, ensure_ascii=True, default=str, separators=(",", ":"))
            line = (
                f"{seq:05d} | {ts} | {env} | {action} | {api_name} | "
                f"{method} {url} | {ok} | {status} | {wagon} | {params_json}"
            )
            with DRYRUN_TRACE_PATH.open("a", encoding="utf-8") as handle:
                handle.write(line)
                handle.write("\n")
    except Exception:  # noqa: BLE001
        pass


def _build_mos125_params(row: sqlite3.Row, mode: str = "out", env: str | None = None) -> Dict[str, str]:
    cfgr = _row_value(row, "CFGL", "MFGL")
    level = _hierarchy_level(cfgr)
    base_trtm = 10000 + (level * 1000)
    planned_out_date = _sanitize_yyyymmdd(_row_value(row, "PLAN_OUT_DATE"))
    planned_out_time = _sanitize_hhmmss(_row_value(row, "PLAN_OUT_TIME"))
    planned_in_date = _sanitize_yyyymmdd(_row_value(row, "PLAN_IN_DATE"))
    planned_in_time = _sanitize_hhmmss(_row_value(row, "PLAN_IN_TIME"))
    params = {
        "RITP": "UMA",
        "RESP": "CHRUPP",
        "TRDT": planned_out_date or _format_yyyymmdd(_row_value(row, "UMBAU_DATUM")),
        "WHLO": "ZUM",
        "RSC4": "UMB",
        "TRTM": planned_out_time or str(base_trtm),
    }
    if mode == "in":
        new_baureihe = _row_value(row, "NEW_BAUREIHE")
        new_sern = _row_value(row, "NEW_SERN")
        cfgr_value = str(cfgr or "").strip()
        top_level = cfgr_value.isdigit() and int(cfgr_value) in {1, 2, 3, 4}
        parent_itno = _row_value(row, "MTRL")
        parent_sern = _row_value(row, "SERN")
        if top_level:
            parent_itno = new_baureihe or _row_value(row, "WAGEN_ITNO") or parent_itno
            parent_sern = new_sern or _row_value(row, "WAGEN_SERN") or parent_sern
        else:
            parent_row = {
                "ITNO": parent_itno,
                "SER2": parent_sern,
                "WAGEN_ITNO": _row_value(row, "WAGEN_ITNO"),
                "WAGEN_SERN": _row_value(row, "WAGEN_SERN"),
            }
            updated_parent_itno, updated_parent_sern = _compute_part_updates(parent_row, new_sern, new_baureihe)
            parent_itno = updated_parent_itno or parent_itno
            parent_sern = updated_parent_sern or parent_sern
        part_itno = _row_value(row, "NEW_PART_ITNO") or _row_value(row, "ITNO")
        part_ser2 = _row_value(row, "NEW_PART_SER2") or _row_value(row, "SER2")
        params["RITP"] = "UME"
        params["TRDT"] = planned_in_date or params["TRDT"]
        params["TRTM"] = planned_in_time or str(base_trtm + 10)
        params["CFGL"] = cfgr
        params["TWSL"] = "EINBAU"
        params["NHAI"] = parent_itno
        params["NHSI"] = parent_sern
        params["ITNI"] = part_itno
        params["BANI"] = part_ser2
    elif mode == "rollback":
        rmts_raw = _row_value(row, "RMTS")
        trtm_value = ""
        if rmts_raw:
            rmts_str = "".join(ch for ch in str(rmts_raw) if ch.isdigit())
            if rmts_str:
                rmts_str = rmts_str.zfill(6)
                try:
                    hours = int(rmts_str[0:2])
                    minutes = int(rmts_str[2:4])
                    seconds = int(rmts_str[4:6])
                    total_seconds = (hours * 3600) + (minutes * 60) + seconds + 60
                    hours = (total_seconds // 3600) % 24
                    minutes = (total_seconds % 3600) // 60
                    seconds = total_seconds % 60
                    trtm_value = str((hours * 10000) + (minutes * 100) + seconds)
                except ValueError:
                    trtm_value = ""
        params["RITP"] = "UME"
        params["TRTM"] = trtm_value or str(base_trtm + 10)
        params["CFGL"] = cfgr
        params["TWSL"] = "EINBAU"
        params["NHAI"] = _row_value(row, "MTRL")
        params["NHSI"] = _row_value(row, "SERN")
        params["ITNI"] = _row_value(row, "ITNO")
        params["BANI"] = _row_value(row, "SER2")
    else:
        params["CFGR"] = cfgr
        params["TWSL"] = "AUSBAU"
        params["NHAR"] = _row_value(row, "MTRL")
        params["NHSR"] = _row_value(row, "SERN")
        params["ITNR"] = _row_value(row, "ITNO")
        params["BANR"] = _row_value(row, "SER2")
    return params


def _build_mos170_params(row: sqlite3.Row, env: str | None = None) -> Dict[str, str]:
    umbau_datum = _format_yyyymmdd(_row_value(row, "UMBAU_DATUM"))
    params = {
        "ITNO": _row_value(row, "ITNO"),
        "BANO": _row_value(row, "SER2"),
        "STRT": "002",
        "SUFI": "SERIENNUMMER AENDERN",
        "STDT": umbau_datum,
        "FIDT": umbau_datum,
        "RESP": "CHRUPP",
        "WHLO": "ZUM",
    }
    return params


# BEGIN WAGON RENNUMBERING
def _build_mos170_wagon_params(
    itno: str,
    sern: str,
    umbau_datum: str,
    whlo: str,
    env: str | None = None,
) -> Dict[str, str]:
    umbau_value = _format_yyyymmdd(umbau_datum)
    params = {
        "ITNO": itno,
        "BANO": sern,
        "STRT": "002",
        "SUFI": "SERIENNUMMER AENDERN",
        "STDT": umbau_value,
        "FIDT": umbau_value,
        "RESP": "CHRUPP",
        "WHLO": whlo,
    }
    return params
# END WAGON RENNUMBERING

def _extract_plpn(response: Any) -> str:
    if not isinstance(response, dict):
        return ""
    rows = _extract_mi_rows({"response": response})
    for row in rows:
        value = row.get("PLPN") or row.get("plpn") or row.get("PlannedOrder")
        if value:
            return str(value)
    for key in ("PLPN", "plpn", "PlannedOrder"):
        value = response.get(key)
        if value:
            return str(value)
    return ""


def _build_cms100_params(plpn: str, env: str | None = None) -> Dict[str, str]:
    params = {
        "QOPLPN": plpn,
        "QOPLPS": "0",
        "QOPLP2": "0",
    }
    return params


def _default_mi_cono_for_env(env: str | None) -> str:
    normalized = _normalize_env(env or DEFAULT_ENV)
    if normalized == "tst":
        return "881"
    if normalized == "prd":
        return "860"
    return ""


def _build_mos450_lstcomponent_params(
    motp: str,
    env: str | None = None,
) -> Dict[str, str]:
    # Fachvorgabe: MOS450MI/LstComponent immer mit MOTP prüfen (nicht HITN).
    cono = _default_mi_cono_for_env(env).strip()
    params: Dict[str, str] = {"MOTP": str(motp or "").strip()}
    if cono:
        params["CONO"] = cono
    return params


def _build_mos450_addcomponent_params(
    motp: str,
    mtrl: str,
    cfgl: str,
    env: str | None = None,
) -> Dict[str, str]:
    cono = _default_mi_cono_for_env(env).strip()
    params: Dict[str, str] = {
        "MOTP": str(motp or "").strip(),
        "MTRL": str(mtrl or "").strip(),
        "CFGL": str(cfgl or "").strip(),
    }
    if cono:
        params["CONO"] = cono
    return params


def _row_contains_item_number(row: Mapping[str, Any], item_number: str) -> bool:
    target = str(item_number or "").strip().upper()
    if not target:
        return False
    for key, raw_value in row.items():
        value = str(raw_value or "").strip().upper()
        if not value:
            continue
        key_upper = str(key or "").strip().upper()
        if value == target and (
            key_upper in {"ITNO", "CMIT", "CMITNO", "COMP", "COMPONENT", "CPIT", "MITN", "MTRL"}
            or "ITN" in key_upper
            or "COMP" in key_upper
            or "MTRL" in key_upper
        ):
            return True
    return False


def _mos450_allows_component(response: Any, candidate_itno: str) -> bool:
    rows = _extract_mi_rows({"response": response})
    if not rows:
        return False
    return any(_row_contains_item_number(row, candidate_itno) for row in rows)


def _mos450_component_status20(response: Any, component_itno: str) -> tuple[bool, str]:
    rows = _extract_mi_rows({"response": response})
    target = str(component_itno or "").strip().upper()
    if not rows or not target:
        return False, "Komponente nicht gefunden."

    seen_statuses: List[str] = []
    for row in rows:
        mtrl = str(row.get("MTRL") or row.get("mtrl") or "").strip().upper()
        if mtrl != target:
            continue
        stat = str(row.get("STAT") or row.get("stat") or "").strip()
        if stat == "20":
            return True, "OK: Komponente vorhanden mit Status 20."
        if stat:
            seen_statuses.append(stat)

    if seen_statuses:
        statuses = ", ".join(sorted(set(seen_statuses)))
        return False, f"Komponente gefunden, aber Status ist nicht 20 ({statuses})."
    return False, "Komponente nicht gefunden."


def _is_mos450_soft_validation_failure(message: str) -> bool:
    text = str(message or "").lower()
    return (
        "komponente nicht gefunden" in text
        or "status ist nicht 20" in text
        or ("komponente" in text and "motp" in text)
    )


def _build_ips_mos100_params(row: sqlite3.Row) -> Dict[str, str]:
    itno = _row_value(row, "ITNO")
    ser2 = _row_value(row, "SER2")
    new_itno = _row_value(row, "NEW_PART_ITNO") or itno
    new_ser2 = _row_value(row, "NEW_PART_SER2") or ser2
    return {
        "WorkOrderNumber": _row_value(row, "MWNO"),
        "Product": itno,
        "NewItemNumber": new_itno,
        "NewLotNumber": new_ser2,
    }


def _build_mos180_params(row: sqlite3.Row, env: str | None = None) -> Dict[str, str]:
    params = {
        "FACI": MOS180_FACI,
        "MWNO": _row_value(row, "MWNO"),
        "RESP": MOS180_RESP,
        "APRB": MOS180_APRB,
    }
    return params


def _build_mos050_params(row: sqlite3.Row) -> Dict[str, str]:
    product = _row_value(row, "ITNO")
    mwno = _row_value(row, "MWNO")
    return {
        "WHFACI": "100",
        "WHMWNO": mwno,
        "WHPRNO": product,
        "WHWHSL": MOS050_LOCATION,
        "WWRPDT": _format_ddmmyy(_row_value(row, "UMBAU_DATUM")),
        "WHBANO": "",
        "WLBREF": "",
    }


def _sanitize_yyyymmdd(value: Any) -> str:
    digits = "".join(ch for ch in str(value or "") if ch.isdigit())
    if len(digits) < 8:
        return ""
    return digits[:8]


def _sanitize_hhmmss(value: Any) -> str:
    digits = "".join(ch for ch in str(value or "") if ch.isdigit())
    if not digits:
        return ""
    return digits.zfill(6)[:6]


def _parse_delay_minutes(value: Any) -> int | None:
    raw = str(value or "").strip()
    if raw == "":
        return None
    try:
        minutes = int(raw)
    except (TypeError, ValueError):
        raise HTTPException(status_code=400, detail="Verzögerung muss eine ganze Zahl sein.") from None
    if minutes < 0:
        raise HTTPException(status_code=400, detail="Verzögerung darf nicht negativ sein.")
    return minutes


def _resolve_teilenummer_base_datetime(row: Mapping[str, Any]) -> datetime:
    rgdt = _sanitize_yyyymmdd(_row_value(row, "RGDT"))
    rgtm = _sanitize_hhmmss(_row_value(row, "RGTM")) or "000000"
    if rgdt:
        try:
            return datetime.strptime(f"{rgdt}{rgtm}", "%Y%m%d%H%M%S")
        except ValueError:
            pass
    return datetime.now()


def _compute_teilenummer_plan_times(
    row: Mapping[str, Any],
    out_delay_min: int | None,
    in_delay_min: int | None,
) -> Dict[str, str]:
    out_minutes = max(0, int(out_delay_min or 0))
    in_minutes = max(0, int(in_delay_min or 0))
    if out_minutes == 0 and in_minutes == 0:
        now_dt = datetime.now()
        out_dt = now_dt
        in_dt = now_dt + timedelta(minutes=1)
    else:
        base_dt = _resolve_teilenummer_base_datetime(row)
        out_dt = base_dt + timedelta(minutes=out_minutes)
        in_dt = out_dt + timedelta(minutes=in_minutes)
    return {
        "PLAN_OUT_DATE": out_dt.strftime("%Y%m%d"),
        "PLAN_OUT_TIME": out_dt.strftime("%H%M%S"),
        "PLAN_IN_DATE": in_dt.strftime("%Y%m%d"),
        "PLAN_IN_TIME": in_dt.strftime("%H%M%S"),
    }


def _resolve_teilenummer_umbau_datum(row: sqlite3.Row) -> str:
    return (
        _row_value(row, "PLAN_OUT_DATE")
        or _row_value(row, "RGDT")
        or _row_value(row, "TIMESTAMP")
        or _row_value(row, "A_BIRT")
        or datetime.utcnow().strftime("%Y%m%d")
    )


def _build_teilenummer_row_map(row: sqlite3.Row, new_itno: str, new_sern: str) -> Dict[str, str]:
    umbau_datum = _resolve_teilenummer_umbau_datum(row)
    parent_itno = _row_value(row, "C_MTRL")
    parent_sern = _row_value(row, "C_SERN")
    return {
        "CFGL": _row_value(row, "C_CFGL"),
        "ITNO": _row_value(row, "A_ITNO"),
        "SER2": _row_value(row, "A_SERN"),
        "MTRL": parent_itno,
        "SERN": parent_sern,
        "UMBAU_DATUM": umbau_datum,
        "NEW_PART_ITNO": new_itno,
        "NEW_PART_SER2": new_sern,
        "WAGEN_ITNO": parent_itno,
        "WAGEN_SERN": parent_sern,
        "OUT_DELAY_MIN": _row_value(row, "OUT_DELAY_MIN"),
        "IN_DELAY_MIN": _row_value(row, "IN_DELAY_MIN"),
        "PLAN_OUT_DATE": _row_value(row, "PLAN_OUT_DATE"),
        "PLAN_OUT_TIME": _row_value(row, "PLAN_OUT_TIME"),
        "PLAN_IN_DATE": _row_value(row, "PLAN_IN_DATE"),
        "PLAN_IN_TIME": _row_value(row, "PLAN_IN_TIME"),
        "RGDT": _row_value(row, "RGDT"),
        "RGTM": _row_value(row, "RGTM"),
    }


def _is_teilenummer_leading_object(row: Mapping[str, Any]) -> bool:
    parent_itno = _row_value(row, "MTRL", "C_MTRL")
    parent_sern = _row_value(row, "SERN", "C_SERN")
    itno = _row_value(row, "ITNO", "A_ITNO")
    sern = _row_value(row, "SER2", "A_SERN")
    if not parent_itno or not parent_sern:
        return True
    return parent_itno == itno and parent_sern == sern


def _teilenummer_log_context(row: sqlite3.Row, new_itno: str, new_sern: str) -> Dict[str, str]:
    return {
        "itno": _row_value(row, "A_ITNO"),
        "sern": _row_value(row, "A_SERN"),
        "new_itno": new_itno,
        "new_sern": new_sern,
    }


def _update_teilenummer_row(
    conn: sqlite3.Connection,
    table_name: str,
    seq: int,
    updates: Dict[str, Any],
) -> None:
    if not updates:
        return
    columns = ", ".join(f'"{key}"=?' for key in updates.keys())
    values = list(updates.values())
    values.append(seq)
    conn.execute(f'UPDATE "{table_name}" SET {columns} WHERE rowid=?', values)
    conn.commit()


def _build_mms240_params(new_itno: str, new_sern: str) -> Dict[str, str]:
    clean_serial = re.sub(r"[^0-9]", "", new_sern)
    return {
        "ITNO": new_itno,
        "SERN": new_sern,
        "FLNO": clean_serial,
    }


def _build_cusext_params(new_itno: str, new_sern: str) -> Dict[str, str]:
    return {
        "FILE": "MILOIN",
        "PK01": new_itno,
        "PK02": new_sern,
        "A130": "2 - DOT-Übertragung aktiv",
    }


def _extract_mwno(response: Any) -> str:
    if not isinstance(response, dict):
        return ""
    rows = _extract_mi_rows({"response": response})
    for row in rows:
        value = (
            row.get("QOMWNO")
            or row.get("qomwno")
            or row.get("MWNO")
            or row.get("mwno")
            or row.get("WorkOrderNumber")
        )
        if value:
            return str(value)
    for key in ("QOMWNO", "qomwno", "MWNO", "mwno", "WorkOrderNumber"):
        value = response.get(key)
        if value:
            return str(value)
    return ""


def _build_crs335_params(acrf: str, new_sern: str, new_baureihe: str) -> Dict[str, str]:
    return {
        "ACRF": acrf,
        "TX40": new_sern,
        "TX15": new_baureihe,
    }


def _build_sts046_params(whlo: str, geit: str, itno: str, bano: str) -> Dict[str, str]:
    return {
        "WHLO": whlo,
        "GEIT": geit,
        "ITNO": itno,
        "BANO": bano,
    }




def _ensure_wagon_data(table: str, env: str) -> str:
    env_table = _table_for(table, env)
    teilenummer_sql = _teilenummer_sql_file(env) if table == TEILENUMMER_TABLE else None
    if table == TEILENUMMER_TABLE and (teilenummer_sql and teilenummer_sql.exists()):
        sql_file = teilenummer_sql
    else:
        sql_file = _wagons_sql_file(env)

    try:
        with _connect(timeout=2.0, busy_timeout_ms=2000) as conn:
            if _table_exists(conn, env_table):
                if table == TEILENUMMER_TABLE:
                    columns = _columns_from_sql_file(teilenummer_sql or TEILENUMMER_SQL_FILE)
                    _ensure_columns(conn, env_table, columns + ["CHECKED"])
                    conn.commit()
                return env_table

            # Kein Auto-Reload mehr beim Öffnen der Seite:
            # Fehlt die Tabelle, wird nur ein leeres Schema angelegt.
            columns = _columns_from_sql_file(sql_file)
            if not columns:
                columns = ["INIT_PLACEHOLDER"]
            _create_table_from_columns(conn, env_table, columns)
            if table == TEILENUMMER_TABLE:
                _ensure_columns(conn, env_table, columns + ["CHECKED"])
                conn.commit()
    except sqlite3.OperationalError as exc:
        # Bei temporären Locks beim Start nicht hart abbrechen.
        if not _is_sqlite_locked_error(exc):
            raise

    try:
        _record_cache_status(
            table,
            env,
            "empty_init",
            connect_timeout=0.75,
            busy_timeout_ms=750,
            max_attempts=1,
            retry_sleep_base=0.0,
        )
    except sqlite3.OperationalError as exc:
        if not _is_sqlite_locked_error(exc):
            raise
    return env_table


def _create_teilenummer_tausch_table(
    conn: sqlite3.Connection,
    source_table: str,
    target_table: str,
) -> List[str]:
    columns = _table_columns(conn, source_table)
    if not columns:
        raise HTTPException(status_code=400, detail=f"Tabelle {source_table} hat keine Spalten.")
    extra_columns = [col for col in TEILENUMMER_TAUSCH_EXTRA_COLUMNS if col not in columns]
    all_columns = columns + extra_columns
    conn.execute(f'DROP TABLE IF EXISTS "{target_table}"')
    column_defs = ", ".join(f'"{col}" TEXT' for col in all_columns)
    conn.execute(f'CREATE TABLE "{target_table}" ({column_defs})')
    return all_columns


def _extract_mi_rows(payload: dict) -> List[Dict[str, Any]]:
    response = payload.get("response") or {}
    records = (
        response.get("MIRecord")
        or response.get("MIRecords")
        or response.get("MIResponse")
        or []
    )
    if not isinstance(records, list):
        records = [records]
    rows: List[Dict[str, Any]] = []
    for record in records:
        if not isinstance(record, dict):
            continue
        name_values = record.get("NameValue") or record.get("nameValue") or []
        if not isinstance(name_values, list):
            name_values = [name_values]
        row: Dict[str, Any] = {}
        for entry in name_values:
            if not isinstance(entry, dict):
                continue
            name = entry.get("Name") or entry.get("name")
            if not name:
                continue
            value = entry.get("Value") if entry.get("Value") is not None else entry.get("value")
            row[str(name)] = "" if value is None else value
        if row:
            rows.append(row)
    return rows


def _store_mi_rows(
    table_base: str,
    env: str,
    rows: List[Dict[str, Any]],
    wagon_itno: str | None = None,
    wagon_sern: str | None = None,
) -> str:
    base = _validate_table(table_base)
    table_name = _table_for(base, env)
    existing_records: List[Dict[str, Any]] = []
    if base == RENUMBER_WAGON_TABLE and rows:
        if wagon_itno or wagon_sern:
            for row in rows:
                if wagon_itno:
                    row["WAGEN_ITNO"] = wagon_itno
                if wagon_sern:
                    row["WAGEN_SERN"] = wagon_sern
        with _connect() as conn:
            if _table_exists(conn, table_name):
                _ensure_renumber_schema(conn, table_name)
                existing_records = [
                    dict(row)
                    for row in conn.execute(f'SELECT * FROM "{table_name}" ORDER BY rowid').fetchall()
                ]
        if existing_records:
            existing_map: Dict[str, List[Dict[str, Any]]] = {}
            for record in existing_records:
                key = _renumber_row_key(record)
                existing_map.setdefault(key, []).append(record)
            for row in rows:
                key = _renumber_row_key(row)
                candidates = existing_map.get(key)
                if not candidates:
                    continue
                existing = candidates.pop(0)
                if _is_ok_status(existing.get("OUT")):
                    row["OUT"] = existing.get("OUT") or ""
                    row["UPDATED_AT"] = existing.get("UPDATED_AT") or ""
                if existing.get("IN"):
                    row["IN"] = existing.get("IN") or ""
                    row["TIMESTAMP_IN"] = existing.get("TIMESTAMP_IN") or ""
                if existing.get("PLPN"):
                    row["PLPN"] = existing.get("PLPN") or ""
                if existing.get("MWNO"):
                    row["MWNO"] = existing.get("MWNO") or ""
                if existing.get("MOS100_STATUS"):
                    row["MOS100_STATUS"] = existing.get("MOS100_STATUS") or ""
                if existing.get("MOS180_STATUS"):
                    row["MOS180_STATUS"] = existing.get("MOS180_STATUS") or ""
                if existing.get("MOS050_STATUS"):
                    row["MOS050_STATUS"] = existing.get("MOS050_STATUS") or ""
                if existing.get("CRS335_STATUS"):
                    row["CRS335_STATUS"] = existing.get("CRS335_STATUS") or ""
                if existing.get("STS046_STATUS"):
                    row["STS046_STATUS"] = existing.get("STS046_STATUS") or ""
                if existing.get("STS046_ADD_STATUS"):
                    row["STS046_ADD_STATUS"] = existing.get("STS046_ADD_STATUS") or ""
                if existing.get("MMS240_STATUS"):
                    row["MMS240_STATUS"] = existing.get("MMS240_STATUS") or ""
                if existing.get("CUSEXT_STATUS"):
                    row["CUSEXT_STATUS"] = existing.get("CUSEXT_STATUS") or ""
                if existing.get("NEW_PART_ITNO"):
                    row["NEW_PART_ITNO"] = existing.get("NEW_PART_ITNO") or ""
                if existing.get("NEW_PART_SER2"):
                    row["NEW_PART_SER2"] = existing.get("NEW_PART_SER2") or ""
        for index, row in enumerate(rows, start=1):
            row["SEQ"] = str(index)
    columns: List[str] = []
    if rows:
        seen = set()
        for key in rows[0].keys():
            if key not in seen:
                columns.append(key)
                seen.add(key)
        for row in rows[1:]:
            for key in row.keys():
                if key not in seen:
                    columns.append(key)
                    seen.add(key)
    else:
        columns = ["EMPTY"]
    if base == RENUMBER_WAGON_TABLE:
        for extra in RENUMBER_EXTRA_COLUMNS:
            if extra not in columns:
                columns.append(extra)
    with _connect() as conn:
        conn.execute(f'DROP TABLE IF EXISTS "{table_name}"')
        column_list = ", ".join(f'"{col}" TEXT' for col in columns)
        conn.execute(f'CREATE TABLE "{table_name}" ({column_list})')
        if rows:
            placeholders = ", ".join("?" for _ in columns)
            column_names = ", ".join(f'"{c}"' for c in columns)
            insert_sql = f'INSERT INTO "{table_name}" ({column_names}) VALUES ({placeholders})'
            data = [[row.get(col, "") for col in columns] for row in rows]
            conn.executemany(insert_sql, data)
        conn.commit()
    return table_name


@app.get("/api/meta/targets")
def meta_targets(env: str = Query(DEFAULT_ENV)) -> dict:
    normalized = _normalize_env(env)
    rsrd_tables = _rsrd_tables(env)
    sqlite_url = f"sqlite:///{DB_PATH.resolve().as_posix()}"
    return {
        "env": normalized,
        "urls": {
            "compass": _safe_ionapi_url(env, "compass"),
            "mi": _safe_ionapi_url(env, "mi"),
            "rsrd_wsdl": _resolve_rsrd_wsdl(env),
            "sqlite": sqlite_url,
        },
        "tables": {
            "wagons": _table_for(DEFAULT_TABLE, env),
            "wagenumbau_wagons": _table_for(WAGENUMBAU_TABLE, env),
            "renumber_wagon": _table_for(RENUMBER_WAGON_TABLE, env),
            "spareparts": _table_for(SPAREPARTS_TABLE, env),
            "sparepart_swaps": _table_for(SPAREPARTS_SWAP_TABLE, env),
            "teilenummer": _table_for(TEILENUMMER_TABLE, env),
            "wagensuche": _table_for(WAGENSUCHE_TABLE, env),
            "rsrd_erp_numbers": _table_for(RSRD_ERP_TABLE, env),
            "rsrd_erp_full": _table_for(RSRD_ERP_FULL_TABLE, env),
            "rsrd_upload": _table_for(RSRD_UPLOAD_TABLE, env),
            "rsrd": {
                "wagons": rsrd_tables.wagons,
                "snapshots": rsrd_tables.snapshots,
                "json": rsrd_tables.json,
                "detail": rsrd_tables.detail,
            },
        },
    }


@app.get("/api/meta/cache_status")
def meta_cache_status(table: str = Query(TEILENUMMER_TABLE)) -> dict:
    base_table = _validate_table(table)
    return {
        "table": base_table,
        "prd": _cache_status_for(base_table, "prd"),
        "tst": _cache_status_for(base_table, "tst"),
    }


@app.get("/api/datalake-sync/datalake/tables")
def datalake_tables(
    env: str = Query("live"),
    autostart: bool = Query(False),
) -> dict:
    normalized = _normalize_env(env)
    # Refresh intentionally only starts via explicit POST /refresh.
    # GET must stay read-only so opening the page never starts a long-running job.
    _ = autostart
    return _datalake_snapshot(normalized)


@app.post("/api/datalake-sync/datalake/tables/refresh")
def datalake_tables_refresh(
    env: str = Query("live"),
    force: bool = Query(False),
) -> dict:
    normalized = _normalize_env(env)
    return _datalake_start_refresh(normalized, force=force)


@app.get("/api/datalake-sync/table-detail")
def datalake_table_detail(
    table: str = Query(..., min_length=1),
    env: str = Query("live"),
) -> dict:
    normalized = _normalize_env(env)
    return _datalake_table_detail(normalized, table)


@app.post("/api/datalake-sync/table-sync-selection")
def datalake_table_sync_selection(
    payload: dict = Body(...),
    env: str = Query("live"),
) -> dict:
    normalized = _normalize_env(env)
    table_name = str(payload.get("table_name") or "").strip().lower()
    enabled = bool(payload.get("enabled"))
    if not table_name or not _DATALAKE_TABLE_NAME_PATTERN.fullmatch(table_name):
        raise HTTPException(status_code=400, detail="Ungueltiger Tabellenname.")
    _datalake_set_sync_selection(normalized, table_name, enabled)
    return {
        "env": _datalake_env_label(normalized),
        "table_name": table_name,
        "sync_selected": enabled,
    }


@app.get("/api/wagons/count")
def wagons_count(
    table: str = DEFAULT_TABLE,
    env: str = Query(DEFAULT_ENV),
) -> dict:
    table_name = _table_for(table, env)
    normalized = _normalize_env(env)
    for _ in range(5):
        try:
            table_name = _ensure_wagon_data(table, env)
            with _connect(timeout=2.0, busy_timeout_ms=2000) as conn:
                total = conn.execute(f'SELECT COUNT(*) FROM "{table_name}"').fetchone()[0]
            return {"table": table_name, "total": total, "env": normalized}
        except sqlite3.OperationalError as exc:
            if _is_sqlite_locked_error(exc):
                time.sleep(0.25)
                continue
            if _is_sqlite_missing_table_error(exc):
                return {"table": table_name, "total": 0, "env": normalized}
            raise
    return {"table": table_name, "total": 0, "env": normalized}


@app.get("/api/wagons/chunk")
def wagons_chunk(
    offset: int = Query(0, ge=0),
    limit: int = Query(200, ge=1, le=1000),
    table: str = DEFAULT_TABLE,
    env: str = Query(DEFAULT_ENV),
) -> dict:
    table_name = _table_for(table, env)
    normalized = _normalize_env(env)
    for _ in range(5):
        try:
            table_name = _ensure_wagon_data(table, env)
            with _connect(timeout=2.0, busy_timeout_ms=2000) as conn:
                if table_name == _table_for(TEILENUMMER_TABLE, env):
                    cursor = conn.execute(
                        f'SELECT rowid AS "ROWID", * FROM "{table_name}" LIMIT ? OFFSET ?',
                        (limit, offset),
                    )
                else:
                    cursor = conn.execute(
                        f'SELECT * FROM "{table_name}" LIMIT ? OFFSET ?',
                        (limit, offset),
                    )
                rows = [dict(row) for row in cursor.fetchall()]
                total = conn.execute(f'SELECT COUNT(*) FROM "{table_name}"').fetchone()[0]
            return {
                "table": table_name,
                "rows": rows,
                "offset": offset,
                "limit": limit,
                "returned": len(rows),
                "total": total,
                "env": normalized,
            }
        except sqlite3.OperationalError as exc:
            if _is_sqlite_locked_error(exc):
                time.sleep(0.25)
                continue
            if _is_sqlite_missing_table_error(exc):
                return {
                    "table": table_name,
                    "rows": [],
                    "offset": offset,
                    "limit": limit,
                    "returned": 0,
                    "total": 0,
                    "env": normalized,
                }
            raise
    return {
        "table": table_name,
        "rows": [],
        "offset": offset,
        "limit": limit,
        "returned": 0,
        "total": 0,
        "env": normalized,
    }


@app.get("/api/wagons/exists")
def wagons_exists(
    sern: str = Query(..., min_length=1),
    table: str = DEFAULT_TABLE,
    env: str = Query(DEFAULT_ENV),
) -> dict:
    env_table = _table_for(table, env)
    template = None if table == DEFAULT_TABLE else _table_for(DEFAULT_TABLE, env)
    with _connect() as conn:
        table_name = _ensure_table(conn, env_table, template)
        row = conn.execute(
            f'SELECT 1 FROM "{table_name}" WHERE "SERIENNUMMER" = ? LIMIT 1',
            (sern,),
        ).fetchone()
    return {"table": table_name, "exists": row is not None, "env": _normalize_env(env)}


@app.get("/api/health")
def health() -> dict:
    return {"status": "ok"}


@app.get("/favicon.ico")
def favicon() -> Response:
    return Response(status_code=204)


def _run_compass_to_sqlite(
    sql_file: Path,
    table: str,
    env: str,
    timeout_seconds: int | None = None,
    mode: str = "replace",
) -> subprocess.CompletedProcess[str]:
    ionapi = _ionapi_path(env, "compass")
    normalized = _normalize_env(env)
    cmd = [
        sys.executable,
        str(PROJECT_ROOT / "python" / "compass_to_sqlite.py"),
        "--scheme",
        DEFAULT_SCHEME,
        "--sql-file",
        str(sql_file),
        "--table",
        table,
        "--sqlite-db",
        str(DB_PATH),
        "--mode",
        mode,
        "--ionapi",
        str(ionapi),
    ]
    if normalized == "tst" and TST_COMPASS_JDBC.exists():
        cmd.extend(["--jdbc-jar", str(TST_COMPASS_JDBC)])
    run_timeout = int(timeout_seconds) if timeout_seconds and timeout_seconds > 0 else None
    return subprocess.run(cmd, capture_output=True, text=True, timeout=run_timeout)


def _run_compass_query_internal(sql: str, env: str, timeout_seconds: int | None = None) -> Dict[str, Any]:
    ionapi = _ionapi_path(env, "compass")
    normalized = _normalize_env(env)
    cmd = [
        sys.executable,
        str(PROJECT_ROOT / "python" / "compass_query.py"),
        "--scheme",
        DEFAULT_SCHEME,
        "--sql",
        sql,
        "--output",
        "json",
        "--ionapi",
        str(ionapi),
    ]
    if DEFAULT_CATALOG:
        cmd.extend(["--catalog", DEFAULT_CATALOG])
    if DEFAULT_COLLECTION:
        cmd.extend(["--default-collection", DEFAULT_COLLECTION])
    if normalized == "tst" and TST_COMPASS_JDBC.exists():
        cmd.extend(["--jdbc-jar", str(TST_COMPASS_JDBC)])
    run_timeout = int(timeout_seconds) if timeout_seconds and timeout_seconds > 0 else None
    try:
        result = subprocess.run(cmd, capture_output=True, text=True, timeout=run_timeout)
    except subprocess.TimeoutExpired as exc:
        timeout_label = f"{run_timeout}s" if run_timeout else "unbekannt"
        raise TimeoutError(f"Compass-Query Timeout nach {timeout_label}") from exc
    if result.returncode != 0:
        raise RuntimeError(f"Compass-Query fehlgeschlagen: {result.stderr or result.stdout}")
    try:
        payload = json.loads(result.stdout)
    except json.JSONDecodeError as exc:
        raise RuntimeError(f"Ungültiges Compass-JSON: {exc}") from exc
    return payload.get("result") or {}


def _run_compass_query(sql: str, env: str) -> Dict[str, Any]:
    try:
        return _run_compass_query_internal(sql, env)
    except TimeoutError as exc:
        raise HTTPException(status_code=500, detail=str(exc)) from exc
    except RuntimeError as exc:
        raise HTTPException(status_code=500, detail=str(exc)) from exc


def _precheck_compass_connection(env: str, timeout_seconds: int = COMPASS_PRECHECK_TIMEOUT_SEC) -> None:
    probe_sql = "SELECT 1 AS PING"
    try:
        _run_compass_query_internal(probe_sql, env, timeout_seconds=timeout_seconds)
    except TimeoutError as exc:
        raise HTTPException(
            status_code=500,
            detail=f"Compass-Verbindungsprüfung Timeout nach {timeout_seconds}s.",
        ) from exc
    except RuntimeError as exc:
        raise HTTPException(
            status_code=500,
            detail=f"Compass-Verbindungsprüfung fehlgeschlagen: {exc}",
        ) from exc


def _datalake_env_label(normalized_env: str) -> str:
    return "live" if normalized_env == "prd" else "tst"


def _datalake_safe_value(value: Any) -> Any:
    if value is None or isinstance(value, (str, int, float, bool)):
        return value
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    iso = getattr(value, "isoformat", None)
    if callable(iso):
        try:
            return iso()
        except Exception:
            pass
    return str(value)


def _datalake_jdbc_path(normalized_env: str) -> Path:
    if normalized_env == "tst" and TST_COMPASS_JDBC.exists():
        return TST_COMPASS_JDBC
    jdbc_dir = CREDENTIALS_ROOT / "jdbc"
    preferred = sorted(jdbc_dir.glob("infor-compass-jdbc-*.jar"), reverse=True)
    if preferred:
        return preferred[0]
    fallback = sorted(jdbc_dir.glob("*.jar"))
    if fallback:
        return fallback[0]
    raise FileNotFoundError(f"Compass JDBC JAR nicht gefunden unter: {jdbc_dir}")


def _datalake_support_jars(jdbc_path: Path) -> List[str]:
    jars = [str(jdbc_path)]
    for extra in sorted(jdbc_path.parent.glob("slf4j-*.jar")):
        if extra.resolve() != jdbc_path.resolve():
            jars.append(str(extra))
    return jars


def _datalake_ensure_driver_ionapi(ionapi_path: Path, jdbc_path: Path) -> None:
    target = jdbc_path.parent / ionapi_path.name
    try:
        if ionapi_path.resolve() == target.resolve():
            return
    except OSError:
        pass
    if target.exists():
        return
    target.write_text(ionapi_path.read_text(encoding="utf-8-sig"), encoding="utf-8")


def _datalake_rows_to_dicts(description: Any, rows: List[Any]) -> List[Dict[str, Any]]:
    columns = [desc[0].strip("\"'") if isinstance(desc[0], str) else str(desc[0]) for desc in (description or [])]
    return [
        {col: _datalake_safe_value(val) for col, val in zip(columns, row)}
        for row in rows
    ]


def _datalake_choose_timestamp_column(columns: List[Dict[str, Any]]) -> str | None:
    by_name: Dict[str, str] = {}
    for col in columns:
        col_name = str(col.get("COLUMN_NAME") or "").strip()
        if not col_name:
            continue
        by_name[col_name.lower()] = col_name
    for preferred in _DATALAKE_TIMESTAMP_PREFERRED_COLUMNS:
        match = by_name.get(preferred)
        if match:
            return match
    for col in columns:
        col_name = str(col.get("COLUMN_NAME") or "").strip()
        if not col_name:
            continue
        data_type = str(col.get("DATA_TYPE") or "").lower()
        if any(token in data_type for token in _DATALAKE_TIMESTAMP_TYPE_HINTS):
            return col_name
    return None


def _datalake_quote_identifier(identifier: str) -> str:
    value = str(identifier or "").strip()
    if not value or not DATALAKE_SAFE_IDENTIFIER.match(value):
        raise ValueError(f"Ungueltiger SQL Identifier: {identifier}")
    return f'"{value}"'


def _fabric_quote_identifier(identifier: str) -> str:
    value = str(identifier or "").strip()
    if not value or not DATALAKE_SAFE_IDENTIFIER.match(value):
        raise ValueError(f"Ungueltiger SQL Identifier: {identifier}")
    return f"[{value}]"


def _datalake_connect_fabric_sql():
    try:
        mod = importlib.import_module("app")
    except Exception as exc:
        raise RuntimeError(f"Fabric-Modul konnte nicht geladen werden: {exc}") from exc
    connect_fn = getattr(mod, "_connect_fabric_sql", None)
    if not callable(connect_fn):
        raise RuntimeError("Fabric-SQL-Verbindung nicht verfügbar (_connect_fabric_sql fehlt).")
    return connect_fn()


def _normalize_timestamp_for_compare(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).strip()
    if not text:
        return ""
    try:
        parsed = datetime.fromisoformat(text.replace(" ", "T").replace("Z", ""))
        return parsed.strftime("%Y-%m-%d %H:%M:%S")
    except Exception:
        return text.lower()


def _datalake_normalize_timeout_tier(value: Any) -> str:
    tier = str(value or "").strip().lower()
    if tier not in _DATALAKE_TIMEOUT_TIERS:
        return "normal"
    return tier


def _datalake_timeout_seconds_for_tier(tier: str) -> int:
    normalized = _datalake_normalize_timeout_tier(tier)
    if normalized == "timeout":
        return _DATALAKE_TIMEOUT_RETRY_SECONDS
    if normalized == "superheavy":
        return _DATALAKE_TIMEOUT_SUPERHEAVY_SECONDS
    return _DATALAKE_TIMEOUT_DEFAULT_SECONDS


def _datalake_next_timeout_tier(tier: str) -> str:
    normalized = _datalake_normalize_timeout_tier(tier)
    if normalized == "normal":
        return "timeout"
    if normalized == "timeout":
        return "superheavy"
    return "superheavy"


def _is_timeout_error(exc: Any) -> bool:
    if isinstance(exc, TimeoutError):
        return True
    message = str(exc or "").lower()
    return (
        "timeout" in message
        or "time out" in message
        or "query timeout" in message
        or "statement timeout" in message
        or "hyt00" in message
    )


def _datalake_record_get(row: Dict[str, Any], key: str, default: Any = None) -> Any:
    if key in row:
        return row.get(key)
    key_lower = str(key).lower()
    for row_key, row_value in row.items():
        if str(row_key).lower() == key_lower:
            return row_value
    return default


def _datalake_compare_status(
    datalake_status: str,
    fabric_status: str,
    datalake_row_count: int | None,
    datalake_field_count: int | None,
    datalake_last_update: str | None,
    fabric_row_count: int | None,
    fabric_field_count: int | None,
    fabric_last_update: str | None,
) -> str:
    if datalake_status == "timeout" or fabric_status == "timeout":
        return "timeout"
    if datalake_status == "error" or fabric_status == "error":
        return "error"
    if fabric_status == "missing":
        return "missing"
    if datalake_status != "ok" or fabric_status != "ok":
        return "pending"
    same_counts = datalake_row_count == fabric_row_count
    same_fields = datalake_field_count == fabric_field_count
    same_update = (
        _normalize_timestamp_for_compare(datalake_last_update)
        == _normalize_timestamp_for_compare(fabric_last_update)
    )
    return "equal" if (same_counts and same_fields and same_update) else "diff"


def _datalake_normalize_data_type(value: Any) -> str:
    text = str(value or "").strip().lower()
    if not text:
        return ""
    text = re.sub(r"\s+", " ", text).strip()
    base = text.split("(", 1)[0].strip()
    if base.startswith("timestamp") or base.startswith("datetime"):
        return "timestamp"
    if base in {"character varying", "varchar", "nvarchar", "string"}:
        return "varchar"
    if base in {"character", "char", "nchar"}:
        return "char"
    if base in {"integer", "int", "int4", "signed", "bigint", "int8"}:
        return "int_family"
    if base in {"smallint", "int2", "tinyint"}:
        return "smallint"
    if base in {"decimal", "numeric", "number"}:
        return "numeric"
    if base in {"double precision", "double", "float8"}:
        return "double"
    if base in {"real", "float4"}:
        return "real"
    if base in {"boolean", "bool", "bit"}:
        return "boolean"
    return base


def _datalake_sync_selection_map(normalized_env: str) -> Dict[str, bool]:
    env_label = _datalake_env_label(normalized_env)
    try:
        with _connect() as conn:
            _ensure_datalake_cache_tables(conn)
            rows = conn.execute(
                f"""
                SELECT table_name, enabled
                FROM {DATALAKE_SYNC_SELECTION_TABLE}
                WHERE env = ?
                """,
                (env_label,),
            ).fetchall()
    except Exception:
        return {}
    result: Dict[str, bool] = {}
    for row in rows:
        table_name = str(row["table_name"] or "").strip().lower()
        if not table_name:
            continue
        enabled_raw = row["enabled"]
        enabled = str(enabled_raw or "").strip().lower() in {"1", "true", "yes", "y"}
        result[table_name] = enabled
    return result


def _datalake_set_sync_selection(normalized_env: str, table_name: str, enabled: bool) -> None:
    env_label = _datalake_env_label(normalized_env)
    now_utc = datetime.utcnow().isoformat() + "Z"
    with _connect() as conn:
        _ensure_datalake_cache_tables(conn)
        conn.execute(
            f"""
            INSERT INTO {DATALAKE_SYNC_SELECTION_TABLE} (env, table_name, enabled, updated_at_utc)
            VALUES (?, ?, ?, ?)
            ON CONFLICT(env, table_name) DO UPDATE SET
                enabled=excluded.enabled,
                updated_at_utc=excluded.updated_at_utc
            """,
            (env_label, table_name, 1 if enabled else 0, now_utc),
        )
        conn.commit()


def _datalake_table_detail(normalized_env: str, table_name: str) -> Dict[str, Any]:
    clean_table = str(table_name or "").strip().lower()
    if not clean_table or not _DATALAKE_TABLE_NAME_PATTERN.fullmatch(clean_table):
        raise HTTPException(status_code=400, detail="Ungueltiger Tabellenname.")

    snapshot = _datalake_snapshot(normalized_env)
    table_entry = next(
        (item for item in (snapshot.get("tables") or []) if str(item.get("table_name") or "").lower() == clean_table),
        None,
    )
    if table_entry is None:
        raise HTTPException(status_code=404, detail=f"Tabelle '{clean_table}' nicht im Snapshot gefunden.")

    sync_selection = _datalake_sync_selection_map(normalized_env)
    if str(table_entry.get("fabric_status") or "").strip().lower() == "missing":
        return {
            "env": _datalake_env_label(normalized_env),
            "table": table_entry,
            "sync_selected": bool(sync_selection.get(clean_table)),
            "fabric_exists": False,
            "fabric_error": None,
            "field_counts": {
                "datalake": int(table_entry.get("field_count") or 0),
                "fabric": 0,
            },
            "field_status_counts": {
                "missing": 0,
                "type_mismatch": 0,
                "match": 0,
            },
            "field_differences": [],
        }

    try:
        datalake_result = _run_compass_query_internal(
            "SELECT COLUMN_NAME, DATA_TYPE, ORDINAL_POSITION "
            "FROM INFORMATION_SCHEMA.COLUMNS "
            "WHERE TABLE_SCHEMA = 'default' "
            f"AND TABLE_NAME = '{clean_table}' "
            "ORDER BY ORDINAL_POSITION",
            normalized_env,
            timeout_seconds=20,
        )
        datalake_rows = datalake_result.get("rows") or []
    except Exception as exc:
        raise HTTPException(status_code=500, detail=f"DataLake-Spalten konnten nicht geladen werden: {exc}") from exc

    datalake_columns: List[Dict[str, Any]] = []
    for row in datalake_rows:
        col_name = str(_datalake_record_get(row, "COLUMN_NAME") or "").strip()
        if not col_name:
            continue
        datalake_columns.append(
            {
                "column_name": col_name,
                "data_type": str(_datalake_record_get(row, "DATA_TYPE") or "").strip(),
                "ordinal_position": int(_datalake_record_get(row, "ORDINAL_POSITION") or 0),
            }
        )

    fabric_columns: List[Dict[str, Any]] = []
    fabric_exists = False
    fabric_error = None
    fabric_conn = None
    try:
        fabric_conn = _datalake_connect_fabric_sql()
        fabric_cur = fabric_conn.cursor()
        try:
            setattr(fabric_cur, "timeout", 20)
        except Exception:
            pass
        fabric_cur.execute(
            "SELECT TABLE_NAME FROM INFORMATION_SCHEMA.TABLES "
            "WHERE TABLE_SCHEMA = 'landing' AND TABLE_TYPE = 'BASE TABLE' AND TABLE_NAME = ?",
            (clean_table,),
        )
        fabric_exists = fabric_cur.fetchone() is not None
        if fabric_exists:
            fabric_cur.execute(
                "SELECT COLUMN_NAME, DATA_TYPE, ORDINAL_POSITION "
                "FROM INFORMATION_SCHEMA.COLUMNS "
                "WHERE TABLE_SCHEMA = 'landing' AND TABLE_NAME = ? "
                "ORDER BY ORDINAL_POSITION",
                (clean_table,),
            )
            fabric_rows = _datalake_rows_to_dicts(fabric_cur.description, fabric_cur.fetchall())
            for row in fabric_rows:
                col_name = str(_datalake_record_get(row, "COLUMN_NAME") or "").strip()
                if not col_name:
                    continue
                fabric_columns.append(
                    {
                        "column_name": col_name,
                        "data_type": str(_datalake_record_get(row, "DATA_TYPE") or "").strip(),
                        "ordinal_position": int(_datalake_record_get(row, "ORDINAL_POSITION") or 0),
                    }
                )
    except Exception as exc:
        fabric_error = str(exc)
    finally:
        try:
            if fabric_conn is not None:
                fabric_conn.close()
        except Exception:
            pass

    field_differences: List[Dict[str, Any]] = []
    if not fabric_error:
        datalake_by_name = {str(col.get("column_name") or "").lower(): col for col in datalake_columns}
        fabric_by_name = {str(col.get("column_name") or "").lower(): col for col in fabric_columns}
        all_names = sorted(set(datalake_by_name) | set(fabric_by_name))
        for lower_name in all_names:
            dl_col = datalake_by_name.get(lower_name)
            fb_col = fabric_by_name.get(lower_name)
            if dl_col is None and fb_col is not None:
                field_differences.append(
                    {
                        "column_name": fb_col.get("column_name") or lower_name,
                        "datalake_type": None,
                        "fabric_type": fb_col.get("data_type"),
                        "status": "missing_in_datalake",
                    }
                )
                continue
            if fb_col is None and dl_col is not None:
                field_differences.append(
                    {
                        "column_name": dl_col.get("column_name") or lower_name,
                        "datalake_type": dl_col.get("data_type"),
                        "fabric_type": None,
                        "status": "missing_in_fabric",
                    }
                )
                continue
            dl_type = _datalake_normalize_data_type(dl_col.get("data_type")) if dl_col else ""
            fb_type = _datalake_normalize_data_type(fb_col.get("data_type")) if fb_col else ""
            field_differences.append(
                {
                    "column_name": (dl_col or fb_col or {}).get("column_name") or lower_name,
                    "datalake_type": dl_col.get("data_type") if dl_col else None,
                    "fabric_type": fb_col.get("data_type") if fb_col else None,
                    "status": "match" if dl_type == fb_type else "type_mismatch",
                }
            )

        priority = {
            "missing_in_fabric": 0,
            "missing_in_datalake": 0,
            "type_mismatch": 1,
            "match": 2,
        }
        field_differences.sort(
            key=lambda item: (
                priority.get(str(item.get("status") or "match"), 9),
                str(item.get("column_name") or "").lower(),
            )
        )

    missing_count = sum(
        1
        for item in field_differences
        if str(item.get("status") or "") in {"missing_in_fabric", "missing_in_datalake"}
    )
    mismatch_count = sum(1 for item in field_differences if str(item.get("status") or "") == "type_mismatch")
    match_count = sum(1 for item in field_differences if str(item.get("status") or "") == "match")

    return {
        "env": _datalake_env_label(normalized_env),
        "table": table_entry,
        "sync_selected": bool(sync_selection.get(clean_table)),
        "fabric_exists": fabric_exists,
        "fabric_error": fabric_error,
        "field_counts": {
            "datalake": len(datalake_columns),
            "fabric": len(fabric_columns),
        },
        "field_status_counts": {
            "missing": missing_count,
            "type_mismatch": mismatch_count,
            "match": match_count,
        },
        "field_differences": field_differences,
    }


def _ensure_datalake_cache_tables(conn: sqlite3.Connection) -> None:
    conn.execute(
        f"""
        CREATE TABLE IF NOT EXISTS {DATALAKE_CACHE_SNAPSHOT_TABLE} (
            env TEXT PRIMARY KEY,
            running INTEGER NOT NULL DEFAULT 0,
            current_table TEXT,
            phase TEXT,
            phase_detail TEXT,
            started_at_utc TEXT,
            finished_at_utc TEXT,
            last_error TEXT,
            total_tables INTEGER NOT NULL DEFAULT 0,
            completed_tables INTEGER NOT NULL DEFAULT 0,
            error_tables INTEGER NOT NULL DEFAULT 0,
            updated_at_utc TEXT NOT NULL
        )
        """
    )
    conn.execute(
        f"""
        CREATE TABLE IF NOT EXISTS {DATALAKE_CACHE_TABLES_TABLE} (
            env TEXT NOT NULL,
            table_name TEXT NOT NULL,
            row_count INTEGER,
            field_count INTEGER,
            last_update TEXT,
            timeout_tier TEXT NOT NULL DEFAULT 'normal',
            timeout_seconds INTEGER,
            timeout_hits INTEGER NOT NULL DEFAULT 0,
            datalake_status TEXT,
            datalake_error TEXT,
            fabric_row_count INTEGER,
            fabric_field_count INTEGER,
            fabric_last_update TEXT,
            fabric_status TEXT,
            fabric_error TEXT,
            compare_status TEXT,
            status TEXT NOT NULL,
            error TEXT,
            duration_ms INTEGER,
            updated_at_utc TEXT NOT NULL,
            PRIMARY KEY (env, table_name)
        )
        """
    )
    conn.execute(
        f"""
        CREATE INDEX IF NOT EXISTS idx_{DATALAKE_CACHE_TABLES_TABLE.lower()}_status
        ON {DATALAKE_CACHE_TABLES_TABLE}(env, status)
        """
    )
    conn.execute(
        f"""
        CREATE TABLE IF NOT EXISTS {DATALAKE_SYNC_SELECTION_TABLE} (
            env TEXT NOT NULL,
            table_name TEXT NOT NULL,
            enabled INTEGER NOT NULL DEFAULT 0,
            updated_at_utc TEXT NOT NULL,
            PRIMARY KEY (env, table_name)
        )
        """
    )
    conn.execute(
        f"""
        CREATE INDEX IF NOT EXISTS idx_{DATALAKE_SYNC_SELECTION_TABLE.lower()}_enabled
        ON {DATALAKE_SYNC_SELECTION_TABLE}(env, enabled)
        """
    )
    snapshot_cols = {
        str(row[1]).lower() for row in conn.execute(f"PRAGMA table_info({DATALAKE_CACHE_SNAPSHOT_TABLE})")
    }
    if "current_table" not in snapshot_cols:
        conn.execute(f"ALTER TABLE {DATALAKE_CACHE_SNAPSHOT_TABLE} ADD COLUMN current_table TEXT")
    if "phase" not in snapshot_cols:
        conn.execute(f"ALTER TABLE {DATALAKE_CACHE_SNAPSHOT_TABLE} ADD COLUMN phase TEXT")
    if "phase_detail" not in snapshot_cols:
        conn.execute(f"ALTER TABLE {DATALAKE_CACHE_SNAPSHOT_TABLE} ADD COLUMN phase_detail TEXT")
    table_cols = {
        str(row[1]).lower() for row in conn.execute(f"PRAGMA table_info({DATALAKE_CACHE_TABLES_TABLE})")
    }
    add_columns = {
        "datalake_status": "TEXT",
        "datalake_error": "TEXT",
        "fabric_row_count": "INTEGER",
        "fabric_field_count": "INTEGER",
        "fabric_last_update": "TEXT",
        "fabric_status": "TEXT",
        "fabric_error": "TEXT",
        "compare_status": "TEXT",
        "timeout_tier": "TEXT NOT NULL DEFAULT 'normal'",
        "timeout_seconds": "INTEGER",
        "timeout_hits": "INTEGER NOT NULL DEFAULT 0",
    }
    for col_name, col_type in add_columns.items():
        if col_name in table_cols:
            continue
        conn.execute(f"ALTER TABLE {DATALAKE_CACHE_TABLES_TABLE} ADD COLUMN {col_name} {col_type}")


def _datalake_state_copy(normalized_env: str) -> Dict[str, Any]:
    with _datalake_tables_lock:
        state = _datalake_tables_state[normalized_env]
        return {
            "running": bool(state.get("running")),
            "current_table": state.get("current_table"),
            "phase": state.get("phase"),
            "phase_detail": state.get("phase_detail"),
            "started_at_utc": state.get("started_at_utc"),
            "finished_at_utc": state.get("finished_at_utc"),
            "last_error": state.get("last_error"),
            "total_tables": int(state.get("total_tables") or 0),
            "completed_tables": int(state.get("completed_tables") or 0),
            "error_tables": int(state.get("error_tables") or 0),
            "tables": [dict(item) for item in (state.get("tables") or {}).values()],
        }


def _datalake_persist_state(normalized_env: str) -> None:
    payload = _datalake_state_copy(normalized_env)
    env_label = _datalake_env_label(normalized_env)
    now_utc = datetime.utcnow().isoformat() + "Z"
    try:
        with _connect() as conn:
            _ensure_datalake_cache_tables(conn)
            conn.execute(
                f"""
                INSERT INTO {DATALAKE_CACHE_SNAPSHOT_TABLE} (
                    env, running, current_table, phase, phase_detail,
                    started_at_utc, finished_at_utc, last_error,
                    total_tables, completed_tables, error_tables, updated_at_utc
                )
                VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                ON CONFLICT(env) DO UPDATE SET
                    running=excluded.running,
                    current_table=excluded.current_table,
                    phase=excluded.phase,
                    phase_detail=excluded.phase_detail,
                    started_at_utc=excluded.started_at_utc,
                    finished_at_utc=excluded.finished_at_utc,
                    last_error=excluded.last_error,
                    total_tables=excluded.total_tables,
                    completed_tables=excluded.completed_tables,
                    error_tables=excluded.error_tables,
                    updated_at_utc=excluded.updated_at_utc
                """,
                (
                    env_label,
                    1 if payload["running"] else 0,
                    str(payload.get("current_table") or "").strip() or None,
                    str(payload.get("phase") or "").strip() or None,
                    str(payload.get("phase_detail") or "").strip() or None,
                    payload["started_at_utc"],
                    payload["finished_at_utc"],
                    payload["last_error"],
                    payload["total_tables"],
                    payload["completed_tables"],
                    payload["error_tables"],
                    now_utc,
                ),
            )
            conn.execute(f"DELETE FROM {DATALAKE_CACHE_TABLES_TABLE} WHERE env = ?", (env_label,))
            if payload["tables"]:
                conn.executemany(
                    f"""
                    INSERT INTO {DATALAKE_CACHE_TABLES_TABLE} (
                        env, table_name, row_count, field_count, last_update,
                        timeout_tier, timeout_seconds, timeout_hits,
                        datalake_status, datalake_error,
                        fabric_row_count, fabric_field_count, fabric_last_update,
                        fabric_status, fabric_error, compare_status,
                        status, error, duration_ms, updated_at_utc
                    ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?, ?)
                    """,
                    [
                        (
                            env_label,
                            str(item.get("table_name") or ""),
                            item.get("row_count"),
                            item.get("field_count"),
                            item.get("last_update"),
                            _datalake_normalize_timeout_tier(item.get("timeout_tier")),
                            int(item.get("timeout_seconds") or 0) or _datalake_timeout_seconds_for_tier(
                                _datalake_normalize_timeout_tier(item.get("timeout_tier"))
                            ),
                            int(item.get("timeout_hits") or 0),
                            item.get("datalake_status"),
                            item.get("datalake_error"),
                            item.get("fabric_row_count"),
                            item.get("fabric_field_count"),
                            item.get("fabric_last_update"),
                            item.get("fabric_status"),
                            item.get("fabric_error"),
                            item.get("compare_status"),
                            str(item.get("status") or "pending"),
                            item.get("error"),
                            item.get("duration_ms"),
                            now_utc,
                        )
                        for item in payload["tables"]
                        if item.get("table_name")
                    ],
                )
            conn.commit()
    except Exception as exc:
        logging.getLogger("datalake-sync").warning("Persistiere DataLake-Snapshot fehlgeschlagen: %s", exc)


def _datalake_load_state_from_db(normalized_env: str) -> None:
    with _datalake_tables_lock:
        state = _datalake_tables_state[normalized_env]
        if state.get("running") or state.get("tables") or int(state.get("total_tables") or 0) > 0:
            return

    env_label = _datalake_env_label(normalized_env)
    try:
        with _connect() as conn:
            _ensure_datalake_cache_tables(conn)
            snapshot_row = conn.execute(
                f"""
                SELECT
                    running,
                    current_table,
                    phase,
                    phase_detail,
                    started_at_utc,
                    finished_at_utc,
                    last_error,
                    total_tables,
                    completed_tables,
                    error_tables
                FROM {DATALAKE_CACHE_SNAPSHOT_TABLE}
                WHERE env = ?
                """,
                (env_label,),
            ).fetchone()
            if snapshot_row is None:
                return
            table_rows = conn.execute(
                f"""
                SELECT
                    table_name,
                    row_count,
                    field_count,
                    last_update,
                    timeout_tier,
                    timeout_seconds,
                    timeout_hits,
                    datalake_status,
                    datalake_error,
                    fabric_row_count,
                    fabric_field_count,
                    fabric_last_update,
                    fabric_status,
                    fabric_error,
                    compare_status,
                    status,
                    error,
                    duration_ms
                FROM {DATALAKE_CACHE_TABLES_TABLE}
                WHERE env = ?
                ORDER BY table_name
                """,
                (env_label,),
            ).fetchall()
    except Exception as exc:
        logging.getLogger("datalake-sync").warning("Lade DataLake-Snapshot aus SQLite fehlgeschlagen: %s", exc)
        return

    loaded_tables = {
        str(row["table_name"]): {
            "table_name": str(row["table_name"]),
            "row_count": row["row_count"],
            "field_count": row["field_count"],
            "last_update": row["last_update"],
            "timeout_tier": _datalake_normalize_timeout_tier(row["timeout_tier"]),
            "timeout_seconds": int(row["timeout_seconds"] or 0)
            or _datalake_timeout_seconds_for_tier(row["timeout_tier"]),
            "timeout_hits": int(row["timeout_hits"] or 0),
            "datalake_status": row["datalake_status"],
            "datalake_error": row["datalake_error"],
            "fabric_row_count": row["fabric_row_count"],
            "fabric_field_count": row["fabric_field_count"],
            "fabric_last_update": row["fabric_last_update"],
            "fabric_status": row["fabric_status"],
            "fabric_error": row["fabric_error"],
            "compare_status": row["compare_status"],
            "status": row["status"] or "pending",
            "error": row["error"],
            "duration_ms": row["duration_ms"],
        }
        for row in table_rows
        if row["table_name"]
    }

    with _datalake_tables_lock:
        state = _datalake_tables_state[normalized_env]
        if state.get("running") or state.get("tables") or int(state.get("total_tables") or 0) > 0:
            return
        # Persisted "running" can be stale after process restarts; never restore active-running state.
        state["running"] = False
        state["current_table"] = None
        state["phase"] = snapshot_row["phase"]
        state["phase_detail"] = snapshot_row["phase_detail"]
        state["started_at_utc"] = snapshot_row["started_at_utc"]
        state["finished_at_utc"] = snapshot_row["finished_at_utc"]
        state["last_error"] = snapshot_row["last_error"]
        state["total_tables"] = int(snapshot_row["total_tables"] or len(loaded_tables))
        state["completed_tables"] = int(snapshot_row["completed_tables"] or 0)
        state["error_tables"] = int(snapshot_row["error_tables"] or 0)
        state["tables"] = loaded_tables
        _datalake_reconcile_idle_state(state)


def _datalake_snapshot(normalized_env: str) -> Dict[str, Any]:
    _datalake_load_state_from_db(normalized_env)
    sync_selection = _datalake_sync_selection_map(normalized_env)
    with _datalake_tables_lock:
        state = _datalake_tables_state[normalized_env]
        _datalake_reconcile_idle_state(state)
        table_items = []
        for item in sorted(state["tables"].values(), key=lambda table_item: table_item["table_name"]):
            row = dict(item)
            table_name = str(row.get("table_name") or "").lower()
            row["sync_selected"] = bool(sync_selection.get(table_name))
            table_items.append(row)
        return {
            "env": _datalake_env_label(normalized_env),
            "running": state["running"],
            "current_table": state.get("current_table"),
            "phase": state.get("phase"),
            "phase_detail": state.get("phase_detail"),
            "started_at_utc": state["started_at_utc"],
            "finished_at_utc": state["finished_at_utc"],
            "last_error": state["last_error"],
            "total_tables": state["total_tables"],
            "completed_tables": state["completed_tables"],
            "error_tables": state["error_tables"],
            "nightly_refresh_prepared": {
                "enabled": False,
                "time_local": _DATALAKE_NIGHTLY_PREPARED_TIME,
            },
            "tables": table_items,
        }


def _datalake_set_phase(
    normalized_env: str,
    phase: str | None,
    phase_detail: str | None = None,
    *,
    persist: bool = True,
) -> None:
    with _datalake_tables_lock:
        state = _datalake_tables_state[normalized_env]
        state["phase"] = str(phase or "").strip() or None
        state["phase_detail"] = str(phase_detail or "").strip() or None
    if persist:
        _datalake_persist_state(normalized_env)


def _datalake_reconcile_idle_state(state: Dict[str, Any]) -> None:
    if bool(state.get("running")):
        return
    state["current_table"] = None
    total_tables = int(state.get("total_tables") or 0)
    completed_tables = int(state.get("completed_tables") or 0)
    phase = str(state.get("phase") or "").strip().lower()
    if phase in {"queued", "init", "discover_datalake", "discover_fabric", "discover_columns", "processing"}:
        if total_tables > 0 and completed_tables >= total_tables:
            state["phase"] = "finished"
            state["phase_detail"] = "Aktualisierung abgeschlossen."
        elif total_tables > 0:
            state["phase"] = "idle"
            state["phase_detail"] = "Aktualisierung pausiert. Bitte 'Neu laden' klicken."
        else:
            state["phase"] = None
            state["phase_detail"] = None
    for table in (state.get("tables") or {}).values():
        status = str(table.get("status") or "").strip().lower()
        if status == "running":
            table["status"] = "queued"
        if str(table.get("datalake_status") or "").strip().lower() == "running":
            table["datalake_status"] = "pending"
        if str(table.get("fabric_status") or "").strip().lower() == "running":
            table["fabric_status"] = "pending"


def _datalake_refresh_worker(normalized_env: str) -> None:
    try:
        _datalake_set_phase(normalized_env, "init", "Verbindungen werden aufgebaut ...")
        try:
            import jaydebeapi  # type: ignore
        except Exception as exc:
            raise RuntimeError(f"jaydebeapi nicht verfügbar: {exc}") from exc
        ionapi_path = _ionapi_path(normalized_env, "compass")
        jdbc_path = _datalake_jdbc_path(normalized_env)
        _datalake_ensure_driver_ionapi(ionapi_path, jdbc_path)
        ion_cfg = json.loads(ionapi_path.read_text(encoding="utf-8-sig"))
        tenant = str(ion_cfg.get("ti") or "").strip()
        if not tenant:
            raise RuntimeError(f"Tenant 'ti' fehlt in {ionapi_path}")
        props = {
            "ION_API_CREDENTIALS": json.dumps(ion_cfg),
            "TENANT": tenant,
        }
        jdbc_url = f"jdbc:infordatalake://{tenant}"
        datalake_conn = jaydebeapi.connect(
            "com.infor.idl.jdbc.Driver",
            jdbc_url,
            props,
            jars=_datalake_support_jars(jdbc_path),
        )
        fabric_conn = None
        try:
            _datalake_set_phase(normalized_env, "discover_datalake", "DataLake Tabellen werden geladen ...")
            datalake_cur = datalake_conn.cursor()
            datalake_cur.execute(
                "SELECT TABLE_NAME FROM INFORMATION_SCHEMA.TABLES "
                "WHERE TABLE_SCHEMA = 'default' AND TABLE_TYPE = 'BASE TABLE' "
                "ORDER BY TABLE_NAME"
            )
            table_rows = _datalake_rows_to_dicts(datalake_cur.description, datalake_cur.fetchall())
            table_names = []
            for row in table_rows:
                raw_name = str(row.get("TABLE_NAME", "")).strip()
                if not raw_name:
                    continue
                if not _DATALAKE_TABLE_NAME_PATTERN.fullmatch(raw_name):
                    continue
                table_names.append(raw_name.lower())
            with _datalake_tables_lock:
                state = _datalake_tables_state[normalized_env]
                existing_tables = dict(state.get("tables") or {})
                state["current_table"] = None
                state["total_tables"] = len(table_names)
                state["tables"] = {
                    table_name: {
                        "timeout_tier": _datalake_normalize_timeout_tier(
                            existing_tables.get(table_name, {}).get("timeout_tier")
                        ),
                        "timeout_seconds": int(existing_tables.get(table_name, {}).get("timeout_seconds") or 0)
                        or _datalake_timeout_seconds_for_tier(
                            existing_tables.get(table_name, {}).get("timeout_tier")
                        ),
                        "timeout_hits": int(existing_tables.get(table_name, {}).get("timeout_hits") or 0),
                        "table_name": table_name,
                        "row_count": existing_tables.get(table_name, {}).get("row_count"),
                        "field_count": existing_tables.get(table_name, {}).get("field_count"),
                        "last_update": existing_tables.get(table_name, {}).get("last_update"),
                        "datalake_status": existing_tables.get(table_name, {}).get("datalake_status"),
                        "datalake_error": existing_tables.get(table_name, {}).get("datalake_error"),
                        "fabric_row_count": existing_tables.get(table_name, {}).get("fabric_row_count"),
                        "fabric_field_count": existing_tables.get(table_name, {}).get("fabric_field_count"),
                        "fabric_last_update": existing_tables.get(table_name, {}).get("fabric_last_update"),
                        "fabric_status": existing_tables.get(table_name, {}).get("fabric_status"),
                        "fabric_error": existing_tables.get(table_name, {}).get("fabric_error"),
                        "compare_status": existing_tables.get(table_name, {}).get("compare_status"),
                        "status": "pending",
                        "error": None,
                        "duration_ms": None,
                    }
                    for table_name in table_names
                }
            _datalake_persist_state(normalized_env)

            fabric_tables: set[str] = set()
            fabric_columns_by_table: Dict[str, List[Dict[str, Any]]] = {}
            fabric_global_error: str | None = None
            fabric_cur = None
            columns_by_table: Dict[str, List[Dict[str, Any]]] = {}
            table_metadata: Dict[str, Dict[str, Any]] = {}
            matching_tables: List[str] = []
            matching_set: set[str] = set()
            try:
                _datalake_set_phase(normalized_env, "discover_fabric", "Fabric Tabellen werden geladen ...")
                fabric_conn = _datalake_connect_fabric_sql()
                fabric_cur = fabric_conn.cursor()
                fabric_cur.execute(
                    "SELECT TABLE_NAME FROM INFORMATION_SCHEMA.TABLES "
                    "WHERE TABLE_SCHEMA = 'landing' AND TABLE_TYPE = 'BASE TABLE'"
                )
                fabric_table_rows = _datalake_rows_to_dicts(fabric_cur.description, fabric_cur.fetchall())
                fabric_tables = {
                    str(row.get("TABLE_NAME") or "").strip().lower()
                    for row in fabric_table_rows
                    if row.get("TABLE_NAME")
                }
                matching_tables = [table_name for table_name in table_names if table_name in fabric_tables]
                matching_set = set(matching_tables)
                if matching_tables:
                    _datalake_set_phase(
                        normalized_env,
                        "discover_columns",
                        f"Spaltenmetadaten werden geladen ({len(matching_tables)} Tabellen) ...",
                    )
                    placeholders = ", ".join("?" for _ in matching_tables)
                    datalake_cur.execute(
                        "SELECT TABLE_NAME, COLUMN_NAME, DATA_TYPE "
                        "FROM INFORMATION_SCHEMA.COLUMNS "
                        "WHERE TABLE_SCHEMA = 'default' "
                        f"AND TABLE_NAME IN ({placeholders}) "
                        "ORDER BY TABLE_NAME, ORDINAL_POSITION",
                        matching_tables,
                    )
                    datalake_column_rows = _datalake_rows_to_dicts(
                        datalake_cur.description, datalake_cur.fetchall()
                    )
                    for row in datalake_column_rows:
                        table_name = str(row.get("TABLE_NAME") or "").lower()
                        if not table_name:
                            continue
                        columns_by_table.setdefault(table_name, []).append(row)

                    fabric_cur.execute(
                        "SELECT TABLE_NAME, COLUMN_NAME, DATA_TYPE "
                        "FROM INFORMATION_SCHEMA.COLUMNS "
                        "WHERE TABLE_SCHEMA = 'landing' "
                        f"AND TABLE_NAME IN ({placeholders}) "
                        "ORDER BY TABLE_NAME, ORDINAL_POSITION",
                        matching_tables,
                    )
                    fabric_column_rows = _datalake_rows_to_dicts(
                        fabric_cur.description, fabric_cur.fetchall()
                    )
                    for row in fabric_column_rows:
                        table_name = str(row.get("TABLE_NAME") or "").strip().lower()
                        if not table_name:
                            continue
                        fabric_columns_by_table.setdefault(table_name, []).append(row)

                    for table_name in matching_tables:
                        cols = columns_by_table.get(table_name, [])
                        table_metadata[table_name] = {
                            "timestamp_col": _datalake_choose_timestamp_column(cols),
                        }
            except Exception as exc:
                fabric_global_error = str(exc)
                matching_tables = []
                matching_set = set()

            with _datalake_tables_lock:
                state = _datalake_tables_state[normalized_env]
                if fabric_global_error:
                    state["current_table"] = None
                    state["phase"] = "error"
                    state["phase_detail"] = "Fabric Metadaten konnten nicht geladen werden."
                    for table_name in table_names:
                        state["completed_tables"] += 1
                        state["error_tables"] += 1
                        state["tables"][table_name] = {
                            "timeout_tier": _datalake_normalize_timeout_tier(
                                state["tables"].get(table_name, {}).get("timeout_tier")
                            ),
                            "timeout_seconds": int(state["tables"].get(table_name, {}).get("timeout_seconds") or 0)
                            or _datalake_timeout_seconds_for_tier(
                                state["tables"].get(table_name, {}).get("timeout_tier")
                            ),
                            "timeout_hits": int(state["tables"].get(table_name, {}).get("timeout_hits") or 0),
                            "table_name": table_name,
                            "row_count": None,
                            "field_count": None,
                            "last_update": None,
                            "datalake_status": "skipped",
                            "datalake_error": None,
                            "fabric_row_count": None,
                            "fabric_field_count": None,
                            "fabric_last_update": None,
                            "fabric_status": "error",
                            "fabric_error": fabric_global_error,
                            "compare_status": "error",
                            "status": "error",
                            "error": f"FB: {fabric_global_error}",
                            "duration_ms": 0,
                        }
                else:
                    state["phase"] = "processing"
                    state["phase_detail"] = "Tabellenvergleich wird vorbereitet ..."
                    for table_name in table_names:
                        if table_name in matching_set:
                            state["tables"][table_name]["status"] = "queued"
                            state["tables"][table_name]["compare_status"] = "pending"
                            state["tables"][table_name]["datalake_status"] = "pending"
                            state["tables"][table_name]["fabric_status"] = "pending"
                            continue
                        state["completed_tables"] += 1
                        state["tables"][table_name] = {
                            "timeout_tier": _datalake_normalize_timeout_tier(
                                state["tables"].get(table_name, {}).get("timeout_tier")
                            ),
                            "timeout_seconds": int(state["tables"].get(table_name, {}).get("timeout_seconds") or 0)
                            or _datalake_timeout_seconds_for_tier(
                                state["tables"].get(table_name, {}).get("timeout_tier")
                            ),
                            "timeout_hits": int(state["tables"].get(table_name, {}).get("timeout_hits") or 0),
                            "table_name": table_name,
                            "row_count": None,
                            "field_count": None,
                            "last_update": None,
                            "datalake_status": "skipped",
                            "datalake_error": None,
                            "fabric_row_count": None,
                            "fabric_field_count": None,
                            "fabric_last_update": None,
                            "fabric_status": "missing",
                            "fabric_error": None,
                            "compare_status": "missing",
                            "status": "missing",
                            "error": None,
                            "duration_ms": 0,
                        }
            _datalake_persist_state(normalized_env)

            for table_name in matching_tables:
                with _datalake_tables_lock:
                    state = _datalake_tables_state[normalized_env]
                    state["current_table"] = table_name
                    state["phase"] = "processing"
                    state["phase_detail"] = f"Tabelle {table_name} wird geprüft ..."
                    item = dict((state.get("tables") or {}).get(table_name) or {})
                    item["table_name"] = table_name
                    item["status"] = "running"
                    item["datalake_status"] = "running"
                    item["fabric_status"] = "running"
                    item["compare_status"] = "pending"
                    item["error"] = None
                    item["duration_ms"] = None
                    state["tables"][table_name] = item
                _datalake_persist_state(normalized_env)

                started = time.perf_counter()
                persist_now = False
                with _datalake_tables_lock:
                    current_item = dict(
                        (_datalake_tables_state[normalized_env].get("tables") or {}).get(table_name) or {}
                    )
                timeout_tier = _datalake_normalize_timeout_tier(current_item.get("timeout_tier"))
                timeout_hits = int(current_item.get("timeout_hits") or 0)
                timeout_seconds = _datalake_timeout_seconds_for_tier(timeout_tier)
                datalake_status = "pending"
                datalake_error = None
                datalake_timed_out = False
                row_count: int | None = None
                field_count = len(columns_by_table.get(table_name, []))
                last_update = None
                try:
                    if not DATALAKE_SAFE_IDENTIFIER.match(table_name):
                        raise ValueError(f"Ungueltiger Tabellenname: {table_name}")
                    meta = table_metadata.get(table_name, {})
                    timestamp_col = str(meta.get("timestamp_col") or "").strip() or None
                    last_update_sql = (
                        f'MAX({_datalake_quote_identifier(timestamp_col)}) AS last_update'
                        if timestamp_col
                        else "NULL AS last_update"
                    )
                    table_sql = _datalake_quote_identifier(table_name)
                    datalake_result = _run_compass_query_internal(
                        "SELECT "
                        f"COUNT(*) AS row_count, {last_update_sql} "
                        f"FROM default.{table_sql}",
                        normalized_env,
                        timeout_seconds=timeout_seconds,
                    )
                    count_rows = datalake_result.get("rows") or []
                    stats = count_rows[0] if count_rows else {}
                    row_count_raw = _datalake_record_get(stats, "row_count", 0)
                    row_count = int(row_count_raw or 0)
                    last_update_raw = _datalake_record_get(stats, "last_update")
                    last_update = str(last_update_raw).strip() if last_update_raw is not None else None
                    if last_update == "":
                        last_update = None
                    datalake_status = "ok"
                except Exception as exc:
                    datalake_status = "timeout" if _is_timeout_error(exc) else "error"
                    datalake_error = str(exc)
                    datalake_timed_out = datalake_status == "timeout"

                fabric_row_count: int | None = None
                fabric_field_count: int | None = None
                fabric_last_update: str | None = None
                fabric_status = "pending"
                fabric_error = None
                fabric_timed_out = False
                fabric_cols = fabric_columns_by_table.get(table_name, [])
                fabric_field_count = len(fabric_cols)
                fabric_timestamp_col = _datalake_choose_timestamp_column(fabric_cols)
                try:
                    table_sql = _fabric_quote_identifier(table_name)
                    fabric_last_sql = (
                        f"MAX({_fabric_quote_identifier(fabric_timestamp_col)}) AS last_update"
                        if fabric_timestamp_col
                        else "NULL AS last_update"
                    )
                    if fabric_cur is None:
                        raise RuntimeError("Fabric Cursor nicht verfügbar.")
                    try:
                        setattr(fabric_cur, "timeout", int(timeout_seconds))
                    except Exception:
                        pass
                    fabric_cur.execute(
                        "SELECT "
                        f"COUNT_BIG(*) AS row_count, {fabric_last_sql} "
                        f"FROM landing.{table_sql}"
                    )
                    fabric_rows = _datalake_rows_to_dicts(fabric_cur.description, fabric_cur.fetchall())
                    fabric_stats = fabric_rows[0] if fabric_rows else {}
                    fabric_row_count = int(_datalake_record_get(fabric_stats, "row_count") or 0)
                    fabric_last_update_raw = _datalake_record_get(fabric_stats, "last_update")
                    fabric_last_update = (
                        str(fabric_last_update_raw).strip()
                        if fabric_last_update_raw is not None
                        else None
                    )
                    if fabric_last_update == "":
                        fabric_last_update = None
                    fabric_status = "ok"
                except Exception as exc:
                    fabric_status = "timeout" if _is_timeout_error(exc) else "error"
                    fabric_error = str(exc)
                    fabric_timed_out = fabric_status == "timeout"

                compare_status = _datalake_compare_status(
                    datalake_status=datalake_status,
                    fabric_status=fabric_status,
                    datalake_row_count=row_count,
                    datalake_field_count=int(field_count) if field_count is not None else None,
                    datalake_last_update=last_update,
                    fabric_row_count=fabric_row_count,
                    fabric_field_count=fabric_field_count,
                    fabric_last_update=fabric_last_update,
                )
                if compare_status == "timeout":
                    overall_status = "timeout"
                elif compare_status == "error":
                    overall_status = "error"
                elif compare_status == "equal":
                    overall_status = "ok"
                elif compare_status == "diff":
                    overall_status = "done"
                else:
                    overall_status = "pending"

                if datalake_timed_out or fabric_timed_out:
                    timeout_tier = _datalake_next_timeout_tier(timeout_tier)
                    timeout_hits += 1
                elif datalake_status == "ok" and fabric_status == "ok":
                    timeout_tier = "normal"
                timeout_seconds = _datalake_timeout_seconds_for_tier(timeout_tier)

                error_parts = []
                if datalake_error:
                    error_parts.append(f"DL: {datalake_error}")
                if fabric_error:
                    error_parts.append(f"FB: {fabric_error}")
                overall_error = " | ".join(error_parts) if error_parts else None
                duration_ms = int((time.perf_counter() - started) * 1000)
                with _datalake_tables_lock:
                    state = _datalake_tables_state[normalized_env]
                    state["current_table"] = None
                    state["phase"] = "processing"
                    state["phase_detail"] = (
                        f"{state['completed_tables'] + 1}/{state['total_tables']} Tabellen verarbeitet"
                    )
                    state["completed_tables"] += 1
                    if overall_status in {"error", "timeout"}:
                        state["error_tables"] += 1
                    state["tables"][table_name] = {
                        "timeout_tier": timeout_tier,
                        "timeout_seconds": timeout_seconds,
                        "timeout_hits": timeout_hits,
                        "table_name": table_name,
                        "row_count": row_count,
                        "field_count": field_count,
                        "last_update": last_update,
                        "datalake_status": datalake_status,
                        "datalake_error": datalake_error,
                        "fabric_row_count": fabric_row_count,
                        "fabric_field_count": fabric_field_count,
                        "fabric_last_update": fabric_last_update,
                        "fabric_status": fabric_status,
                        "fabric_error": fabric_error,
                        "compare_status": compare_status,
                        "status": overall_status,
                        "error": overall_error,
                        "duration_ms": duration_ms,
                    }
                    persist_now = (state["completed_tables"] % 10 == 0) or overall_status in {"error", "timeout"}
                if persist_now:
                    _datalake_persist_state(normalized_env)
        finally:
            try:
                if fabric_conn is not None:
                    fabric_conn.close()
            except Exception:
                pass
            datalake_conn.close()
    except Exception as exc:
        with _datalake_tables_lock:
            state = _datalake_tables_state[normalized_env]
            state["last_error"] = str(exc)
            state["current_table"] = None
            state["phase"] = "error"
            state["phase_detail"] = str(exc)
    finally:
        with _datalake_tables_lock:
            state = _datalake_tables_state[normalized_env]
            state["running"] = False
            state["current_table"] = None
            if state.get("phase") != "error":
                state["phase"] = "finished"
                state["phase_detail"] = "Aktualisierung abgeschlossen."
            state["finished_at_utc"] = datetime.utcnow().isoformat() + "Z"
        _datalake_persist_state(normalized_env)


def _datalake_start_refresh(normalized_env: str, force: bool = False) -> Dict[str, Any]:
    should_start = False
    with _datalake_tables_lock:
        state = _datalake_tables_state[normalized_env]
        if state["running"]:
            pass
        elif state["total_tables"] > 0 and not force:
            pass
        else:
            state["running"] = True
            state["current_table"] = None
            state["phase"] = "queued"
            state["phase_detail"] = "Job wurde gestartet ..."
            state["started_at_utc"] = datetime.utcnow().isoformat() + "Z"
            state["finished_at_utc"] = None
            state["last_error"] = None
            state["completed_tables"] = 0
            state["error_tables"] = 0
            existing_tables = state.get("tables") or {}
            if existing_tables:
                state["total_tables"] = len(existing_tables)
                for table_name, table in existing_tables.items():
                    table["table_name"] = table_name
                    tier = _datalake_normalize_timeout_tier(table.get("timeout_tier"))
                    table["timeout_tier"] = tier
                    table["timeout_seconds"] = int(table.get("timeout_seconds") or 0) or _datalake_timeout_seconds_for_tier(tier)
                    table["timeout_hits"] = int(table.get("timeout_hits") or 0)
                    table["status"] = "queued"
                    table["error"] = None
                    table["datalake_status"] = "pending"
                    table["datalake_error"] = None
                    table["fabric_status"] = "pending"
                    table["fabric_error"] = None
                    table["compare_status"] = "pending"
                    table["duration_ms"] = None
            else:
                state["total_tables"] = 0
                state["tables"] = {}
            should_start = True
    if should_start:
        _datalake_persist_state(normalized_env)
        threading.Thread(
            target=_datalake_refresh_worker,
            args=(normalized_env,),
            daemon=True,
        ).start()
    return _datalake_snapshot(normalized_env)


def _fetch_msy_text(txid: str, env: str) -> str:
    txid_value = re.sub(r"\\D", "", str(txid))
    if not txid_value:
        return ""
    cached = _msy_text_cache.get(txid_value)
    if cached is not None:
        return cached
    sql = f"""
        SELECT TX60, LINO
        FROM (
            SELECT
                TX60,
                LINO,
                ROW_NUMBER() OVER (PARTITION BY LINO ORDER BY Timestamp DESC) AS rn
            FROM MSYTXL
            WHERE TXID = {txid_value}
              AND LINO BETWEEN 1 AND 5
        ) t
        WHERE rn = 1
        ORDER BY LINO
    """
    result = _run_compass_query(sql, env)
    rows = result.get("rows") or []
    parts = [str(row.get("TX60") or "") for row in rows]
    text = "".join(parts)
    _msy_text_cache[txid_value] = text
    return text


def _fetch_wg_tsi_text(sern: str, env: str) -> str:
    digits = re.sub(r"\\D", "", str(sern))
    if not digits:
        return ""
    cached = _wg_tsi_text_cache.get(digits)
    if cached is not None:
        return cached
    sql = f"""
        WITH Tx AS (
            SELECT A.TXID, 1 AS prio, A.ATNR
            FROM MIATTR A
            WHERE A.ATID = 'WG-TSI_ZUS_ZERT'
              AND REPLACE(REPLACE(CAST(A.BANO AS VARCHAR(100)), ' ', ''), '-', '') = '{digits}'
            UNION ALL
            SELECT A.TXID, 2 AS prio, A.ATNR
            FROM MIATTR A
            JOIN MROUHI M
              ON A.ITNO = M.ITNO
             AND A.BANO = M.SERN
             AND M.REDN = 0
             AND M.REMD = 0
            WHERE A.ATID = 'WG-TSI_ZUS_ZERT'
              AND REPLACE(REPLACE(CAST(M.HISN AS VARCHAR(100)), ' ', ''), '-', '') = '{digits}'
        ),
        Picked AS (
            SELECT TXID
            FROM Tx
            ORDER BY prio, ATNR DESC
            LIMIT 1
        ),
        Lines AS (
            SELECT
                X.TX60,
                X.LINO,
                ROW_NUMBER() OVER (PARTITION BY X.LINO ORDER BY X.Timestamp DESC) AS rn
            FROM MSYTXL X
            JOIN Picked P ON X.TXID = P.TXID
            WHERE X.LINO BETWEEN 1 AND 5
        )
        SELECT TX60, LINO
        FROM Lines
        WHERE rn = 1
        ORDER BY LINO
    """
    result = _run_compass_query(sql, env)
    rows = result.get("rows") or []
    parts = [str(row.get("TX60") or "") for row in rows]
    text = "".join(parts)
    _wg_tsi_text_cache[digits] = text
    return text


def _fetch_wg_tsi_txid(sern: str, env: str) -> str:
    digits = re.sub(r"\D", "", str(sern))
    if not digits:
        return ""
    cached = _wg_tsi_txid_cache.get(digits)
    if cached is not None:
        return cached
    sql = f"""
        SELECT A.TXID
        FROM MIATTR A
        WHERE A.ATID = 'WG-TSI_ZUS_ZERT'
          AND REPLACE(REPLACE(CAST(A.BANO AS VARCHAR(100)), ' ', ''), '-', '') = '{digits}'
        ORDER BY A.ATNR DESC
        LIMIT 1
    """
    result = _run_compass_query(sql, env)
    rows = result.get("rows") or []
    if not rows:
        # fallback: try MROUHI mapping
        sql2 = f"""
            SELECT A.TXID
            FROM MIATTR A
            JOIN MROUHI M
              ON A.ITNO = M.ITNO
             AND A.BANO = M.SERN
             AND M.REDN = 0
             AND M.REMD = 0
            WHERE A.ATID = 'WG-TSI_ZUS_ZERT'
              AND REPLACE(REPLACE(CAST(M.HISN AS VARCHAR(100)), ' ', ''), '-', '') = '{digits}'
            ORDER BY A.ATNR DESC
            LIMIT 1
        """
        result2 = _run_compass_query(sql2, env)
        rows = result2.get("rows") or []
        if not rows:
            return ""
    txid = str(rows[0].get("TXID") or "")
    _wg_tsi_txid_cache[digits] = txid
    return txid


def _build_load_erp_cmd(env: str) -> List[str]:
    ionapi = _ionapi_path(env, "compass")
    table_name = _table_for(RSRD_ERP_TABLE, env)
    return [
        sys.executable,
        str(PROJECT_ROOT / "python" / "load_erp_wagons.py"),
        "--scheme",
        DEFAULT_SCHEME,
        "--sqlite-db",
        str(DB_PATH),
        "--table",
        table_name,
        "--ionapi",
        str(ionapi),
    ]


def _build_erp_full_cmd(env: str) -> List[str]:
    if not RSRD_ERP_SQL_FILE.exists():
        raise FileNotFoundError(f"SQL-Datei nicht gefunden: {RSRD_ERP_SQL_FILE}")
    ionapi = _ionapi_path(env, "compass")
    table_name = _table_for(RSRD_ERP_FULL_TABLE, env)
    return [
        sys.executable,
        str(PROJECT_ROOT / "python" / "compass_to_sqlite.py"),
        "--scheme",
        DEFAULT_SCHEME,
        "--sql-file",
        str(RSRD_ERP_SQL_FILE),
        "--table",
        table_name,
        "--sqlite-db",
        str(DB_PATH),
        "--mode",
        "replace",
        "--ionapi",
        str(ionapi),
    ]


def _build_teilenummer_reload_cmd(env: str, sqlite_db_path: Path | None = None) -> List[str]:
    sql_file = _teilenummer_sql_file(env)
    if not sql_file.exists():
        raise FileNotFoundError(f"SQL-Datei nicht gefunden: {sql_file}")
    ionapi = _ionapi_path(env, "compass")
    table_name = _table_for(TEILENUMMER_TABLE, env)
    target_db = sqlite_db_path or DB_PATH
    cmd = [
        sys.executable,
        str(PROJECT_ROOT / "python" / "compass_to_sqlite.py"),
        "--scheme",
        DEFAULT_SCHEME,
        "--sql-file",
        str(sql_file),
        "--table",
        table_name,
        "--sqlite-db",
        str(target_db),
        "--mode",
        "truncate",
        "--ionapi",
        str(ionapi),
    ]
    if _normalize_env(env) == "tst" and TST_COMPASS_JDBC.exists():
        cmd.extend(["--jdbc-jar", str(TST_COMPASS_JDBC)])
    return cmd


def _create_job(job_type: str, env: str) -> Dict[str, Any]:
    job_id = uuid.uuid4().hex
    job = {
        "id": job_id,
        "type": job_type,
        "env": _normalize_env(env),
        "status": "running",
        "logs": [],
        "result": None,
        "error": None,
        "started": datetime.utcnow().isoformat(),
        "finished": None,
    }
    with _jobs_lock:
        _jobs[job_id] = job
    return job


def _find_running_job(job_type: str, env: str) -> Dict[str, Any] | None:
    normalized_env = _normalize_env(env)
    with _jobs_lock:
        now = datetime.utcnow()
        for job in _jobs.values():
            if (
                job.get("type") == job_type
                and job.get("env") == normalized_env
                and job.get("status") == "running"
            ):
                started_raw = str(job.get("started") or "")
                stale_limit = TEILENUMMER_RELOAD_TIMEOUT_SEC + JOB_STALE_MARGIN_SEC
                if started_raw:
                    try:
                        started_at = datetime.fromisoformat(started_raw)
                        if (now - started_at).total_seconds() > stale_limit:
                            job["status"] = "error"
                            job["error"] = "Job stale (Timeout überschritten)."
                            job["finished"] = now.isoformat()
                            continue
                    except Exception:
                        pass
                return dict(job)
    return None


def _append_job_log(job_id: str, message: str) -> None:
    if not message:
        return
    with _jobs_lock:
        job = _jobs.get(job_id)
        if not job:
            return
        logs = job.setdefault("logs", [])
        logs.append(message)
        if len(logs) > JOB_LOG_LIMIT:
            del logs[: len(logs) - JOB_LOG_LIMIT]


def _finish_job(job_id: str, status: str, result: Dict[str, Any] | None = None, error: str | None = None) -> None:
    with _jobs_lock:
        job = _jobs.get(job_id)
        if not job:
            return
        job["status"] = status
        job["result"] = result
        job["error"] = error
        job["finished"] = datetime.utcnow().isoformat()


def _job_snapshot(job_id: str) -> Dict[str, Any]:
    with _jobs_lock:
        job = _jobs.get(job_id)
        if not job:
            raise HTTPException(status_code=404, detail="Job nicht gefunden.")
        snapshot = dict(job)
        snapshot["logs"] = list(job.get("logs", []))
    return snapshot


def _goldenview_safe_name(name: str) -> str:
    cleaned = re.sub(r"[^a-zA-Z0-9_-]+", "_", name.strip()) or "query"
    return cleaned.strip("_")


def _goldenview_write_excel(path: Path, columns: List[str], rows: List[List[Any]]) -> None:
    if rows and isinstance(rows[0], dict):
        rows = [[row.get(col) for col in columns] for row in rows]  # type: ignore[index]
    wb = Workbook(write_only=True)
    ws = wb.create_sheet(title="Data")
    ws.append(columns)
    for row in rows:
        ws.append([None if v is None else v for v in row])
    path.parent.mkdir(parents=True, exist_ok=True)
    wb.save(path)


def _goldenview_write_md(path: Path, query: dict, fields: List[dict]) -> None:
    lines = [
        f"# {query.get('name') or 'SQL'}",
        "",
        f"Erstellt am: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S UTC')}",
        "",
    ]
    if query.get("description"):
        lines.extend([query["description"], ""])
    lines.extend(["## SQL", "", "```sql", query.get("sql_text", ""), "```", ""])
    if fields:
        lines.append("## Felder")
        lines.append("")
        lines.append("| Feld | Beschreibung | Verbundene Felder |")
        lines.append("| --- | --- | --- |")
        for field in fields:
            connected = ", ".join(field.get("connected_fields") or [])
            desc = field.get("field_description") or ""
            lines.append(f"| {field.get('field_name')} | {desc} | {connected} |")
        lines.append("")
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_text("\n".join(lines), encoding="utf-8")


def _goldenview_write_latest_readme(latest_dir: Path) -> None:
    files = sorted([p for p in latest_dir.glob("*") if p.is_file() and p.name != "README.md"])
    lines = [
        "# Latest Exports",
        "",
        f"Aktualisiert: {datetime.utcnow().strftime('%Y-%m-%d %H:%M:%S UTC')}",
        "",
    ]
    if not files:
        lines.append("Keine Dateien vorhanden.")
    else:
        for file in files:
            lines.append(f"- {file.name}")
    (latest_dir / "README.md").write_text("\n".join(lines), encoding="utf-8")


def _goldenview_job(query_id: int, job_id: str) -> None:
    try:
        _append_job_log(job_id, "Lade SQL aus SQLite ...")
        with _connect() as conn:
            _init_goldenview_db(conn)
            query = conn.execute(
                f"""
                SELECT id, name, sql_text, description
                FROM {GOLDENVIEW_QUERIES_TABLE}
                WHERE id = ?
                """,
                (query_id,),
            ).fetchone()
            if not query:
                raise ValueError("Eintrag nicht gefunden.")
            fields = conn.execute(
                f"""
                SELECT field_name, field_description, connected_fields
                FROM {GOLDENVIEW_FIELDS_TABLE}
                WHERE query_id = ?
                ORDER BY id ASC
                """,
                (query_id,),
            ).fetchall()
        query_dict = dict(query)
        field_list = [
            {
                "field_name": row["field_name"],
                "field_description": row["field_description"] or "",
                "connected_fields": json.loads(row["connected_fields"] or "[]"),
            }
            for row in fields
        ]
        _append_job_log(job_id, "Führe SQL gegen M3 aus ...")
        result = _run_compass_query(query_dict["sql_text"], "prd")
        columns = result.get("columns") or []
        rows = result.get("rows") or []
        safe_name = _goldenview_safe_name(query_dict.get("name") or f"sql_{query_id}")
        stamp = datetime.utcnow().strftime("%Y%m%d_%H%M%S")
        day_folder = datetime.utcnow().strftime("%Y-%m-%d")
        base_exports = GOLDENVIEW_EXPORT_DIR / "exports"
        latest_dir = base_exports / "latest"
        archive_dir = base_exports / "archive" / day_folder
        archive_excel = archive_dir / f"{safe_name}_{stamp}.xlsx"
        archive_md = archive_dir / f"{safe_name}_{stamp}.md"
        latest_excel = latest_dir / f"{safe_name}.xlsx"
        latest_md = latest_dir / f"{safe_name}.md"
        _append_job_log(job_id, f"Schreibe Excel ({len(rows)} Zeilen) ...")
        _goldenview_write_excel(archive_excel, columns, rows)
        _append_job_log(job_id, "Schreibe Markdown ...")
        _goldenview_write_md(archive_md, query_dict, field_list)
        latest_dir.mkdir(parents=True, exist_ok=True)
        shutil.copy2(archive_excel, latest_excel)
        shutil.copy2(archive_md, latest_md)
        _goldenview_write_latest_readme(latest_dir)
        with _connect() as conn:
            _init_goldenview_db(conn)
            conn.execute(
                f"""
                UPDATE {GOLDENVIEW_QUERIES_TABLE}
                SET excel_path = ?, md_path = ?, generated_at = datetime('now')
                WHERE id = ?
                """,
                (str(latest_excel), str(latest_md), query_id),
            )
            conn.commit()
        _finish_job(job_id, "success", result={"excel": str(latest_excel), "md": str(latest_md)})
    except Exception as exc:  # noqa: BLE001
        _append_job_log(job_id, f"Fehler: {exc}")
        _finish_job(job_id, "error", error=str(exc))


def _start_subprocess_job(
    job_type: str,
    cmd: List[str],
    env: str,
    finalize_fn,
    max_runtime_sec: int | None = None,
    prepare_fn=None,
) -> Dict[str, Any]:
    job = _create_job(job_type, env)

    def runner() -> None:
        process: subprocess.Popen[str] | None = None
        run_cmd = cmd
        run_finalize_fn = finalize_fn
        exclusive_reload_gate = job_type in {"reload_teilenummer"}
        gate_owner_id = threading.get_ident()
        if exclusive_reload_gate:
            with _reload_merge_state_lock:
                global _reload_merge_thread_id
                _reload_merge_thread_id = gate_owner_id
            _append_job_log(job["id"], "Exklusiver SQLite-Reload aktiv ...")
        try:
            if prepare_fn is not None:
                try:
                    prepared = prepare_fn(job["id"])
                    if prepared:
                        run_cmd, run_finalize_fn = prepared
                except Exception as exc:  # noqa: BLE001
                    _append_job_log(job["id"], f"Vorbereitung fehlgeschlagen: {exc}")
                    _finish_job(job["id"], "error", error=str(exc))
                    return
            try:
                process = subprocess.Popen(
                    run_cmd,
                    stdout=subprocess.PIPE,
                    stderr=subprocess.STDOUT,
                    text=True,
                )
            except Exception as exc:  # noqa: BLE001
                _append_job_log(job["id"], f"Start fehlgeschlagen: {exc}")
                _finish_job(job["id"], "error", error=str(exc))
                return

            assert process.stdout is not None
            reader_error: Exception | None = None
            keep_progress_lines = job_type in {"reload_teilenummer"}

            def consume_stdout() -> None:
                nonlocal reader_error
                try:
                    for line in process.stdout:
                        text = line.strip()
                        if not text or (PROGRESS_LINE.match(text) and not keep_progress_lines):
                            continue
                        _append_job_log(job["id"], text)
                except Exception as exc:  # noqa: BLE001
                    reader_error = exc

            reader = threading.Thread(target=consume_stdout, daemon=True)
            reader.start()

            try:
                if max_runtime_sec and max_runtime_sec > 0:
                    returncode = process.wait(timeout=max_runtime_sec)
                else:
                    returncode = process.wait()
                reader.join(timeout=2.0)
                if reader_error is not None:
                    raise reader_error
                if returncode != 0:
                    message = f"Prozess endete mit Code {returncode}"
                    _append_job_log(job["id"], message)
                    _finish_job(job["id"], "error", error=message)
                    return
                result = run_finalize_fn(job["id"])
                _finish_job(job["id"], "success", result=result)
            except subprocess.TimeoutExpired:
                try:
                    process.kill()
                    process.wait(timeout=5)
                except Exception:
                    pass
                timeout_label = int(max_runtime_sec) if max_runtime_sec else 0
                message = f"Prozess-Timeout nach {timeout_label}s"
                _append_job_log(job["id"], message)
                _finish_job(job["id"], "error", error=message)
            except Exception as exc:  # noqa: BLE001
                _append_job_log(job["id"], f"Fehler: {exc}")
                _finish_job(job["id"], "error", error=str(exc))
            finally:
                try:
                    if process is not None and process.stdout is not None:
                        process.stdout.close()
                except Exception:
                    pass
        finally:
            if exclusive_reload_gate:
                with _reload_merge_state_lock:
                    if _reload_merge_thread_id == gate_owner_id:
                        _reload_merge_thread_id = None

    threading.Thread(target=runner, daemon=True).start()
    return job


def _finalize_load_erp(job_id: str, env: str) -> Dict[str, Any]:
    with _connect() as conn:
        numbers_table = _ensure_table(conn, _table_for(RSRD_ERP_TABLE, env), None)
        count_wagons = conn.execute(f"SELECT COUNT(*) FROM {numbers_table}").fetchone()[0]
    message = f"ERP-Wagennummern geladen: {count_wagons}."
    _append_job_log(job_id, message)
    return {"count_wagons": count_wagons}


def _finalize_load_erp_full(job_id: str, env: str) -> Dict[str, Any]:
    with _connect() as conn:
        full_table = _ensure_table(conn, _table_for(RSRD_ERP_FULL_TABLE, env), None)
        count_full = conn.execute(f"SELECT COUNT(*) FROM {full_table}").fetchone()[0]
    message = f"ERP-Wagenattribute geladen: {count_full}."
    _append_job_log(job_id, message)
    return {"count_full": count_full}


def _finalize_teilenummer_reload(job_id: str, env: str) -> Dict[str, Any]:
    table_name = _table_for(TEILENUMMER_TABLE, env)
    with _connect() as conn:
        _ensure_columns(conn, table_name, ["CHECKED"])
        conn.execute(f'UPDATE "{table_name}" SET "CHECKED" = ""')
        count_rows = int(conn.execute(f'SELECT COUNT(*) FROM "{table_name}"').fetchone()[0] or 0)
        conn.commit()
    _record_cache_status(TEILENUMMER_TABLE, env, "manual_reload", row_count=count_rows)
    _append_job_log(job_id, f"Reload abgeschlossen: {count_rows} Datensätze.")
    return {"rows": count_rows}


def _prepare_teilenummer_work_db(env: str) -> Path:
    normalized = _normalize_env(env)
    tmp_root = Path(os.getenv("MFDAPPS_TMP_ROOT", "/tmp")).expanduser()
    work_dir = tmp_root / "mfdapps-reload"
    work_dir.mkdir(parents=True, exist_ok=True)
    now = time.time()
    for stale in work_dir.glob(f"{DB_PATH.stem}_reload_{normalized}_*.db"):
        try:
            if now - stale.stat().st_mtime > 12 * 3600:
                stale.unlink()
        except OSError:
            continue
    work_db = work_dir / f"{DB_PATH.stem}_reload_{normalized}_{uuid.uuid4().hex[:8]}.db"
    if work_db.exists():
        work_db.unlink()
    with sqlite3.connect(str(work_db), timeout=5):
        pass
    return work_db


def _finalize_teilenummer_reload_work(job_id: str, env: str, work_db: Path) -> Dict[str, Any]:
    table_name = _table_for(TEILENUMMER_TABLE, env)
    with sqlite3.connect(str(work_db), timeout=30) as source_conn:
        source_conn.row_factory = sqlite3.Row
        source_conn.execute("PRAGMA busy_timeout = 30000")
        source_columns = [
            str(row["name"] or "")
            for row in source_conn.execute(f'PRAGMA table_info("{table_name}")').fetchall()
            if str(row["name"] or "")
        ]
        if not source_columns:
            raise RuntimeError(f"Reload-Tabelle {table_name} fehlt in Arbeitsdatenbank.")
    count_rows = 0
    merge_attempts = 40
    with _reload_merge_state_lock:
        global _reload_merge_thread_id
        _reload_merge_thread_id = threading.get_ident()
    try:
        for attempt in range(merge_attempts):
            try:
                with _connect(timeout=60.0, busy_timeout_ms=60000) as conn:
                    if not _table_exists(conn, table_name):
                        column_defs = ", ".join(f'"{col}" TEXT' for col in source_columns + ["CHECKED"])
                        conn.execute(f'CREATE TABLE IF NOT EXISTS "{table_name}" ({column_defs})')
                    _ensure_columns(conn, table_name, source_columns + ["CHECKED"])
                    work_db_sql = str(work_db).replace("'", "''")
                    conn.execute(f"ATTACH DATABASE '{work_db_sql}' AS workdb")
                    try:
                        column_list = ", ".join(f'"{col}"' for col in source_columns)
                        conn.execute(f'DELETE FROM "{table_name}"')
                        conn.execute(
                            f'INSERT INTO "{table_name}" ({column_list}) '
                            f'SELECT {column_list} FROM workdb."{table_name}"'
                        )
                    finally:
                        conn.execute("DETACH DATABASE workdb")
                    conn.execute(f'UPDATE "{table_name}" SET "CHECKED" = ""')
                    count_rows = int(conn.execute(f'SELECT COUNT(*) FROM "{table_name}"').fetchone()[0] or 0)
                    conn.commit()
                break
            except sqlite3.OperationalError as exc:
                if not _is_sqlite_locked_error(exc) or attempt >= (merge_attempts - 1):
                    raise
                wait_sec = 0.25 * (attempt + 1)
                _append_job_log(
                    job_id,
                    f"SQLite lock beim Merge ({attempt + 1}/{merge_attempts}) – retry in {wait_sec:.2f}s ...",
                )
                time.sleep(wait_sec)
    finally:
        with _reload_merge_state_lock:
            _reload_merge_thread_id = None
    for suffix in ("", "-wal", "-shm"):
        try:
            Path(f"{work_db}{suffix}").unlink(missing_ok=True)
        except Exception:
            pass
    _record_cache_status(TEILENUMMER_TABLE, env, "manual_reload", row_count=count_rows)
    _append_job_log(job_id, f"Reload abgeschlossen: {count_rows} Datensätze.")
    return {"rows": count_rows}


def _reload_spareparts_table(env: str) -> None:
    if not SPAREPARTS_SQL_FILE.exists():
        return
    table_name = _table_for(SPAREPARTS_TABLE, env)
    result = _run_compass_to_sqlite(SPAREPARTS_SQL_FILE, table_name, env)
    if result.returncode != 0:
        print(
            f"Ersatzteil-Reload fehlgeschlagen: {result.stderr or result.stdout}",
            file=sys.stderr,
        )
        return
    _record_cache_status(SPAREPARTS_TABLE, env, "background_reload")


@app.post("/api/reload")
def reload_database(
    background_tasks: BackgroundTasks,
    table: str = DEFAULT_TABLE,
    env: str = Query(DEFAULT_ENV),
) -> dict:
    wagons_sql = _wagons_sql_file(env)
    if not wagons_sql.exists():
        raise HTTPException(status_code=500, detail=f"SQL-Datei nicht gefunden: {wagons_sql}")

    table = _validate_table(table)
    table_name = _table_for(table, env)
    result = _run_compass_to_sqlite(wagons_sql, table_name, env)
    if result.returncode != 0:
        raise HTTPException(
            status_code=500,
            detail=f"Reload fehlgeschlagen: {result.stderr or result.stdout}",
        )

    if table == DEFAULT_TABLE:
        background_tasks.add_task(_reload_spareparts_table, env)
    _record_cache_status(table, env, "manual_reload")
    return {"message": "Reload erfolgreich", "stdout": result.stdout, "env": _normalize_env(env)}


@app.post("/api/teilenummer/reload")
def reload_teilenummer(env: str = Query(DEFAULT_ENV)) -> dict:
    sql_file = _teilenummer_sql_file(env)
    if not sql_file.exists():
        raise HTTPException(status_code=500, detail=f"SQL-Datei nicht gefunden: {sql_file}")
    _precheck_compass_connection(env)
    table_name = _table_for(TEILENUMMER_TABLE, env)
    try:
        result = _run_compass_to_sqlite(
            sql_file,
            table_name,
            env,
            timeout_seconds=TEILENUMMER_RELOAD_TIMEOUT_SEC,
            mode="truncate",
        )
    except subprocess.TimeoutExpired as exc:
        timeout_label = int(TEILENUMMER_RELOAD_TIMEOUT_SEC)
        stderr_text = (exc.stderr or exc.output or "").strip() if isinstance(exc.stderr or exc.output, str) else ""
        detail = f"Reload Timeout nach {timeout_label}s."
        if stderr_text:
            detail = f"{detail} {stderr_text}"
        raise HTTPException(status_code=500, detail=detail) from exc
    if result.returncode != 0:
        raise HTTPException(
            status_code=500,
            detail=f"Reload fehlgeschlagen: {result.stderr or result.stdout}",
        )
    with _connect() as conn:
        _ensure_columns(conn, table_name, ["CHECKED"])
        conn.execute(f'UPDATE "{table_name}" SET "CHECKED" = ""')
        conn.commit()
    _record_cache_status(TEILENUMMER_TABLE, env, "manual_reload")
    return {"message": "Reload erfolgreich", "stdout": result.stdout, "env": _normalize_env(env)}


@app.post("/api/teilenummer/reload_job")
def reload_teilenummer_job(env: str = Query(DEFAULT_ENV)) -> dict:
    existing = _find_running_job("reload_teilenummer", env)
    if existing is not None:
        return {
            "job_id": existing["id"],
            "status": "running",
            "env": existing.get("env") or _normalize_env(env),
            "reused": True,
        }
    def prepare(job_id: str):
        _append_job_log(job_id, "Compass-Verbindung wird geprüft ...")
        _precheck_compass_connection(env)
        _append_job_log(job_id, "SQLite wird vorbereitet ...")
        wal_mode = _ensure_sqlite_wal_ready()
        _append_job_log(job_id, f"SQLite journal_mode: {wal_mode}")
        _append_job_log(job_id, "Reload wird direkt auf SQLite ausgeführt ...")

    try:
        cmd = _build_teilenummer_reload_cmd(env, sqlite_db_path=DB_PATH)
    except FileNotFoundError as exc:
        raise HTTPException(status_code=500, detail=str(exc)) from exc
    job = _start_subprocess_job(
        "reload_teilenummer",
        cmd,
        env,
        lambda job_id: _finalize_teilenummer_reload(job_id, env),
        max_runtime_sec=TEILENUMMER_RELOAD_TIMEOUT_SEC,
        prepare_fn=prepare,
    )
    return {"job_id": job["id"], "status": job["status"], "env": job["env"]}


@app.post("/api/wagensuche/reload")
def reload_wagensuche(env: str = Query(DEFAULT_ENV)) -> dict:
    sql_file = _wagensuche_sql_file(env)
    if not sql_file.exists():
        raise HTTPException(status_code=500, detail=f"SQL-Datei nicht gefunden: {sql_file}")
    table_name = _table_for(WAGENSUCHE_TABLE, env)
    result = _run_compass_to_sqlite(sql_file, table_name, env)
    if result.returncode != 0:
        raise HTTPException(
            status_code=500,
            detail=f"Reload fehlgeschlagen: {result.stderr or result.stdout}",
        )
    with _connect() as conn:
        count = conn.execute(f'SELECT COUNT(*) FROM "{table_name}"').fetchone()[0]
    _record_cache_status(WAGENSUCHE_TABLE, env, "manual_reload", row_count=int(count))
    return {
        "message": "Reload erfolgreich",
        "count": count,
        "stdout": result.stdout,
        "env": _normalize_env(env),
    }


@app.get("/api/wagensuche/suggest")
def wagensuche_suggest(
    q: str = Query(..., min_length=1),
    limit: int = Query(15, ge=1, le=50),
    env: str = Query(DEFAULT_ENV),
) -> dict:
    table_name = _table_for(WAGENSUCHE_TABLE, env)
    query = q.strip()
    if not query:
        return {"items": [], "env": _normalize_env(env)}
    query_compact = re.sub(r"\D", "", query)
    if "*" in query:
        like_raw = query.replace("*", "%")
    else:
        like_raw = f"{query}%"
    if "*" in query_compact:
        like_compact = query_compact.replace("*", "%")
    else:
        like_compact = f"{query_compact}%"
    with _connect() as conn:
        if not _table_exists(conn, table_name):
            raise HTTPException(status_code=404, detail=f"Tabelle {table_name} nicht gefunden.")
        rows = conn.execute(
            f'''SELECT ITNO, SERN, ITDS
                FROM "{table_name}"
                WHERE SERN LIKE ?
                   OR REPLACE(REPLACE(SERN, ' ', ''), '-', '') LIKE ?
                ORDER BY SERN
                LIMIT ?''',
            (like_raw, like_compact, limit),
        ).fetchall()
    items = [
        {"itno": (row["ITNO"] or "").strip(), "sern": (row["SERN"] or "").strip(), "itds": (row["ITDS"] or "").strip()}
        for row in rows
    ]
    return {"items": items, "env": _normalize_env(env)}


@app.get("/api/wagensuche/position")
def wagensuche_position(
    sern: str = Query(..., min_length=1),
) -> dict:
    value = sern.strip()
    if not value:
        raise HTTPException(status_code=400, detail="Seriennummer fehlt.")
    result = _wagensuche_latest_position(value)
    if not result:
        return {"found": False, "sern": value}
    return {"found": True, "data": result}


@app.get("/api/wagensuche/maps_key")
def wagensuche_maps_key() -> dict:
    key = os.getenv("VITE_GOOGLE_MAPS_API_KEY", "").strip()
    if not key:
        raise HTTPException(status_code=500, detail="Google Maps API Key fehlt.")
    return {"key": key}


@app.post("/api/teilenummer/check")
def teilenummer_check(payload: dict = Body(...), env: str = Query(DEFAULT_ENV)) -> dict:
    rowid = payload.get("ROWID") or payload.get("rowid")
    birt = payload.get("A_BIRT")
    itno = payload.get("A_ITNO")
    sern = payload.get("A_SERN")
    checked = payload.get("checked")
    out_delay_min = _parse_delay_minutes(payload.get("out_delay_min"))
    in_delay_min = _parse_delay_minutes(payload.get("in_delay_min"))
    table_name = _table_for(TEILENUMMER_TABLE, env)
    with _connect() as conn:
        if not _table_exists(conn, table_name):
            raise HTTPException(status_code=404, detail=f"Tabelle {table_name} nicht gefunden.")
        _ensure_columns(
            conn,
            table_name,
            ["CHECKED", "OUT_DELAY_MIN", "IN_DELAY_MIN", "PLAN_OUT_DATE", "PLAN_OUT_TIME", "PLAN_IN_DATE", "PLAN_IN_TIME"],
        )
        value = "1" if bool(checked) else ""
        plan_times = None
        out_delay_value = ""
        in_delay_value = ""
        if rowid not in (None, ""):
            try:
                rowid_value = int(rowid)
            except (TypeError, ValueError):
                raise HTTPException(status_code=400, detail="ROWID ist ungültig.") from None
            row = conn.execute(
                f'SELECT rowid AS ROWID, * FROM "{table_name}" WHERE rowid = ? LIMIT 1',
                (rowid_value,),
            ).fetchone()
            if row is None:
                raise HTTPException(status_code=404, detail="Datensatz nicht gefunden.")
            if value:
                plan_times = _compute_teilenummer_plan_times(dict(row), out_delay_min, in_delay_min)
                out_delay_value = "" if out_delay_min is None else str(out_delay_min)
                in_delay_value = "" if in_delay_min is None else str(in_delay_min)
            cursor = conn.execute(
                f'''UPDATE "{table_name}"
                    SET "CHECKED" = ?,
                        "OUT_DELAY_MIN" = ?,
                        "IN_DELAY_MIN" = ?,
                        "PLAN_OUT_DATE" = ?,
                        "PLAN_OUT_TIME" = ?,
                        "PLAN_IN_DATE" = ?,
                        "PLAN_IN_TIME" = ?
                    WHERE rowid = ?''',
                (
                    value,
                    out_delay_value,
                    in_delay_value,
                    (plan_times or {}).get("PLAN_OUT_DATE", ""),
                    (plan_times or {}).get("PLAN_OUT_TIME", ""),
                    (plan_times or {}).get("PLAN_IN_DATE", ""),
                    (plan_times or {}).get("PLAN_IN_TIME", ""),
                    rowid_value,
                ),
            )
        else:
            if not birt or not itno or not sern:
                raise HTTPException(status_code=400, detail="ROWID oder A_BIRT/A_ITNO/A_SERN fehlt.")
            row = conn.execute(
                f'''SELECT rowid AS ROWID, * FROM "{table_name}"
                    WHERE "A_BIRT" = ? AND "A_ITNO" = ? AND "A_SERN" = ?
                    ORDER BY rowid ASC
                    LIMIT 1''',
                (birt, itno, sern),
            ).fetchone()
            if row is None:
                raise HTTPException(status_code=404, detail="Datensatz nicht gefunden.")
            if value:
                plan_times = _compute_teilenummer_plan_times(dict(row), out_delay_min, in_delay_min)
                out_delay_value = "" if out_delay_min is None else str(out_delay_min)
                in_delay_value = "" if in_delay_min is None else str(in_delay_min)
            cursor = conn.execute(
                f'''UPDATE "{table_name}"
                    SET "CHECKED" = ?,
                        "OUT_DELAY_MIN" = ?,
                        "IN_DELAY_MIN" = ?,
                        "PLAN_OUT_DATE" = ?,
                        "PLAN_OUT_TIME" = ?,
                        "PLAN_IN_DATE" = ?,
                        "PLAN_IN_TIME" = ?
                    WHERE "A_BIRT" = ? AND "A_ITNO" = ? AND "A_SERN" = ?''',
                (
                    value,
                    out_delay_value,
                    in_delay_value,
                    (plan_times or {}).get("PLAN_OUT_DATE", ""),
                    (plan_times or {}).get("PLAN_OUT_TIME", ""),
                    (plan_times or {}).get("PLAN_IN_DATE", ""),
                    (plan_times or {}).get("PLAN_IN_TIME", ""),
                    birt,
                    itno,
                    sern,
                ),
            )
        conn.commit()
    return {
        "table": table_name,
        "matched": cursor.rowcount,
        "checked": bool(checked),
        "out_delay_min": out_delay_min,
        "in_delay_min": in_delay_min,
        "plan_times": plan_times or {},
        "env": _normalize_env(env),
    }


@app.post("/api/teilenummer/excel/validate")
def teilenummer_validate_excel(
    payload: dict = Body(...),
    env: str = Query(DEFAULT_ENV),
) -> dict:
    filename = str(payload.get("filename") or "").strip()
    if not filename:
        raise HTTPException(status_code=400, detail="Dateiname fehlt.")

    suffix = Path(filename).suffix.lower()
    if suffix not in {".xlsx", ".xlsm", ".xls"}:
        raise HTTPException(status_code=400, detail="Ungültiges Dateiformat. Bitte Excel-Datei hochladen.")

    content_b64 = str(payload.get("content_b64") or "").strip()
    if not content_b64:
        raise HTTPException(status_code=400, detail="Excel-Inhalt fehlt.")
    try:
        import base64
        content = base64.b64decode(content_b64, validate=True)
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=400, detail=f"Excel-Inhalt ist ungültig: {exc}") from exc

    if not content:
        raise HTTPException(status_code=400, detail="Excel-Datei ist leer.")
    out_delay_min = _parse_delay_minutes(payload.get("out_delay_min"))
    in_delay_min = _parse_delay_minutes(payload.get("in_delay_min"))

    try:
        workbook = load_workbook(io.BytesIO(content), data_only=True, read_only=True)
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=400, detail=f"Excel konnte nicht gelesen werden: {exc}") from exc

    sheet = workbook.active
    mappings: List[Dict[str, str]] = []
    seen_old_keys: set[str] = set()
    for row_index, row in enumerate(sheet.iter_rows(min_row=2, max_col=4, values_only=True), start=2):
        old_itno = str(row[0] or "").strip()
        old_sern = str(row[1] or "").strip()
        new_itno = str(row[2] or "").strip()
        new_sern = str(row[3] or "").strip()
        if not old_itno and not old_sern and not new_itno and not new_sern:
            continue
        if not old_itno or not old_sern or not new_itno or not new_sern:
            raise HTTPException(
                status_code=400,
                detail=f"Excel ungültig: Zeile {row_index}: Spalten A-D müssen befüllt sein.",
            )
        old_key = f"{old_itno}||{old_sern}"
        if old_key in seen_old_keys:
            raise HTTPException(
                status_code=400,
                detail=f"Excel ungültig: Zeile {row_index}: Duplikat für alte Position {old_itno}/{old_sern}.",
            )
        seen_old_keys.add(old_key)
        mappings.append(
            {
                "old_itno": old_itno,
                "old_sern": old_sern,
                "new_itno": new_itno,
                "new_sern": new_sern,
                "row": str(row_index),
            }
        )

    if not mappings:
        raise HTTPException(status_code=400, detail="Excel enthält keine Datenzeilen ab Zeile 2.")

    table_name = _ensure_wagon_data(TEILENUMMER_TABLE, env)
    with _connect() as conn:
        rows = conn.execute(
            f'SELECT "A_ITNO", "A_SERN", "C_MTRL", "RGDT", "RGTM" FROM "{table_name}"'
        ).fetchall()
    rows = [dict(row) for row in rows]

    def _norm_itno(value: Any) -> str:
        return str(value or "").strip().upper()

    def _norm_sern(value: Any) -> str:
        return re.sub(r"[\s\-]+", "", str(value or "").strip().upper())

    def _norm_pair(itno_value: Any, sern_value: Any) -> str:
        return f"{_norm_itno(itno_value)}||{_norm_sern(sern_value)}"

    existing_itnos = {_norm_itno(row["A_ITNO"]) for row in rows if row["A_ITNO"]}
    existing_serns = {_norm_sern(row["A_SERN"]) for row in rows if row["A_SERN"]}
    existing_pairs = {
        _norm_pair(row["A_ITNO"], row["A_SERN"])
        for row in rows
        if row["A_ITNO"] and row["A_SERN"]
    }
    rows_by_key = {
        _norm_pair(row["A_ITNO"], row["A_SERN"]): row
        for row in rows
        if row["A_ITNO"] and row["A_SERN"]
    }

    validated: List[Dict[str, str]] = []
    validation_checks: List[Dict[str, Any]] = []
    validation_errors: List[str] = []
    normalized_env = _normalize_env(env)
    env_label = normalized_env.upper()
    cono_value = _default_mi_cono_for_env(env).strip()
    ionapi_path = _ionapi_path(env, "mi")
    ion_cfg = load_ionapi_config(str(ionapi_path))
    base_url = build_base_url(ion_cfg)
    token = get_access_token_service_account(ion_cfg)
    mos450_cache: Dict[str, tuple[bool, str]] = {}
    for entry in mappings:
        row_label = entry["row"]
        old_itno = entry["old_itno"]
        old_sern = entry["old_sern"]
        new_itno = entry["new_itno"]
        new_sern = entry["new_sern"]
        old_key = _norm_pair(old_itno, old_sern)

        if old_key not in existing_pairs:
            validation_error = (
                f"Zeile {row_label}: alte Position {old_itno}/{old_sern} existiert nicht im System."
            )
            validation_checks.append(
                {
                    "row": int(row_label),
                    "old_itno": old_itno,
                    "old_sern": old_sern,
                    "new_itno": new_itno,
                    "new_sern": new_sern,
                    "allowed": False,
                    "cono": cono_value,
                    "message": validation_error,
                }
            )
            validation_errors.append(validation_error)
            continue
        if _norm_itno(new_itno) not in existing_itnos:
            validation_error = f"Zeile {row_label}: neue ITNO '{new_itno}' existiert nicht im System."
            validation_checks.append(
                {
                    "row": int(row_label),
                    "old_itno": old_itno,
                    "old_sern": old_sern,
                    "new_itno": new_itno,
                    "new_sern": new_sern,
                    "allowed": False,
                    "cono": cono_value,
                    "message": validation_error,
                }
            )
            validation_errors.append(validation_error)
            continue
        if _norm_sern(new_sern) in existing_serns:
            validation_error = f"Zeile {row_label}: neue SERN '{new_sern}' existiert bereits im System."
            validation_checks.append(
                {
                    "row": int(row_label),
                    "old_itno": old_itno,
                    "old_sern": old_sern,
                    "new_itno": new_itno,
                    "new_sern": new_sern,
                    "allowed": False,
                    "cono": cono_value,
                    "message": validation_error,
                }
            )
            validation_errors.append(validation_error)
            continue
        source_row = rows_by_key.get(old_key)
        hitn_value = str((source_row or {}).get("C_MTRL") or "").strip()
        if not hitn_value:
            validation_error = (
                f"Zeile {row_label}: HITN fehlt für {old_itno}/{old_sern}. "
                "Einbauprüfung über MOS450MI nicht möglich."
            )
            validation_checks.append(
                {
                    "row": int(row_label),
                    "old_itno": old_itno,
                    "old_sern": old_sern,
                    "hitn": hitn_value,
                    "new_itno": new_itno,
                    "new_sern": new_sern,
                    "allowed": False,
                    "cono": cono_value,
                    "message": validation_error,
                }
            )
            validation_errors.append(validation_error)
            continue
        # Cache pro MOTP + NEUER ITNO, damit bei großen Excel-Dateien nicht pro Zeile neu geprüft wird.
        cache_key = f"{_norm_itno(hitn_value)}||{_norm_itno(new_itno)}"
        if cache_key in mos450_cache:
            allowed, check_message = mos450_cache[cache_key]
        else:
            params = _build_mos450_lstcomponent_params(hitn_value, env=env)
            request_url = _build_m3_request_url(base_url, "MOS450MI", "LstComponent", params)
            try:
                response = call_m3_mi_get(base_url, token, "MOS450MI", "LstComponent", params)
                mi_error = _mi_error_message(response)
                if mi_error:
                    allowed = False
                    check_message = f"MOS450MI Fehler (MOTP): {mi_error}"
                else:
                    allowed, status_message = _mos450_component_status20(response, new_itno)
                    check_message = (
                        "OK: Komponente vorhanden mit Status 20."
                        if allowed
                        else f"Komponente {new_itno} in MOTP {hitn_value}: {status_message}"
                    )
                _append_api_log(
                    "teilenummer_mos450_lstcomponent",
                    params,
                    response,
                    allowed,
                    "" if allowed else check_message,
                    env=env_label,
                    wagon={
                        "itno": old_itno,
                        "sern": old_sern,
                        "new_itno": new_itno,
                        "new_sern": new_sern,
                    },
                    dry_run=False,
                    request_url=request_url,
                    program="MOS450MI",
                    transaction="LstComponent",
                )
            except Exception as exc:  # noqa: BLE001
                allowed = False
                check_message = f"MOS450MI Request fehlgeschlagen (MOTP): {exc}"
            mos450_cache[cache_key] = (allowed, check_message)
        # Excel-Validierung bleibt read-only:
        # Wenn MOS450 nur "Komponente fehlt / Status !=20" meldet, lassen wir den Datensatz zu.
        # Der echte Run holt das per MOS450 AddComponent + Recheck nach.
        if not allowed and _is_mos450_soft_validation_failure(check_message):
            allowed = True
            check_message = (
                f"WARN: {check_message} "
                "(wird im Run via MOS450MI/AddComponent abgesichert)."
            )

        validation_checks.append(
            {
                "row": int(row_label),
                "old_itno": old_itno,
                "old_sern": old_sern,
                "hitn": hitn_value,
                "new_itno": new_itno,
                "new_sern": new_sern,
                "allowed": bool(allowed),
                "cono": cono_value,
                "message": check_message,
            }
        )
        if not allowed:
            validation_errors.append(f"Zeile {row_label}: {check_message}")
            continue
        plan_times = _compute_teilenummer_plan_times(
            source_row or {},
            out_delay_min,
            in_delay_min,
        )

        validated.append(
            {
                "old_itno": old_itno,
                "old_sern": old_sern,
                "new_itno": new_itno,
                "new_sern": new_sern,
                "out_delay_min": "" if out_delay_min is None else str(out_delay_min),
                "in_delay_min": "" if in_delay_min is None else str(in_delay_min),
                "PLAN_OUT_DATE": plan_times["PLAN_OUT_DATE"],
                "PLAN_OUT_TIME": plan_times["PLAN_OUT_TIME"],
                "PLAN_IN_DATE": plan_times["PLAN_IN_DATE"],
                "PLAN_IN_TIME": plan_times["PLAN_IN_TIME"],
            }
        )

    if validation_errors:
        raise HTTPException(
            status_code=400,
            detail={
                "message": "Excel ungültig: Einbauprüfung fehlgeschlagen.",
                "errors": validation_errors,
                "checks": validation_checks,
                "env": env_label,
            },
        )

    with _connect() as conn:
        _ensure_columns(
            conn,
            table_name,
            ["OUT_DELAY_MIN", "IN_DELAY_MIN", "PLAN_OUT_DATE", "PLAN_OUT_TIME", "PLAN_IN_DATE", "PLAN_IN_TIME"],
        )
        for entry in validated:
            conn.execute(
                f'''UPDATE "{table_name}"
                    SET "OUT_DELAY_MIN" = ?,
                        "IN_DELAY_MIN" = ?,
                        "PLAN_OUT_DATE" = ?,
                        "PLAN_OUT_TIME" = ?,
                        "PLAN_IN_DATE" = ?,
                        "PLAN_IN_TIME" = ?
                    WHERE "A_ITNO" = ? AND "A_SERN" = ?''',
                (
                    entry["out_delay_min"],
                    entry["in_delay_min"],
                    entry["PLAN_OUT_DATE"],
                    entry["PLAN_OUT_TIME"],
                    entry["PLAN_IN_DATE"],
                    entry["PLAN_IN_TIME"],
                    entry["old_itno"],
                    entry["old_sern"],
                ),
            )
        conn.commit()

    return {
        "filename": filename,
        "count": len(validated),
        "mappings": validated,
        "checks": validation_checks,
        "validation": {
            "checked": len(validation_checks),
            "ok": len(validation_checks) - len(validation_errors),
            "failed": len(validation_errors),
        },
        "out_delay_min": out_delay_min if out_delay_min is not None else "",
        "in_delay_min": in_delay_min if in_delay_min is not None else "",
        "env": normalized_env,
    }


@app.post("/api/teilenummer/prepare")
def teilenummer_prepare(payload: dict = Body(...), env: str = Query(DEFAULT_ENV)) -> dict:
    legacy_new_itno = (payload.get("new_itno") or "").strip()
    legacy_new_sern = (payload.get("new_sern") or "").strip()
    mappings_payload = payload.get("mappings") or []
    out_delay_min = _parse_delay_minutes(payload.get("out_delay_min"))
    in_delay_min = _parse_delay_minutes(payload.get("in_delay_min"))

    mapping_by_key: Dict[str, Dict[str, str]] = {}
    if isinstance(mappings_payload, list):
        for entry in mappings_payload:
            if not isinstance(entry, dict):
                continue
            old_itno = str(entry.get("old_itno") or entry.get("A_ITNO") or "").strip()
            old_sern = str(entry.get("old_sern") or entry.get("A_SERN") or "").strip()
            new_itno = str(entry.get("new_itno") or entry.get("NITNO") or "").strip()
            new_sern = str(entry.get("new_sern") or entry.get("NSERN") or "").strip()
            if not old_itno or not old_sern or not new_itno:
                continue
            key = f"{old_itno}||{old_sern}"
            mapping_by_key[key] = {
                "old_itno": old_itno,
                "old_sern": old_sern,
                "new_itno": new_itno,
                "new_sern": new_sern,
            }

    if not mapping_by_key and not legacy_new_itno:
        raise HTTPException(status_code=400, detail="Neue ITNO fehlt.")

    source_table = _ensure_wagon_data(TEILENUMMER_TABLE, env)
    target_table = _table_for(TEILENUMMER_TAUSCH_TABLE, env)
    timestamp = datetime.now().isoformat(timespec="seconds")
    with _connect() as conn:
        if not _table_exists(conn, source_table):
            raise HTTPException(status_code=404, detail=f"Tabelle {source_table} nicht gefunden.")
        _ensure_columns(conn, source_table, ["CHECKED"])
        # Auswahl erfolgt immer über Checkboxen (CHECKED=1).
        # Excel-Mappings liefern nur die Zielwerte pro markierter Position.
        rows = [
            dict(row)
            for row in conn.execute(f'SELECT rowid AS "ROWID", * FROM "{source_table}" WHERE "CHECKED" = "1"')
            .fetchall()
        ]
        if not rows:
            raise HTTPException(
                status_code=400,
                detail="Keine Position markiert. Bitte gewünschte Zeilen per Checkbox auswählen.",
            )

        if mapping_by_key:
            rows = [
                row
                for row in rows
                if f"{row.get('A_ITNO', '')}||{row.get('A_SERN', '')}" in mapping_by_key
            ]
            if not rows:
                raise HTTPException(
                    status_code=400,
                    detail="Keine markierte Position entspricht dem Excel-Mapping.",
                )

        if not mapping_by_key and legacy_new_sern and len(rows) > 1:
            raise HTTPException(
                status_code=400,
                detail="Mit Neue SERN ist nur eine markierte Position erlaubt.",
            )

        if mapping_by_key:
            candidate_itnos = {
                mapping_by_key[f"{row.get('A_ITNO', '')}||{row.get('A_SERN', '')}"]["new_itno"]
                for row in rows
                if f"{row.get('A_ITNO', '')}||{row.get('A_SERN', '')}" in mapping_by_key
            }
        else:
            candidate_itnos = {legacy_new_itno}
        missing_itnos: List[str] = []
        for candidate in sorted(candidate_itnos):
            exists = conn.execute(
                f'SELECT 1 FROM "{source_table}" WHERE "A_ITNO" = ? LIMIT 1',
                (candidate,),
            ).fetchone()
            if not exists:
                missing_itnos.append(candidate)
        if missing_itnos:
            raise HTTPException(
                status_code=400,
                detail=f"Neue ITNO nicht vorhanden: {', '.join(missing_itnos)}",
            )

        columns = _table_columns(conn, source_table)
        insert_columns = _create_teilenummer_tausch_table(conn, source_table, target_table)
        if rows:
            placeholders = ", ".join("?" for _ in insert_columns)
            column_list = ", ".join(f'"{col}"' for col in insert_columns)
            insert_sql = f'INSERT INTO "{target_table}" ({column_list}) VALUES ({placeholders})'
            data: List[List[Any]] = []
            for row in rows:
                key = f"{row.get('A_ITNO', '')}||{row.get('A_SERN', '')}"
                mapping = mapping_by_key.get(key)
                resolved_new_itno = (mapping or {}).get("new_itno") or legacy_new_itno
                resolved_new_sern = (mapping or {}).get("new_sern") or legacy_new_sern
                # Beim Vorbereiten immer mit den aktuell gesetzten Verzögerungsfeldern rechnen.
                # So werden alte Planwerte aus früheren Checks nicht versehentlich weiterverwendet.
                effective_out_delay = out_delay_min
                effective_in_delay = in_delay_min
                plan_times = _compute_teilenummer_plan_times(
                    row,
                    effective_out_delay,
                    effective_in_delay,
                )
                rowid_value = row.get("ROWID")
                if rowid_value not in (None, ""):
                    try:
                        conn.execute(
                            f'''UPDATE "{source_table}"
                                SET "OUT_DELAY_MIN" = ?,
                                    "IN_DELAY_MIN" = ?,
                                    "PLAN_OUT_DATE" = ?,
                                    "PLAN_OUT_TIME" = ?,
                                    "PLAN_IN_DATE" = ?,
                                    "PLAN_IN_TIME" = ?
                                WHERE rowid = ?''',
                            (
                                "" if effective_out_delay is None else str(effective_out_delay),
                                "" if effective_in_delay is None else str(effective_in_delay),
                                plan_times["PLAN_OUT_DATE"],
                                plan_times["PLAN_OUT_TIME"],
                                plan_times["PLAN_IN_DATE"],
                                plan_times["PLAN_IN_TIME"],
                                int(rowid_value),
                            ),
                        )
                    except (TypeError, ValueError):
                        pass
                extra_values = {
                    "NITNO": resolved_new_itno,
                    "NSERN": resolved_new_sern,
                    "OUT_DELAY_MIN": "" if effective_out_delay is None else str(effective_out_delay),
                    "IN_DELAY_MIN": "" if effective_in_delay is None else str(effective_in_delay),
                    "PLAN_OUT_DATE": plan_times["PLAN_OUT_DATE"],
                    "PLAN_OUT_TIME": plan_times["PLAN_OUT_TIME"],
                    "PLAN_IN_DATE": plan_times["PLAN_IN_DATE"],
                    "PLAN_IN_TIME": plan_times["PLAN_IN_TIME"],
                    "UMGEBAUT": "",
                    "TIMESTAMP": timestamp,
                    "OUT_STATUS": "",
                    "MOS170_STATUS": "",
                    "PLPN": "",
                    "CMS100_STATUS": "",
                    "MWNO": "",
                    "MOS100_STATUS": "",
                    "MOS180_STATUS": "",
                    "MOS050_STATUS": "",
                    "IN_STATUS": "",
                }
                row_values = dict(row)
                row_values.update(extra_values)
                values = [row_values.get(col, "") for col in insert_columns]
                data.append(values)
            conn.executemany(insert_sql, data)
        conn.commit()
    return {
        "table": target_table,
        "count": len(rows),
        "mappings": len(mapping_by_key),
        "out_delay_min": out_delay_min if out_delay_min is not None else "",
        "in_delay_min": in_delay_min if in_delay_min is not None else "",
        "env": _normalize_env(env),
    }


@app.post("/api/teilenummer/clear_tausch")
def teilenummer_clear_tausch(env: str = Query(DEFAULT_ENV)) -> dict:
    table_name = _table_for(TEILENUMMER_TAUSCH_TABLE, env)
    with _connect() as conn:
        if not _table_exists(conn, table_name):
            return {"table": table_name, "cleared": 0, "env": _normalize_env(env)}
        count_row = conn.execute(f'SELECT COUNT(1) AS CNT FROM "{table_name}"').fetchone()
        cleared = int((count_row["CNT"] if count_row and "CNT" in count_row.keys() else 0) or 0)
        conn.execute(f'DELETE FROM "{table_name}"')
        conn.commit()
    return {"table": table_name, "cleared": cleared, "env": _normalize_env(env)}


@app.get("/api/teilenummer/api_log")
def teilenummer_api_log(limit: int = Query(400, ge=1, le=5000)) -> dict:
    entries: List[Dict[str, Any]] = []
    if not API_LOG_PATH.exists():
        return {"path": str(API_LOG_PATH), "count": 0, "entries": entries}
    with API_LOG_PATH.open("r", encoding="utf-8") as handle:
        lines = handle.readlines()
    selected = lines[-limit:]
    for line in selected:
        raw = line.strip()
        if not raw:
            continue
        try:
            parsed = json.loads(raw)
            if isinstance(parsed, dict):
                entries.append(parsed)
            else:
                entries.append({"raw": raw})
        except Exception:  # noqa: BLE001
            entries.append({"raw": raw})
    return {"path": str(API_LOG_PATH), "count": len(entries), "entries": entries}


@app.post("/api/teilenummer/api_log/clear")
def teilenummer_api_log_clear() -> dict:
    _clear_api_log()
    return {"path": str(API_LOG_PATH), "cleared": True}


@app.get("/api/teilenummer/api_log.csv")
def teilenummer_api_log_csv() -> Response:
    import csv

    output = io.StringIO()
    writer = csv.writer(output, delimiter=";")
    writer.writerow(
        [
            "ts",
            "env",
            "action",
            "program",
            "transaction",
            "ok",
            "error",
            "itno",
            "sern",
            "new_itno",
            "new_sern",
            "request_url",
        ]
    )
    if API_LOG_PATH.exists():
        with API_LOG_PATH.open("r", encoding="utf-8") as handle:
            for line in handle:
                raw = line.strip()
                if not raw:
                    continue
                try:
                    entry = json.loads(raw)
                except Exception:  # noqa: BLE001
                    continue
                wagon = entry.get("wagon") if isinstance(entry, dict) else {}
                request = entry.get("request") if isinstance(entry, dict) else {}
                writer.writerow(
                    [
                        entry.get("ts", ""),
                        entry.get("env", ""),
                        entry.get("action", ""),
                        entry.get("program", ""),
                        entry.get("transaction", ""),
                        entry.get("ok", ""),
                        entry.get("error", ""),
                        (wagon or {}).get("itno", ""),
                        (wagon or {}).get("sern", ""),
                        (wagon or {}).get("new_itno", ""),
                        (wagon or {}).get("new_sern", ""),
                        (request or {}).get("url", ""),
                    ]
                )
    content = output.getvalue()
    output.close()
    filename = f"API_{datetime.now().strftime('%Y%m%d_%H%M%S')}.csv"
    headers = {"Content-Disposition": f"attachment; filename*=UTF-8''{quote(filename)}"}
    return Response(content=content, media_type="text/csv; charset=utf-8", headers=headers)


@app.get("/api/teilenummer/api_log/view", response_class=HTMLResponse)
def teilenummer_api_log_view() -> HTMLResponse:
    html = """<!doctype html>
<html lang="de">
<head>
  <meta charset="utf-8" />
  <meta name="viewport" content="width=device-width,initial-scale=1" />
  <title>API-Log</title>
  <style>
    body { font-family: Arial, sans-serif; margin: 14px; background: #f3f5f8; color: #1f2937; }
    .bar { display: flex; gap: 8px; align-items: center; margin-bottom: 10px; }
    button, a.btn { border: 1px solid #cbd5e1; background: #fff; border-radius: 6px; padding: 6px 10px; font-size: 12px; font-weight: 700; color: #1f2937; text-decoration: none; cursor: pointer; }
    .log-wrap { margin: 0; border: 1px solid #dbe3ee; border-radius: 8px; background: #fff; padding: 10px; max-height: 78vh; overflow: auto; font-size: 11px; line-height: 1.35; }
    .log-entry { border-bottom: 1px solid #eef2f7; padding: 6px 0; }
    .log-entry:last-child { border-bottom: none; }
    .log-entry.ok { color: #166534; }
    .log-entry.warn { color: #a16207; }
    .log-entry.nok { color: #b91c1c; }
    .log-main { white-space: pre-wrap; font-family: ui-monospace, SFMono-Regular, Menlo, Consolas, monospace; }
    .log-url { white-space: pre-wrap; font-family: ui-monospace, SFMono-Regular, Menlo, Consolas, monospace; margin-top: 2px; opacity: 0.9; }
    .empty { color: #64748b; }
    .hint { font-size: 11px; color: #64748b; margin-left: auto; }
  </style>
</head>
<body>
  <div class="bar">
    <button id="refreshBtn" type="button">Aktualisieren</button>
    <button id="clearBtn" type="button">Log löschen</button>
    <a id="csvBtn" class="btn" href="/api/teilenummer/api_log.csv" target="_blank" rel="noopener">Als CSV exportieren</a>
    <span class="hint" id="countInfo">Lade ...</span>
  </div>
  <div id="logContent" class="log-wrap"><div class="empty">Lade API-Log ...</div></div>
  <script>
    const params = new URLSearchParams(window.location.search);
    const env = params.get("env") || "tst";
    const logContent = document.getElementById("logContent");
    const countInfo = document.getElementById("countInfo");
    const csvBtn = document.getElementById("csvBtn");
    csvBtn.href = `/api/teilenummer/api_log.csv?env=${encodeURIComponent(env)}`;

    const toBool = (value) => {
      if (value === true || value === false) return value;
      const text = String(value ?? "").trim().toLowerCase();
      if (text === "true") return true;
      if (text === "false") return false;
      return null;
    };

    const classify = (entry) => {
      if (!entry || typeof entry !== "object") return "";
      const ok = toBool(entry.ok);
      const action = String(entry.action || "").toLowerCase();
      const error = String(entry.error || "").trim().toLowerCase();
      const status = String(entry.status || "").trim().toLowerCase();

      // Waiting/retry states should be highlighted in yellow.
      const isWaitCase =
        action.includes("cms100") &&
        (error.includes("mwno fehlt") || error.includes("plpn fehlt") || status === "skipped");
      if (isWaitCase) return "warn";

      if (ok === true) return "ok";
      if (ok === false) return "nok";
      if (error) return "nok";
      return "";
    };

    const format = (entry) => {
      if (!entry || typeof entry !== "object") {
        return { klass: "", main: String(entry || ""), url: "" };
      }
      const ts = entry.ts || "";
      const action = entry.action || "";
      const ok = entry.ok === true ? "OK" : (entry.ok === false ? "NOK" : "");
      const error = entry.error || "";
      const wagon = entry.wagon || {};
      const itno = wagon.itno || "";
      const sern = wagon.sern || "";
      const url = (entry.request || {}).url || "";
      return {
        klass: classify(entry),
        main: `${ts} | ${action} | ${ok} | ${itno}/${sern}${error ? ` | ${error}` : ""}`,
        url: url,
      };
    };

    const escapeHtml = (value) =>
      String(value || "")
        .replaceAll("&", "&amp;")
        .replaceAll("<", "&lt;")
        .replaceAll(">", "&gt;")
        .replaceAll('"', "&quot;");

    const renderEntries = (entries) => {
      if (!Array.isArray(entries) || !entries.length) {
        logContent.innerHTML = '<div class="empty">API-Log ist leer.</div>';
        return;
      }
      const html = entries.map((entry) => {
        const row = format(entry);
        const urlHtml = row.url ? `<div class="log-url">${escapeHtml(row.url)}</div>` : "";
        return `<div class="log-entry ${row.klass}"><div class="log-main">${escapeHtml(row.main)}</div>${urlHtml}</div>`;
      }).join("");
      logContent.innerHTML = html;
    };

    async function loadLog() {
      logContent.innerHTML = '<div class="empty">Lade API-Log ...</div>';
      const res = await fetch(`/api/teilenummer/api_log?env=${encodeURIComponent(env)}&limit=3000`);
      if (!res.ok) {
        const txt = await res.text();
        logContent.innerHTML = `<div class="empty">${escapeHtml(txt || "API-Log konnte nicht geladen werden.")}</div>`;
        countInfo.textContent = "Fehler";
        return;
      }
      const data = await res.json();
      const entries = Array.isArray(data.entries) ? data.entries : [];
      countInfo.textContent = `${entries.length} Einträge`;
      renderEntries(entries);
    }

    document.getElementById("refreshBtn").addEventListener("click", () => { void loadLog(); });
    document.getElementById("clearBtn").addEventListener("click", async () => {
      await fetch(`/api/teilenummer/api_log/clear?env=${encodeURIComponent(env)}`, { method: "POST" });
      await loadLog();
    });
    void loadLog();
  </script>
</body>
</html>"""
    return HTMLResponse(content=html)


@app.post("/api/teilenummer/export_xlsx")
def teilenummer_export_xlsx(payload: dict = Body(...), env: str = Query(DEFAULT_ENV)) -> Response:
    columns_raw = payload.get("columns")
    rows_raw = payload.get("rows")
    filename_raw = str(payload.get("filename") or "").strip()
    if not isinstance(columns_raw, list) or not columns_raw:
        raise HTTPException(status_code=400, detail="Spalten fehlen.")
    if not isinstance(rows_raw, list):
        raise HTTPException(status_code=400, detail="Zeilen fehlen.")

    columns = [str(col or "").strip() for col in columns_raw if str(col or "").strip()]
    if not columns:
        raise HTTPException(status_code=400, detail="Spalten fehlen.")

    wb = Workbook(write_only=True)
    ws = wb.create_sheet(title="Wagenuebersicht")
    default_font = Font(name="Arial", size=11)
    header_font = Font(name="Arial", size=11, bold=True)
    wagon_warning_font = Font(name="Arial", size=11, bold=True, color="9C0006")
    part_warning_fill = PatternFill(fill_type="solid", fgColor="FDECEC")
    widths = [len(str(col)) for col in columns]

    header_cells: List[WriteOnlyCell] = []
    for col in columns:
        cell = WriteOnlyCell(ws, value=col)
        cell.font = header_font
        header_cells.append(cell)
    ws.append(header_cells)

    for row in rows_raw:
        if not isinstance(row, dict):
            continue
        wagon_status = str(row.get("__W_STAT", "") or "").strip()
        part_status = str(row.get("__A_STAT", "") or "").strip()
        out_cells: List[WriteOnlyCell] = []
        for idx, col in enumerate(columns):
            value = row.get(col, "")
            text_value = "" if value is None else str(value)
            widths[idx] = max(widths[idx], len(text_value))
            cell = WriteOnlyCell(ws, value=value)
            cell.font = default_font
            # ITNO/SERN: nur Hintergrund, wenn Teilstatus kritisch ist.
            if col in {"ITNO", "SERN"} and part_status and part_status != "20":
                cell.fill = part_warning_fill
            # Spalte N/O (W_ITNO/W_SERN): fett + rote Schrift, wenn Wagenstatus kritisch ist.
            if col in {"W_ITNO", "W_SERN"} and wagon_status and wagon_status != "20":
                cell.font = wagon_warning_font
            out_cells.append(cell)
        ws.append(out_cells)

    for idx, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(idx)].width = min(max(width + 2, 8), 60)

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    normalized_env = _normalize_env(env).upper()
    safe_name = re.sub(
        r"[^A-Za-z0-9._-]+",
        "_",
        filename_raw,
    ) or f"Wagenuebersicht_{normalized_env}_{timestamp}.xlsx"
    if not safe_name.lower().endswith(".xlsx"):
        safe_name += ".xlsx"

    headers = {
        "Content-Disposition": f"attachment; filename*=UTF-8''{quote(safe_name)}"
    }
    return Response(
        content=output.getvalue(),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers=headers,
    )


@app.get("/api/teilenummer/mode")
def teilenummer_mode(env: str = Query(DEFAULT_ENV)) -> dict:
    normalized = _normalize_env(env)
    dry_run = _effective_dry_run(env)
    return {
        "env": normalized,
        "dry_run": dry_run,
        "mode": "DRYRUN" if dry_run else "LIVE",
    }


@app.post("/api/teilenummer/run")
def teilenummer_run(env: str = Query(DEFAULT_ENV)) -> dict:
    job = _create_job("teilenummer_run", env)
    normalized_env = _normalize_env(env)
    run_dry_run = _effective_dry_run(env)

    def _worker() -> None:
        try:
            table_name = _table_for(TEILENUMMER_TAUSCH_TABLE, env)
            with _connect() as conn:
                if not _table_exists(conn, table_name):
                    raise HTTPException(status_code=404, detail=f"Tabelle {table_name} nicht gefunden.")
                _ensure_columns(conn, table_name, TEILENUMMER_TAUSCH_EXTRA_COLUMNS)
                rows = conn.execute(
                    f'SELECT rowid AS seq, * FROM "{table_name}" ORDER BY rowid ASC'
                ).fetchall()

            if not rows:
                raise HTTPException(status_code=404, detail="Keine Daten in TEILENUMMER_TAUSCH.")

            total_steps = len(rows) * 9
            _update_job(job["id"], total=total_steps, processed=0, results=[])
            _append_job_log(job["id"], f"Teilenummer-Ablauf startet: {len(rows)} Datensätze.")

            dry_run = run_dry_run
            ionapi_path = _ionapi_path(env, "mi")
            ion_cfg = load_ionapi_config(str(ionapi_path))
            base_url = build_base_url(ion_cfg)
            token = ""
            if not dry_run:
                token = get_access_token_service_account(ion_cfg)

            env_label = _normalize_env(env).upper()
            mode_label = "DRYRUN" if dry_run else "LIVE"
            _append_job_log(job["id"], f"Betriebsmodus: {env_label} · {mode_label}")
            processed = 0
            row_contexts: List[Dict[str, Any]] = []
            if dry_run:
                _reset_dryrun_trace("teilenummer_run", env_label)
                _append_job_log(job["id"], f"DryRun-Trace aktiviert: {DRYRUN_TRACE_PATH}")

            with _connect() as conn:
                for index, row in enumerate(rows, start=1):
                    old_itno = _row_value(row, "A_ITNO")
                    old_sern = _row_value(row, "A_SERN")
                    new_itno = _row_value(row, "NITNO") or old_itno
                    new_sern = _row_value(row, "NSERN") or old_sern
                    row_map = _build_teilenummer_row_map(row, new_itno, new_sern)
                    wagon_ctx = _teilenummer_log_context(row, new_itno, new_sern)

                    row_is_leading = _is_teilenummer_leading_object(row_map)
                    row_hard_failed = False
                    skip_after_error = "SKIPPED: vorheriger Schritt fehlgeschlagen"
                    params: Dict[str, Any] = {}
                    request_url = ""

                    # MOS125MI Ausbau
                    if row_is_leading:
                        out_status = "SKIPPED: führendes Objekt (kein Ausbau)"
                        _append_api_log(
                            "teilenummer_ausbau",
                            {},
                            {"skipped": out_status},
                            True,
                            None,
                            env=env_label,
                            wagon=wagon_ctx,
                            dry_run=dry_run,
                            request_url="",
                            program="MOS125MI",
                            transaction="RemoveInstall",
                            status="SKIPPED",
                        )
                    else:
                        params = _build_mos125_params(row_map, mode="out", env=env)
                        request_url = _build_m3_request_url(base_url, "MOS125MI", "RemoveInstall", params)
                        if not params.get("TRDT"):
                            ok = False
                            status_label = "ERROR"
                            error_message = "UMBAU_DATUM fehlt"
                            response = {"error": error_message}
                        elif dry_run:
                            ok = True
                            status_label = "DRYRUN"
                            error_message = None
                            response = {"dry_run": True}
                        else:
                            try:
                                response = call_m3_mi_get(base_url, token, "MOS125MI", "RemoveInstall", params)
                                ok, status_label, error_message = _mi_status(response)
                            except Exception as exc:  # noqa: BLE001
                                ok = False
                                status_label = "ERROR"
                                error_message = str(exc)
                                response = {"error": error_message}
                        out_status = status_label if ok else f"{status_label}: {error_message}"
                        _append_api_log(
                            "teilenummer_ausbau",
                            params,
                            response,
                            ok,
                            error_message,
                            env=env_label,
                            wagon=wagon_ctx,
                            dry_run=dry_run,
                            request_url=request_url,
                            program="MOS125MI",
                            transaction="RemoveInstall",
                        )
                        if not ok and not dry_run:
                            row_hard_failed = True
                    _update_teilenummer_row(conn, table_name, row["seq"], {"OUT_STATUS": out_status})
                    processed += 1
                    _update_job(job["id"], processed=processed)

                    # MOS170MI AddProp (mit Retry)
                    plpn = ""
                    mos170_status = ""
                    params = {}
                    request_url = ""
                    if row_hard_failed:
                        mos170_status = skip_after_error
                        _append_api_log(
                            "teilenummer_mos170_addprop",
                            {},
                            {"skipped": skip_after_error},
                            True,
                            None,
                            env=env_label,
                            wagon=wagon_ctx,
                            dry_run=dry_run,
                            request_url="",
                            program="MOS170MI",
                            transaction="AddProp",
                            status="SKIPPED",
                        )
                        _update_teilenummer_row(
                            conn,
                            table_name,
                            row["seq"],
                            {"MOS170_STATUS": mos170_status, "PLPN": plpn},
                        )
                        processed += 1
                        _update_job(job["id"], processed=processed)
                    else:
                        attempt = 1
                        while True:
                            params = _build_mos170_params(row_map, env=env)
                            request_url = _build_m3_request_url(base_url, "MOS170MI", "AddProp", params)
                            required_missing = not params.get("ITNO") or not params.get("BANO") or not params.get("STDT")
                            if required_missing:
                                ok = False
                                status_label = "ERROR"
                                error_message = "Pflichtfelder fehlen"
                                response = {"error": error_message}
                            elif dry_run:
                                ok = True
                                status_label = "DRYRUN"
                                error_message = None
                                response = {"dry_run": True}
                            else:
                                try:
                                    response = call_m3_mi_get(base_url, token, "MOS170MI", "AddProp", params)
                                    ok, status_label, error_message = _mi_status(response)
                                except Exception as exc:  # noqa: BLE001
                                    ok = False
                                    status_label = "ERROR"
                                    error_message = str(exc)
                                    response = {"error": error_message}
                            plpn = _extract_plpn(response) if ok else ""
                            if ok and not plpn:
                                ok = False
                                status_label = "ERROR"
                                error_message = "PLPN fehlt"
                            mos170_status = status_label if ok else f"{status_label}: {error_message}"
                            _append_api_log(
                                "teilenummer_mos170_addprop",
                                params,
                                {"plpn": plpn, "response": response},
                                ok,
                                error_message,
                                env=env_label,
                                wagon=wagon_ctx,
                                dry_run=dry_run,
                                request_url=request_url,
                                program="MOS170MI",
                                transaction="AddProp",
                            )
                            _update_teilenummer_row(
                                conn,
                                table_name,
                                row["seq"],
                                {"MOS170_STATUS": mos170_status, "PLPN": plpn},
                            )
                            processed += 1
                            _update_job(job["id"], processed=processed)
                            if plpn or dry_run or required_missing:
                                break
                            if MOS170_RETRY_MAX and attempt >= MOS170_RETRY_MAX:
                                break
                            total_steps += 1
                            _update_job(job["id"], total=total_steps)
                            _append_job_log(
                                job["id"],
                                f"MOS170MI AddProp: Warte {MOS170_RETRY_DELAY_SEC} Sekunden auf ERP ...",
                            )
                            time.sleep(MOS170_RETRY_DELAY_SEC)
                            attempt += 1
                        if not plpn and not dry_run:
                            row_hard_failed = True

                    # MOS170 PLPN (Log)
                    if row_hard_failed and not plpn:
                        _append_api_log(
                            "teilenummer_mos170_plpn",
                            params,
                            {"plpn": "", "skipped": skip_after_error},
                            True,
                            None,
                            env=env_label,
                            wagon=wagon_ctx,
                            dry_run=dry_run,
                            request_url=request_url,
                            program="MOS170MI",
                            transaction="AddProp",
                            status="SKIPPED",
                        )
                    else:
                        _append_api_log(
                            "teilenummer_mos170_plpn",
                            params,
                            {"plpn": plpn},
                            bool(plpn),
                            None if plpn else "PLPN fehlt",
                            env=env_label,
                            wagon=wagon_ctx,
                            dry_run=dry_run,
                            request_url=request_url,
                            program="MOS170MI",
                            transaction="AddProp",
                        )
                    processed += 1
                    _update_job(job["id"], processed=processed)
                    row_contexts.append(
                        {
                            "index": index,
                            "row": row,
                            "old_itno": old_itno,
                            "old_sern": old_sern,
                            "new_itno": new_itno,
                            "new_sern": new_sern,
                            "row_map": row_map,
                            "wagon_ctx": wagon_ctx,
                            "row_is_leading": row_is_leading,
                            "row_hard_failed": row_hard_failed,
                            "skip_after_error": skip_after_error,
                            "plpn": plpn,
                        }
                    )
                    continue

                    # CMS100MI Lst_PLPN_MWNO (mit Retry)
                    mwno = ""
                    if row_hard_failed:
                        cms_status = skip_after_error
                        _append_api_log(
                            "teilenummer_cms100_lst_plpn_mwno",
                            {},
                            {"qomwno": "", "skipped": skip_after_error},
                            True,
                            None,
                            env=env_label,
                            wagon=wagon_ctx,
                            dry_run=dry_run,
                            request_url="",
                            program="CMS100MI",
                            transaction="Lst_PLPN_MWNO",
                            status="SKIPPED",
                        )
                        _update_teilenummer_row(
                            conn,
                            table_name,
                            row["seq"],
                            {"CMS100_STATUS": cms_status, "MWNO": mwno},
                        )
                        processed += 1
                        _update_job(job["id"], processed=processed)
                    else:
                        cms_status = ""
                        attempt = 1
                        while True:
                            params = _build_cms100_params(plpn, env=env)
                            request_url = _build_m3_request_url(base_url, "CMS100MI", "Lst_PLPN_MWNO", params)
                            if not plpn:
                                ok = False
                                status_label = "ERROR"
                                error_message = "PLPN fehlt"
                                response = {"error": error_message}
                                mwno = ""
                            elif dry_run:
                                ok = True
                                status_label = "DRYRUN"
                                error_message = None
                                response = {"dry_run": True}
                                mwno = "DRYRUN"
                            else:
                                try:
                                    response = call_m3_mi_get(base_url, token, "CMS100MI", "Lst_PLPN_MWNO", params)
                                    ok, status_label, error_message = _mi_status(response)
                                    mwno = _extract_mwno(response) if ok else ""
                                except Exception as exc:  # noqa: BLE001
                                    ok = False
                                    status_label = "ERROR"
                                    error_message = str(exc)
                                    response = {"error": error_message}
                                    mwno = ""
                            if ok and not mwno:
                                ok = False
                                status_label = "ERROR"
                                error_message = "MWNO fehlt"
                            cms_status = status_label if ok else f"{status_label}: {error_message}"
                            _append_api_log(
                                "teilenummer_cms100_lst_plpn_mwno",
                                params,
                                {"qomwno": mwno, "response": response},
                                ok,
                                error_message,
                                env=env_label,
                                wagon=wagon_ctx,
                                dry_run=dry_run,
                                request_url=request_url,
                                program="CMS100MI",
                                transaction="Lst_PLPN_MWNO",
                            )
                            _update_teilenummer_row(
                                conn,
                                table_name,
                                row["seq"],
                                {"CMS100_STATUS": cms_status, "MWNO": mwno},
                            )
                            processed += 1
                            _update_job(job["id"], processed=processed)
                            if mwno or dry_run or not plpn:
                                break
                            if CMS100_RETRY_MAX and attempt >= CMS100_RETRY_MAX:
                                break
                            total_steps += 1
                            _update_job(job["id"], total=total_steps)
                            _append_job_log(
                                job["id"],
                                f"CMS100MI: Warte {CMS100_RETRY_DELAY_SEC} Sekunden auf ERP ...",
                            )
                            time.sleep(CMS100_RETRY_DELAY_SEC)
                            attempt += 1
                        if not mwno and not dry_run:
                            row_hard_failed = True

                    # IPS MOS100 Chg_SERN (ITNO + SERN)
                    params = {
                        "WorkOrderNumber": mwno,
                        "Product": old_itno,
                        "NewItemNumber": new_itno,
                        "NewLotNumber": new_sern,
                    }
                    request_url = _build_ips_request_url(base_url, "MOS100")
                    if row_hard_failed:
                        status_label = "SKIPPED"
                        _append_api_log(
                            "teilenummer_ips_mos100_chgsern",
                            params,
                            {"skipped": skip_after_error},
                            True,
                            None,
                            env=env_label,
                            wagon=wagon_ctx,
                            dry_run=dry_run,
                            request_url=request_url,
                            program="MOS100",
                            transaction="Chg_SERN",
                            request_method="POST",
                            status=status_label,
                        )
                    else:
                        attempt = 1
                        ok = False
                        error_message = None
                        response = {}
                        status_label = "NOK"
                        while True:
                            if not mwno:
                                ok = False
                                status_label = "NOK"
                                error_message = "MWNO fehlt"
                                response = {"error": error_message}
                            elif dry_run:
                                ok = True
                                status_label = "DRYRUN"
                                error_message = None
                                response = {"dry_run": True}
                            else:
                                try:
                                    response = _call_ips_service(
                                        base_url,
                                        token,
                                        "MOS100",
                                        "Chg_SERN",
                                        params,
                                        env=env,
                                    )
                                    ok = int(response.get("status_code") or 0) < 400
                                    if ok:
                                        status_label = "OK"
                                        error_message = None
                                    elif _ips_mos100_already_exists_fault(response):
                                        ok = True
                                        status_label = "OK_ALREADY_EXISTS"
                                        error_message = None
                                    else:
                                        status_label = "NOK"
                                        error_message = f"HTTP {response.get('status_code')}"
                                except Exception as exc:  # noqa: BLE001
                                    ok = False
                                    status_label = "NOK"
                                    error_message = str(exc)
                                    response = {"error": error_message}
                            _append_api_log(
                                "teilenummer_ips_mos100_chgsern",
                                params,
                                response,
                                ok,
                                error_message,
                                env=env_label,
                                wagon=wagon_ctx,
                                dry_run=dry_run,
                                request_url=request_url,
                                program="MOS100",
                                transaction="Chg_SERN",
                                request_method="POST",
                                status=status_label,
                            )
                            if ok or dry_run:
                                break
                            if MOS100_RETRY_MAX and attempt >= MOS100_RETRY_MAX:
                                break
                            if MOS100_RETRY_DELAY_SEC:
                                time.sleep(MOS100_RETRY_DELAY_SEC)
                            attempt += 1
                        if not ok and not dry_run:
                            row_hard_failed = True
                    _update_teilenummer_row(
                        conn,
                        table_name,
                        row["seq"],
                        {"MOS100_STATUS": status_label},
                    )
                    processed += 1
                    _update_job(job["id"], processed=processed)

                    # MOS180MI Approve (mit Retry)
                    mos180_row = dict(row_map)
                    mos180_row["MWNO"] = mwno
                    if row_hard_failed:
                        mos180_status = skip_after_error
                        _append_api_log(
                            "teilenummer_mos180_approve",
                            {},
                            {"skipped": skip_after_error},
                            True,
                            None,
                            env=env_label,
                            wagon=wagon_ctx,
                            dry_run=dry_run,
                            request_url="",
                            program="MOS180MI",
                            transaction="Approve",
                            status="SKIPPED",
                        )
                        _update_teilenummer_row(
                            conn,
                            table_name,
                            row["seq"],
                            {"MOS180_STATUS": mos180_status},
                        )
                        processed += 1
                        _update_job(job["id"], processed=processed)
                    else:
                        attempt = 1
                        while True:
                            params = _build_mos180_params(mos180_row, env=env)
                            request_url = _build_m3_request_url(base_url, "MOS180MI", "Approve", params)
                            if not mwno:
                                ok = False
                                status_label = "ERROR"
                                error_message = "MWNO fehlt"
                                response = {"error": error_message}
                            elif dry_run:
                                ok = True
                                status_label = "DRYRUN"
                                error_message = None
                                response = {"dry_run": True}
                            else:
                                try:
                                    response = call_m3_mi_get(base_url, token, "MOS180MI", "Approve", params)
                                    ok, status_label, error_message = _mi_status(response)
                                except Exception as exc:  # noqa: BLE001
                                    ok = False
                                    status_label = "ERROR"
                                    error_message = str(exc)
                                    response = {"error": error_message}
                            mos180_status = status_label if ok else f"{status_label}: {error_message}"
                            _append_api_log(
                                "teilenummer_mos180_approve",
                                params,
                                response,
                                ok,
                                error_message,
                                env=env_label,
                                wagon=wagon_ctx,
                                dry_run=dry_run,
                                request_url=request_url,
                                program="MOS180MI",
                                transaction="Approve",
                            )
                            _update_teilenummer_row(
                                conn,
                                table_name,
                                row["seq"],
                                {"MOS180_STATUS": mos180_status},
                            )
                            processed += 1
                            _update_job(job["id"], processed=processed)
                            if ok or dry_run or not mwno:
                                break
                            if MOS180_RETRY_MAX and attempt >= MOS180_RETRY_MAX:
                                break
                            total_steps += 1
                            _update_job(job["id"], total=total_steps)
                            _append_job_log(
                                job["id"],
                                f"MOS180MI Approve: Warte {MOS180_RETRY_DELAY_SEC} Sekunden auf ERP ...",
                            )
                            if MOS180_RETRY_DELAY_SEC:
                                time.sleep(MOS180_RETRY_DELAY_SEC)
                            attempt += 1
                        if not ok and not dry_run:
                            row_hard_failed = True

                    # IPS MOS050 Montage (mit Retry)
                    mos050_row = dict(row_map)
                    mos050_row["MWNO"] = mwno
                    if row_is_leading:
                        mos050_status = "SKIPPED: führendes Objekt (kein Montage-Schritt)"
                        _append_api_log(
                            "teilenummer_ips_mos50_montage",
                            {},
                            {"skipped": mos050_status},
                            True,
                            None,
                            env=env_label,
                            wagon=wagon_ctx,
                            dry_run=dry_run,
                            request_url="",
                            program=MOS050_SERVICE or "MOS050",
                            transaction=MOS050_OPERATION or "Montage",
                            request_method="POST",
                            status="SKIPPED",
                        )
                        _update_teilenummer_row(
                            conn,
                            table_name,
                            row["seq"],
                            {"MOS050_STATUS": mos050_status},
                        )
                        processed += 1
                        _update_job(job["id"], processed=processed)
                    elif row_hard_failed:
                        mos050_status = skip_after_error
                        _append_api_log(
                            "teilenummer_ips_mos50_montage",
                            {},
                            {"skipped": skip_after_error},
                            True,
                            None,
                            env=env_label,
                            wagon=wagon_ctx,
                            dry_run=dry_run,
                            request_url="",
                            program=MOS050_SERVICE or "MOS050",
                            transaction=MOS050_OPERATION or "Montage",
                            request_method="POST",
                            status="SKIPPED",
                        )
                        _update_teilenummer_row(
                            conn,
                            table_name,
                            row["seq"],
                            {"MOS050_STATUS": mos050_status},
                        )
                        processed += 1
                        _update_job(job["id"], processed=processed)
                    else:
                        attempt = 1
                        while True:
                            params = _build_mos050_params(mos050_row)
                            request_url = _build_ips_request_url(base_url, MOS050_SERVICE)
                            if not mwno:
                                ok = False
                                status_label = "NOK"
                                error_message = "MWNO fehlt"
                                response = {"error": error_message}
                            elif dry_run:
                                ok = True
                                status_label = "DRYRUN"
                                error_message = None
                                response = {"dry_run": True}
                            else:
                                try:
                                    response = _call_ips_service(
                                        base_url,
                                        token,
                                        MOS050_SERVICE,
                                        MOS050_OPERATION,
                                        params,
                                        namespace_override=MOS050_NAMESPACE or None,
                                        body_tag_override=MOS050_BODY_TAG or None,
                                        env=env,
                                    )
                                    ok = int(response.get("status_code") or 0) < 400
                                    status_label = "OK" if ok else "NOK"
                                    error_message = None if ok else f"HTTP {response.get('status_code')}"
                                except Exception as exc:  # noqa: BLE001
                                    ok = False
                                    status_label = "NOK"
                                    error_message = str(exc)
                                    response = {"error": error_message}
                            _append_api_log(
                                "teilenummer_ips_mos50_montage",
                                params,
                                response,
                                ok,
                                error_message,
                                env=env_label,
                                wagon=wagon_ctx,
                                dry_run=dry_run,
                                request_url=request_url,
                                program=MOS050_SERVICE or "MOS050",
                                transaction=MOS050_OPERATION or "Montage",
                                request_method="POST",
                                status=status_label,
                            )
                            _update_teilenummer_row(
                                conn,
                                table_name,
                                row["seq"],
                                {"MOS050_STATUS": status_label},
                            )
                            processed += 1
                            _update_job(job["id"], processed=processed)
                            if ok or dry_run or not mwno:
                                break
                            if MOS050_RETRY_MAX and attempt >= MOS050_RETRY_MAX:
                                break
                            total_steps += 1
                            _update_job(job["id"], total=total_steps)
                            _append_job_log(
                                job["id"],
                                f"IPS MOS050: Warte {MOS050_RETRY_DELAY_SEC} Sekunden auf ERP ...",
                            )
                            if MOS050_RETRY_DELAY_SEC:
                                time.sleep(MOS050_RETRY_DELAY_SEC)
                            attempt += 1
                        if not ok and not dry_run:
                            row_hard_failed = True

                    # MOS125MI Einbau
                    if row_is_leading:
                        in_status = "SKIPPED: führendes Objekt (kein Einbau)"
                        _append_api_log(
                            "teilenummer_einbau",
                            {},
                            {"skipped": in_status},
                            True,
                            None,
                            env=env_label,
                            wagon=wagon_ctx,
                            dry_run=dry_run,
                            request_url="",
                            program="MOS125MI",
                            transaction="RemoveInstall",
                            status="SKIPPED",
                        )
                    elif row_hard_failed:
                        in_status = skip_after_error
                        _append_api_log(
                            "teilenummer_einbau",
                            {},
                            {"skipped": skip_after_error},
                            True,
                            None,
                            env=env_label,
                            wagon=wagon_ctx,
                            dry_run=dry_run,
                            request_url="",
                            program="MOS125MI",
                            transaction="RemoveInstall",
                            status="SKIPPED",
                        )
                    else:
                        params = _build_mos125_params(row_map, mode="in", env=env)
                        request_url = _build_m3_request_url(base_url, "MOS125MI", "RemoveInstall", params)
                        if not params.get("TRDT"):
                            ok = False
                            status_label = "ERROR"
                            error_message = "UMBAU_DATUM fehlt"
                            response = {"error": error_message}
                        elif dry_run:
                            ok = True
                            status_label = "DRYRUN"
                            error_message = None
                            response = {"dry_run": True}
                        else:
                            try:
                                response = call_m3_mi_get(base_url, token, "MOS125MI", "RemoveInstall", params)
                                ok, status_label, error_message = _mi_status(response)
                            except Exception as exc:  # noqa: BLE001
                                ok = False
                                status_label = "ERROR"
                                error_message = str(exc)
                                response = {"error": error_message}
                        in_status = status_label if ok else f"{status_label}: {error_message}"
                        _append_api_log(
                            "teilenummer_einbau",
                            params,
                            response,
                            ok,
                            error_message,
                            env=env_label,
                            wagon=wagon_ctx,
                            dry_run=dry_run,
                            request_url=request_url,
                            program="MOS125MI",
                            transaction="RemoveInstall",
                        )
                    _update_teilenummer_row(conn, table_name, row["seq"], {"IN_STATUS": in_status})
                    processed += 1
                    _update_job(job["id"], processed=processed)

                    row_result = "OK"
                    if row_hard_failed:
                        row_result = "FEHLER"
                    elif row_is_leading:
                        row_result = "OK_FUEHRENDES_OBJEKT"
                    _append_job_log(
                        job["id"],
                        f"{index}/{len(rows)} abgeschlossen [{row_result}]: {old_itno} {old_sern} -> {new_itno} {new_sern}",
                    )

            if row_contexts and not dry_run:
                _append_job_log(
                    job["id"],
                    f"Phase 1 abgeschlossen ({len(row_contexts)} Datensätze). "
                    f"Warte {TEILENUMMER_PHASE_SPLIT_WAIT_SEC} Sekunden vor CMS100 ...",
                )
                time.sleep(TEILENUMMER_PHASE_SPLIT_WAIT_SEC)

            mos450_runtime_cache: Dict[str, tuple[bool, str, Dict[str, str], str, Any]] = {}
            with _connect() as conn:
                for ctx in row_contexts:
                    index = int(ctx["index"])
                    row = ctx["row"]
                    old_itno = str(ctx["old_itno"] or "")
                    old_sern = str(ctx["old_sern"] or "")
                    new_itno = str(ctx["new_itno"] or "")
                    new_sern = str(ctx["new_sern"] or "")
                    row_map = dict(ctx["row_map"])
                    wagon_ctx = dict(ctx["wagon_ctx"])
                    row_is_leading = bool(ctx["row_is_leading"])
                    row_hard_failed = bool(ctx["row_hard_failed"])
                    skip_after_error = str(ctx["skip_after_error"] or "SKIPPED: vorheriger Schritt fehlgeschlagen")
                    plpn = str(ctx["plpn"] or "")

                    # MOS450MI Precheck (verhindert MO90804 beim Einbau)
                    if row_is_leading:
                        _append_api_log(
                            "teilenummer_mos450_runtime_precheck",
                            {},
                            {"skipped": "SKIPPED: führendes Objekt (kein Komponenten-Precheck)"},
                            True,
                            None,
                            env=env_label,
                            wagon=wagon_ctx,
                            dry_run=dry_run,
                            request_url="",
                            program="MOS450MI",
                            transaction="LstComponent",
                            status="SKIPPED",
                        )
                    elif row_hard_failed:
                        _append_api_log(
                            "teilenummer_mos450_runtime_precheck",
                            {},
                            {"skipped": skip_after_error},
                            True,
                            None,
                            env=env_label,
                            wagon=wagon_ctx,
                            dry_run=dry_run,
                            request_url="",
                            program="MOS450MI",
                            transaction="LstComponent",
                            status="SKIPPED",
                        )
                    elif dry_run:
                        params = _build_mos450_lstcomponent_params(_row_value(row_map, "MTRL"), env=env)
                        request_url = _build_m3_request_url(base_url, "MOS450MI", "LstComponent", params)
                        _append_api_log(
                            "teilenummer_mos450_runtime_precheck",
                            params,
                            {"dry_run": True},
                            True,
                            None,
                            env=env_label,
                            wagon=wagon_ctx,
                            dry_run=dry_run,
                            request_url=request_url,
                            program="MOS450MI",
                            transaction="LstComponent",
                            status="DRYRUN",
                        )
                    else:
                        hitn_value = _row_value(row_map, "MTRL")
                        cfgl_value = _row_value(row_map, "CFGL")
                        hitn_key = str(hitn_value or "").strip().upper()
                        new_itno_key = str(new_itno or "").strip().upper()
                        cache_key = f"{hitn_key}||{new_itno_key}"
                        if cache_key in mos450_runtime_cache:
                            allowed, check_message, params, request_url, cached_response = mos450_runtime_cache[cache_key]
                            response = {"cached": True, "response": cached_response}
                        else:
                            params = _build_mos450_lstcomponent_params(hitn_value, env=env)
                            request_url = _build_m3_request_url(base_url, "MOS450MI", "LstComponent", params)
                            if not hitn_key:
                                allowed = False
                                check_message = "MOTP/HITN fehlt."
                                response = {"error": check_message}
                            elif not new_itno_key:
                                allowed = False
                                check_message = "Neue ITNO fehlt."
                                response = {"error": check_message}
                            else:
                                try:
                                    response = call_m3_mi_get(base_url, token, "MOS450MI", "LstComponent", params)
                                    mi_error = _mi_error_message(response)
                                    if mi_error:
                                        allowed = False
                                        check_message = f"MOS450MI Fehler (MOTP): {mi_error}"
                                    else:
                                        allowed, status_message = _mos450_component_status20(response, new_itno)
                                        check_message = (
                                            "OK: Komponente vorhanden mit Status 20."
                                            if allowed
                                            else f"Komponente {new_itno} in MOTP {hitn_value}: {status_message}"
                                        )
                                except Exception as exc:  # noqa: BLE001
                                    allowed = False
                                    check_message = f"MOS450MI Request fehlgeschlagen (MOTP): {exc}"
                                    response = {"error": check_message}

                            # Fallback: Wenn Komponente laut LstComponent fehlt/nicht Status 20,
                            # dann AddComponent versuchen und danach erneut validieren.
                            if not allowed and hitn_key and new_itno_key:
                                add_params = _build_mos450_addcomponent_params(
                                    hitn_value,
                                    new_itno,
                                    cfgl_value,
                                    env=env,
                                )
                                add_request_url = _build_m3_request_url(
                                    base_url,
                                    "MOS450MI",
                                    "AddComponent",
                                    add_params,
                                )
                                if not add_params.get("CFGL"):
                                    add_ok = False
                                    add_status = "ERROR"
                                    add_error = "CFGL fehlt für MOS450MI/AddComponent."
                                    add_response: Any = {"error": add_error}
                                else:
                                    try:
                                        add_response = call_m3_mi_get(
                                            base_url,
                                            token,
                                            "MOS450MI",
                                            "AddComponent",
                                            add_params,
                                        )
                                        add_ok, add_status, add_error = _mi_status(add_response)
                                    except Exception as exc:  # noqa: BLE001
                                        add_ok = False
                                        add_status = "ERROR"
                                        add_error = str(exc)
                                        add_response = {"error": add_error}
                                _append_api_log(
                                    "teilenummer_mos450_addcomponent",
                                    add_params,
                                    add_response,
                                    add_ok,
                                    add_error,
                                    env=env_label,
                                    wagon=wagon_ctx,
                                    dry_run=dry_run,
                                    request_url=add_request_url,
                                    program="MOS450MI",
                                    transaction="AddComponent",
                                    status=add_status if add_status else ("OK" if add_ok else "NOK"),
                                )
                                if add_ok:
                                    # Direkt nach AddComponent nochmals LstComponent prüfen.
                                    try:
                                        verify_response = call_m3_mi_get(
                                            base_url,
                                            token,
                                            "MOS450MI",
                                            "LstComponent",
                                            params,
                                        )
                                        verify_error = _mi_error_message(verify_response)
                                        if verify_error:
                                            allowed = False
                                            check_message = f"MOS450MI Fehler nach AddComponent (MOTP): {verify_error}"
                                        else:
                                            allowed, status_message = _mos450_component_status20(
                                                verify_response,
                                                new_itno,
                                            )
                                            check_message = (
                                                "OK: Komponente nach AddComponent vorhanden mit Status 20."
                                                if allowed
                                                else f"Komponente {new_itno} in MOTP {hitn_value} nach AddComponent: {status_message}"
                                            )
                                        response = {
                                            "initial_precheck": response,
                                            "addcomponent": add_response,
                                            "postcheck": verify_response,
                                        }
                                    except Exception as exc:  # noqa: BLE001
                                        allowed = False
                                        check_message = f"MOS450MI LstComponent nach AddComponent fehlgeschlagen: {exc}"
                                        response = {
                                            "initial_precheck": response,
                                            "addcomponent": add_response,
                                            "postcheck_error": str(exc),
                                        }
                                else:
                                    check_message = f"{check_message} | AddComponent fehlgeschlagen: {add_error}"
                                    response = {
                                        "initial_precheck": response,
                                        "addcomponent": add_response,
                                    }
                                _append_job_log(
                                    job["id"],
                                    f"{index}/{len(rows)} MOS450 AddComponent "
                                    f"{'OK' if add_ok else 'NOK'}: {hitn_value} / {new_itno} / {cfgl_value or '-'}",
                                )
                            mos450_runtime_cache[cache_key] = (allowed, check_message, params, request_url, response)
                        _append_api_log(
                            "teilenummer_mos450_runtime_precheck",
                            params,
                            response,
                            bool(allowed),
                            "" if allowed else check_message,
                            env=env_label,
                            wagon=wagon_ctx,
                            dry_run=dry_run,
                            request_url=request_url,
                            program="MOS450MI",
                            transaction="LstComponent",
                            status="OK" if allowed else "NOK",
                        )
                        if not allowed:
                            row_hard_failed = True
                            skip_after_error = f"SKIPPED: MOS450-Komponentenprüfung fehlgeschlagen ({check_message})"
                            _append_job_log(
                                job["id"],
                                f"{index}/{len(rows)} MOS450-Precheck NOK: "
                                f"{old_itno}/{old_sern} -> {new_itno}/{new_sern} | {check_message}",
                            )
                    processed += 1
                    _update_job(job["id"], processed=processed)

                    # CMS100MI Lst_PLPN_MWNO (mit Retry, max. 10 Versuche, 3s)
                    mwno = ""
                    if row_hard_failed:
                        cms_status = skip_after_error
                        _append_api_log(
                            "teilenummer_cms100_lst_plpn_mwno",
                            {},
                            {"qomwno": "", "skipped": skip_after_error},
                            True,
                            None,
                            env=env_label,
                            wagon=wagon_ctx,
                            dry_run=dry_run,
                            request_url="",
                            program="CMS100MI",
                            transaction="Lst_PLPN_MWNO",
                            status="SKIPPED",
                        )
                        _update_teilenummer_row(
                            conn,
                            table_name,
                            row["seq"],
                            {"CMS100_STATUS": cms_status, "MWNO": mwno},
                        )
                        processed += 1
                        _update_job(job["id"], processed=processed)
                    else:
                        cms_status = ""
                        attempt = 1
                        while True:
                            params = _build_cms100_params(plpn, env=env)
                            request_url = _build_m3_request_url(base_url, "CMS100MI", "Lst_PLPN_MWNO", params)
                            if not plpn:
                                ok = False
                                status_label = "ERROR"
                                error_message = "PLPN fehlt"
                                response = {"error": error_message}
                                mwno = ""
                            elif dry_run:
                                ok = True
                                status_label = "DRYRUN"
                                error_message = None
                                response = {"dry_run": True}
                                mwno = "DRYRUN"
                            else:
                                try:
                                    response = call_m3_mi_get(base_url, token, "CMS100MI", "Lst_PLPN_MWNO", params)
                                    ok, status_label, error_message = _mi_status(response)
                                    mwno = _extract_mwno(response) if ok else ""
                                except Exception as exc:  # noqa: BLE001
                                    ok = False
                                    status_label = "ERROR"
                                    error_message = str(exc)
                                    response = {"error": error_message}
                                    mwno = ""
                            if ok and not mwno:
                                ok = False
                                status_label = "ERROR"
                                error_message = "MWNO fehlt"
                            cms_status = status_label if ok else f"{status_label}: {error_message}"
                            _append_api_log(
                                "teilenummer_cms100_lst_plpn_mwno",
                                params,
                                {"qomwno": mwno, "response": response},
                                ok,
                                error_message,
                                env=env_label,
                                wagon=wagon_ctx,
                                dry_run=dry_run,
                                request_url=request_url,
                                program="CMS100MI",
                                transaction="Lst_PLPN_MWNO",
                            )
                            _update_teilenummer_row(
                                conn,
                                table_name,
                                row["seq"],
                                {"CMS100_STATUS": cms_status, "MWNO": mwno},
                            )
                            processed += 1
                            _update_job(job["id"], processed=processed)
                            if mwno or dry_run or not plpn:
                                break
                            if attempt >= TEILENUMMER_CMS100_RETRY_MAX:
                                break
                            total_steps += 1
                            _update_job(job["id"], total=total_steps)
                            _append_job_log(
                                job["id"],
                                f"CMS100MI: Warte {TEILENUMMER_CMS100_RETRY_DELAY_SEC:g} Sekunden auf ERP ...",
                            )
                            time.sleep(TEILENUMMER_CMS100_RETRY_DELAY_SEC)
                            attempt += 1
                        if not mwno and not dry_run:
                            row_hard_failed = True

                    # IPS MOS100 Chg_SERN (ITNO + SERN)
                    params = {
                        "WorkOrderNumber": mwno,
                        "Product": old_itno,
                        "NewItemNumber": new_itno,
                        "NewLotNumber": new_sern,
                    }
                    request_url = _build_ips_request_url(base_url, "MOS100")
                    if row_hard_failed:
                        status_label = "SKIPPED"
                        _append_api_log(
                            "teilenummer_ips_mos100_chgsern",
                            params,
                            {"skipped": skip_after_error},
                            True,
                            None,
                            env=env_label,
                            wagon=wagon_ctx,
                            dry_run=dry_run,
                            request_url=request_url,
                            program="MOS100",
                            transaction="Chg_SERN",
                            request_method="POST",
                            status=status_label,
                        )
                    else:
                        attempt = 1
                        ok = False
                        error_message = None
                        response = {}
                        status_label = "NOK"
                        while True:
                            if not mwno:
                                ok = False
                                status_label = "NOK"
                                error_message = "MWNO fehlt"
                                response = {"error": error_message}
                            elif dry_run:
                                ok = True
                                status_label = "DRYRUN"
                                error_message = None
                                response = {"dry_run": True}
                            else:
                                try:
                                    response = _call_ips_service(
                                        base_url,
                                        token,
                                        "MOS100",
                                        "Chg_SERN",
                                        params,
                                        env=env,
                                    )
                                    ok = int(response.get("status_code") or 0) < 400
                                    if ok:
                                        status_label = "OK"
                                        error_message = None
                                    elif _ips_mos100_already_exists_fault(response):
                                        ok = True
                                        status_label = "OK_ALREADY_EXISTS"
                                        error_message = None
                                    else:
                                        status_label = "NOK"
                                        error_message = f"HTTP {response.get('status_code')}"
                                except Exception as exc:  # noqa: BLE001
                                    ok = False
                                    status_label = "NOK"
                                    error_message = str(exc)
                                    response = {"error": error_message}
                            _append_api_log(
                                "teilenummer_ips_mos100_chgsern",
                                params,
                                response,
                                ok,
                                error_message,
                                env=env_label,
                                wagon=wagon_ctx,
                                dry_run=dry_run,
                                request_url=request_url,
                                program="MOS100",
                                transaction="Chg_SERN",
                                request_method="POST",
                                status=status_label,
                            )
                            if ok or dry_run:
                                break
                            if MOS100_RETRY_MAX and attempt >= MOS100_RETRY_MAX:
                                break
                            if MOS100_RETRY_DELAY_SEC:
                                time.sleep(MOS100_RETRY_DELAY_SEC)
                            attempt += 1
                        if not ok and not dry_run:
                            row_hard_failed = True
                    _update_teilenummer_row(
                        conn,
                        table_name,
                        row["seq"],
                        {"MOS100_STATUS": status_label},
                    )
                    processed += 1
                    _update_job(job["id"], processed=processed)

                    # MOS180MI Approve (mit Retry)
                    mos180_row = dict(row_map)
                    mos180_row["MWNO"] = mwno
                    if row_hard_failed:
                        mos180_status = skip_after_error
                        _append_api_log(
                            "teilenummer_mos180_approve",
                            {},
                            {"skipped": skip_after_error},
                            True,
                            None,
                            env=env_label,
                            wagon=wagon_ctx,
                            dry_run=dry_run,
                            request_url="",
                            program="MOS180MI",
                            transaction="Approve",
                            status="SKIPPED",
                        )
                        _update_teilenummer_row(
                            conn,
                            table_name,
                            row["seq"],
                            {"MOS180_STATUS": mos180_status},
                        )
                        processed += 1
                        _update_job(job["id"], processed=processed)
                    else:
                        attempt = 1
                        while True:
                            params = _build_mos180_params(mos180_row, env=env)
                            request_url = _build_m3_request_url(base_url, "MOS180MI", "Approve", params)
                            if not mwno:
                                ok = False
                                status_label = "ERROR"
                                error_message = "MWNO fehlt"
                                response = {"error": error_message}
                            elif dry_run:
                                ok = True
                                status_label = "DRYRUN"
                                error_message = None
                                response = {"dry_run": True}
                            else:
                                try:
                                    response = call_m3_mi_get(base_url, token, "MOS180MI", "Approve", params)
                                    ok, status_label, error_message = _mi_status(response)
                                except Exception as exc:  # noqa: BLE001
                                    ok = False
                                    status_label = "ERROR"
                                    error_message = str(exc)
                                    response = {"error": error_message}
                            mos180_status = status_label if ok else f"{status_label}: {error_message}"
                            _append_api_log(
                                "teilenummer_mos180_approve",
                                params,
                                response,
                                ok,
                                error_message,
                                env=env_label,
                                wagon=wagon_ctx,
                                dry_run=dry_run,
                                request_url=request_url,
                                program="MOS180MI",
                                transaction="Approve",
                            )
                            _update_teilenummer_row(
                                conn,
                                table_name,
                                row["seq"],
                                {"MOS180_STATUS": mos180_status},
                            )
                            processed += 1
                            _update_job(job["id"], processed=processed)
                            if ok or dry_run or not mwno:
                                break
                            if MOS180_RETRY_MAX and attempt >= MOS180_RETRY_MAX:
                                break
                            total_steps += 1
                            _update_job(job["id"], total=total_steps)
                            _append_job_log(
                                job["id"],
                                f"MOS180MI Approve: Warte {MOS180_RETRY_DELAY_SEC} Sekunden auf ERP ...",
                            )
                            if MOS180_RETRY_DELAY_SEC:
                                time.sleep(MOS180_RETRY_DELAY_SEC)
                            attempt += 1
                        if not ok and not dry_run:
                            row_hard_failed = True

                    # IPS MOS050 Montage (mit Retry)
                    mos050_row = dict(row_map)
                    mos050_row["MWNO"] = mwno
                    if row_is_leading:
                        mos050_status = "SKIPPED: führendes Objekt (kein Montage-Schritt)"
                        _append_api_log(
                            "teilenummer_ips_mos50_montage",
                            {},
                            {"skipped": mos050_status},
                            True,
                            None,
                            env=env_label,
                            wagon=wagon_ctx,
                            dry_run=dry_run,
                            request_url="",
                            program=MOS050_SERVICE or "MOS050",
                            transaction=MOS050_OPERATION or "Montage",
                            request_method="POST",
                            status="SKIPPED",
                        )
                        _update_teilenummer_row(
                            conn,
                            table_name,
                            row["seq"],
                            {"MOS050_STATUS": mos050_status},
                        )
                        processed += 1
                        _update_job(job["id"], processed=processed)
                    elif row_hard_failed:
                        mos050_status = skip_after_error
                        _append_api_log(
                            "teilenummer_ips_mos50_montage",
                            {},
                            {"skipped": skip_after_error},
                            True,
                            None,
                            env=env_label,
                            wagon=wagon_ctx,
                            dry_run=dry_run,
                            request_url="",
                            program=MOS050_SERVICE or "MOS050",
                            transaction=MOS050_OPERATION or "Montage",
                            request_method="POST",
                            status="SKIPPED",
                        )
                        _update_teilenummer_row(
                            conn,
                            table_name,
                            row["seq"],
                            {"MOS050_STATUS": mos050_status},
                        )
                        processed += 1
                        _update_job(job["id"], processed=processed)
                    else:
                        attempt = 1
                        while True:
                            params = _build_mos050_params(mos050_row)
                            request_url = _build_ips_request_url(base_url, MOS050_SERVICE)
                            if not mwno:
                                ok = False
                                status_label = "NOK"
                                error_message = "MWNO fehlt"
                                response = {"error": error_message}
                            elif dry_run:
                                ok = True
                                status_label = "DRYRUN"
                                error_message = None
                                response = {"dry_run": True}
                            else:
                                try:
                                    response = _call_ips_service(
                                        base_url,
                                        token,
                                        MOS050_SERVICE,
                                        MOS050_OPERATION,
                                        params,
                                        namespace_override=MOS050_NAMESPACE or None,
                                        body_tag_override=MOS050_BODY_TAG or None,
                                        env=env,
                                    )
                                    ok = int(response.get("status_code") or 0) < 400
                                    status_label = "OK" if ok else "NOK"
                                    error_message = None if ok else f"HTTP {response.get('status_code')}"
                                except Exception as exc:  # noqa: BLE001
                                    ok = False
                                    status_label = "NOK"
                                    error_message = str(exc)
                                    response = {"error": error_message}
                            _append_api_log(
                                "teilenummer_ips_mos50_montage",
                                params,
                                response,
                                ok,
                                error_message,
                                env=env_label,
                                wagon=wagon_ctx,
                                dry_run=dry_run,
                                request_url=request_url,
                                program=MOS050_SERVICE or "MOS050",
                                transaction=MOS050_OPERATION or "Montage",
                                request_method="POST",
                                status=status_label,
                            )
                            _update_teilenummer_row(
                                conn,
                                table_name,
                                row["seq"],
                                {"MOS050_STATUS": status_label},
                            )
                            processed += 1
                            _update_job(job["id"], processed=processed)
                            if ok or dry_run or not mwno:
                                break
                            if MOS050_RETRY_MAX and attempt >= MOS050_RETRY_MAX:
                                break
                            total_steps += 1
                            _update_job(job["id"], total=total_steps)
                            _append_job_log(
                                job["id"],
                                f"IPS MOS050: Warte {MOS050_RETRY_DELAY_SEC} Sekunden auf ERP ...",
                            )
                            if MOS050_RETRY_DELAY_SEC:
                                time.sleep(MOS050_RETRY_DELAY_SEC)
                            attempt += 1
                        if not ok and not dry_run:
                            row_hard_failed = True

                    # MOS125MI Einbau
                    if row_is_leading:
                        in_status = "SKIPPED: führendes Objekt (kein Einbau)"
                        _append_api_log(
                            "teilenummer_einbau",
                            {},
                            {"skipped": in_status},
                            True,
                            None,
                            env=env_label,
                            wagon=wagon_ctx,
                            dry_run=dry_run,
                            request_url="",
                            program="MOS125MI",
                            transaction="RemoveInstall",
                            status="SKIPPED",
                        )
                    elif row_hard_failed:
                        in_status = skip_after_error
                        _append_api_log(
                            "teilenummer_einbau",
                            {},
                            {"skipped": skip_after_error},
                            True,
                            None,
                            env=env_label,
                            wagon=wagon_ctx,
                            dry_run=dry_run,
                            request_url="",
                            program="MOS125MI",
                            transaction="RemoveInstall",
                            status="SKIPPED",
                        )
                    else:
                        params = _build_mos125_params(row_map, mode="in", env=env)
                        request_url = _build_m3_request_url(base_url, "MOS125MI", "RemoveInstall", params)
                        if not params.get("TRDT"):
                            ok = False
                            status_label = "ERROR"
                            error_message = "UMBAU_DATUM fehlt"
                            response = {"error": error_message}
                        elif dry_run:
                            ok = True
                            status_label = "DRYRUN"
                            error_message = None
                            response = {"dry_run": True}
                        else:
                            try:
                                response = call_m3_mi_get(base_url, token, "MOS125MI", "RemoveInstall", params)
                                ok, status_label, error_message = _mi_status(response)
                            except Exception as exc:  # noqa: BLE001
                                ok = False
                                status_label = "ERROR"
                                error_message = str(exc)
                                response = {"error": error_message}
                        in_status = status_label if ok else f"{status_label}: {error_message}"
                        _append_api_log(
                            "teilenummer_einbau",
                            params,
                            response,
                            ok,
                            error_message,
                            env=env_label,
                            wagon=wagon_ctx,
                            dry_run=dry_run,
                            request_url=request_url,
                            program="MOS125MI",
                            transaction="RemoveInstall",
                        )
                    _update_teilenummer_row(conn, table_name, row["seq"], {"IN_STATUS": in_status})
                    processed += 1
                    _update_job(job["id"], processed=processed)

                    row_result = "OK"
                    if row_hard_failed:
                        row_result = "FEHLER"
                    elif row_is_leading:
                        row_result = "OK_FUEHRENDES_OBJEKT"
                    _append_job_log(
                        job["id"],
                        f"{index}/{len(rows)} abgeschlossen [{row_result}]: {old_itno} {old_sern} -> {new_itno} {new_sern}",
                    )

            with _connect() as conn:
                source_table = _table_for(TEILENUMMER_TABLE, env)
                if _table_exists(conn, source_table):
                    _ensure_columns(conn, source_table, ["CHECKED"])
                    conn.execute(f'UPDATE "{source_table}" SET "CHECKED" = ""')
                conn.execute(f'DROP TABLE IF EXISTS "{table_name}"')
                conn.commit()

            _finish_job(
                job["id"],
                "success",
                result={"total": total_steps, "ok": processed, "error": 0},
            )
        except Exception as exc:  # noqa: BLE001
            _append_job_log(job["id"], f"Fehler: {exc}")
            _finish_job(job["id"], "error", error=str(exc))

    threading.Thread(target=_worker, daemon=True).start()
    return {
        "job_id": job["id"],
        "status": job["status"],
        "env": normalized_env,
        "dry_run": run_dry_run,
        "mode": "DRYRUN" if run_dry_run else "LIVE",
    }


@app.get("/api/objstrk")
def objstrk(
    mtrl: str = Query(..., min_length=1),
    sern: str = Query(..., min_length=1),
    store_table: str | None = Query(None),
    env: str = Query(DEFAULT_ENV),
) -> dict:
    """Lädt Objektstruktur über MOS256MI (Debug: rohe Antwort zurückgeben)."""
    ionapi = _ionapi_path(env, "mi")
    cmd = [
        sys.executable,
        str(PROJECT_ROOT / "python" / "m3_api_call.py"),
        "--program",
        "MOS256MI",
        "--transaction",
        "LstAsBuild",
        "--params-json",
        json.dumps({"MTRL": mtrl, "SERN": sern, "EXPA": "1", "MEVA": "1"}),
        "--ionapi",
        str(ionapi),
    ]
    result = subprocess.run(cmd, capture_output=True, text=True)
    if result.returncode != 0:
        raise HTTPException(status_code=500, detail=result.stderr or result.stdout or "MOS256 fehlgeschlagen")
    try:
        payload = json.loads(result.stdout)
    except json.JSONDecodeError as exc:
        raise HTTPException(status_code=500, detail=f"Ungültige MOS256 Antwort: {exc}") from exc

    if store_table:
        rows = _extract_mi_rows(payload)
        try:
            if store_table == RENUMBER_WAGON_TABLE:
                _clear_api_log()
            if store_table == RENUMBER_WAGON_TABLE and not rows:
                _clear_table_rows(store_table, env)
            else:
                _store_mi_rows(store_table, env, rows, wagon_itno=mtrl, wagon_sern=sern)
        except Exception as exc:  # noqa: BLE001
            raise HTTPException(status_code=500, detail=f"Objektstruktur speichern fehlgeschlagen: {exc}") from exc

    return payload


@app.post("/api/renumber/import_mrouhi")
def renumber_import_mrouhi(payload: dict = Body(...), env: str = Query(DEFAULT_ENV)) -> dict:
    rows = payload.get("rows") if isinstance(payload, dict) else None
    if not isinstance(rows, list) or not rows:
        raise HTTPException(status_code=400, detail="rows fehlt oder ist leer.")

    mapped_rows: List[Dict[str, Any]] = []
    wagon_itno = ""
    wagon_sern = ""
    for entry in rows:
        if not isinstance(entry, dict):
            continue
        hiit = (entry.get("HIIT") or "").strip()
        hisn = (entry.get("HISN") or "").strip()
        cfgl = (entry.get("CFGL") or "").strip()
        itno = (entry.get("ITNO") or "").strip()
        sern = (entry.get("SERN") or "").strip()
        remd = (entry.get("REMD") or "").strip()
        rmts = (entry.get("RMTS") or "").strip()
        if not hiit or not hisn or not itno:
            continue
        if not wagon_itno:
            wagon_itno = hiit
        if not wagon_sern:
            wagon_sern = hisn
        mapped_rows.append(
            {
                "WAGEN_ITNO": hiit,
                "WAGEN_SERN": hisn,
                "MTRL": hiit,
                "SERN": hisn,
                "ITNO": itno,
                "SER2": sern,
                "CFGL": cfgl,
                "MFGL": cfgl,
                "UMBAU_DATUM": remd,
                "RMTS": rmts,
                "OUT": "OK",
            }
        )

    if not mapped_rows:
        raise HTTPException(status_code=400, detail="Keine gueltigen Zeilen gefunden.")

    _store_mi_rows(RENUMBER_WAGON_TABLE, env, mapped_rows, wagon_itno=wagon_itno, wagon_sern=wagon_sern)
    return {
        "message": "RENUMBER_WAGON importiert.",
        "rows": len(mapped_rows),
        "env": _normalize_env(env),
        "wagon_itno": wagon_itno,
        "wagon_sern": wagon_sern,
    }


@app.get("/api/renumber/objstrk")
def renumber_objstrk(env: str = Query(DEFAULT_ENV)) -> dict:
    table_name = _table_for(RENUMBER_WAGON_TABLE, env)
    with _connect() as conn:
        if not _table_exists(conn, table_name):
            raise HTTPException(status_code=404, detail=f"Tabelle {table_name} nicht gefunden.")
        _ensure_renumber_schema(conn, table_name)
        rows = conn.execute(
            f"""SELECT rowid AS seq, * FROM "{table_name}"
            ORDER BY CASE
              WHEN "SEQ" IS NULL OR "SEQ" = '' THEN rowid
              ELSE CAST("SEQ" AS INTEGER)
            END ASC"""
        ).fetchall()

    wagon_itno = ""
    wagon_sern = ""
    records = []
    for row in rows:
        row_dict = dict(row)
        row_dict.pop("seq", None)
        if not wagon_itno:
            wagon_itno = row_dict.get("WAGEN_ITNO") or row_dict.get("MTRL") or ""
        if not wagon_sern:
            wagon_sern = row_dict.get("WAGEN_SERN") or row_dict.get("SERN") or ""
        name_values = [
            {"Name": str(key), "Value": "" if value is None else str(value)}
            for key, value in row_dict.items()
        ]
        records.append({"NameValue": name_values})

    return {"response": {"MIRecord": records}, "wagon_itno": wagon_itno, "wagon_sern": wagon_sern}


def _run_compass_sql(sql: str, env: str, table_name: str) -> List[Dict[str, Any]]:
    ionapi = _ionapi_path(env, "compass")
    cmd = [
        sys.executable,
        str(PROJECT_ROOT / "python" / "compass_to_sqlite.py"),
        "--ionapi",
        str(ionapi),
        "--sql",
        sql,
        "--sqlite-db",
        str(DB_PATH),
        "--table",
        table_name,
        "--mode",
        "replace",
    ]
    if _normalize_env(env) == "tst" and TST_COMPASS_JDBC.exists():
        cmd.extend(["--jdbc-jar", str(TST_COMPASS_JDBC)])
    result = subprocess.run(cmd, capture_output=True, text=True)
    if result.returncode != 0:
        raise HTTPException(
            status_code=500,
            detail=result.stderr or result.stdout or "Compass SQL fehlgeschlagen",
        )
    with _connect() as conn:
        if not _table_exists(conn, table_name):
            return []
        rows = [dict(row) for row in conn.execute(f'SELECT * FROM "{table_name}"').fetchall()]
        conn.execute(f'DROP TABLE IF EXISTS "{table_name}"')
        conn.commit()
    return rows


def _load_mrouhi_rows(hisn: str, env: str) -> List[Dict[str, Any]]:
    cleaned = (hisn or "").strip()
    if not cleaned:
        return []
    safe = cleaned.replace("'", "''")
    sql = (
        "SELECT "
        "a.HIIT, a.HISN, a.CFGL, a.ITNO, a.SERN, a.REMD, a.RMTS, "
        "b.EQTP, b.STAT "
        "FROM MROUHI a "
        "LEFT OUTER JOIN MILOIN b ON a.SERN = b.SERN "
        f"WHERE a.HISN = '{safe}' ORDER BY a.CFGL"
    )
    table_name = f"mrouhi_tmp_{uuid.uuid4().hex[:10]}"
    return _run_compass_sql(sql, env, table_name)


def _fetch_objstrk_rows(mtrl: str, sern: str, env: str) -> List[Dict[str, Any]]:
    ionapi = _ionapi_path(env, "mi")
    cmd = [
        sys.executable,
        str(PROJECT_ROOT / "python" / "m3_api_call.py"),
        "--program",
        "MOS256MI",
        "--transaction",
        "LstAsBuild",
        "--params-json",
        json.dumps({"MTRL": mtrl, "SERN": sern, "EXPA": "1", "MEVA": "1"}),
        "--ionapi",
        str(ionapi),
    ]
    result = subprocess.run(cmd, capture_output=True, text=True, timeout=120)
    if result.returncode != 0:
        raise HTTPException(status_code=500, detail=result.stderr or result.stdout or "MOS256 fehlgeschlagen")
    try:
        payload = json.loads(result.stdout)
    except json.JSONDecodeError as exc:
        raise HTTPException(status_code=500, detail=f"Ungültige MOS256 Antwort: {exc}") from exc
    return _extract_mi_rows(payload)


def _remd_is_blank(value: str) -> bool:
    normalized = str(value or "").strip()
    return normalized in {"", "0", "00000000"}


def _cfgl_segments(value: str) -> List[int]:
    parts = [part for part in str(value or "").split("-") if part != ""]
    segments: List[int] = []
    for part in parts:
        if part.isdigit():
            segments.append(int(part))
        else:
            digits = "".join(ch for ch in part if ch.isdigit())
            segments.append(int(digits) if digits else 0)
    return segments


def _cfgl_sort_key_desc(value: str) -> tuple:
    segments = _cfgl_segments(value)
    return (len(segments), segments)


def _parent_cfgl_for(cfgl: str) -> str:
    segments = [part for part in str(cfgl or "").split("-") if part != ""]
    if len(segments) == 4 and segments[2] == "01":
        return f"{segments[0]}-{segments[1]}-{segments[3]}"
    return cfgl.rsplit("-", 1)[0] if "-" in cfgl else ""


def _build_mrouhi_preview_rows(hisn: str, env: str) -> List[Dict[str, str]]:
    rows = _load_mrouhi_rows(hisn, env)
    entries: List[Dict[str, str]] = []
    for idx, entry in enumerate(rows):
        cfgl = (entry.get("CFGL") or "").strip()
        if not cfgl:
            continue
        remd = (entry.get("REMD") or "").strip()
        if _remd_is_blank(remd):
            continue
        entries.append(
            {
                "idx": str(idx),
                "HIIT": (entry.get("HIIT") or "").strip(),
                "HISN": (entry.get("HISN") or "").strip(),
                "CFGL": cfgl,
                "ITNO": (entry.get("ITNO") or "").strip(),
                "SERN": (entry.get("SERN") or "").strip(),
                "REMD": (entry.get("REMD") or "").strip(),
                "RMTS": (entry.get("RMTS") or "").strip(),
                "EQTP": (entry.get("EQTP") or "").strip(),
                "STAT": (entry.get("STAT") or "").strip(),
            }
        )
    cfgl_map: Dict[str, List[Dict[str, str]]] = {}
    for entry in entries:
        cfgl_map.setdefault(entry["CFGL"], []).append(entry)
    cfgl_counts = {cfgl: len(items) for cfgl, items in cfgl_map.items()}
    sorted_entries = sorted(
        entries,
        key=lambda entry: (
            _cfgl_sort_key_desc(entry["CFGL"]),
            entry["idx"],
        ),
        reverse=True,
    )
    eqtp_parent_map = {"110": "106"}
    child_indexes: Dict[str, int] = {}
    preview: List[Dict[str, str]] = []
    for entry in sorted_entries:
        cfgl = entry["CFGL"]
        parent_cfgl = _parent_cfgl_for(cfgl)
        parent_itno = entry["HIIT"]
        parent_sern = entry["HISN"]
        parent_candidates = []
        if parent_cfgl:
            parent_candidates = cfgl_map.get(parent_cfgl) or []
            child_count = cfgl_counts.get(cfgl, 0)
            if child_count > len(parent_candidates):
                parent_prefix = parent_cfgl.rsplit("-", 1)[0] if "-" in parent_cfgl else ""
                if parent_prefix:
                    target_depth = len(_cfgl_segments(parent_cfgl))
                    for cfgl_key, candidates in cfgl_map.items():
                        if cfgl_key == parent_cfgl:
                            continue
                        if not cfgl_key.startswith(parent_prefix + "-"):
                            continue
                        if len(_cfgl_segments(cfgl_key)) != target_depth:
                            continue
                        parent_candidates.extend(candidates)
        if parent_candidates:
            expected_parent_eqtp = eqtp_parent_map.get(entry.get("EQTP", ""))
            if expected_parent_eqtp:
                filtered = [
                    candidate
                    for candidate in parent_candidates
                    if candidate.get("EQTP", "") == expected_parent_eqtp
                ]
                if filtered:
                    parent_candidates = filtered
            parent_candidates = sorted(
                parent_candidates,
                key=lambda candidate: (candidate.get("SERN") or "", candidate.get("ITNO") or ""),
            )
            child_index = child_indexes.get(cfgl, 0)
            child_indexes[cfgl] = child_index + 1
            if len(parent_candidates) >= cfgl_counts.get(cfgl, 0):
                chosen = parent_candidates[child_index]
            else:
                chosen = parent_candidates[child_index % len(parent_candidates)]
            parent_itno = chosen.get("ITNO", "") or parent_itno
            parent_sern = chosen.get("SERN", "") or parent_sern
        preview.append(
            {
                "CFGL": cfgl,
                "ITNO": entry["ITNO"],
                "SERN": entry["SERN"],
                "REMD": entry["REMD"],
                "RMTS": entry["RMTS"],
                "PARENT_CFGL": parent_cfgl,
                "PARENT_ITNO": parent_itno,
                "PARENT_SERN": parent_sern,
            }
        )
    return preview


def _build_mrouhi_parent_candidates(hisn: str, env: str) -> Dict[str, List[Dict[str, str]]]:
    rows = _load_mrouhi_rows(hisn, env)
    entries: List[Dict[str, str]] = []
    for entry in rows:
        cfgl = (entry.get("CFGL") or "").strip()
        if not cfgl:
            continue
        remd = (entry.get("REMD") or "").strip()
        if _remd_is_blank(remd):
            continue
        entries.append(
            {
                "CFGL": cfgl,
                "ITNO": (entry.get("ITNO") or "").strip(),
                "SERN": (entry.get("SERN") or "").strip(),
                "EQTP": (entry.get("EQTP") or "").strip(),
            }
        )
    cfgl_map: Dict[str, List[Dict[str, str]]] = {}
    for entry in entries:
        cfgl_map.setdefault(entry["CFGL"], []).append(entry)
    cfgl_counts = {cfgl: len(items) for cfgl, items in cfgl_map.items()}
    eqtp_parent_map = {"110": "106"}
    candidate_map: Dict[str, List[Dict[str, str]]] = {}
    for entry in entries:
        cfgl = entry["CFGL"]
        parent_cfgl = _parent_cfgl_for(cfgl)
        if not parent_cfgl:
            continue
        parent_candidates = list(cfgl_map.get(parent_cfgl) or [])
        child_count = cfgl_counts.get(cfgl, 0)
        if child_count > len(parent_candidates):
            parent_prefix = parent_cfgl.rsplit("-", 1)[0] if "-" in parent_cfgl else ""
            if parent_prefix:
                target_depth = len(_cfgl_segments(parent_cfgl))
                for cfgl_key, candidates in cfgl_map.items():
                    if cfgl_key == parent_cfgl:
                        continue
                    if not cfgl_key.startswith(parent_prefix + "-"):
                        continue
                    if len(_cfgl_segments(cfgl_key)) != target_depth:
                        continue
                    parent_candidates.extend(candidates)
        expected_parent_eqtp = eqtp_parent_map.get(entry.get("EQTP", ""))
        if expected_parent_eqtp:
            filtered = [
                candidate
                for candidate in parent_candidates
                if candidate.get("EQTP", "") == expected_parent_eqtp
            ]
            if filtered:
                parent_candidates = filtered
        parent_candidates = sorted(
            parent_candidates,
            key=lambda candidate: (candidate.get("SERN") or "", candidate.get("ITNO") or ""),
        )
        candidate_map[cfgl] = parent_candidates
    return candidate_map


def _run_rollback_job(job: dict, env: str) -> None:
    try:
        table_name = _table_for(RENUMBER_WAGON_TABLE, env)
        with _connect() as conn:
            if not _table_exists(conn, table_name):
                raise HTTPException(status_code=404, detail=f"Tabelle {table_name} nicht gefunden.")
            _ensure_renumber_schema(conn, table_name)
            rows = conn.execute(
                f"""SELECT rowid AS seq, * FROM "{table_name}"
                ORDER BY CASE
                  WHEN "SEQ" IS NULL OR "SEQ" = '' THEN rowid
                  ELSE CAST("SEQ" AS INTEGER)
                END ASC"""
            ).fetchall()

        target_rows = [row for row in rows if _row_value(row, "OUT") in {"OK", "DRYRUN"}]
        target_rows = sorted(target_rows, key=lambda row: row["seq"])
        total = len(target_rows)
        _update_job(job["id"], total=total, processed=0, results=[])
        _append_job_log(job["id"], f"Starte Roll-Back Einbau: {total} Positionen.")

        dry_run = _effective_dry_run(env)
        parent_candidates_map: Dict[str, List[Dict[str, str]]] = {}
        if target_rows:
            wagon_sern = (_row_value(target_rows[0], "WAGEN_SERN") or "").strip()
            if wagon_sern:
                parent_candidates_map = _build_mrouhi_parent_candidates(wagon_sern, env)
        ionapi_path = _ionapi_path(env, "mi")
        ion_cfg = load_ionapi_config(str(ionapi_path))
        base_url = build_base_url(ion_cfg)
        token = ""
        if not dry_run:
            token = get_access_token_service_account(ion_cfg)

        ok_count = 0
        error_count = 0
        env_label = _normalize_env(env).upper()
        with _connect() as conn:
            for idx, row in enumerate(target_rows, start=1):
                params = _build_mos125_params(row, mode="in", env=env)
                log_params = {
                    "ITNO": params.get("ITNI", ""),
                    "SERN": params.get("BANI", ""),
                    "PARENT_ITNO": params.get("NHAI", ""),
                    "PARENT_SERN": params.get("NHSI", ""),
                }
                wagon_ctx = _wagon_log_context(row)
                request_url = ""
                if not params["TRDT"]:
                    status = "ERROR: UMBAU_DATUM fehlt"
                    ok = False
                    _append_api_log(
                        "rollback",
                        log_params,
                        {"error": "UMBAU_DATUM fehlt"},
                        ok,
                        "UMBAU_DATUM fehlt",
                        env=env_label,
                        wagon=wagon_ctx,
                        dry_run=dry_run,
                        request_url=request_url,
                    )
                elif dry_run:
                    status = "DRYRUN"
                    ok = True
                    _append_api_log(
                        "rollback",
                        log_params,
                        {"dry_run": True},
                        ok,
                        None,
                        env=env_label,
                        wagon=wagon_ctx,
                        dry_run=dry_run,
                        request_url=request_url,
                    )
                else:
                    try:
                        response = call_m3_mi_get(base_url, token, "MOS125MI", "RemoveInstall", params)
                        ok, status_label, error_message = _mi_status(response)
                        code, _ = _mi_extract_code_message(response)
                        status = status_label if ok else f"ERROR: {status_label}"
                        if ok and status_label == "OK_IDEMPOTENT":
                            status = "OK: bereits installiert"
                        _append_api_log(
                            "rollback",
                            log_params,
                            response,
                            ok,
                            error_message,
                            env=env_label,
                            wagon=wagon_ctx,
                            dry_run=dry_run,
                            request_url=request_url,
                            status=status_label,
                        )
                        retry_codes = {"MO12524", "MO12528"}
                        cfgl = params.get("CFGL", "")
                        candidates = parent_candidates_map.get(cfgl, []) if cfgl else []
                        if code in retry_codes and len(candidates) > 1:
                            current_parent = (params.get("NHAI", ""), params.get("NHSI", ""))
                            for candidate in candidates:
                                candidate_parent = (candidate.get("ITNO", ""), candidate.get("SERN", ""))
                                if candidate_parent == current_parent:
                                    continue
                                params["NHAI"] = candidate_parent[0]
                                params["NHSI"] = candidate_parent[1]
                                log_params = {
                                    "ITNO": params.get("ITNI", ""),
                                    "SERN": params.get("BANI", ""),
                                    "PARENT_ITNO": params.get("NHAI", ""),
                                    "PARENT_SERN": params.get("NHSI", ""),
                                }
                                response = call_m3_mi_get(base_url, token, "MOS125MI", "RemoveInstall", params)
                                ok, status_label, error_message = _mi_status(response)
                                code, _ = _mi_extract_code_message(response)
                                status = status_label if ok else f"ERROR: {status_label}"
                                if ok and status_label == "OK_IDEMPOTENT":
                                    status = "OK: bereits installiert"
                                _append_api_log(
                                    "rollback",
                                    log_params,
                                    response,
                                    ok,
                                    error_message,
                                    env=env_label,
                                    wagon=wagon_ctx,
                                    dry_run=dry_run,
                                    request_url=request_url,
                                    status=status_label,
                                )
                                if code not in retry_codes:
                                    break
                    except Exception as exc:  # noqa: BLE001
                        status = f"ERROR: {exc}"
                        ok = False
                        _append_api_log(
                            "rollback",
                            log_params,
                            {"error": str(exc)},
                            ok,
                            str(exc),
                            env=env_label,
                            wagon=wagon_ctx,
                            dry_run=dry_run,
                            request_url=request_url,
                            status="ERROR",
                        )

                conn.execute(
                    f'UPDATE "{table_name}" SET "ROLLBACK"=?, "TIMESTAMP_ROLLBACK"=? WHERE rowid=?',
                    (status, datetime.utcnow().isoformat(sep=" ", timespec="seconds"), row["seq"]),
                )
                conn.commit()

                result = {
                    "seq": row["seq"],
                    "cfgr": params.get("CFGL") or params.get("CFGR") or "",
                    "itno": params.get("ITNI") or params.get("ITNR") or _row_value(row, "ITNO"),
                    "ser2": _row_value(row, "SER2"),
                    "rollback": status,
                    "ok": ok,
                }
                with _jobs_lock:
                    job_ref = _jobs.get(job["id"])
                    if job_ref is not None:
                        job_ref["processed"] = idx
                        job_ref.setdefault("results", []).append(result)
                if ok:
                    ok_count += 1
                else:
                    error_count += 1
                status_label = "ERROR" if not ok else ("DRYRUN" if dry_run else "OK")
                _append_job_log(
                    job["id"],
                    f"{idx}/{total} {status_label} CFGL={params.get('CFGL', '')} ITNI={params.get('ITNI', '')} "
                    f"BANI={params.get('BANI', '')}",
                )

        _finish_job(
            job["id"],
            "success",
            result={"total": total, "ok": ok_count, "error": error_count},
        )
    except Exception as exc:  # noqa: BLE001
        _append_job_log(job["id"], f"Fehler: {exc}")
        _finish_job(job["id"], "error", error=str(exc))


@app.post("/api/renumber/rollback_from_mrouhi")
def renumber_rollback_from_mrouhi(
    payload: dict | None = Body(None),
    hisn: str = Query(""),
    env: str = Query(DEFAULT_ENV),
) -> dict:
    resolved = (hisn or "").strip()
    if not resolved and isinstance(payload, dict):
        resolved = (payload.get("hisn") or "").strip()
    if resolved.startswith("INFO:"):
        resolved = ""
    hisn = resolved
    if not hisn:
        raise HTTPException(status_code=400, detail="HISN fehlt.")
    job = _create_job("renumber_rollback_mrouhi", env)

    def _worker() -> None:
        try:
            _append_job_log(job["id"], f"MROUHI Import für HISN {hisn} ...")
            _append_api_log(
                "mrouhi_import_start",
                {"hisn": hisn},
                {"status": "starting"},
                True,
                None,
                env=_normalize_env(env).upper(),
                wagon={"sern": hisn},
                dry_run=_effective_dry_run(env),
                program="COMPASS",
                transaction="MROUHI",
            )
            rows = _load_mrouhi_rows(hisn, env)
            if not rows:
                raise HTTPException(status_code=404, detail="Keine MROUHI Daten gefunden.")
            entries: List[Dict[str, str]] = []
            wagon_itno = ""
            wagon_sern = ""
            for entry in rows:
                hiit = (entry.get("HIIT") or "").strip()
                hisn_value = (entry.get("HISN") or "").strip()
                cfgl = (entry.get("CFGL") or "").strip()
                itno = (entry.get("ITNO") or "").strip()
                sern = (entry.get("SERN") or "").strip()
                remd = (entry.get("REMD") or "").strip()
                rmts = (entry.get("RMTS") or "").strip()
                if not hiit or not hisn_value or not itno:
                    continue
                if _remd_is_blank(remd):
                    continue
                if not wagon_itno:
                    wagon_itno = hiit
                if not wagon_sern:
                    wagon_sern = hisn_value
                entries.append(
                    {
                        "HIIT": hiit,
                        "HISN": hisn_value,
                        "CFGL": cfgl,
                        "ITNO": itno,
                        "SERN": sern,
                        "REMD": remd,
                        "RMTS": rmts,
                    }
                )

            if not entries:
                raise HTTPException(status_code=400, detail="Keine gültigen MROUHI Zeilen gefunden.")

            mapped_rows: List[Dict[str, Any]] = []
            preview_rows = _build_mrouhi_preview_rows(hisn, env)
            missing_remd = 0
            for entry in preview_rows:
                cfgl = entry.get("CFGL", "").strip()
                itno = entry.get("ITNO", "").strip()
                sern = entry.get("SERN", "").strip()
                parent_itno = entry.get("PARENT_ITNO", "").strip()
                parent_sern = entry.get("PARENT_SERN", "").strip()
                remd = entry.get("REMD", "").strip()
                rmts = entry.get("RMTS", "").strip()
                if not remd:
                    missing_remd += 1
                mapped_rows.append(
                    {
                        "WAGEN_ITNO": wagon_itno,
                        "WAGEN_SERN": wagon_sern,
                        "MTRL": parent_itno,
                        "SERN": parent_sern,
                        "ITNO": itno,
                        "SER2": sern,
                        "CFGL": cfgl,
                        "MFGL": cfgl,
                        "UMBAU_DATUM": remd or "",
                        "RMTS": rmts or "",
                        "OUT": "OK",
                    }
                )

            if not mapped_rows:
                raise HTTPException(status_code=400, detail="Keine gültigen MROUHI Zeilen für Rollback.")

            _store_mi_rows(
                RENUMBER_WAGON_TABLE,
                env,
                mapped_rows,
                wagon_itno=wagon_itno,
                wagon_sern=wagon_sern,
            )
            _append_api_log(
                "mrouhi_import",
                {"hisn": hisn},
                {
                    "rows": len(mapped_rows),
                    "wagon_itno": wagon_itno,
                    "wagon_sern": wagon_sern,
                    "missing_remd": missing_remd,
                    "preview_rows": len(preview_rows),
                },
                True,
                None,
                env=_normalize_env(env).upper(),
                wagon={"itno": wagon_itno, "sern": wagon_sern},
                dry_run=_effective_dry_run(env),
                program="COMPASS",
                transaction="MROUHI",
            )
            _append_job_log(job["id"], f"Import abgeschlossen: {len(mapped_rows)} Positionen.")
        except Exception as exc:  # noqa: BLE001
            _append_job_log(job["id"], f"Fehler: {exc}")
            _finish_job(job["id"], "error", error=str(exc))
            return
        _run_rollback_job(job, env)

    threading.Thread(target=_worker, daemon=True).start()
    return {"job_id": job["id"], "status": job["status"], "env": job["env"]}


@app.get("/api/renumber/rollback_preview")
def renumber_rollback_preview(hisn: str = Query(..., min_length=1), env: str = Query(DEFAULT_ENV)) -> dict:
    preview_rows = _build_mrouhi_preview_rows(hisn, env)
    if not preview_rows:
        raise HTTPException(status_code=404, detail="Keine Preview-Zeilen gefunden.")
    env_label = _normalize_env(env).upper()
    for idx, row in enumerate(preview_rows, start=1):
        _append_api_log(
            "rollback_preview",
            {"hisn": hisn},
            {"index": idx, **row},
            True,
            None,
            env=env_label,
            wagon={"sern": hisn},
            dry_run=_effective_dry_run(env),
            program="MROUHI",
            transaction="PREVIEW",
        )
    return {"rows": len(preview_rows), "env": _normalize_env(env)}


@app.post("/api/renumber/update")
def renumber_update(payload: dict = Body(...), env: str = Query(DEFAULT_ENV)) -> dict:
    table_name = _table_for(RENUMBER_WAGON_TABLE, env)
    new_sern = (payload.get("new_sern") or "").strip()
    new_baureihe = (payload.get("new_baureihe") or "").strip()
    umbau_datum = (payload.get("umbau_datum") or "").strip()
    umbau_art = (payload.get("umbau_art") or "").strip()
    if not new_sern or not new_baureihe or not umbau_datum or not umbau_art:
        raise HTTPException(status_code=400, detail="Pflichtfelder fehlen.")
    timestamp = datetime.utcnow().isoformat(sep=" ", timespec="seconds")
    with _connect() as conn:
        if not _table_exists(conn, table_name):
            raise HTTPException(status_code=404, detail=f"Tabelle {table_name} nicht gefunden.")
        _ensure_renumber_schema(conn, table_name)
        conn.execute(
            f"""UPDATE "{table_name}" SET
            "NEW_SERN"=?,
            "NEW_BAUREIHE"=?,
            "UMBAU_DATUM"=?,
            "UMBAU_ART"=?,
            "NEW_PART_ITNO"=?,
            "NEW_PART_SER2"=?,
            "PLPN"=?,
            "MWNO"=?,
            "MOS100_STATUS"=?,
            "MOS180_STATUS"=?,
            "MOS050_STATUS"=?,
            "CRS335_STATUS"=?,
            "STS046_STATUS"=?,
            "STS046_ADD_STATUS"=?,
            "MMS240_STATUS"=?,
            "CUSEXT_STATUS"=?,
            "OUT"=?,
            "UPDATED_AT"=?,
            "IN"=?,
            "TIMESTAMP_IN"=?,
            "ROLLBACK"=?,
            "TIMESTAMP_ROLLBACK"=?
            """,
            (
                new_sern,
                new_baureihe,
                umbau_datum,
                umbau_art,
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                "",
                timestamp,
                "",
                "",
                "",
                "",
            ),
        )
        rows = conn.execute(f'SELECT rowid AS seq, * FROM "{table_name}"').fetchall()
        for row in rows:
            new_itno, new_ser2 = _compute_part_updates(row, new_sern, new_baureihe)
            conn.execute(
                f'UPDATE "{table_name}" SET "NEW_PART_ITNO"=?, "NEW_PART_SER2"=? WHERE rowid=?',
                (new_itno, new_ser2, row["seq"]),
            )
        conn.commit()
        total = conn.execute(f'SELECT COUNT(*) FROM "{table_name}"').fetchone()[0]
    return {"table": table_name, "updated": total, "env": _normalize_env(env)}


def _renumber_pending_count(conn: sqlite3.Connection, table_name: str, mode: str) -> int:
    needs_renumber_clause = (
        'IFNULL("SER2", \'\') <> \'\' AND '
        '(IFNULL("NEW_PART_ITNO", \'\') <> \'\' OR IFNULL("NEW_PART_SER2", \'\') <> \'\')'
    )
    if mode == "out":
        query = f'SELECT COUNT(*) FROM "{table_name}" WHERE IFNULL("OUT", \'\') = \'\''
    elif mode == "in":
        query = f'SELECT COUNT(*) FROM "{table_name}" WHERE IFNULL("IN", \'\') = \'\''
    elif mode == "mos170":
        query = (
            f'SELECT COUNT(*) FROM "{table_name}" WHERE {needs_renumber_clause} '
            'AND IFNULL("PLPN", \'\') = \'\''
        )
    elif mode == "mos170_plpn":
        query = (
            f'SELECT COUNT(*) FROM "{table_name}" WHERE IFNULL("PLPN", \'\') <> \'\' '
            'AND IFNULL("MWNO", \'\') = \'\''
        )
    elif mode == "mos100":
        query = (
            f'SELECT COUNT(*) FROM "{table_name}" WHERE IFNULL("MWNO", \'\') <> \'\' '
            'AND (IFNULL("NEW_PART_ITNO", \'\') <> \'\' OR IFNULL("NEW_PART_SER2", \'\') <> \'\') '
            'AND IFNULL("MOS100_STATUS", \'\') = \'\''
        )
    elif mode == "mos180":
        query = (
            f'SELECT COUNT(*) FROM "{table_name}" WHERE IFNULL("MWNO", \'\') <> \'\' '
            'AND IFNULL("MOS180_STATUS", \'\') = \'\''
        )
    elif mode == "mos050":
        query = (
            f'SELECT COUNT(*) FROM "{table_name}" WHERE IFNULL("MWNO", \'\') <> \'\' '
            f'AND {needs_renumber_clause} AND IFNULL("MOS050_STATUS", \'\') = \'\''
        )
    elif mode == "crs335":
        query = f'SELECT COUNT(*) FROM "{table_name}" WHERE IFNULL("CRS335_STATUS", \'\') = \'\''
    elif mode == "sts046":
        query = f'SELECT COUNT(*) FROM "{table_name}" WHERE IFNULL("STS046_STATUS", \'\') = \'\''
    elif mode == "sts046_add":
        query = f'SELECT COUNT(*) FROM "{table_name}" WHERE IFNULL("STS046_ADD_STATUS", \'\') = \'\''
    elif mode == "mms240":
        query = f'SELECT COUNT(*) FROM "{table_name}" WHERE IFNULL("MMS240_STATUS", \'\') = \'\''
    elif mode == "cusext":
        query = f'SELECT COUNT(*) FROM "{table_name}" WHERE IFNULL("CUSEXT_STATUS", \'\') = \'\''
    elif mode == "rollback":
        query = (
            f'SELECT COUNT(*) FROM "{table_name}" '
            'WHERE IFNULL("OUT", \'\') IN (\'OK\', \'DRYRUN\') '
            'AND IFNULL("ROLLBACK", \'\') = \'\''
        )
    elif mode == "wagon_renumber":
        return 0
    else:
        raise ValueError(f"Unbekannter Modus: {mode}")
    return int(conn.execute(query).fetchone()[0] or 0)


@app.get("/api/renumber/pending")
def renumber_pending(mode: str, env: str = Query(DEFAULT_ENV)) -> dict:
    normalized = (mode or "").strip().lower()
    if not normalized:
        raise HTTPException(status_code=400, detail="Mode fehlt.")
    table_name = _table_for(RENUMBER_WAGON_TABLE, env)
    with _connect() as conn:
        if not _table_exists(conn, table_name):
            raise HTTPException(status_code=404, detail=f"Tabelle {table_name} nicht gefunden.")
        _ensure_renumber_schema(conn, table_name)
        try:
            pending = _renumber_pending_count(conn, table_name, normalized)
        except ValueError as exc:
            raise HTTPException(status_code=400, detail=str(exc)) from exc
    return {"mode": normalized, "pending": pending, "env": _normalize_env(env)}


@app.post("/api/renumber/run")
def renumber_run(env: str = Query(DEFAULT_ENV)) -> dict:
    job = _create_job("renumber_run", env)

    def _worker() -> None:
        try:
            table_name = _table_for(RENUMBER_WAGON_TABLE, env)
            with _connect() as conn:
                if not _table_exists(conn, table_name):
                    raise HTTPException(status_code=404, detail=f"Tabelle {table_name} nicht gefunden.")
                _ensure_renumber_schema(conn, table_name)
                rows = conn.execute(
                    f"""SELECT rowid AS seq, * FROM "{table_name}"
                    ORDER BY CASE
                      WHEN "SEQ" IS NULL OR "SEQ" = '' THEN rowid
                      ELSE CAST("SEQ" AS INTEGER)
                    END ASC"""
                ).fetchall()

            total = len(rows)
            _update_job(job["id"], total=total, processed=0, results=[])
            _append_job_log(job["id"], "Teile werden ausgebaut")

            dry_run = _effective_dry_run(env)
            ionapi_path = _ionapi_path(env, "mi")
            ion_cfg = load_ionapi_config(str(ionapi_path))
            base_url = build_base_url(ion_cfg)
            token = ""
            if not dry_run:
                token = get_access_token_service_account(ion_cfg)

            ok_count = 0
            error_count = 0
            env_label = _normalize_env(env).upper()
            with _connect() as conn:
                for idx, row in enumerate(rows, start=1):
                    params = _build_mos125_params(row, mode="out", env=env)
                    wagon_ctx = _wagon_log_context(row)
                    request_url = _build_m3_request_url(base_url, "MOS125MI", "RemoveInstall", params)
                    if not params["TRDT"]:
                        out = "ERROR: UMBAU_DATUM fehlt"
                        ok = False
                        _append_api_log(
                            "ausbau",
                            params,
                            {"error": "UMBAU_DATUM fehlt"},
                            ok,
                            "UMBAU_DATUM fehlt",
                            env=env_label,
                            wagon=wagon_ctx,
                            dry_run=dry_run,
                            request_url=request_url,
                        )
                    elif dry_run:
                        out = "DRYRUN"
                        ok = True
                        _append_api_log(
                            "ausbau",
                            params,
                            {"dry_run": True},
                            ok,
                            None,
                            env=env_label,
                            wagon=wagon_ctx,
                            dry_run=dry_run,
                            request_url=request_url,
                        )
                    else:
                        try:
                            response = call_m3_mi_get(
                                base_url, token, "MOS125MI", "RemoveInstall", params
                            )
                            error_message = _mi_error_message(response)
                            if error_message:
                                out = f"ERROR: {error_message}"
                                ok = False
                            else:
                                out = "OK"
                                ok = True
                            _append_api_log(
                                "ausbau",
                                params,
                                response,
                                ok,
                                error_message,
                                env=env_label,
                                wagon=wagon_ctx,
                                dry_run=dry_run,
                                request_url=request_url,
                            )
                        except Exception as exc:  # noqa: BLE001
                            out = f"ERROR: {exc}"
                            ok = False
                            _append_api_log(
                                "ausbau",
                                params,
                                {"error": str(exc)},
                                ok,
                                str(exc),
                                env=env_label,
                                wagon=wagon_ctx,
                                dry_run=dry_run,
                                request_url=request_url,
                            )

                    conn.execute(
                        f'UPDATE "{table_name}" SET "OUT"=?, "UPDATED_AT"=? WHERE rowid=?',
                        (out, datetime.utcnow().isoformat(sep=" ", timespec="seconds"), row["seq"]),
                    )
                    conn.commit()

                    result = {
                        "seq": row["seq"],
                        "cfgr": params["CFGR"],
                        "itno": params["ITNR"],
                        "ser2": params["BANR"],
                        "out": out,
                        "ok": ok,
                    }
                    with _jobs_lock:
                        job_ref = _jobs.get(job["id"])
                        if job_ref is not None:
                            job_ref["processed"] = idx
                            job_ref.setdefault("results", []).append(result)
                    if ok:
                        ok_count += 1
                    else:
                        error_count += 1
                    status = "ERROR" if not ok else ("DRYRUN" if dry_run else "OK")
                    _append_job_log(
                        job["id"],
                        f"Teile werden ausgebaut: {idx}/{len(rows)} {status}",
                    )

            _finish_job(
                job["id"],
                "success",
                result={"total": total, "ok": ok_count, "error": error_count},
            )
        except Exception as exc:  # noqa: BLE001
            _append_job_log(job["id"], f"Fehler: {exc}")
            _finish_job(job["id"], "error", error=str(exc))

    threading.Thread(target=_worker, daemon=True).start()
    return {"job_id": job["id"], "status": job["status"], "env": job["env"]}


@app.post("/api/renumber/mos170")
def renumber_mos170(env: str = Query(DEFAULT_ENV)) -> dict:
    job = _create_job("mos170_addprop", env)

    def _worker() -> None:
        try:
            table_name = _table_for(RENUMBER_WAGON_TABLE, env)
            with _connect() as conn:
                if not _table_exists(conn, table_name):
                    raise HTTPException(status_code=404, detail=f"Tabelle {table_name} nicht gefunden.")
                _ensure_renumber_schema(conn, table_name)
                rows = conn.execute(
                    f"""SELECT rowid AS seq, * FROM "{table_name}"
                    ORDER BY CASE
                      WHEN "SEQ" IS NULL OR "SEQ" = '' THEN rowid
                      ELSE CAST("SEQ" AS INTEGER)
                    END ASC"""
                ).fetchall()

            target_rows = [row for row in rows if _needs_renumber(row)]
            total = len(target_rows)
            _update_job(job["id"], total=total, processed=0, results=[])
            _append_job_log(job["id"], f"MOS170MI AddProp: {total} Positionen.")

            dry_run = _effective_dry_run(env)
            ionapi_path = _ionapi_path(env, "mi")
            ion_cfg = load_ionapi_config(str(ionapi_path))
            base_url = build_base_url(ion_cfg)
            token = ""
            if not dry_run:
                token = get_access_token_service_account(ion_cfg)

            env_label = _normalize_env(env).upper()
            ok_count = 0
            error_count = 0
            processed = 0
            pending_rows = list(target_rows)
            attempt = 1
            while pending_rows:
                if MOS170_RETRY_MAX and attempt > MOS170_RETRY_MAX:
                    _append_job_log(
                        job["id"],
                        f"MOS170MI AddProp: Abbruch nach {MOS170_RETRY_MAX} Versuchen, {len(pending_rows)} ohne PLPN.",
                    )
                    break
                _append_job_log(
                    job["id"],
                    f"MOS170MI AddProp: Versuch {attempt} für {len(pending_rows)} Positionen.",
                )
                next_pending = []
                with _connect() as conn:
                    for row in pending_rows:
                        params = _build_mos170_params(row, env=env)
                        request_url = _build_m3_request_url(base_url, "MOS170MI", "AddProp", params)
                        required_missing = not params.get("ITNO") or not params.get("BANO") or not params.get("STDT")
                        if required_missing:
                            ok = False
                            error_message = "Pflichtfelder fehlen"
                            response = {"error": error_message}
                        elif dry_run:
                            ok = True
                            error_message = None
                            response = {"dry_run": True}
                        else:
                            try:
                                response = call_m3_mi_get(base_url, token, "MOS170MI", "AddProp", params)
                                error_message = _mi_error_message(response)
                                ok = not bool(error_message)
                            except Exception as exc:  # noqa: BLE001
                                response = {"error": str(exc)}
                                error_message = str(exc)
                                ok = False

                        plpn = _extract_plpn(response) if ok else ""
                        log_response = {"plpn": plpn, "response": response}
                        conn.execute(
                            f'UPDATE "{table_name}" SET "PLPN"=? WHERE rowid=?',
                            (plpn, row["seq"]),
                        )
                        conn.commit()
                        _append_api_log(
                            "ih_addprop",
                            params,
                            log_response,
                            ok,
                            error_message,
                            env=env_label,
                            wagon=_wagon_log_context(row),
                            dry_run=dry_run,
                            request_url=request_url,
                            program="MOS170MI",
                            transaction="AddProp",
                        )
                        if not plpn:
                            _append_api_log(
                                "ih_addprop_missing_plpn",
                                params,
                                log_response,
                                False,
                                "PLPN fehlt",
                                env=env_label,
                                wagon=_wagon_log_context(row),
                                dry_run=dry_run,
                                request_url=request_url,
                                program="MOS170MI",
                                transaction="AddProp",
                            )
                            next_pending.append(row)

                        processed += 1
                        with _jobs_lock:
                            job_ref = _jobs.get(job["id"])
                            if job_ref is not None:
                                job_ref["processed"] = processed
                        if ok:
                            ok_count += 1
                        else:
                            error_count += 1

                if not next_pending:
                    break
                if dry_run:
                    _append_job_log(job["id"], "MOS170MI AddProp: Dry-Run aktiv, keine weiteren Versuche.")
                    break
                total = total + len(next_pending)
                _update_job(job["id"], total=total)
                _append_job_log(job["id"], f"Warte {MOS170_RETRY_DELAY_SEC} Sekunden auf ERP ...")
                time.sleep(MOS170_RETRY_DELAY_SEC)
                attempt += 1
                pending_rows = next_pending

            _finish_job(
                job["id"],
                "success",
                result={"total": total, "ok": ok_count, "error": error_count},
            )
        except Exception as exc:  # noqa: BLE001
            _append_job_log(job["id"], f"Fehler: {exc}")
            _finish_job(job["id"], "error", error=str(exc))

    threading.Thread(target=_worker, daemon=True).start()
    return {"job_id": job["id"], "status": job["status"], "env": job["env"]}


@app.post("/api/renumber/cms100")
@app.post("/api/renumber/mos170/plpn")
def renumber_cms100(env: str = Query(DEFAULT_ENV)) -> dict:
    job = _create_job("mos170_plpn", env)

    def _worker() -> None:
        try:
            table_name = _table_for(RENUMBER_WAGON_TABLE, env)
            with _connect() as conn:
                if not _table_exists(conn, table_name):
                    raise HTTPException(status_code=404, detail=f"Tabelle {table_name} nicht gefunden.")
                _ensure_renumber_schema(conn, table_name)

            def _load_rows(conn: sqlite3.Connection) -> list[sqlite3.Row]:
                return conn.execute(
                    f"""SELECT rowid AS seq, * FROM "{table_name}"
                    WHERE IFNULL("PLPN", '') <> '' AND IFNULL("MWNO", '') = ''
                    ORDER BY CASE
                      WHEN "SEQ" IS NULL OR "SEQ" = '' THEN rowid
                      ELSE CAST("SEQ" AS INTEGER)
                    END ASC"""
                ).fetchall()

            with _connect() as conn:
                cms_rows = _load_rows(conn)
            total = len(cms_rows)
            _update_job(job["id"], total=total, processed=0, results=[])
            _append_job_log(job["id"], f"MOS170 PLPN: {total} Positionen.")

            dry_run = _effective_dry_run(env)
            ionapi_path = _ionapi_path(env, "mi")
            ion_cfg = load_ionapi_config(str(ionapi_path))
            base_url = build_base_url(ion_cfg)
            token = ""
            if not dry_run:
                token = get_access_token_service_account(ion_cfg)

            env_label = _normalize_env(env).upper()
            ok_count = 0
            error_count = 0
            processed = 0
            attempt = 1
            while cms_rows:
                if CMS100_RETRY_MAX and attempt > CMS100_RETRY_MAX:
                    _append_job_log(
                        job["id"],
                        f"MOS170 PLPN: Abbruch nach {CMS100_RETRY_MAX} Versuchen, {len(cms_rows)} ohne MWNO.",
                    )
                    break
                _append_job_log(job["id"], f"MOS170 PLPN: Versuch {attempt} für {len(cms_rows)} Positionen.")
                with _connect() as conn:
                    for row in cms_rows:
                        plpn = _row_value(row, "PLPN")
                        params = _build_cms100_params(plpn, env=env)
                        request_url = _build_m3_request_url(base_url, "CMS100MI", "Lst_PLPN_MWNO", params)
                        if not plpn:
                            ok = False
                            error_message = "PLPN fehlt"
                            response = {"error": error_message}
                        elif dry_run:
                            ok = True
                            error_message = None
                            response = {"dry_run": True}
                        else:
                            try:
                                response = call_m3_mi_get(base_url, token, "CMS100MI", "Lst_PLPN_MWNO", params)
                                error_message = _mi_error_message(response)
                                ok = not bool(error_message)
                            except Exception as exc:  # noqa: BLE001
                                response = {"error": str(exc)}
                                error_message = str(exc)
                                ok = False

                        mwno = _extract_mwno(response) if ok else ""
                        log_response = {"qomwno": mwno, "response": response}
                        conn.execute(
                            f'UPDATE "{table_name}" SET "MWNO"=? WHERE rowid=?',
                            (mwno, row["seq"]),
                        )
                        conn.commit()
                        _append_api_log(
                            "mos170_plpn",
                            params,
                            log_response,
                            ok,
                            error_message,
                            env=env_label,
                            wagon=_wagon_log_context(row),
                            dry_run=dry_run,
                            request_url=request_url,
                            program="CMS100MI",
                            transaction="Lst_PLPN_MWNO",
                        )
                        if not mwno:
                            _append_api_log(
                                "mos170_plpn_missing_mwno",
                                params,
                                log_response,
                                False,
                                "QOMWNO fehlt",
                                env=env_label,
                                wagon=_wagon_log_context(row),
                                dry_run=dry_run,
                                request_url=request_url,
                                program="CMS100MI",
                                transaction="Lst_PLPN_MWNO",
                            )

                        processed += 1
                        with _jobs_lock:
                            job_ref = _jobs.get(job["id"])
                            if job_ref is not None:
                                job_ref["processed"] = processed
                        if ok:
                            ok_count += 1
                        else:
                            error_count += 1

                with _connect() as conn:
                    cms_rows = _load_rows(conn)
                if not cms_rows:
                    break
                if dry_run:
                    _append_job_log(job["id"], "MOS170 PLPN: Dry-Run aktiv, keine weiteren Versuche.")
                    break
                if CMS100_RETRY_MAX and attempt >= CMS100_RETRY_MAX:
                    _append_job_log(
                        job["id"],
                        f"MOS170 PLPN: keine MWNO nach {CMS100_RETRY_MAX} Versuchen.",
                    )
                    break
                total = total + len(cms_rows)
                _update_job(job["id"], total=total)
                _append_job_log(job["id"], f"Warte {CMS100_RETRY_DELAY_SEC} Sekunden auf ERP ...")
                time.sleep(CMS100_RETRY_DELAY_SEC)
                attempt += 1

            _finish_job(
                job["id"],
                "success",
                result={"total": total, "ok": ok_count, "error": error_count},
            )
        except Exception as exc:  # noqa: BLE001
            _append_job_log(job["id"], f"Fehler: {exc}")
            _finish_job(job["id"], "error", error=str(exc))

    threading.Thread(target=_worker, daemon=True).start()
    return {"job_id": job["id"], "status": job["status"], "env": job["env"]}


@app.post("/api/renumber/mos100")
def renumber_mos100(env: str = Query(DEFAULT_ENV)) -> dict:
    job = _create_job("ips_mos100_chgsern", env)

    def _worker() -> None:
        try:
            table_name = _table_for(RENUMBER_WAGON_TABLE, env)
            with _connect() as conn:
                if not _table_exists(conn, table_name):
                    raise HTTPException(status_code=404, detail=f"Tabelle {table_name} nicht gefunden.")
                _ensure_renumber_schema(conn, table_name)
                rows = conn.execute(
                    f"""SELECT rowid AS seq, * FROM "{table_name}"
                    WHERE IFNULL("MWNO", '') <> ''
                      AND (IFNULL("NEW_PART_ITNO", '') <> '' OR IFNULL("NEW_PART_SER2", '') <> '')
                      AND IFNULL("MOS100_STATUS", '') = ''
                    ORDER BY CASE
                      WHEN "SEQ" IS NULL OR "SEQ" = '' THEN rowid
                      ELSE CAST("SEQ" AS INTEGER)
                    END ASC"""
                ).fetchall()

            total = len(rows)
            _update_job(job["id"], total=total, processed=0, results=[])
            _append_job_log(job["id"], f"IPS MOS100 Chg_SERN: {total} Positionen.")

            dry_run = _effective_dry_run(env)
            ionapi_path = _ionapi_path(env, "mi")
            ion_cfg = load_ionapi_config(str(ionapi_path))
            base_url = build_base_url(ion_cfg)
            token = ""
            if not dry_run:
                token = get_access_token_service_account(ion_cfg)

            env_label = _normalize_env(env).upper()
            ok_count = 0
            error_count = 0
            processed = 0
            with _connect() as conn:
                for row in rows:
                    params = _build_ips_mos100_params(row)
                    request_url = _build_ips_request_url(base_url, "MOS100")
                    mwno = params.get("WorkOrderNumber") or ""
                    attempt = 1
                    ok = False
                    error_message = None
                    response: Any = {}
                    status_label = "NOK"
                    while True:
                        if not mwno:
                            ok = False
                            error_message = "MWNO fehlt"
                            response = {"error": error_message}
                            status_label = "NOK"
                        elif dry_run:
                            ok = True
                            error_message = None
                            response = {"dry_run": True}
                            status_label = "DRYRUN"
                        else:
                            try:
                                response = _call_ips_service(
                                    base_url,
                                    token,
                                    "MOS100",
                                    "Chg_SERN",
                                    params,
                                    env=env,
                                )
                                ok = int(response.get("status_code") or 0) < 400
                                if ok:
                                    error_message = None
                                    status_label = "OK"
                                elif _ips_mos100_already_exists_fault(response):
                                    ok = True
                                    error_message = None
                                    status_label = "OK_ALREADY_EXISTS"
                                else:
                                    error_message = f"HTTP {response.get('status_code')}"
                                    status_label = "NOK"
                            except Exception as exc:  # noqa: BLE001
                                response = {"error": str(exc)}
                                error_message = str(exc)
                                ok = False
                                status_label = "NOK"

                        _append_api_log(
                            "ips_mos100_chgsern",
                            params,
                            response,
                            ok,
                            error_message,
                            env=env_label,
                            wagon=_wagon_log_context(row),
                            dry_run=dry_run,
                            request_url=request_url,
                            program="MOS100",
                            transaction="Chg_SERN",
                            request_method="POST",
                            status=status_label,
                        )
                        if ok or dry_run:
                            break
                        if MOS100_RETRY_MAX and attempt >= MOS100_RETRY_MAX:
                            break
                        if MOS100_RETRY_DELAY_SEC:
                            time.sleep(MOS100_RETRY_DELAY_SEC)
                        attempt += 1

                    with _connect() as conn:
                        conn.execute(
                            f'UPDATE "{table_name}" SET "MOS100_STATUS"=? WHERE rowid=?',
                            (status_label, row["seq"]),
                        )
                        conn.commit()
                    processed += 1
                    with _jobs_lock:
                        job_ref = _jobs.get(job["id"])
                        if job_ref is not None:
                            job_ref["processed"] = processed
                    if ok:
                        ok_count += 1
                    else:
                        error_count += 1

            _finish_job(
                job["id"],
                "success",
                result={"total": total, "ok": ok_count, "error": error_count},
            )
        except Exception as exc:  # noqa: BLE001
            _append_job_log(job["id"], f"Fehler: {exc}")
            _finish_job(job["id"], "error", error=str(exc))

    threading.Thread(target=_worker, daemon=True).start()
    return {"job_id": job["id"], "status": job["status"], "env": job["env"]}


# BEGIN WAGON RENNUMBERING
@app.post("/api/renumber/wagon")
def renumber_wagon(env: str = Query(DEFAULT_ENV)) -> dict:
    job = _create_job("wagon_renumber", env)

    def _worker() -> None:
        try:
            table_name = _table_for(RENUMBER_WAGON_TABLE, env)
            with _connect() as conn:
                if not _table_exists(conn, table_name):
                    raise HTTPException(status_code=404, detail=f"Tabelle {table_name} nicht gefunden.")
                _ensure_renumber_schema(conn, table_name)
                row = conn.execute(
                    f"""SELECT rowid AS seq, * FROM "{table_name}"
                    ORDER BY CASE
                      WHEN "SEQ" IS NULL OR "SEQ" = '' THEN rowid
                      ELSE CAST("SEQ" AS INTEGER)
                    END ASC LIMIT 1"""
                ).fetchone()

            if not row:
                raise HTTPException(status_code=404, detail="Keine Daten in der Umnummerierungs-Tabelle.")

            old_itno = _row_value(row, "WAGEN_ITNO") or _row_value(row, "ITNO")
            old_sern = _row_value(row, "WAGEN_SERN") or _row_value(row, "SERN")
            new_itno = _row_value(row, "NEW_BAUREIHE") or old_itno
            new_sern = _row_value(row, "NEW_SERN") or old_sern
            umbau_datum = _row_value(row, "UMBAU_DATUM")
            if not old_itno or not old_sern or not new_itno or not new_sern or not umbau_datum:
                raise HTTPException(status_code=400, detail="Pflichtfelder für Wagen fehlen.")

            whlo = ""
            wagon_table = _table_for(WAGENUMBAU_TABLE, env)
            with _connect() as conn:
                if _table_exists(conn, wagon_table):
                    columns = {
                        row[1]
                        for row in conn.execute(f'PRAGMA table_info("{wagon_table}")').fetchall()
                        if row and len(row) > 1
                    }
                    if "LAGERORT" in columns:
                        result = conn.execute(
                            f'SELECT "LAGERORT" FROM "{wagon_table}" WHERE "BAUREIHE"=? AND "SERIENNUMMER"=? LIMIT 1',
                            (old_itno, old_sern),
                        ).fetchone()
                        if result and result[0]:
                            whlo = str(result[0])
            if not whlo:
                raise HTTPException(status_code=400, detail="LAGERORT fehlt für den Wagen.")

            _update_job(job["id"], total=5, processed=0, results=[])
            _append_job_log(job["id"], "Starte Wagen-Umnummerierung")

            dry_run = _effective_dry_run(env)
            ionapi_path = _ionapi_path(env, "mi")
            ion_cfg = load_ionapi_config(str(ionapi_path))
            base_url = build_base_url(ion_cfg)
            token = ""
            if not dry_run:
                token = get_access_token_service_account(ion_cfg)

            env_label = _normalize_env(env).upper()
            processed = 0

            if WAGON_RENUMBER_FIXED_PLPN:
                plpn = WAGON_RENUMBER_FIXED_PLPN
                with _connect() as conn:
                    conn.execute(
                        f'UPDATE "{table_name}" SET "PLPN"=? WHERE rowid=?',
                        (plpn, row["seq"]),
                    )
                    conn.commit()
            elif WAGON_RENUMBER_SKIP_MOS170:
                mwno = _row_value(row, "MWNO")
                if not mwno:
                    raise HTTPException(status_code=400, detail="MWNO fehlt fuer MOS100 (MOS170/CMS100 uebersprungen).")
            else:
                # MOS170 AddProp
                plpn = ""
                params = _build_mos170_wagon_params(old_itno, old_sern, umbau_datum, whlo, env=env)
                request_url = _build_m3_request_url(base_url, "MOS170MI", "AddProp", params)
                if not params.get("ITNO") or not params.get("BANO") or not params.get("STDT"):
                    ok = False
                    error_message = "Pflichtfelder fehlen"
                    response = {"error": error_message}
                elif dry_run:
                    ok = True
                    error_message = None
                    response = {"dry_run": True}
                else:
                    try:
                        response = call_m3_mi_get(base_url, token, "MOS170MI", "AddProp", params)
                        error_message = _mi_error_message(response)
                        ok = not bool(error_message)
                        plpn = _extract_plpn(response) if ok else ""
                    except Exception as exc:  # noqa: BLE001
                        response = {"error": str(exc)}
                        error_message = str(exc)
                        ok = False

                _append_api_log(
                    "wagon_mos170_addprop",
                    params,
                    response,
                    ok,
                    error_message,
                    env=env_label,
                    wagon={"itno": old_itno, "sern": old_sern, "new_itno": new_itno, "new_sern": new_sern},
                    dry_run=dry_run,
                    request_url=request_url,
                    program="MOS170MI",
                    transaction="AddProp",
                )
                processed += 1
                _update_job(job["id"], processed=processed)
                if not ok and not dry_run:
                    raise HTTPException(status_code=500, detail="MOS170 AddProp fehlgeschlagen.")

                # MOS170 PLPN (aus AddProp Response)
                if dry_run:
                    plpn = "DRYRUN"
                _append_api_log(
                    "wagon_mos170_plpn",
                    params,
                    {"plpn": plpn},
                    bool(plpn),
                    None if plpn else "PLPN fehlt nach MOS170",
                    env=env_label,
                    wagon={"itno": old_itno, "sern": old_sern, "new_itno": new_itno, "new_sern": new_sern},
                    dry_run=dry_run,
                    request_url=request_url,
                    program="MOS170MI",
                    transaction="AddProp",
                )
                processed += 1
                _update_job(job["id"], processed=processed)
                if not plpn:
                    raise HTTPException(status_code=500, detail="PLPN fehlt nach MOS170.")
                with _connect() as conn:
                    conn.execute(
                        f'UPDATE "{table_name}" SET "PLPN"=? WHERE rowid=?',
                        (plpn, row["seq"]),
                    )
                    conn.commit()

            if not WAGON_RENUMBER_SKIP_MOS170:
                # CMS100 MWNO
                mwno = ""
                attempt = 1
                while True:
                    if WAGON_CMS100_RETRY_MAX and attempt > WAGON_CMS100_RETRY_MAX:
                        break
                    params = _build_cms100_params(plpn, env=env)
                    request_url = _build_m3_request_url(base_url, "CMS100MI", "Lst_PLPN_MWNO", params)
                    if dry_run:
                        ok = True
                        error_message = None
                        response = {"dry_run": True}
                        mwno = "DRYRUN"
                    else:
                        try:
                            response = call_m3_mi_get(base_url, token, "CMS100MI", "Lst_PLPN_MWNO", params)
                            error_message = _mi_error_message(response)
                            ok = not bool(error_message)
                            mwno = _extract_mwno(response) if ok else ""
                        except Exception as exc:  # noqa: BLE001
                            response = {"error": str(exc)}
                            error_message = str(exc)
                            ok = False

                    log_response = {"qomwno": mwno, "response": response}
                    _append_api_log(
                        "wagon_cms100_lst_plpn_mwno",
                        params,
                        log_response,
                        ok,
                        error_message,
                        env=env_label,
                        wagon={"itno": old_itno, "sern": old_sern, "new_itno": new_itno, "new_sern": new_sern},
                        dry_run=dry_run,
                        request_url=request_url,
                        program="CMS100MI",
                        transaction="Lst_PLPN_MWNO",
                    )
                    processed += 1
                    _update_job(job["id"], processed=processed)
                    if mwno:
                        break
                    if dry_run:
                        break
                    if WAGON_CMS100_RETRY_DELAY_SEC:
                        time.sleep(WAGON_CMS100_RETRY_DELAY_SEC)
                    attempt += 1

                if not mwno:
                    raise HTTPException(status_code=500, detail="MWNO fehlt nach CMS100.")
                with _connect() as conn:
                    conn.execute(
                        f'UPDATE "{table_name}" SET "MWNO"=? WHERE rowid=?',
                        (mwno, row["seq"]),
                    )
                    conn.commit()

            # IPS MOS100 Chg_SERN
            params = {
                "WorkOrderNumber": mwno,
                "Product": old_itno,
                "NewItemNumber": new_itno,
                "NewLotNumber": new_sern,
            }
            request_url = _build_ips_request_url(base_url, "MOS100")
            attempt = 1
            ok = False
            error_message = None
            response: Any = {}
            status_label = "NOK"
            while True:
                if dry_run:
                    ok = True
                    error_message = None
                    response = {"dry_run": True}
                    status_label = "DRYRUN"
                else:
                    try:
                        response = _call_ips_service(
                            base_url,
                            token,
                            "MOS100",
                            "Chg_SERN",
                            params,
                            env=env,
                        )
                        ok = int(response.get("status_code") or 0) < 400
                        error_message = None if ok else f"HTTP {response.get('status_code')}"
                        status_label = "OK" if ok else "NOK"
                    except Exception as exc:  # noqa: BLE001
                        response = {"error": str(exc)}
                        error_message = str(exc)
                        ok = False
                        status_label = "NOK"

                _append_api_log(
                    "wagon_ips_mos100_chgsern",
                    params,
                    response,
                    ok,
                    error_message,
                    env=env_label,
                    wagon={"itno": old_itno, "sern": old_sern, "new_itno": new_itno, "new_sern": new_sern},
                    dry_run=dry_run,
                    request_url=request_url,
                    program="MOS100",
                    transaction="Chg_SERN",
                    request_method="POST",
                    status=status_label,
                )
                if ok or dry_run:
                    break
                if WAGON_MOS100_RETRY_MAX and attempt >= WAGON_MOS100_RETRY_MAX:
                    break
                if MOS100_RETRY_DELAY_SEC:
                    time.sleep(MOS100_RETRY_DELAY_SEC)
                attempt += 1

            processed += 1
            _update_job(job["id"], processed=processed)
            if not ok:
                raise HTTPException(status_code=500, detail="MOS100 Chg_SERN fehlgeschlagen.")

            # MOS180 Approve
            params = {
                "FACI": MOS180_FACI,
                "MWNO": mwno,
                "RESP": MOS180_RESP,
                "APRB": MOS180_APRB,
            }
            request_url = _build_m3_request_url(base_url, "MOS180MI", "Approve", params)
            if dry_run:
                ok = True
                error_message = None
                response = {"dry_run": True}
            else:
                try:
                    response = call_m3_mi_get(base_url, token, "MOS180MI", "Approve", params)
                    error_message = _mi_error_message(response)
                    ok = not bool(error_message)
                except Exception as exc:  # noqa: BLE001
                    response = {"error": str(exc)}
                    error_message = str(exc)
                    ok = False

            _append_api_log(
                "wagon_mos180_approve",
                params,
                response,
                ok,
                error_message,
                env=env_label,
                wagon={"itno": old_itno, "sern": old_sern, "new_itno": new_itno, "new_sern": new_sern},
                dry_run=dry_run,
                request_url=request_url,
                program="MOS180MI",
                transaction="Approve",
            )
            processed += 1
            _update_job(job["id"], processed=processed)
            if not ok:
                raise HTTPException(status_code=500, detail="MOS180 Approve fehlgeschlagen.")

            # CRS335 UpdCtrlObj
            acrf_value = ""
            with _connect() as conn:
                if _table_exists(conn, wagon_table):
                    columns = {
                        row[1]
                        for row in conn.execute(f'PRAGMA table_info("{wagon_table}")').fetchall()
                        if row and len(row) > 1
                    }
                    if "ACRF" in columns:
                        result = conn.execute(
                            f'SELECT "ACRF" FROM "{wagon_table}" WHERE "BAUREIHE"=? AND "SERIENNUMMER"=? LIMIT 1',
                            (old_itno, old_sern),
                        ).fetchone()
                        if result and result[0]:
                            acrf_value = str(result[0])
            acrf_value = acrf_value or CRS335_ACRF
            params = _build_crs335_params(acrf_value, new_sern, new_itno)
            request_url = _build_m3_request_url(base_url, "CRS335MI", "UpdCtrlObj", params)
            if not params.get("ACRF"):
                ok = False
                error_message = "ACRF fehlt"
                response = {"error": error_message}
            elif dry_run:
                ok = True
                error_message = None
                response = {"dry_run": True}
            else:
                try:
                    response = call_m3_mi_get(base_url, token, "CRS335MI", "UpdCtrlObj", params)
                    error_message = _mi_error_message(response)
                    ok = not bool(error_message)
                except Exception as exc:  # noqa: BLE001
                    response = {"error": str(exc)}
                    error_message = str(exc)
                    ok = False

            _append_api_log(
                "wagon_crs335_updctrlobj",
                params,
                response,
                ok,
                error_message,
                env=env_label,
                wagon={"itno": old_itno, "sern": old_sern, "new_itno": new_itno, "new_sern": new_sern},
                dry_run=dry_run,
                request_url=request_url,
                program="CRS335MI",
                transaction="UpdCtrlObj",
            )
            processed += 1
            _update_job(job["id"], processed=processed)
            if not ok:
                raise HTTPException(status_code=500, detail="CRS335 fehlgeschlagen.")

            _finish_job(
                job["id"],
                "success",
                result={"total": processed, "ok": processed, "error": 0},
            )
        except Exception as exc:  # noqa: BLE001
            _append_job_log(job["id"], f"Fehler: {exc}")
            _finish_job(job["id"], "error", error=str(exc))

    threading.Thread(target=_worker, daemon=True).start()
    return {"job_id": job["id"], "status": job["status"], "env": job["env"]}
# END WAGON RENNUMBERING

@app.post("/api/renumber/install")
def renumber_install(env: str = Query(DEFAULT_ENV)) -> dict:
    job = _create_job("renumber_install", env)

    def _worker() -> None:
        try:
            table_name = _table_for(RENUMBER_WAGON_TABLE, env)
            with _connect() as conn:
                if not _table_exists(conn, table_name):
                    raise HTTPException(status_code=404, detail=f"Tabelle {table_name} nicht gefunden.")
                _ensure_renumber_schema(conn, table_name)
                rows = conn.execute(
                    f"""SELECT rowid AS seq, * FROM "{table_name}"
                    ORDER BY CASE
                      WHEN "SEQ" IS NULL OR "SEQ" = '' THEN rowid
                      ELSE CAST("SEQ" AS INTEGER)
                    END DESC"""
                ).fetchall()

            total = len(rows)
            _update_job(job["id"], total=total, processed=0, results=[])
            _append_job_log(job["id"], f"Starte MOS125MI Einbau: {total} Positionen.")

            dry_run = _effective_dry_run(env)
            ionapi_path = _ionapi_path(env, "mi")
            ion_cfg = load_ionapi_config(str(ionapi_path))
            base_url = build_base_url(ion_cfg)
            token = ""
            if not dry_run:
                token = get_access_token_service_account(ion_cfg)

            ok_count = 0
            error_count = 0
            env_label = _normalize_env(env).upper()
            with _connect() as conn:
                for idx, row in enumerate(rows, start=1):
                    params = _build_mos125_params(row, mode="in", env=env)
                    wagon_ctx = _wagon_log_context(row)
                    request_url = _build_m3_request_url(base_url, "MOS125MI", "RemoveInstall", params)
                    if not params["TRDT"]:
                        status = "ERROR: UMBAU_DATUM fehlt"
                        ok = False
                        _append_api_log(
                            "einbau",
                            params,
                            {"error": "UMBAU_DATUM fehlt"},
                            ok,
                            "UMBAU_DATUM fehlt",
                            env=env_label,
                            wagon=wagon_ctx,
                            dry_run=dry_run,
                            request_url=request_url,
                        )
                    elif dry_run:
                        status = "DRYRUN"
                        ok = True
                        _append_api_log(
                            "einbau",
                            params,
                            {"dry_run": True},
                            ok,
                            None,
                            env=env_label,
                            wagon=wagon_ctx,
                            dry_run=dry_run,
                            request_url=request_url,
                        )
                    else:
                        try:
                            response = call_m3_mi_get(
                                base_url, token, "MOS125MI", "RemoveInstall", params
                            )
                            error_message = _mi_error_message(response)
                            if error_message:
                                status = f"ERROR: {error_message}"
                                ok = False
                            else:
                                status = "OK"
                                ok = True
                            _append_api_log(
                                "einbau",
                                params,
                                response,
                                ok,
                                error_message,
                                env=env_label,
                                wagon=wagon_ctx,
                                dry_run=dry_run,
                                request_url=request_url,
                            )
                        except Exception as exc:  # noqa: BLE001
                            status = f"ERROR: {exc}"
                            ok = False
                            _append_api_log(
                                "einbau",
                                params,
                                {"error": str(exc)},
                                ok,
                                str(exc),
                                env=env_label,
                                wagon=wagon_ctx,
                                dry_run=dry_run,
                                request_url=request_url,
                            )

                    conn.execute(
                        f'UPDATE "{table_name}" SET "IN"=?, "TIMESTAMP_IN"=? WHERE rowid=?',
                        (status, datetime.utcnow().isoformat(sep=" ", timespec="seconds"), row["seq"]),
                    )
                    conn.commit()

                    result = {
                        "seq": row["seq"],
                        "cfgr": params.get("CFGL") or params.get("CFGR") or "",
                        "itno": params.get("ITNI") or params.get("ITNR") or _row_value(row, "ITNO"),
                        "ser2": _row_value(row, "SER2"),
                        "in": status,
                        "ok": ok,
                    }
                    with _jobs_lock:
                        job_ref = _jobs.get(job["id"])
                        if job_ref is not None:
                            job_ref["processed"] = idx
                            job_ref.setdefault("results", []).append(result)
                    if ok:
                        ok_count += 1
                    else:
                        error_count += 1
                    status_label = "ERROR" if not ok else ("DRYRUN" if dry_run else "OK")
                    _append_job_log(
                        job["id"],
                        f"{idx}/{total} {status_label} CFGL={params.get('CFGL', '')} ITNI={params.get('ITNI', '')} BANI={params.get('BANI', '')}",
                    )

            _finish_job(
                job["id"],
                "success",
                result={"total": total, "ok": ok_count, "error": error_count},
            )
        except Exception as exc:  # noqa: BLE001
            _append_job_log(job["id"], f"Fehler: {exc}")
            _finish_job(job["id"], "error", error=str(exc))

    threading.Thread(target=_worker, daemon=True).start()
    return {"job_id": job["id"], "status": job["status"], "env": job["env"]}


@app.post("/api/renumber/rollback")
def renumber_rollback(env: str = Query(DEFAULT_ENV)) -> dict:
    job = _create_job("renumber_rollback", env)

    threading.Thread(target=_run_rollback_job, args=(job, env), daemon=True).start()
    return {"job_id": job["id"], "status": job["status"], "env": job["env"]}


@app.post("/api/renumber/mos180")
def renumber_mos180(env: str = Query(DEFAULT_ENV)) -> dict:
    job = _create_job("mos180_approve", env)

    def _worker() -> None:
        try:
            table_name = _table_for(RENUMBER_WAGON_TABLE, env)
            with _connect() as conn:
                if not _table_exists(conn, table_name):
                    raise HTTPException(status_code=404, detail=f"Tabelle {table_name} nicht gefunden.")
                _ensure_renumber_schema(conn, table_name)
                rows = conn.execute(
                    f"""SELECT rowid AS seq, * FROM "{table_name}"
                    WHERE IFNULL("MWNO", '') <> ''
                    ORDER BY CASE
                      WHEN "SEQ" IS NULL OR "SEQ" = '' THEN rowid
                      ELSE CAST("SEQ" AS INTEGER)
                    END ASC"""
                ).fetchall()

            mwno_map: Dict[str, Dict[str, Any]] = {}
            for row in rows:
                mwno = _row_value(row, "MWNO")
                if not mwno:
                    continue
                entry = mwno_map.get(mwno)
                if entry is None:
                    mwno_map[mwno] = {"rowids": [row["seq"]], "row": row}
                else:
                    entry["rowids"].append(row["seq"])

            total = len(mwno_map)
            _update_job(job["id"], total=total, processed=0, results=[])
            _append_job_log(job["id"], f"MOS180MI Approve: {total} MWNO.")

            dry_run = _effective_dry_run(env)
            ionapi_path = _ionapi_path(env, "mi")
            ion_cfg = load_ionapi_config(str(ionapi_path))
            base_url = build_base_url(ion_cfg)
            token = ""
            if not dry_run:
                token = get_access_token_service_account(ion_cfg)

            env_label = _normalize_env(env).upper()
            ok_count = 0
            error_count = 0
            for idx, (mwno, entry) in enumerate(mwno_map.items(), start=1):
                row = entry["row"]
                params = _build_mos180_params(row, env=env)
                request_url = _build_m3_request_url(base_url, "MOS180MI", "Approve", params)
                mwno = params.get("MWNO") or mwno
                if not mwno:
                    ok = False
                    error_message = "MWNO fehlt"
                    response = {"error": error_message}
                    status_label = "NOK"
                elif dry_run:
                    ok = True
                    error_message = None
                    response = {"dry_run": True}
                    status_label = "OK"
                else:
                    try:
                        response = call_m3_mi_get(base_url, token, "MOS180MI", "Approve", params)
                        error_message = _mi_error_message(response)
                        ok = not bool(error_message)
                        status_label = "OK" if ok else "NOK"
                    except Exception as exc:  # noqa: BLE001
                        response = {"error": str(exc)}
                        error_message = str(exc)
                        ok = False
                        status_label = "NOK"

                _append_api_log(
                    "mos180_approve",
                    params,
                    response,
                    ok,
                    error_message,
                    env=env_label,
                    wagon=_wagon_log_context(row),
                    dry_run=dry_run,
                    request_url=request_url,
                    program="MOS180MI",
                    transaction="Approve",
                    status=status_label,
                )
                _append_job_result(
                    job["id"],
                    {
                        "itno": _row_value(row, "NEW_BAUREIHE")
                        or _row_value(row, "WAGEN_ITNO")
                        or _row_value(row, "ITNO"),
                        "sern": _row_value(row, "NEW_SERN")
                        or _row_value(row, "WAGEN_SERN")
                        or _row_value(row, "SERN"),
                        "mwno": mwno,
                        "status": status_label,
                    },
                )
                with _connect() as conn:
                    for rowid in entry["rowids"]:
                        conn.execute(
                            f'UPDATE "{table_name}" SET "MOS180_STATUS"=? WHERE rowid=?',
                            (status_label, rowid),
                        )
                    conn.commit()
                with _jobs_lock:
                    job_ref = _jobs.get(job["id"])
                    if job_ref is not None:
                        job_ref["processed"] = idx
                if ok:
                    ok_count += 1
                else:
                    error_count += 1

            _finish_job(
                job["id"],
                "success",
                result={"total": total, "ok": ok_count, "error": error_count},
            )
        except Exception as exc:  # noqa: BLE001
            _append_job_log(job["id"], f"Fehler: {exc}")
            _finish_job(job["id"], "error", error=str(exc))

    threading.Thread(target=_worker, daemon=True).start()
    return {"job_id": job["id"], "status": job["status"], "env": job["env"]}


@app.post("/api/renumber/mos050")
def renumber_mos050(env: str = Query(DEFAULT_ENV)) -> dict:
    job = _create_job("mos050_montage", env)

    def _worker() -> None:
        try:
            table_name = _table_for(RENUMBER_WAGON_TABLE, env)
            with _connect() as conn:
                if not _table_exists(conn, table_name):
                    raise HTTPException(status_code=404, detail=f"Tabelle {table_name} nicht gefunden.")
                _ensure_renumber_schema(conn, table_name)
                rows = conn.execute(
                    f"""SELECT rowid AS seq, * FROM "{table_name}"
                    WHERE IFNULL("MWNO", '') <> ''
                    ORDER BY CASE
                      WHEN "SEQ" IS NULL OR "SEQ" = '' THEN rowid
                      ELSE CAST("SEQ" AS INTEGER)
                    END ASC"""
                ).fetchall()

            target_rows = [row for row in rows if _needs_renumber(row)]
            total = len(target_rows)
            _update_job(job["id"], total=total, processed=0, results=[])
            _append_job_log(job["id"], f"MOS050 Montage: {total} Positionen.")

            dry_run = _effective_dry_run(env)
            ionapi_path = _ionapi_path(env, "mi")
            ion_cfg = load_ionapi_config(str(ionapi_path))
            base_url = build_base_url(ion_cfg)
            token = ""
            if not dry_run:
                token = get_access_token_service_account(ion_cfg)

            env_label = _normalize_env(env).upper()
            ok_count = 0
            error_count = 0
            for idx, row in enumerate(target_rows, start=1):
                params = _build_mos050_params(row)
                request_url = _build_ips_request_url(base_url, MOS050_SERVICE)
                mwno = params.get("WHMWNO") or params.get("WorkOrderNumber") or ""
                if not mwno:
                    ok = False
                    error_message = "MWNO fehlt"
                    response = {"error": error_message}
                    status_label = "NOK"
                elif dry_run:
                    ok = True
                    error_message = None
                    response = {"dry_run": True}
                    status_label = "OK"
                else:
                    try:
                        response = _call_ips_service(
                            base_url,
                            token,
                            MOS050_SERVICE,
                            MOS050_OPERATION,
                            params,
                            namespace_override=MOS050_NAMESPACE or None,
                            body_tag_override=MOS050_BODY_TAG or None,
                            env=env,
                        )
                        ok = int(response.get("status_code") or 0) < 400
                        error_message = None if ok else f"HTTP {response.get('status_code')}"
                        status_label = "OK" if ok else "NOK"
                    except Exception as exc:  # noqa: BLE001
                        response = {"error": str(exc)}
                        error_message = str(exc)
                        ok = False
                        status_label = "NOK"

                _append_api_log(
                    "ips_mos050_montage",
                    params,
                    response,
                    ok,
                    error_message,
                    env=env_label,
                    wagon=_wagon_log_context(row),
                    dry_run=dry_run,
                    request_url=request_url,
                    program=MOS050_SERVICE or "MOS050",
                    transaction=MOS050_OPERATION or "Montage",
                    request_method="POST",
                    status=status_label,
                )
                _append_job_result(
                    job["id"],
                    {
                        "itno": _row_value(row, "NEW_BAUREIHE")
                        or _row_value(row, "WAGEN_ITNO")
                        or _row_value(row, "ITNO"),
                        "sern": _row_value(row, "NEW_SERN")
                        or _row_value(row, "WAGEN_SERN")
                        or _row_value(row, "SERN"),
                        "status": status_label,
                    },
                )
                with _connect() as conn:
                    conn.execute(
                        f'UPDATE "{table_name}" SET "MOS050_STATUS"=? WHERE rowid=?',
                        (status_label, row["seq"]),
                    )
                    conn.commit()
                with _jobs_lock:
                    job_ref = _jobs.get(job["id"])
                    if job_ref is not None:
                        job_ref["processed"] = idx
                if ok:
                    ok_count += 1
                else:
                    error_count += 1

            _finish_job(
                job["id"],
                "success",
                result={"total": total, "ok": ok_count, "error": error_count},
            )
        except Exception as exc:  # noqa: BLE001
            _append_job_log(job["id"], f"Fehler: {exc}")
            _finish_job(job["id"], "error", error=str(exc))

    threading.Thread(target=_worker, daemon=True).start()
    return {"job_id": job["id"], "status": job["status"], "env": job["env"]}


@app.post("/api/renumber/mms240")
def renumber_mms240(env: str = Query(DEFAULT_ENV)) -> dict:
    job = _create_job("mms240_upd", env)

    def _worker() -> None:
        try:
            table_name = _table_for(RENUMBER_WAGON_TABLE, env)
            with _connect() as conn:
                if not _table_exists(conn, table_name):
                    raise HTTPException(status_code=404, detail=f"Tabelle {table_name} nicht gefunden.")
                _ensure_renumber_schema(conn, table_name)
                rows = conn.execute(
                    f"""SELECT rowid AS seq, * FROM "{table_name}"
                    ORDER BY CASE
                      WHEN "SEQ" IS NULL OR "SEQ" = '' THEN rowid
                      ELSE CAST("SEQ" AS INTEGER)
                    END ASC"""
                ).fetchall()

            if not rows:
                _finish_job(job["id"], "success", result={"total": 0, "ok": 0, "error": 0})
                return

            wagons: Dict[tuple[str, str], Dict[str, Any]] = {}
            for row in rows:
                wagon_itno = _row_value(row, "WAGEN_ITNO") or _row_value(row, "ITNO")
                wagon_sern = _row_value(row, "WAGEN_SERN") or _row_value(row, "SERN")
                wagon_key = (wagon_itno, wagon_sern)
                entry = wagons.get(wagon_key)
                if not entry:
                    wagons[wagon_key] = {
                        "new_itno": _row_value(row, "NEW_BAUREIHE") or wagon_itno,
                        "new_sern": _row_value(row, "NEW_SERN") or wagon_sern,
                        "rowids": [row["seq"]],
                    }
                else:
                    entry["rowids"].append(row["seq"])

            total = len(wagons)
            _update_job(job["id"], total=total, processed=0, results=[])
            _append_job_log(job["id"], f"MMS240MI Upd: {total} Wagen.")

            dry_run = _effective_dry_run(env)
            ionapi_path = _ionapi_path(env, "mi")
            ion_cfg = load_ionapi_config(str(ionapi_path))
            base_url = build_base_url(ion_cfg)
            token = ""
            if not dry_run:
                token = get_access_token_service_account(ion_cfg)

            env_label = _normalize_env(env).upper()
            ok_count = 0
            error_count = 0
            for idx, (wagon_key, entry) in enumerate(wagons.items(), start=1):
                new_itno = entry["new_itno"]
                new_sern = entry["new_sern"]
                params = _build_mms240_params(new_itno, new_sern)
                request_url = _build_m3_request_url(base_url, "MMS240MI", "Upd", params)

                if not new_itno or not new_sern:
                    ok = False
                    error_message = "ITNO/SERN fehlt"
                    response = {"error": error_message}
                    status_label = "NOK"
                elif dry_run:
                    ok = True
                    error_message = None
                    response = {"dry_run": True}
                    status_label = "OK"
                else:
                    try:
                        response = call_m3_mi_get(base_url, token, "MMS240MI", "Upd", params)
                        error_message = _mi_error_message(response)
                        ok = not bool(error_message)
                        status_label = "OK" if ok else "NOK"
                    except Exception as exc:  # noqa: BLE001
                        response = {"error": str(exc)}
                        error_message = str(exc)
                        ok = False
                        status_label = "NOK"

                _append_api_log(
                    "mms240_upd",
                    params,
                    response,
                    ok,
                    error_message,
                    env=env_label,
                    wagon={
                        "itno": wagon_key[0],
                        "sern": wagon_key[1],
                        "new_itno": new_itno,
                        "new_sern": new_sern,
                    },
                    dry_run=dry_run,
                    request_url=request_url,
                    program="MMS240MI",
                    transaction="Upd",
                    status=status_label,
                )
                _append_job_result(
                    job["id"],
                    {
                        "itno": new_itno,
                        "sern": new_sern,
                        "status": status_label,
                    },
                )
                with _connect() as conn:
                    for rowid in entry["rowids"]:
                        conn.execute(
                            f'UPDATE "{table_name}" SET "MMS240_STATUS"=? WHERE rowid=?',
                            (status_label, rowid),
                        )
                    conn.commit()

                with _jobs_lock:
                    job_ref = _jobs.get(job["id"])
                    if job_ref is not None:
                        job_ref["processed"] = idx
                if ok:
                    ok_count += 1
                else:
                    error_count += 1

            _finish_job(
                job["id"],
                "success",
                result={"total": total, "ok": ok_count, "error": error_count},
            )
        except Exception as exc:  # noqa: BLE001
            _append_job_log(job["id"], f"Fehler: {exc}")
            _finish_job(job["id"], "error", error=str(exc))

    threading.Thread(target=_worker, daemon=True).start()
    return {"job_id": job["id"], "status": job["status"], "env": job["env"]}


@app.post("/api/renumber/cusext")
def renumber_cusext(env: str = Query(DEFAULT_ENV)) -> dict:
    job = _create_job("cusext_addfieldvalue", env)

    def _worker() -> None:
        try:
            table_name = _table_for(RENUMBER_WAGON_TABLE, env)
            with _connect() as conn:
                if not _table_exists(conn, table_name):
                    raise HTTPException(status_code=404, detail=f"Tabelle {table_name} nicht gefunden.")
                _ensure_renumber_schema(conn, table_name)
                rows = conn.execute(
                    f"""SELECT rowid AS seq, * FROM "{table_name}"
                    ORDER BY CASE
                      WHEN "SEQ" IS NULL OR "SEQ" = '' THEN rowid
                      ELSE CAST("SEQ" AS INTEGER)
                    END ASC"""
                ).fetchall()

            if not rows:
                _finish_job(job["id"], "success", result={"total": 0, "ok": 0, "error": 0})
                return

            wagons: Dict[tuple[str, str], Dict[str, Any]] = {}
            for row in rows:
                wagon_itno = _row_value(row, "WAGEN_ITNO") or _row_value(row, "ITNO")
                wagon_sern = _row_value(row, "WAGEN_SERN") or _row_value(row, "SERN")
                wagon_key = (wagon_itno, wagon_sern)
                entry = wagons.get(wagon_key)
                if not entry:
                    wagons[wagon_key] = {
                        "new_itno": _row_value(row, "NEW_BAUREIHE") or wagon_itno,
                        "new_sern": _row_value(row, "NEW_SERN") or wagon_sern,
                        "rowids": [row["seq"]],
                    }
                else:
                    entry["rowids"].append(row["seq"])

            total = len(wagons)
            _update_job(job["id"], total=total, processed=0, results=[])
            _append_job_log(job["id"], f"CUSEXTMI AddFieldValue: {total} Wagen.")

            dry_run = _effective_dry_run(env)
            ionapi_path = _ionapi_path(env, "mi")
            ion_cfg = load_ionapi_config(str(ionapi_path))
            base_url = build_base_url(ion_cfg)
            token = ""
            if not dry_run:
                token = get_access_token_service_account(ion_cfg)

            env_label = _normalize_env(env).upper()
            ok_count = 0
            error_count = 0
            for idx, (wagon_key, entry) in enumerate(wagons.items(), start=1):
                new_itno = entry["new_itno"]
                new_sern = entry["new_sern"]
                params = _build_cusext_params(new_itno, new_sern)
                request_url = _build_m3_request_url(base_url, "CUSEXTMI", "AddFieldValue", params)

                if not new_itno or not new_sern:
                    ok = False
                    error_message = "ITNO/SERN fehlt"
                    response = {"error": error_message}
                    status_label = "NOK"
                elif dry_run:
                    ok = True
                    error_message = None
                    response = {"dry_run": True}
                    status_label = "OK"
                else:
                    try:
                        response = call_m3_mi_get(base_url, token, "CUSEXTMI", "AddFieldValue", params)
                        error_message = _mi_error_message(response)
                        ok = not bool(error_message)
                        status_label = "OK" if ok else "NOK"
                    except Exception as exc:  # noqa: BLE001
                        response = {"error": str(exc)}
                        error_message = str(exc)
                        ok = False
                        status_label = "NOK"

                _append_api_log(
                    "cusext_addfieldvalue",
                    params,
                    response,
                    ok,
                    error_message,
                    env=env_label,
                    wagon={
                        "itno": wagon_key[0],
                        "sern": wagon_key[1],
                        "new_itno": new_itno,
                        "new_sern": new_sern,
                    },
                    dry_run=dry_run,
                    request_url=request_url,
                    program="CUSEXTMI",
                    transaction="AddFieldValue",
                    status=status_label,
                )
                _append_job_result(
                    job["id"],
                    {
                        "itno": new_itno,
                        "sern": new_sern,
                        "status": status_label,
                    },
                )
                with _connect() as conn:
                    for rowid in entry["rowids"]:
                        conn.execute(
                            f'UPDATE "{table_name}" SET "CUSEXT_STATUS"=? WHERE rowid=?',
                            (status_label, rowid),
                        )
                    conn.commit()

                with _jobs_lock:
                    job_ref = _jobs.get(job["id"])
                    if job_ref is not None:
                        job_ref["processed"] = idx
                if ok:
                    ok_count += 1
                else:
                    error_count += 1

            _finish_job(
                job["id"],
                "success",
                result={"total": total, "ok": ok_count, "error": error_count},
            )
        except Exception as exc:  # noqa: BLE001
            _append_job_log(job["id"], f"Fehler: {exc}")
            _finish_job(job["id"], "error", error=str(exc))

    threading.Thread(target=_worker, daemon=True).start()
    return {"job_id": job["id"], "status": job["status"], "env": job["env"]}


@app.post("/api/renumber/crs335")
def renumber_crs335(env: str = Query(DEFAULT_ENV)) -> dict:
    job = _create_job("crs335_updctrlobj", env)

    def _worker() -> None:
        try:
            table_name = _table_for(RENUMBER_WAGON_TABLE, env)
            with _connect() as conn:
                if not _table_exists(conn, table_name):
                    raise HTTPException(status_code=404, detail=f"Tabelle {table_name} nicht gefunden.")
                _ensure_renumber_schema(conn, table_name)
                rows = conn.execute(
                    f"""SELECT rowid AS seq, * FROM "{table_name}"
                    ORDER BY CASE
                      WHEN "SEQ" IS NULL OR "SEQ" = '' THEN rowid
                      ELSE CAST("SEQ" AS INTEGER)
                    END ASC"""
                ).fetchall()

            if not rows:
                _finish_job(job["id"], "success", result={"total": 0, "ok": 0, "error": 0})
                return

            wagons: Dict[tuple[str, str], Dict[str, str]] = {}
            for row in rows:
                wagon_key = (
                    _row_value(row, "WAGEN_ITNO") or _row_value(row, "ITNO"),
                    _row_value(row, "WAGEN_SERN") or _row_value(row, "SERN"),
                )
                if wagon_key in wagons:
                    continue
                wagons[wagon_key] = {
                    "new_sern": _row_value(row, "NEW_SERN") or _row_value(row, "WAGEN_SERN") or _row_value(row, "SERN"),
                    "new_baureihe": _row_value(row, "NEW_BAUREIHE") or _row_value(row, "WAGEN_ITNO") or _row_value(row, "ITNO"),
                }
            acrf_by_wagon: Dict[tuple[str, str], str] = {}
            wagon_table = _table_for(WAGENUMBAU_TABLE, env)
            with _connect() as conn:
                if _table_exists(conn, wagon_table):
                    columns = {
                        row[1]
                        for row in conn.execute(f'PRAGMA table_info("{wagon_table}")').fetchall()
                        if row and len(row) > 1
                    }
                    if "ACRF" in columns:
                        for wagon_key in wagons.keys():
                            row = conn.execute(
                                f'SELECT "ACRF" FROM "{wagon_table}" WHERE "BAUREIHE"=? AND "SERIENNUMMER"=? LIMIT 1',
                                (wagon_key[0], wagon_key[1]),
                            ).fetchone()
                            if row and row[0]:
                                acrf_by_wagon[wagon_key] = str(row[0])

            total = len(wagons)
            _update_job(job["id"], total=total, processed=0, results=[])
            _append_job_log(job["id"], f"CRS335MI UpdCtrlObj: {total} Wagen.")

            dry_run = _effective_dry_run(env)
            ionapi_path = _ionapi_path(env, "mi")
            ion_cfg = load_ionapi_config(str(ionapi_path))
            base_url = build_base_url(ion_cfg)
            token = ""
            if not dry_run:
                token = get_access_token_service_account(ion_cfg)

            env_label = _normalize_env(env).upper()
            ok_count = 0
            error_count = 0
            for idx, (wagon_key, values) in enumerate(wagons.items(), start=1):
                acrf_value = acrf_by_wagon.get(wagon_key) or CRS335_ACRF
                params = _build_crs335_params(acrf_value, values["new_sern"], values["new_baureihe"])
                request_url = _build_m3_request_url(base_url, "CRS335MI", "UpdCtrlObj", params)
                if not params.get("ACRF"):
                    ok = False
                    error_message = "ACRF fehlt"
                    response = {"error": error_message}
                    status_label = "NOK"
                elif dry_run:
                    ok = True
                    error_message = None
                    response = {"dry_run": True}
                    status_label = "OK"
                else:
                    try:
                        response = call_m3_mi_get(base_url, token, "CRS335MI", "UpdCtrlObj", params)
                        error_message = _mi_error_message(response)
                        ok = not bool(error_message)
                        status_label = "OK" if ok else "NOK"
                    except Exception as exc:  # noqa: BLE001
                        response = {"error": str(exc)}
                        error_message = str(exc)
                        ok = False
                        status_label = "NOK"

                _append_api_log(
                    "crs335_updctrlobj",
                    params,
                    response,
                    ok,
                    error_message,
                    env=env_label,
                    wagon={"itno": wagon_key[0], "sern": wagon_key[1]},
                    dry_run=dry_run,
                    request_url=request_url,
                    program="CRS335MI",
                    transaction="UpdCtrlObj",
                    status=status_label,
                )
                _append_job_result(
                    job["id"],
                    {
                        "itno": values["new_baureihe"],
                        "sern": values["new_sern"],
                        "status": status_label,
                    },
                )
                with _connect() as conn:
                    conn.execute(
                        f'UPDATE "{table_name}" SET "CRS335_STATUS"=? WHERE "WAGEN_ITNO"=? AND "WAGEN_SERN"=?',
                        (status_label, wagon_key[0], wagon_key[1]),
                    )
                    conn.commit()

                with _jobs_lock:
                    job_ref = _jobs.get(job["id"])
                    if job_ref is not None:
                        job_ref["processed"] = idx
                if ok:
                    ok_count += 1
                else:
                    error_count += 1

            _finish_job(
                job["id"],
                "success",
                result={"total": total, "ok": ok_count, "error": error_count},
            )
        except Exception as exc:  # noqa: BLE001
            _append_job_log(job["id"], f"Fehler: {exc}")
            _finish_job(job["id"], "error", error=str(exc))

    threading.Thread(target=_worker, daemon=True).start()
    return {"job_id": job["id"], "status": job["status"], "env": job["env"]}


@app.post("/api/renumber/sts046")
def renumber_sts046(env: str = Query(DEFAULT_ENV)) -> dict:
    job = _create_job("sts046_delgenitem", env)

    def _worker() -> None:
        try:
            table_name = _table_for(RENUMBER_WAGON_TABLE, env)
            wagon_table = _table_for(WAGENUMBAU_TABLE, env)
            with _connect() as conn:
                if not _table_exists(conn, table_name):
                    raise HTTPException(status_code=404, detail=f"Tabelle {table_name} nicht gefunden.")
                _ensure_renumber_schema(conn, table_name)
                rows = conn.execute(
                    f"""SELECT rowid AS seq, * FROM "{table_name}"
                    ORDER BY CASE
                      WHEN "SEQ" IS NULL OR "SEQ" = '' THEN rowid
                      ELSE CAST("SEQ" AS INTEGER)
                    END ASC"""
                ).fetchall()

            if not rows:
                _finish_job(job["id"], "success", result={"total": 0, "ok": 0, "error": 0})
                return

            wagon_meta: Dict[tuple[str, str], Dict[str, str]] = {}
            with _connect() as conn:
                if _table_exists(conn, wagon_table):
                    columns = {
                        row[1]
                        for row in conn.execute(f'PRAGMA table_info("{wagon_table}")').fetchall()
                        if row and len(row) > 1
                    }
                    if {"LAGERORT", "ACMC"} <= columns:
                        wagon_rows = conn.execute(
                            f'SELECT "BAUREIHE","SERIENNUMMER","LAGERORT","ACMC" FROM "{wagon_table}"'
                        ).fetchall()
                        for row in wagon_rows:
                            key = (str(row[0] or ""), str(row[1] or ""))
                            wagon_meta[key] = {
                                "WHLO": str(row[2] or ""),
                                "GEIT": str(row[3] or ""),
                            }

            wagons: Dict[tuple[str, str], Dict[str, Any]] = {}
            for row in rows:
                wagon_itno = _row_value(row, "WAGEN_ITNO") or _row_value(row, "ITNO")
                wagon_sern = _row_value(row, "WAGEN_SERN") or _row_value(row, "SERN")
                wagon_key = (wagon_itno, wagon_sern)
                entry = wagons.get(wagon_key)
                if not entry:
                    wagons[wagon_key] = {
                        "itno": wagon_itno,
                        "bano": wagon_sern,
                        "rowids": [row["seq"]],
                    }
                else:
                    entry["rowids"].append(row["seq"])

            total = len(wagons)
            _update_job(job["id"], total=total, processed=0, results=[])
            _append_job_log(job["id"], f"STS046MI DelGenItem: {total} Wagen.")

            dry_run = _effective_dry_run(env)
            ionapi_path = _ionapi_path(env, "mi")
            ion_cfg = load_ionapi_config(str(ionapi_path))
            base_url = build_base_url(ion_cfg)
            token = ""
            if not dry_run:
                token = get_access_token_service_account(ion_cfg)

            env_label = _normalize_env(env).upper()
            ok_count = 0
            error_count = 0
            for idx, (wagon_key, entry) in enumerate(wagons.items(), start=1):
                meta = wagon_meta.get(wagon_key) or {}
                whlo = meta.get("WHLO", "")
                geit = meta.get("GEIT", "")
                itno = entry["itno"]
                bano = entry["bano"]
                params = _build_sts046_params(whlo, geit, itno, bano)
                request_url = _build_m3_request_url(base_url, "STS046MI", "DelGenItem", params)

                if not whlo or not geit or not itno:
                    ok = False
                    error_message = "WHLO/GEIT/ITNO fehlt"
                    response = {"error": error_message}
                    status_label = "NOK"
                elif dry_run:
                    ok = True
                    error_message = None
                    response = {"dry_run": True}
                    status_label = "OK"
                else:
                    try:
                        response = call_m3_mi_get(base_url, token, "STS046MI", "DelGenItem", params)
                        error_message = _mi_error_message(response)
                        ok = not bool(error_message)
                        status_label = "OK" if ok else "NOK"
                    except Exception as exc:  # noqa: BLE001
                        response = {"error": str(exc)}
                        error_message = str(exc)
                        ok = False
                        status_label = "NOK"

                _append_api_log(
                    "sts046_delgenitem",
                    params,
                    response,
                    ok,
                    error_message,
                    env=env_label,
                    wagon={"itno": wagon_key[0], "sern": wagon_key[1]},
                    dry_run=dry_run,
                    request_url=request_url,
                    program="STS046MI",
                    transaction="DelGenItem",
                    status=status_label,
                )
                _append_job_result(
                    job["id"],
                    {
                        "itno": wagon_key[0],
                        "sern": wagon_key[1],
                        "status": status_label,
                    },
                )
                with _connect() as conn:
                    for rowid in entry["rowids"]:
                        conn.execute(
                            f'UPDATE "{table_name}" SET "STS046_STATUS"=? WHERE rowid=?',
                            (status_label, rowid),
                        )
                    conn.commit()

                with _jobs_lock:
                    job_ref = _jobs.get(job["id"])
                    if job_ref is not None:
                        job_ref["processed"] = idx
                if ok:
                    ok_count += 1
                else:
                    error_count += 1

            _finish_job(
                job["id"],
                "success",
                result={"total": total, "ok": ok_count, "error": error_count},
            )
        except Exception as exc:  # noqa: BLE001
            _append_job_log(job["id"], f"Fehler: {exc}")
            _finish_job(job["id"], "error", error=str(exc))

    threading.Thread(target=_worker, daemon=True).start()
    return {"job_id": job["id"], "status": job["status"], "env": job["env"]}


@app.post("/api/renumber/sts046/add")
def renumber_sts046_add(env: str = Query(DEFAULT_ENV)) -> dict:
    job = _create_job("sts046_addgenitem", env)

    def _worker() -> None:
        try:
            table_name = _table_for(RENUMBER_WAGON_TABLE, env)
            wagon_table = _table_for(WAGENUMBAU_TABLE, env)
            with _connect() as conn:
                if not _table_exists(conn, table_name):
                    raise HTTPException(status_code=404, detail=f"Tabelle {table_name} nicht gefunden.")
                _ensure_renumber_schema(conn, table_name)
                rows = conn.execute(
                    f"""SELECT rowid AS seq, * FROM "{table_name}"
                    ORDER BY CASE
                      WHEN "SEQ" IS NULL OR "SEQ" = '' THEN rowid
                      ELSE CAST("SEQ" AS INTEGER)
                    END ASC"""
                ).fetchall()

            if not rows:
                _finish_job(job["id"], "success", result={"total": 0, "ok": 0, "error": 0})
                return

            wagon_meta: Dict[tuple[str, str], Dict[str, str]] = {}
            acmc_by_baureihe: Dict[str, str] = {}
            with _connect() as conn:
                if _table_exists(conn, wagon_table):
                    columns = {
                        row[1]
                        for row in conn.execute(f'PRAGMA table_info("{wagon_table}")').fetchall()
                        if row and len(row) > 1
                    }
                    if {"LAGERORT", "ACMC"} <= columns:
                        wagon_rows = conn.execute(
                            f'SELECT "BAUREIHE","SERIENNUMMER","LAGERORT","ACMC" FROM "{wagon_table}"'
                        ).fetchall()
                        for row in wagon_rows:
                            baureihe = str(row[0] or "")
                            sern = str(row[1] or "")
                            whlo = str(row[2] or "")
                            acmc = str(row[3] or "")
                            wagon_meta[(baureihe, sern)] = {"WHLO": whlo, "GEIT": acmc}
                            if baureihe and acmc and baureihe not in acmc_by_baureihe:
                                acmc_by_baureihe[baureihe] = acmc

            wagons: Dict[tuple[str, str], Dict[str, Any]] = {}
            for row in rows:
                wagon_itno = _row_value(row, "WAGEN_ITNO") or _row_value(row, "ITNO")
                wagon_sern = _row_value(row, "WAGEN_SERN") or _row_value(row, "SERN")
                wagon_key = (wagon_itno, wagon_sern)
                entry = wagons.get(wagon_key)
                if not entry:
                    wagons[wagon_key] = {
                        "new_itno": _row_value(row, "NEW_BAUREIHE") or wagon_itno,
                        "new_sern": _row_value(row, "NEW_SERN") or wagon_sern,
                        "rowids": [row["seq"]],
                    }
                else:
                    entry["rowids"].append(row["seq"])

            total = len(wagons)
            _update_job(job["id"], total=total, processed=0, results=[])
            _append_job_log(job["id"], f"STS046MI AddGenItem: {total} Wagen.")

            dry_run = _effective_dry_run(env)
            ionapi_path = _ionapi_path(env, "mi")
            ion_cfg = load_ionapi_config(str(ionapi_path))
            base_url = build_base_url(ion_cfg)
            token = ""
            if not dry_run:
                token = get_access_token_service_account(ion_cfg)

            env_label = _normalize_env(env).upper()
            ok_count = 0
            error_count = 0
            for idx, (wagon_key, entry) in enumerate(wagons.items(), start=1):
                meta = wagon_meta.get(wagon_key) or {}
                whlo = meta.get("WHLO", "")
                new_itno = entry["new_itno"]
                new_sern = entry["new_sern"]
                old_geit = meta.get("GEIT", "")
                new_geit = acmc_by_baureihe.get(new_itno, "")
                geits = []
                if new_geit:
                    geits.append(new_geit)
                if old_geit and old_geit != new_geit:
                    geits.append(old_geit)
                if not geits:
                    geits = [""]

                all_ok = True
                status_label = "OK"
                for geit in geits:
                    params = _build_sts046_params(whlo, geit, new_itno, new_sern)
                    request_url = _build_m3_request_url(base_url, "STS046MI", "AddGenItem", params)

                    if not whlo or not geit or not new_itno:
                        ok = False
                        error_message = "WHLO/GEIT/ITNO fehlt"
                        response = {"error": error_message}
                        status_label = "NOK"
                    elif dry_run:
                        ok = True
                        error_message = None
                        response = {"dry_run": True}
                        status_label = "OK"
                    else:
                        try:
                            response = call_m3_mi_get(base_url, token, "STS046MI", "AddGenItem", params)
                            error_message = _mi_error_message(response)
                            ok = not bool(error_message)
                            status_label = "OK" if ok else "NOK"
                        except Exception as exc:  # noqa: BLE001
                            response = {"error": str(exc)}
                            error_message = str(exc)
                            ok = False
                            status_label = "NOK"

                    _append_api_log(
                        "sts046_addgenitem",
                        params,
                        response,
                        ok,
                        error_message,
                        env=env_label,
                        wagon={"itno": wagon_key[0], "sern": wagon_key[1]},
                        dry_run=dry_run,
                        request_url=request_url,
                        program="STS046MI",
                        transaction="AddGenItem",
                        status=status_label,
                    )
                    if not ok:
                        all_ok = False

                status_label = "OK" if all_ok else "NOK"
                _append_job_result(
                    job["id"],
                    {
                        "itno": new_itno,
                        "sern": new_sern,
                        "status": status_label,
                    },
                )
                with _connect() as conn:
                    for rowid in entry["rowids"]:
                        conn.execute(
                            f'UPDATE "{table_name}" SET "STS046_ADD_STATUS"=? WHERE rowid=?',
                            (status_label, rowid),
                        )
                    conn.commit()

                with _jobs_lock:
                    job_ref = _jobs.get(job["id"])
                    if job_ref is not None:
                        job_ref["processed"] = idx
                if ok:
                    ok_count += 1
                else:
                    error_count += 1

            _finish_job(
                job["id"],
                "success",
                result={"total": total, "ok": ok_count, "error": error_count},
            )
        except Exception as exc:  # noqa: BLE001
            _append_job_log(job["id"], f"Fehler: {exc}")
            _finish_job(job["id"], "error", error=str(exc))

    threading.Thread(target=_worker, daemon=True).start()
    return {"job_id": job["id"], "status": job["status"], "env": job["env"]}


@app.get("/api/renumber/debug")
def renumber_debug(env: str = Query(DEFAULT_ENV)) -> dict:
    table_name = _table_for(RENUMBER_WAGON_TABLE, env)
    with _connect() as conn:
        if not _table_exists(conn, table_name):
            raise HTTPException(status_code=404, detail=f"Tabelle {table_name} nicht gefunden.")
        _ensure_columns(conn, table_name, RENUMBER_EXTRA_COLUMNS)
    rows = conn.execute(
        f"""SELECT rowid AS seq, * FROM "{table_name}"
        ORDER BY CASE
          WHEN "SEQ" IS NULL OR "SEQ" = '' THEN rowid
          ELSE CAST("SEQ" AS INTEGER)
        END"""
    ).fetchall()

    calls: List[Dict[str, Any]] = []
    for row in rows:
        cfgr = _row_value(row, "CFGL", "MFGL")
        level = _hierarchy_level(cfgr)
        trtm = str(10000 + (level * 1000))
        umbau_datum = _row_value(row, "UMBAU_DATUM")
        call = {
            "SEQ": row["seq"],
            "RITP": "UMA",
            "RESP": "CHRUPP",
            "TRDT": _format_yyyymmdd(umbau_datum),
            "WHLO": "ZUM",
            "TWSL": "AUSBAU",
            "RSC4": "UMB",
            "TRTM": trtm,
            "CFGR": cfgr,
            "NHAR": _row_value(row, "MTRL"),
            "NHSR": _row_value(row, "SERN"),
            "ITNR": _row_value(row, "ITNO"),
            "BANR": _row_value(row, "SER2"),
        }
        calls.append(call)
    return {"table": table_name, "calls": calls, "env": _normalize_env(env)}


@app.get("/api/spareparts/search")
def spareparts_search(
    eqtp: str = Query(..., min_length=1),
    type_filter: str = Query("", max_length=80, alias="type"),
    item: str = Query("", max_length=80),
    serial: str = Query("", max_length=80),
    facility: str = Query("", max_length=40),
    bin: str = Query("", max_length=80),
    limit: int = Query(50, ge=1, le=200),
    env: str = Query(DEFAULT_ENV),
) -> dict:
    table_name = _table_for(SPAREPARTS_TABLE, env)
    with _connect() as conn:
        _ensure_table(conn, table_name, SPAREPARTS_TABLE)
        clauses = ["TEILEART = ?", "UPPER(IFNULL(LAGERPLATZ, '')) <> 'INSTALLED'"]
        params: list[str] = [eqtp]
        if type_filter:
            clauses.append('"WAGEN-TYP" LIKE ?')
            params.append(f"%{type_filter}%")
        if item:
            clauses.append('"BAUREIHE" LIKE ?')
            params.append(f"%{item}%")
        if serial:
            clauses.append('"SERIENNUMMER" LIKE ?')
            params.append(f"%{serial}%")
        if facility:
            clauses.append('"LAGERORT" LIKE ?')
            params.append(f"%{facility}%")
        if bin:
            clauses.append('"LAGERPLATZ" LIKE ?')
            params.append(f"%{bin}%")
        sql = (
            f'SELECT ID, "BAUREIHE", "ITNO", "SERIENNUMMER", "WAGEN-TYP", LAGERORT, LAGERPLATZ '
            f"FROM {table_name} "
            f"WHERE {' AND '.join(clauses)} "
            f'ORDER BY "BAUREIHE", "SERIENNUMMER" '
            f"LIMIT ?"
        )
        params.append(limit)
        rows = [dict(row) for row in conn.execute(sql, params).fetchall()]
    return {"rows": rows, "eqtp": eqtp, "env": _normalize_env(env)}


@app.get("/api/spareparts/filters")
def spareparts_filters(
    eqtp: str = Query(..., min_length=1),
    env: str = Query(DEFAULT_ENV),
) -> dict:
    table_name = _table_for(SPAREPARTS_TABLE, env)
    with _connect() as conn:
        _ensure_table(conn, table_name, SPAREPARTS_TABLE)

        def fetch(column: str, limit: int = 250) -> list[str]:
            sql = (
                f'SELECT DISTINCT "{column}" '
                f'FROM {table_name} '
                f'WHERE TEILEART = ? '
                f'AND UPPER(IFNULL(LAGERPLATZ, "")) <> "INSTALLED" '
                f'AND IFNULL("{column}", "") <> "" '
                f'ORDER BY "{column}" LIMIT {limit}'
            )
            return [row[0] for row in conn.execute(sql, (eqtp,)).fetchall()]

        return {
            "types": fetch("WAGEN-TYP"),
            "items": fetch("BAUREIHE"),
            "serials": fetch("SERIENNUMMER"),
            "facilities": fetch("LAGERORT"),
            "bins": fetch("LAGERPLATZ"),
        }


@app.get("/api/spareparts/selections")
def spareparts_selections(
    mtrl: str = Query(..., min_length=1),
    sern: str = Query(..., min_length=1),
    env: str = Query(DEFAULT_ENV),
) -> dict:
    table_name = _table_for(SPAREPARTS_SWAP_TABLE, env)
    with _connect() as conn:
        _ensure_swap_table(conn, table_name)
        rows = [
            dict(row)
            for row in conn.execute(
                f"""
                SELECT *
                FROM {table_name}
                WHERE WAGEN_ITNO = ? AND WAGEN_SERN = ?
                """,
                (mtrl, sern),
            ).fetchall()
        ]
    return {"rows": rows, "env": _normalize_env(env)}


@app.post("/api/spareparts/select")
def spareparts_select(
    env: str = Query(DEFAULT_ENV),
    payload: dict = Body(...),
) -> dict:
    required = [
        "WAGEN_ITNO",
        "WAGEN_SERN",
        "ORIGINAL_ITNO",
        "ORIGINAL_SERN",
        "ERSATZ_ITNO",
        "ERSATZ_SERN",
    ]
    for field in required:
        if not payload.get(field):
            raise HTTPException(status_code=400, detail=f"Feld {field} ist erforderlich.")

    user = payload.get("USER") or os.getenv("SPAREPART_USER", "UNBEKANNT")
    upload_flag = payload.get("UPLOAD") or "N"
    timestamp = payload.get("TIMESTAMP") or datetime.utcnow().isoformat(timespec="seconds")

    table_name = _table_for(SPAREPARTS_SWAP_TABLE, env)
    with _connect() as conn:
        _ensure_swap_table(conn, table_name)
        conn.execute(
            f"""
            INSERT INTO {table_name} (
                WAGEN_ITNO, WAGEN_SERN, ORIGINAL_ITNO, ORIGINAL_SERN,
                ERSATZ_ITNO, ERSATZ_SERN, USER, UPLOAD, TIMESTAMP
            ) VALUES (?, ?, ?, ?, ?, ?, ?, ?, ?)
            ON CONFLICT(WAGEN_ITNO, WAGEN_SERN, ORIGINAL_ITNO, ORIGINAL_SERN)
            DO UPDATE SET
                ERSATZ_ITNO=excluded.ERSATZ_ITNO,
                ERSATZ_SERN=excluded.ERSATZ_SERN,
                USER=excluded.USER,
                UPLOAD=excluded.UPLOAD,
                TIMESTAMP=excluded.TIMESTAMP
            """,
            (
                payload["WAGEN_ITNO"],
                payload["WAGEN_SERN"],
                payload["ORIGINAL_ITNO"],
                payload["ORIGINAL_SERN"],
                payload["ERSATZ_ITNO"],
                payload["ERSATZ_SERN"],
                user,
                upload_flag,
                timestamp,
            ),
        )
        conn.commit()
    return {
        "message": "Ersatzteil gespeichert",
        "record": {
            **payload,
            "USER": user,
            "UPLOAD": upload_flag,
            "TIMESTAMP": timestamp,
            "env": _normalize_env(env),
        },
    }


@app.delete("/api/spareparts/select")
def spareparts_delete(
    env: str = Query(DEFAULT_ENV),
    payload: dict = Body(...),
) -> dict:
    required = ["WAGEN_ITNO", "WAGEN_SERN", "ORIGINAL_ITNO", "ORIGINAL_SERN"]
    for field in required:
        if not payload.get(field):
            raise HTTPException(status_code=400, detail=f"Feld {field} ist erforderlich.")
    table_name = _table_for(SPAREPARTS_SWAP_TABLE, env)
    with _connect() as conn:
        _ensure_swap_table(conn, table_name)
        cursor = conn.execute(
            f"""
            DELETE FROM {table_name}
            WHERE WAGEN_ITNO = ? AND WAGEN_SERN = ? AND ORIGINAL_ITNO = ? AND ORIGINAL_SERN = ?
            """,
            (
                payload["WAGEN_ITNO"],
                payload["WAGEN_SERN"],
                payload["ORIGINAL_ITNO"],
                payload["ORIGINAL_SERN"],
            ),
        )
        conn.commit()
    if cursor.rowcount == 0:
        raise HTTPException(status_code=404, detail="Kein Eintrag zum Löschen gefunden.")
    return {"message": "Ersatzteilzuordnung gelöscht", "env": _normalize_env(env)}


@app.get("/api/spareparts/swaps")
def spareparts_swaps(
    upload: str = Query("N"),
    env: str = Query(DEFAULT_ENV),
) -> dict:
    flag = (upload or "").strip().upper()
    table_name = _table_for(SPAREPARTS_SWAP_TABLE, env)
    with _connect() as conn:
        _ensure_swap_table(conn, table_name)
        base_query = f"SELECT rowid AS ID, * FROM {table_name}"
        params: List[str] = []
        if flag:
            base_query += " WHERE UPPER(COALESCE(UPLOAD, '')) = ?"
            params.append(flag)
        base_query += " ORDER BY COALESCE(TIMESTAMP, '') DESC"
        rows = [dict(row) for row in conn.execute(base_query, params).fetchall()]
    return {"rows": rows, "env": _normalize_env(env)}


@app.post("/api/rsrd2/load_erp")
def rsrd2_load_erp(env: str = Query(DEFAULT_ENV)) -> dict:
    job = _start_subprocess_job(
        "load_erp",
        _build_load_erp_cmd(env),
        env,
        lambda job_id: _finalize_load_erp(job_id, env),
    )
    return {"job_id": job["id"], "status": job["status"], "env": job["env"]}


@app.post("/api/rsrd2/load_erp_full")
def rsrd2_load_erp_full(env: str = Query(DEFAULT_ENV)) -> dict:
    job = _start_subprocess_job(
        "load_erp_full",
        _build_erp_full_cmd(env),
        env,
        lambda job_id: _finalize_load_erp_full(job_id, env),
    )
    return {"job_id": job["id"], "status": job["status"], "env": job["env"]}


@app.get("/api/rsrd2/jobs/{job_id}")
def rsrd2_job_status(job_id: str) -> dict:
    return _job_snapshot(job_id)


@app.get("/api/rsrd2/wagons")
def rsrd2_wagons(
    limit: int = Query(50, ge=1, le=200),
    offset: int = Query(0, ge=0),
    env: str = Query(DEFAULT_ENV),
    rsrd_env: str | None = Query(None),
) -> dict:
    rsrd_env_norm = _normalize_rsrd_env(rsrd_env, env)
    with _connect() as conn:
        tables = _ensure_rsrd_tables(conn, rsrd_env_norm)
        rows = [
            {
                "wagon_id": row["wagon_id"],
                "updated_at": row["updated_at"],
                "data": json.loads(row["data_json"]),
            }
            for row in conn.execute(
                f"""
                SELECT wagon_id, data_json, updated_at
                FROM {tables.wagons}
                ORDER BY updated_at DESC
                LIMIT ? OFFSET ?
                """,
                (limit, offset),
            )
        ]
        total = conn.execute(f"SELECT COUNT(*) FROM {tables.wagons}").fetchone()[0]
    return {
        "rows": rows,
        "limit": limit,
        "offset": offset,
        "total": total,
        "erp_env": _normalize_env(env),
        "rsrd_env": rsrd_env_norm,
    }


@app.get("/api/rsrd2/suggestions")
def rsrd2_suggestions(
    field: str = Query(...),
    q: str = Query(""),
    limit: int = Query(20, ge=1, le=200),
    env: str = Query(DEFAULT_ENV),
) -> dict:
    field = field.strip().lower()
    if not q.strip():
        return {"values": [], "env": _normalize_env(env)}

    erp_table = _table_for(RSRD_ERP_FULL_TABLE, env)
    numbers_table = _table_for(RSRD_ERP_TABLE, env)
    with _connect() as conn:
        if not _table_exists(conn, erp_table):
            raise HTTPException(status_code=404, detail=f"Tabelle {erp_table} nicht gefunden.")
        numbers_exists = _table_exists(conn, numbers_table)
        join_numbers = ""
        wagen_typ_expr = "e.WAGEN_TYP"
        if numbers_exists:
            join_numbers = (
                f"LEFT JOIN {numbers_table} n "
                "ON n.wagon_sern_numeric = CAST(e.WAGEN_SERIENNUMMER AS TEXT)"
            )
            wagen_typ_expr = "COALESCE(e.WAGEN_TYP, n.wagon_typ)"

        if field == "sern":
            expr = "e.ERP_SERIENNUMMER"
            where_expr = "REPLACE(REPLACE(e.ERP_SERIENNUMMER, ' ', ''), '-', '')"
            pattern = _sern_filter_pattern(q)
        elif field == "baureihe":
            expr = "e.WG_BAUREIHE"
            where_expr = expr
            pattern = _like_pattern(q)
        elif field == "halter":
            expr = "e.WG_HALTER_CODE"
            where_expr = expr
            pattern = _like_pattern(q)
        elif field in {"wagen_typ", "uic"}:
            expr = wagen_typ_expr
            where_expr = wagen_typ_expr
            pattern = _like_pattern(q)
        else:
            raise HTTPException(status_code=400, detail="Unbekanntes Feld.")

        rows = conn.execute(
            f"""
            SELECT DISTINCT {expr} AS value
            FROM {erp_table} e
            {join_numbers}
            WHERE {where_expr} IS NOT NULL
              AND TRIM({where_expr}) <> ''
              AND {where_expr} LIKE ? ESCAPE '\\'
            ORDER BY value
            LIMIT ?
            """,
            (pattern, limit),
        ).fetchall()
    values = [row["value"] for row in rows if row and row["value"] is not None]
    return {"values": values, "env": _normalize_env(env)}


@app.get("/api/rsrd2/overview")
def rsrd2_overview(
    limit: int = Query(25, ge=1, le=200),
    offset: int = Query(0, ge=0),
    env: str = Query(DEFAULT_ENV),
    rsrd_env: str | None = Query(None),
    sern: str | None = Query(None),
    baureihe: str | None = Query(None),
    halter: str | None = Query(None),
    uic: str | None = Query(None),
    status: str | None = Query(None),
    diff_count: str | None = Query(None),
    docs: str | None = Query(None),
) -> dict:
    rsrd_env_norm = _normalize_rsrd_env(rsrd_env, env)
    erp_table = _table_for(RSRD_ERP_FULL_TABLE, env)
    sync_table = _table_for(RSRD_SYNC_TABLE, env)
    selection_table = _table_for(RSRD_SYNC_SELECTION_TABLE, env)
    numbers_table = _table_for(RSRD_ERP_TABLE, env)
    filters = []
    params: List[Any] = []

    if sern:
        filters.append(
            "REPLACE(REPLACE(CAST(e.WAGEN_SERIENNUMMER AS TEXT), ' ', ''), '-', '') LIKE ? ESCAPE '\\'"
        )
        params.append(_sern_filter_pattern(sern))
    if baureihe:
        filters.append("e.WG_BAUREIHE LIKE ? ESCAPE '\\'")
        params.append(_like_pattern(baureihe))
    if halter:
        filters.append("e.WG_HALTER_CODE LIKE ? ESCAPE '\\'")
        params.append(_like_pattern(halter))

    with _connect() as conn:
        if not _table_exists(conn, erp_table):
            raise HTTPException(status_code=404, detail=f"Tabelle {erp_table} nicht gefunden.")
        _ensure_rsrd_sync_table(conn, env)
        _ensure_rsrd_sync_selection_table(conn, env)
        _ensure_rsrd_document_tables(conn)
        tables = _ensure_rsrd_tables(conn, rsrd_env_norm)
        upload_table = _ensure_rsrd_upload_table(conn, rsrd_env_norm)
        numbers_exists = _table_exists(conn, numbers_table)
        join_numbers = ""
        wagen_typ_expr = "e.WAGEN_TYP"
        if numbers_exists:
            join_numbers = (
                f"LEFT JOIN {numbers_table} n "
                "ON n.wagon_sern_numeric = CAST(e.WAGEN_SERIENNUMMER AS TEXT)"
            )
            wagen_typ_expr = "COALESCE(e.WAGEN_TYP, n.wagon_typ)"

        if uic:
            if numbers_exists:
                filters.append("COALESCE(e.WAGEN_TYP, n.wagon_typ) LIKE ? ESCAPE '\\'")
            else:
                filters.append("e.WAGEN_TYP LIKE ? ESCAPE '\\'")
            params.append(_like_pattern(uic))
        if status:
            status_norm = status.strip().lower()
            if status_norm in {"green", "ok", "present"}:
                filters.append("r.wagon_id IS NOT NULL")
            elif status_norm in {"red", "missing"}:
                filters.append("r.wagon_id IS NULL")
        diff_count_expr = (
            "CAST((LENGTH(u.diff_json) - LENGTH(REPLACE(u.diff_json, '\"equal\": false', ''))) "
            "/ LENGTH('\"equal\": false') AS INTEGER)"
        )
        docs_norm = str(docs or "").strip().lower()
        if docs_norm in {"yes", "ja", "j", "y", "1", "true"}:
            filters.append(
                f"""
                EXISTS (
                    SELECT 1
                    FROM {RSRD_DOCUMENT_ASSIGNMENTS_TABLE} a
                    WHERE (
                        a.assign_type = 'baureihe'
                        AND UPPER(TRIM(a.assign_value)) = UPPER(TRIM(e.WG_BAUREIHE))
                    ) OR (
                        a.assign_type = 'wagen_typ'
                        AND UPPER(TRIM(a.assign_value)) = UPPER(TRIM({wagen_typ_expr}))
                    )
                )
                """
            )
        elif docs_norm in {"no", "nein", "n", "0", "false"}:
            filters.append(
                f"""
                NOT EXISTS (
                    SELECT 1
                    FROM {RSRD_DOCUMENT_ASSIGNMENTS_TABLE} a
                    WHERE (
                        a.assign_type = 'baureihe'
                        AND UPPER(TRIM(a.assign_value)) = UPPER(TRIM(e.WG_BAUREIHE))
                    ) OR (
                        a.assign_type = 'wagen_typ'
                        AND UPPER(TRIM(a.assign_value)) = UPPER(TRIM({wagen_typ_expr}))
                    )
                )
                """
            )

        base_filters = list(filters)
        base_params = list(params)
        diff_count_norm = str(diff_count or "").strip()
        if diff_count_norm:
            try:
                diff_count_value = int(diff_count_norm)
            except ValueError:
                diff_count_value = None
            if diff_count_value is not None and diff_count_value >= 0:
                filters.append(f"u.diff_json IS NOT NULL AND {diff_count_expr} = ?")
                params.append(diff_count_value)

        where_clause = f"WHERE {' AND '.join(filters)}" if filters else ""
        base_where_clause = f"WHERE {' AND '.join(base_filters)}" if base_filters else ""
        diff_count_rows = conn.execute(
            f"""
            SELECT diff_count, COUNT(*) AS item_count
            FROM (
                SELECT
                    CASE
                        WHEN u.diff_json IS NULL OR TRIM(u.diff_json) = '' THEN NULL
                        ELSE {diff_count_expr}
                    END AS diff_count
                FROM {erp_table} e
                {join_numbers}
                LEFT JOIN {upload_table} u
                  ON u.wagon_number_freight = CAST(e.WAGEN_SERIENNUMMER AS TEXT)
                LEFT JOIN {tables.detail} r
                  ON r.wagon_number_freight = REPLACE(REPLACE(CAST(e.WAGEN_SERIENNUMMER AS TEXT), ' ', ''), '-', '')
                {base_where_clause}
            ) x
            WHERE diff_count IS NOT NULL
            GROUP BY diff_count
            ORDER BY diff_count
            """,
            base_params,
        ).fetchall()
        total = conn.execute(
            f"""
            SELECT COUNT(*)
            FROM {erp_table} e
            {join_numbers}
            LEFT JOIN {upload_table} u
              ON u.wagon_number_freight = CAST(e.WAGEN_SERIENNUMMER AS TEXT)
            LEFT JOIN {tables.detail} r
              ON r.wagon_number_freight = REPLACE(REPLACE(CAST(e.WAGEN_SERIENNUMMER AS TEXT), ' ', ''), '-', '')
            {where_clause}
            """,
            params,
        ).fetchone()[0]
        rows = conn.execute(
            f"""
            SELECT
                CAST(e.WAGEN_SERIENNUMMER AS TEXT) AS wagon_number,
                e.WG_BAUREIHE AS baureihe,
                {wagen_typ_expr} AS wagen_typ,
                e.WG_HALTER_CODE AS halter_code,
                s.enabled AS sync_enabled,
                sel.sync_data_env AS sync_data_env,
                sel.sync_km_env AS sync_km_env,
                sel.sync_docs_env AS sync_docs_env,
                sel.one_time_transfer AS one_time_transfer,
                sel.updated_at AS sync_updated_at,
                r.wagon_id AS rsrd_wagon_id,
                CASE
                    WHEN u.diff_json IS NULL OR TRIM(u.diff_json) = '' THEN NULL
                    ELSE {diff_count_expr}
                END AS compare_diff_count,
                u.diff_json AS compare_diff_json,
                u.updated_at AS compare_updated_at
            FROM {erp_table} e
            LEFT JOIN {sync_table} s
              ON s.wagon_number_freight = CAST(e.WAGEN_SERIENNUMMER AS TEXT)
            LEFT JOIN {selection_table} sel
              ON sel.wagon_number_freight = CAST(e.WAGEN_SERIENNUMMER AS TEXT)
            LEFT JOIN {upload_table} u
              ON u.wagon_number_freight = CAST(e.WAGEN_SERIENNUMMER AS TEXT)
            {join_numbers}
            LEFT JOIN {tables.detail} r
              ON r.wagon_number_freight = REPLACE(REPLACE(CAST(e.WAGEN_SERIENNUMMER AS TEXT), ' ', ''), '-', '')
            {where_clause}
            ORDER BY CAST(e.WAGEN_SERIENNUMMER AS TEXT)
            LIMIT ? OFFSET ?
            """,
            params + [limit, offset],
        ).fetchall()
        doc_map = _load_matching_documents(conn, rows)

    out_rows = []
    for row in rows:
        item = dict(row)
        item["rsrd_present"] = bool(item.get("rsrd_wagon_id"))
        wagon = str(item.get("wagon_number") or "").strip()
        docs = doc_map.get(wagon, [])
        item["sync_docs_flag"] = "J" if docs else "N"
        item["sync_docs_items"] = docs
        item["sync_docs_text"] = ", ".join(docs)
        diff_count = item.get("compare_diff_count")
        if diff_count is not None:
            try:
                diff_count = int(diff_count)
            except Exception:
                diff_count = None
        raw_diff = item.get("compare_diff_json")
        if diff_count is None and raw_diff:
            try:
                parsed_diffs = json.loads(raw_diff)
                if isinstance(parsed_diffs, list):
                    diff_count = sum(
                        1
                        for diff in parsed_diffs
                        if not (isinstance(diff, dict) and diff.get("equal"))
                    )
            except Exception:
                diff_count = None
        item.pop("compare_diff_count", None)
        item.pop("compare_diff_json", None)
        item.pop("compare_updated_at", None)
        item["diff_count"] = diff_count
        out_rows.append(item)

    diff_count_options = []
    for entry in diff_count_rows:
        try:
            value = int(entry["diff_count"])
            count = int(entry["item_count"])
        except Exception:
            continue
        if value < 0:
            continue
        diff_count_options.append({"value": value, "count": count})

    return {
        "rows": out_rows,
        "limit": limit,
        "offset": offset,
        "total": total,
        "diff_count_options": diff_count_options,
        "erp_env": _normalize_env(env),
        "rsrd_env": rsrd_env_norm,
    }


@app.get("/api/rsrd2/documents")
def rsrd2_documents(env: str = Query(DEFAULT_ENV)) -> dict:
    erp_table = _table_for(RSRD_ERP_FULL_TABLE, env)
    with _connect() as conn:
        _ensure_rsrd_document_tables(conn)
        docs_rows = conn.execute(
            f"""
            SELECT
                id,
                display_name,
                original_name,
                mime_type,
                size_bytes,
                uploaded_at,
                updated_at
            FROM {RSRD_DOCUMENTS_TABLE}
            ORDER BY uploaded_at DESC, id DESC
            """
        ).fetchall()
        assign_rows = conn.execute(
            f"""
            SELECT document_id, assign_type, assign_value
            FROM {RSRD_DOCUMENT_ASSIGNMENTS_TABLE}
            ORDER BY assign_type, assign_value
            """
        ).fetchall()
        by_doc: Dict[int, Dict[str, List[str]]] = {}
        for row in assign_rows:
            doc_id = int(row["document_id"])
            if doc_id not in by_doc:
                by_doc[doc_id] = {"baureihen": [], "wagen_typen": []}
            target = "baureihen" if row["assign_type"] == "baureihe" else "wagen_typen"
            value = str(row["assign_value"] or "").strip()
            if value and value not in by_doc[doc_id][target]:
                by_doc[doc_id][target].append(value)

        baureihe_options: List[str] = []
        wagen_typ_options: List[str] = []
        if _table_exists(conn, erp_table):
            baureihe_rows = conn.execute(
                f"""
                SELECT DISTINCT WG_BAUREIHE AS value
                FROM {erp_table}
                WHERE WG_BAUREIHE IS NOT NULL AND TRIM(WG_BAUREIHE) <> ''
                ORDER BY WG_BAUREIHE
                LIMIT 2000
                """
            ).fetchall()
            wagen_typ_rows = conn.execute(
                f"""
                SELECT DISTINCT WAGEN_TYP AS value
                FROM {erp_table}
                WHERE WAGEN_TYP IS NOT NULL AND TRIM(WAGEN_TYP) <> ''
                ORDER BY WAGEN_TYP
                LIMIT 2000
                """
            ).fetchall()
            baureihe_options = [str(row["value"]) for row in baureihe_rows]
            wagen_typ_options = [str(row["value"]) for row in wagen_typ_rows]

    documents = []
    for row in docs_rows:
        doc_id = int(row["id"])
        assigns = by_doc.get(doc_id, {"baureihen": [], "wagen_typen": []})
        documents.append(
            {
                "id": doc_id,
                "display_name": row["display_name"],
                "original_name": row["original_name"],
                "mime_type": row["mime_type"],
                "size_bytes": row["size_bytes"],
                "uploaded_at": row["uploaded_at"],
                "updated_at": row["updated_at"],
                "baureihen": assigns["baureihen"],
                "wagen_typen": assigns["wagen_typen"],
                "download_url": f"/api/rsrd2/documents/{doc_id}/download",
            }
        )
    return {
        "documents": documents,
        "baureihe_options": baureihe_options,
        "wagen_typ_options": wagen_typ_options,
        "env": _normalize_env(env),
    }


@app.post("/api/rsrd2/documents/upload")
async def rsrd2_documents_upload(
    request: Request,
    name: str | None = Query(None),
) -> dict:
    raw = await request.body()
    if not raw:
        raise HTTPException(status_code=400, detail="Dateiinhalt fehlt.")
    header_name = str(request.headers.get("x-file-name") or "").strip()
    provided_name = str(name or "").strip()
    decoded_name = unquote(provided_name or header_name or "dokument")
    original_name = Path(decoded_name).name or "dokument"
    safe_name = re.sub(r"[^0-9A-Za-z._-]+", "_", original_name).strip("._")
    if not safe_name:
        safe_name = "dokument"
    stored_name = f"{datetime.utcnow().strftime('%Y%m%d%H%M%S')}_{uuid.uuid4().hex}_{safe_name}"
    content_type = str(request.headers.get("content-type") or "").split(";", 1)[0].strip()
    mime_type = content_type or mimetypes.guess_type(original_name)[0] or "application/octet-stream"
    now = datetime.utcnow().isoformat(timespec="seconds")
    target = _rsrd_documents_dir() / stored_name
    target.write_bytes(raw)

    with _connect() as conn:
        _ensure_rsrd_document_tables(conn)
        cursor = conn.execute(
            f"""
            INSERT INTO {RSRD_DOCUMENTS_TABLE} (
                display_name,
                original_name,
                stored_name,
                mime_type,
                size_bytes,
                uploaded_at,
                updated_at
            ) VALUES (?, ?, ?, ?, ?, ?, ?)
            """,
            (original_name, original_name, stored_name, mime_type, len(raw), now, now),
        )
        doc_id = int(cursor.lastrowid)
        conn.commit()
    return {
        "id": doc_id,
        "display_name": original_name,
        "original_name": original_name,
        "mime_type": mime_type,
        "size_bytes": len(raw),
        "uploaded_at": now,
    }


@app.post("/api/rsrd2/documents/{doc_id}/assignments")
def rsrd2_documents_assignments(
    doc_id: int,
    payload: dict = Body(...),
) -> dict:
    baureihen = _normalize_doc_assign_values(payload.get("baureihen"))
    wagen_typen = _normalize_doc_assign_values(payload.get("wagen_typen"))
    now = datetime.utcnow().isoformat(timespec="seconds")
    with _connect() as conn:
        _ensure_rsrd_document_tables(conn)
        exists = conn.execute(
            f"SELECT id FROM {RSRD_DOCUMENTS_TABLE} WHERE id = ?",
            (doc_id,),
        ).fetchone()
        if not exists:
            raise HTTPException(status_code=404, detail="Dokument nicht gefunden.")
        conn.execute(
            f"DELETE FROM {RSRD_DOCUMENT_ASSIGNMENTS_TABLE} WHERE document_id = ?",
            (doc_id,),
        )
        for value in baureihen:
            conn.execute(
                f"""
                INSERT INTO {RSRD_DOCUMENT_ASSIGNMENTS_TABLE} (
                    document_id, assign_type, assign_value, created_at
                ) VALUES (?, 'baureihe', ?, ?)
                """,
                (doc_id, value, now),
            )
        for value in wagen_typen:
            conn.execute(
                f"""
                INSERT INTO {RSRD_DOCUMENT_ASSIGNMENTS_TABLE} (
                    document_id, assign_type, assign_value, created_at
                ) VALUES (?, 'wagen_typ', ?, ?)
                """,
                (doc_id, value, now),
            )
        conn.execute(
            f"UPDATE {RSRD_DOCUMENTS_TABLE} SET updated_at = ? WHERE id = ?",
            (now, doc_id),
        )
        conn.commit()
    return {
        "id": doc_id,
        "baureihen": baureihen,
        "wagen_typen": wagen_typen,
        "updated_at": now,
    }


@app.get("/api/rsrd2/documents/{doc_id}/download")
def rsrd2_documents_download(doc_id: int) -> FileResponse:
    with _connect() as conn:
        _ensure_rsrd_document_tables(conn)
        row = conn.execute(
            f"""
            SELECT original_name, stored_name, mime_type
            FROM {RSRD_DOCUMENTS_TABLE}
            WHERE id = ?
            """,
            (doc_id,),
        ).fetchone()
    if not row:
        raise HTTPException(status_code=404, detail="Dokument nicht gefunden.")
    path = _rsrd_documents_dir() / str(row["stored_name"])
    if not path.exists():
        raise HTTPException(status_code=404, detail=f"Datei nicht gefunden: {path}")
    return FileResponse(
        path,
        media_type=str(row["mime_type"] or "application/octet-stream"),
        filename=str(row["original_name"] or path.name),
    )


@app.delete("/api/rsrd2/documents/{doc_id}")
def rsrd2_documents_delete(doc_id: int) -> dict:
    with _connect() as conn:
        _ensure_rsrd_document_tables(conn)
        row = conn.execute(
            f"SELECT stored_name FROM {RSRD_DOCUMENTS_TABLE} WHERE id = ?",
            (doc_id,),
        ).fetchone()
        if not row:
            raise HTTPException(status_code=404, detail="Dokument nicht gefunden.")
        conn.execute(
            f"DELETE FROM {RSRD_DOCUMENT_ASSIGNMENTS_TABLE} WHERE document_id = ?",
            (doc_id,),
        )
        conn.execute(
            f"DELETE FROM {RSRD_DOCUMENTS_TABLE} WHERE id = ?",
            (doc_id,),
        )
        conn.commit()
    file_path = _rsrd_documents_dir() / str(row["stored_name"])
    if file_path.exists():
        try:
            file_path.unlink()
        except OSError:
            pass
    return {"deleted": doc_id}


@app.post("/api/rsrd2/sync_flag")
def rsrd2_sync_flag(
    env: str = Query(DEFAULT_ENV),
    payload: dict = Body(...),
) -> dict:
    wagon = str(payload.get("wagon") or "").strip()
    enabled = bool(payload.get("enabled"))
    if not wagon:
        raise HTTPException(status_code=400, detail="Wagennummer fehlt.")
    table_name = _table_for(RSRD_SYNC_TABLE, env)
    timestamp = datetime.utcnow().isoformat(timespec="seconds")
    with _connect() as conn:
        _ensure_table(conn, table_name, None)
        conn.execute(
            f"""
            INSERT INTO {table_name} (wagon_number_freight, enabled, updated_at)
            VALUES (?, ?, ?)
            ON CONFLICT(wagon_number_freight)
            DO UPDATE SET enabled=excluded.enabled, updated_at=excluded.updated_at
            """,
            (wagon, "1" if enabled else "0", timestamp),
        )
        conn.commit()
    return {"wagon": wagon, "enabled": enabled, "updated_at": timestamp, "env": _normalize_env(env)}


@app.post("/api/rsrd2/sync_env")
def rsrd2_sync_env(
    env: str = Query(DEFAULT_ENV),
    payload: dict = Body(...),
) -> dict:
    wagon = str(payload.get("wagon") or "").strip()
    kind = str(payload.get("kind") or "").strip().lower()
    value = str(payload.get("value") or "").strip().upper()
    if not wagon:
        raise HTTPException(status_code=400, detail="Wagennummer fehlt.")
    column_map = {
        "data": "sync_data_env",
        "km": "sync_km_env",
        "docs": "sync_docs_env",
    }
    column = column_map.get(kind)
    if not column:
        raise HTTPException(status_code=400, detail="Ungültiger Sync-Typ.")
    if value in {"J", "T", "P"}:
        value = "J"
    else:
        value = "N"
    table_name = _table_for(RSRD_SYNC_SELECTION_TABLE, env)
    timestamp = datetime.utcnow().isoformat(timespec="seconds")
    with _connect() as conn:
        _ensure_rsrd_sync_selection_table(conn, env)
        conn.execute(
            f"""
            INSERT INTO {table_name} (wagon_number_freight, {column}, updated_at)
            VALUES (?, ?, ?)
            ON CONFLICT(wagon_number_freight)
            DO UPDATE SET {column}=excluded.{column}, updated_at=excluded.updated_at
            """,
            (wagon, value, timestamp),
        )
        conn.commit()
    return {"wagon": wagon, "kind": kind, "value": value, "updated_at": timestamp, "env": _normalize_env(env)}


@app.post("/api/rsrd2/sync_env_bulk")
def rsrd2_sync_env_bulk(
    env: str = Query(DEFAULT_ENV),
    rsrd_env: str | None = Query(None),
    payload: dict = Body(...),
) -> dict:
    rsrd_env_norm = _normalize_rsrd_env(rsrd_env, env)
    kind = str(payload.get("kind") or "").strip().lower()
    value = str(payload.get("value") or "").strip().upper()
    filters = payload.get("filters") if isinstance(payload, dict) else None
    filters = filters if isinstance(filters, dict) else {}
    column_map = {
        "data": "sync_data_env",
        "km": "sync_km_env",
        "docs": "sync_docs_env",
    }
    column = column_map.get(kind)
    if not column:
        raise HTTPException(status_code=400, detail="Ungültiger Sync-Typ.")
    if value in {"J", "T", "P"}:
        value = "J"
    else:
        value = "N"

    sern = str(filters.get("sern") or "").strip()
    baureihe = str(filters.get("baureihe") or "").strip()
    halter = str(filters.get("halter") or "").strip()
    uic = str(filters.get("uic") or "").strip()
    status = str(filters.get("status") or "").strip().lower()
    docs = str(filters.get("docs") or "").strip().lower()

    where = []
    params: List[Any] = []
    if sern:
        where.append(
            "REPLACE(REPLACE(CAST(e.WAGEN_SERIENNUMMER AS TEXT), ' ', ''), '-', '') LIKE ? ESCAPE '\\'"
        )
        params.append(_sern_filter_pattern(sern))
    if baureihe:
        where.append("e.WG_BAUREIHE LIKE ? ESCAPE '\\'")
        params.append(_like_pattern(baureihe))
    if halter:
        where.append("e.WG_HALTER_CODE LIKE ? ESCAPE '\\'")
        params.append(_like_pattern(halter))
    if status in {"green", "ok", "present"}:
        where.append("r.wagon_id IS NOT NULL")
    elif status in {"red", "missing"}:
        where.append("r.wagon_id IS NULL")
    timestamp = datetime.utcnow().isoformat(timespec="seconds")

    erp_table = _table_for(RSRD_ERP_FULL_TABLE, env)
    selection_table = _table_for(RSRD_SYNC_SELECTION_TABLE, env)
    numbers_table = _table_for(RSRD_ERP_TABLE, env)

    with _connect() as conn:
        if not _table_exists(conn, erp_table):
            raise HTTPException(status_code=404, detail=f"Tabelle {erp_table} nicht gefunden.")
        tables = _ensure_rsrd_tables(conn, rsrd_env_norm)
        _ensure_rsrd_sync_selection_table(conn, env)
        _ensure_rsrd_document_tables(conn)
        numbers_exists = _table_exists(conn, numbers_table)
        join_numbers = ""
        wagen_typ_expr = "e.WAGEN_TYP"
        if numbers_exists:
            join_numbers = (
                f"LEFT JOIN {numbers_table} n "
                "ON n.wagon_sern_numeric = CAST(e.WAGEN_SERIENNUMMER AS TEXT)"
            )
            wagen_typ_expr = "COALESCE(e.WAGEN_TYP, n.wagon_typ)"

        if uic:
            if numbers_exists:
                where.append("COALESCE(e.WAGEN_TYP, n.wagon_typ) LIKE ? ESCAPE '\\'")
            else:
                where.append("e.WAGEN_TYP LIKE ? ESCAPE '\\'")
            params.append(_like_pattern(uic))
        if docs in {"yes", "ja", "j", "y", "1", "true"}:
            where.append(
                f"""
                EXISTS (
                    SELECT 1
                    FROM {RSRD_DOCUMENT_ASSIGNMENTS_TABLE} a
                    WHERE (
                        a.assign_type = 'baureihe'
                        AND UPPER(TRIM(a.assign_value)) = UPPER(TRIM(e.WG_BAUREIHE))
                    ) OR (
                        a.assign_type = 'wagen_typ'
                        AND UPPER(TRIM(a.assign_value)) = UPPER(TRIM({wagen_typ_expr}))
                    )
                )
                """
            )
        elif docs in {"no", "nein", "n", "0", "false"}:
            where.append(
                f"""
                NOT EXISTS (
                    SELECT 1
                    FROM {RSRD_DOCUMENT_ASSIGNMENTS_TABLE} a
                    WHERE (
                        a.assign_type = 'baureihe'
                        AND UPPER(TRIM(a.assign_value)) = UPPER(TRIM(e.WG_BAUREIHE))
                    ) OR (
                        a.assign_type = 'wagen_typ'
                        AND UPPER(TRIM(a.assign_value)) = UPPER(TRIM({wagen_typ_expr}))
                    )
                )
                """
            )

        where_clause = f"WHERE {' AND '.join(where)}" if where else ""
        total = conn.execute(
            f"""
            SELECT COUNT(*)
            FROM {erp_table} e
            {join_numbers}
            LEFT JOIN {tables.detail} r
              ON r.wagon_number_freight = REPLACE(REPLACE(CAST(e.WAGEN_SERIENNUMMER AS TEXT), ' ', ''), '-', '')
            {where_clause}
            """,
            params,
        ).fetchone()[0]

        conn.execute(
            f"""
            INSERT INTO {selection_table} (wagon_number_freight, {column}, updated_at)
            SELECT CAST(e.WAGEN_SERIENNUMMER AS TEXT), ?, ?
            FROM {erp_table} e
            {join_numbers}
            LEFT JOIN {tables.detail} r
              ON r.wagon_number_freight = REPLACE(REPLACE(CAST(e.WAGEN_SERIENNUMMER AS TEXT), ' ', ''), '-', '')
            {where_clause}
            ON CONFLICT(wagon_number_freight)
            DO UPDATE SET {column}=excluded.{column}, updated_at=excluded.updated_at
            """,
            [value, timestamp] + params,
        )
        conn.commit()

    return {
        "kind": kind,
        "value": value,
        "updated_at": timestamp,
        "total": total,
        "erp_env": _normalize_env(env),
        "rsrd_env": rsrd_env_norm,
    }


@app.post("/api/rsrd2/one_time_transfer")
def rsrd2_one_time_transfer(
    env: str = Query(DEFAULT_ENV),
    payload: dict = Body(...),
) -> dict:
    wagon = str(payload.get("wagon") or "").strip()
    enabled = bool(payload.get("enabled"))
    if not wagon:
        raise HTTPException(status_code=400, detail="Wagennummer fehlt.")
    table_name = _table_for(RSRD_SYNC_SELECTION_TABLE, env)
    timestamp = datetime.utcnow().isoformat(timespec="seconds")
    with _connect() as conn:
        _ensure_rsrd_sync_selection_table(conn, env)
        conn.execute(
            f"""
            INSERT INTO {table_name} (wagon_number_freight, one_time_transfer, updated_at)
            VALUES (?, ?, ?)
            ON CONFLICT(wagon_number_freight)
            DO UPDATE SET one_time_transfer=excluded.one_time_transfer, updated_at=excluded.updated_at
            """,
            (wagon, "1" if enabled else "0", timestamp),
        )
        conn.commit()
    return {"wagon": wagon, "enabled": enabled, "updated_at": timestamp, "env": _normalize_env(env)}


@app.post("/api/rsrd2/one_time_transfer_bulk")
def rsrd2_one_time_transfer_bulk(
    env: str = Query(DEFAULT_ENV),
    rsrd_env: str | None = Query(None),
    payload: dict = Body(...),
) -> dict:
    rsrd_env_norm = _normalize_rsrd_env(rsrd_env, env)
    value = str(payload.get("value") or "").strip().upper()
    filters = payload.get("filters") if isinstance(payload, dict) else None
    filters = filters if isinstance(filters, dict) else {}
    value = "1" if value == "J" else "0"

    sern = str(filters.get("sern") or "").strip()
    baureihe = str(filters.get("baureihe") or "").strip()
    halter = str(filters.get("halter") or "").strip()
    uic = str(filters.get("uic") or "").strip()
    status = str(filters.get("status") or "").strip().lower()
    docs = str(filters.get("docs") or "").strip().lower()

    where = []
    params: List[Any] = []
    if sern:
        where.append(
            "REPLACE(REPLACE(CAST(e.WAGEN_SERIENNUMMER AS TEXT), ' ', ''), '-', '') LIKE ? ESCAPE '\\'"
        )
        params.append(_sern_filter_pattern(sern))
    if baureihe:
        where.append("e.WG_BAUREIHE LIKE ? ESCAPE '\\'")
        params.append(_like_pattern(baureihe))
    if halter:
        where.append("e.WG_HALTER_CODE LIKE ? ESCAPE '\\'")
        params.append(_like_pattern(halter))
    if status in {"green", "ok", "present"}:
        where.append("r.wagon_id IS NOT NULL")
    elif status in {"red", "missing"}:
        where.append("r.wagon_id IS NULL")
    timestamp = datetime.utcnow().isoformat(timespec="seconds")

    erp_table = _table_for(RSRD_ERP_FULL_TABLE, env)
    selection_table = _table_for(RSRD_SYNC_SELECTION_TABLE, env)
    numbers_table = _table_for(RSRD_ERP_TABLE, env)

    with _connect() as conn:
        if not _table_exists(conn, erp_table):
            raise HTTPException(status_code=404, detail=f"Tabelle {erp_table} nicht gefunden.")
        tables = _ensure_rsrd_tables(conn, rsrd_env_norm)
        _ensure_rsrd_sync_selection_table(conn, env)
        _ensure_rsrd_document_tables(conn)
        numbers_exists = _table_exists(conn, numbers_table)
        join_numbers = ""
        wagen_typ_expr = "e.WAGEN_TYP"
        if numbers_exists:
            join_numbers = (
                f"LEFT JOIN {numbers_table} n "
                "ON n.wagon_sern_numeric = CAST(e.WAGEN_SERIENNUMMER AS TEXT)"
            )
            wagen_typ_expr = "COALESCE(e.WAGEN_TYP, n.wagon_typ)"

        if uic:
            if numbers_exists:
                where.append("COALESCE(e.WAGEN_TYP, n.wagon_typ) LIKE ? ESCAPE '\\'")
            else:
                where.append("e.WAGEN_TYP LIKE ? ESCAPE '\\'")
            params.append(_like_pattern(uic))
        if docs in {"yes", "ja", "j", "y", "1", "true"}:
            where.append(
                f"""
                EXISTS (
                    SELECT 1
                    FROM {RSRD_DOCUMENT_ASSIGNMENTS_TABLE} a
                    WHERE (
                        a.assign_type = 'baureihe'
                        AND UPPER(TRIM(a.assign_value)) = UPPER(TRIM(e.WG_BAUREIHE))
                    ) OR (
                        a.assign_type = 'wagen_typ'
                        AND UPPER(TRIM(a.assign_value)) = UPPER(TRIM({wagen_typ_expr}))
                    )
                )
                """
            )
        elif docs in {"no", "nein", "n", "0", "false"}:
            where.append(
                f"""
                NOT EXISTS (
                    SELECT 1
                    FROM {RSRD_DOCUMENT_ASSIGNMENTS_TABLE} a
                    WHERE (
                        a.assign_type = 'baureihe'
                        AND UPPER(TRIM(a.assign_value)) = UPPER(TRIM(e.WG_BAUREIHE))
                    ) OR (
                        a.assign_type = 'wagen_typ'
                        AND UPPER(TRIM(a.assign_value)) = UPPER(TRIM({wagen_typ_expr}))
                    )
                )
                """
            )

        where_clause = f"WHERE {' AND '.join(where)}" if where else ""

        total = conn.execute(
            f"""
            SELECT COUNT(*)
            FROM {erp_table} e
            {join_numbers}
            LEFT JOIN {tables.detail} r
              ON r.wagon_number_freight = REPLACE(REPLACE(CAST(e.WAGEN_SERIENNUMMER AS TEXT), ' ', ''), '-', '')
            {where_clause}
            """,
            params,
        ).fetchone()[0]

        conn.execute(
            f"""
            INSERT INTO {selection_table} (wagon_number_freight, one_time_transfer, updated_at)
            SELECT CAST(e.WAGEN_SERIENNUMMER AS TEXT), ?, ?
            FROM {erp_table} e
            {join_numbers}
            LEFT JOIN {tables.detail} r
              ON r.wagon_number_freight = REPLACE(REPLACE(CAST(e.WAGEN_SERIENNUMMER AS TEXT), ' ', ''), '-', '')
            {where_clause}
            ON CONFLICT(wagon_number_freight)
            DO UPDATE SET one_time_transfer=excluded.one_time_transfer, updated_at=excluded.updated_at
            """,
            [value, timestamp] + params,
        )
        conn.commit()

    return {
        "value": value,
        "updated_at": timestamp,
        "total": total,
        "erp_env": _normalize_env(env),
        "rsrd_env": rsrd_env_norm,
    }


def _normalize_documents(documents: Any) -> List[Dict[str, Any]]:
    if not documents:
        return []
    docs_obj: Any = documents
    if isinstance(docs_obj, dict):
        docs_obj = (
            docs_obj.get("Document")
            or docs_obj.get("Documents")
            or docs_obj.get("DocumentList")
            or docs_obj
        )
    if isinstance(docs_obj, dict):
        docs_list = [docs_obj]
    elif isinstance(docs_obj, list):
        docs_list = docs_obj
    else:
        return []
    cleaned: List[Dict[str, Any]] = []
    for doc in docs_list:
        if not isinstance(doc, dict):
            continue
        if not doc:
            continue
        cleaned.append(doc)
    return cleaned


@app.post("/api/rsrd2/compare")
def rsrd2_compare(
    limit: int | None = Query(None, gt=0),
    offset: int = Query(0, ge=0),
    create_upload: bool = Query(True),
    include_all: bool = Query(False),
    save_compare: bool = Query(False),
    use_live_erp_text: bool = Query(False),
    env: str = Query(DEFAULT_ENV),
    rsrd_env: str | None = Query(None),
    payload: dict | None = Body(default=None),
) -> dict:
    wagons = payload.get("wagons") if payload else None
    if wagons is not None:
        if not isinstance(wagons, list) or not all(isinstance(item, (str, int)) for item in wagons):
            raise HTTPException(status_code=400, detail="Feld 'wagons' muss eine Liste von Wagennummern sein.")
        normalized = []
        for item in wagons:
            raw = str(item).strip()
            if not raw:
                continue
            digits = re.sub(r"\D", "", raw)
            normalized.append(digits or raw)
        wagons = normalized

    rsrd_env_norm = _normalize_rsrd_env(rsrd_env, env)
    with _connect() as conn:
        tables = _ensure_rsrd_tables(conn, rsrd_env_norm)
        upload_table = _ensure_rsrd_upload_table(conn, rsrd_env_norm)
        erp_full_table = _table_for(RSRD_ERP_FULL_TABLE, env)

        where_clause = ""
        params: List[Any] = []
        if wagons:
            placeholders = ", ".join("?" for _ in wagons)
            where_clause = (
                "WHERE REPLACE(REPLACE(CAST(e.WAGEN_SERIENNUMMER AS TEXT), ' ', ''), '-', '') "
                f"IN ({placeholders})"
            )
            params.extend(wagons)

        total_query = f"SELECT COUNT(*) FROM {erp_full_table} e {where_clause}"
        total = conn.execute(total_query, params).fetchone()[0]

        query = f"""
            SELECT
                e.*,
                r.wagon_id AS rsrd_wagon_id,
                r.wagon_number_freight AS rsrd_wagon_number_freight,
                r.administrative_json,
                r.design_json,
                r.documents_json AS dataset_json,
                j.payload_json AS raw_payload_json
            FROM {erp_full_table} e
            LEFT JOIN {tables.detail} r
                ON r.wagon_number_freight = REPLACE(REPLACE(CAST(e.WAGEN_SERIENNUMMER AS TEXT), ' ', ''), '-', '')
            LEFT JOIN {tables.json} j
                ON j.wagon_id = r.wagon_id
            {where_clause}
            ORDER BY CAST(e.WAGEN_SERIENNUMMER AS TEXT)
        """
        if limit is not None:
            query += " LIMIT ? OFFSET ?"
            params = params + [limit, offset]

        rows = conn.execute(query, params).fetchall()

        results = []
        created = 0
        for row in rows:
            erp_row = dict(row)
            admin = json.loads(row["administrative_json"]) if row["administrative_json"] else {}
            design = json.loads(row["design_json"]) if row["design_json"] else {}
            dataset = json.loads(row["dataset_json"]) if row["dataset_json"] else {}
            meta = dataset.get("RSRD2MetaData") if isinstance(dataset, dict) else {}
            if not meta and row["raw_payload_json"]:
                try:
                    payload = json.loads(row["raw_payload_json"]) if row["raw_payload_json"] else {}
                except Exception:
                    payload = {}
                if isinstance(payload, dict):
                    meta = payload.get("RSRD2MetaData") or {}
            # Default comparison must be DB-vs-RSRD only.
            # Optional live enrichment can be enabled explicitly for diagnostics.
            if use_live_erp_text and wagons and len(wagons) <= 5:
                try:
                    long_text = _fetch_wg_tsi_text(erp_row.get("WAGEN_SERIENNUMMER"), env)
                except HTTPException:
                    long_text = ""
                if long_text:
                    erp_row["WG_TSI_ZUS_ZERT"] = long_text
            diffs = compare_erp_to_rsrd(erp_row, admin, design, meta or {}, include_all=include_all)
            diff_count = sum(1 for diff in diffs if not diff.get("equal"))
            documents = _normalize_documents(dataset) if isinstance(dataset, dict) else []

            payload_obj = build_erp_payload(erp_row)
            wagon_number = (payload_obj.get("AdministrativeDataSet") or {}).get("WagonNumberFreight")
            wagon_number_str = str(wagon_number) if wagon_number is not None else None

            if save_compare and wagon_number_str:
                now = datetime.utcnow().isoformat(timespec="seconds")
                conn.execute(
                    f"""
                    INSERT INTO {upload_table} (
                        wagon_number_freight,
                        rsrd_wagon_id,
                        payload_json,
                        diff_json,
                        rsrd_json,
                        created_at,
                        updated_at
                    ) VALUES (?, ?, ?, ?, ?, ?, ?)
                    ON CONFLICT(wagon_number_freight) DO UPDATE SET
                        rsrd_wagon_id=excluded.rsrd_wagon_id,
                        payload_json=excluded.payload_json,
                        diff_json=excluded.diff_json,
                        rsrd_json=excluded.rsrd_json,
                        updated_at=excluded.updated_at
                    """,
                    (
                        wagon_number_str,
                        row["rsrd_wagon_id"],
                        serialize_payload(payload_obj),
                        serialize_diffs(diffs),
                        row["dataset_json"],
                        now,
                        now,
                    ),
                )

            if create_upload and diff_count > 0 and wagon_number_str:
                now = datetime.utcnow().isoformat(timespec="seconds")
                conn.execute(
                    f"""
                    INSERT INTO {upload_table} (
                        wagon_number_freight,
                        rsrd_wagon_id,
                        payload_json,
                        diff_json,
                        rsrd_json,
                        created_at,
                        updated_at
                    ) VALUES (?, ?, ?, ?, ?, ?, ?)
                    ON CONFLICT(wagon_number_freight) DO UPDATE SET
                        rsrd_wagon_id=excluded.rsrd_wagon_id,
                        payload_json=excluded.payload_json,
                        diff_json=excluded.diff_json,
                        rsrd_json=excluded.rsrd_json,
                        updated_at=excluded.updated_at
                    """,
                    (
                        wagon_number_str,
                        row["rsrd_wagon_id"],
                        serialize_payload(payload_obj),
                        serialize_diffs(diffs),
                        row["dataset_json"],
                        now,
                        now,
                    ),
                )
                created += 1

            results.append(
                {
                    "wagon_number_freight": wagon_number_str,
                    "rsrd_wagon_id": row["rsrd_wagon_id"],
                    "rsrd_missing": not bool(row["administrative_json"] or row["design_json"]),
                    "diff_count": diff_count,
                    "differences": diffs,
                    "documents": documents,
                }
            )

        conn.commit()

    return {
        "rows": results,
        "limit": limit,
        "offset": offset,
        "total": total,
        "created": created,
        "erp_env": _normalize_env(env),
        "rsrd_env": rsrd_env_norm,
    }


@app.post("/api/rsrd2/upload_xml")
def rsrd2_upload_xml(
    env: str = Query(DEFAULT_ENV),
    rsrd_env: str | None = Query(None),
    upload: bool = Query(False),
    payload: dict = Body(...),
) -> dict:
    wagon = str(payload.get("wagon") or payload.get("sern") or payload.get("wagon_number") or "").strip()
    if not wagon:
        raise HTTPException(status_code=400, detail="Wagennummer fehlt.")
    digits = re.sub(r"\D", "", wagon)
    wagon_key = digits or wagon

    with _connect() as conn:
        erp_full_table = _table_for(RSRD_ERP_FULL_TABLE, env)
        _ensure_table(conn, erp_full_table, None)
        row = conn.execute(
            f"""
            SELECT *
            FROM {erp_full_table}
            WHERE REPLACE(REPLACE(CAST(WAGEN_SERIENNUMMER AS TEXT), ' ', ''), '-', '') = ?
            LIMIT 1
            """,
            (wagon_key,),
        ).fetchone()
        if not row and wagon:
            row = conn.execute(
                f"""
                SELECT *
                FROM {erp_full_table}
                WHERE CAST(WAGEN_SERIENNUMMER AS TEXT) = ?
                LIMIT 1
                """,
                (wagon,),
            ).fetchone()

    if not row:
        raise HTTPException(status_code=404, detail="Wagen nicht gefunden.")

    erp_row = dict(row)
    try:
        long_text = _fetch_wg_tsi_text(erp_row.get("WAGEN_SERIENNUMMER"), env)
    except HTTPException:
        long_text = ""
    if long_text:
        erp_row["WG_TSI_ZUS_ZERT"] = long_text

    payload_obj = build_erp_payload(erp_row)
    payload_json = serialize_payload(payload_obj)
    try:
        payload_clean = json.loads(payload_json)
    except json.JSONDecodeError:
        payload_clean = payload_obj
    xml = _rsrd_build_upload_xml(payload_clean)

    response_text = None
    response_status = None
    request_url = None
    rsrd_env_norm = _normalize_rsrd_env(rsrd_env, env)
    if upload:
        request_url = _resolve_rsrd_upload_url(rsrd_env_norm)
        if not request_url:
            raise HTTPException(status_code=500, detail="RSRD WSDL URL fehlt.")
        user, password = _rsrd_upload_credentials(rsrd_env_norm)
        headers = {
            "Accept": "text/xml",
            "Content-Type": "text/xml; charset=utf-8",
        }
        import requests

        resp = requests.post(
            request_url,
            headers=headers,
            data=xml.encode("utf-8"),
            auth=(user, password),
            timeout=60,
        )
        response_status = resp.status_code
        response_text = resp.text

    return {
        "wagon": wagon_key,
        "erp_env": _normalize_env(env),
        "rsrd_env": rsrd_env_norm,
        "upload": upload,
        "xml": xml,
        "request_url": request_url,
        "response_status": response_status,
        "response_text": response_text,
    }


@app.post("/api/rsrd2/sync")
def rsrd2_sync(
    env: str = Query(DEFAULT_ENV),
    rsrd_env: str | None = Query(None),
    payload: dict = Body(...),
) -> dict:
    wagons = payload.get("wagons") or []
    if not isinstance(wagons, list) or not all(isinstance(item, str) for item in wagons):
        raise HTTPException(status_code=400, detail="Feld 'wagons' muss eine Liste von Wagennummern sein.")
    snapshots = bool(payload.get("snapshots", True))
    rsrd_env_norm = _normalize_rsrd_env(rsrd_env, env)
    try:
        tables = _rsrd_tables(rsrd_env_norm)
        rsrd_sync_wagons(wagons, keep_snapshots=snapshots, tables=tables, env=rsrd_env_norm)
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=500, detail=str(exc)) from exc
    return {
        "synced": len(wagons),
        "snapshots": snapshots,
        "erp_env": _normalize_env(env),
        "rsrd_env": rsrd_env_norm,
    }


@app.post("/api/rsrd2/sync_all")
def rsrd2_sync_all(
    limit: int | None = Query(None, gt=0),
    snapshots: bool = Query(True),
    env: str = Query(DEFAULT_ENV),
    rsrd_env: str | None = Query(None),
) -> dict:
    rsrd_env_norm = _normalize_rsrd_env(rsrd_env, env)
    with _connect() as conn:
        wagons = _fetch_erp_wagon_numbers(conn, env, limit)
    if not wagons:
        raise HTTPException(status_code=404, detail="Keine Wagennummern im ERP-Cache gefunden.")
    try:
        tables = _rsrd_tables(rsrd_env_norm)
        stats = rsrd_sync_wagons(
            wagons,
            keep_snapshots=snapshots,
            mode="full",
            tables=tables,
            env=rsrd_env_norm,
        )
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=500, detail=str(exc)) from exc
    return {
        "synced": len(wagons),
        "staged": stats["staged"],
        "processed": stats["processed"],
        "snapshots": snapshots,
        "erp_env": _normalize_env(env),
        "rsrd_env": rsrd_env_norm,
    }


@app.post("/api/rsrd2/fetch_json")
def rsrd2_fetch_json(
    limit: int | None = Query(None, gt=0),
    snapshots: bool = Query(False),
    env: str = Query(DEFAULT_ENV),
    rsrd_env: str | None = Query(None),
) -> dict:
    rsrd_env_norm = _normalize_rsrd_env(rsrd_env, env)
    with _connect() as conn:
        wagons = _fetch_erp_wagon_numbers(conn, env, limit)
    if not wagons:
        raise HTTPException(status_code=404, detail="Keine Wagennummern im ERP-Cache gefunden.")
    try:
        tables = _rsrd_tables(rsrd_env_norm)
        stats = rsrd_sync_wagons(
            wagons,
            keep_snapshots=snapshots,
            mode="stage",
            tables=tables,
            env=rsrd_env_norm,
        )
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=500, detail=str(exc)) from exc
    return {
        "staged": stats["staged"],
        "snapshots": snapshots,
        "erp_env": _normalize_env(env),
        "rsrd_env": rsrd_env_norm,
    }


@app.post("/api/rsrd2/process_json")
def rsrd2_process_json(
    limit: int | None = Query(None, gt=0),
    env: str = Query(DEFAULT_ENV),
    rsrd_env: str | None = Query(None),
) -> dict:
    rsrd_env_norm = _normalize_rsrd_env(rsrd_env, env)
    try:
        tables = _rsrd_tables(rsrd_env_norm)
        stats = rsrd_sync_wagons(
            [],
            keep_snapshots=False,
            mode="process",
            process_limit=limit,
            tables=tables,
            env=rsrd_env_norm,
        )
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=500, detail=str(exc)) from exc
    return {
        "processed": stats["processed"],
        "limit": limit,
        "erp_env": _normalize_env(env),
        "rsrd_env": rsrd_env_norm,
    }


# Hidden M3 bridge (always PRD)
@app.get("/M3BRIDGE.html", include_in_schema=False)
def m3_bridge() -> PlainTextResponse:
    _ionapi_path("prd", "mi")
    return PlainTextResponse("OK")


# Hidden SQL bridge (always PRD)
@app.post("/query", include_in_schema=False)
def m3_sql_bridge(
    payload: dict = Body(...),
    format: str | None = Query(None),
) -> dict:
    sql = str(payload.get("sql") or "").strip()
    if not sql:
        raise HTTPException(status_code=400, detail="SQL fehlt.")
    result = _run_compass_query(sql, "prd")
    columns = result.get("columns") or []
    rows = result.get("rows") or []
    fmt = (format or "").strip().lower()
    if fmt in {"kv", "dict", "map", "object"}:
        if rows and not isinstance(rows[0], dict):
            rows = [dict(zip(columns, row)) for row in rows]
        return {"columns": columns, "rows": rows}
    if rows and isinstance(rows[0], dict):
        rows = [[row.get(col) for col in columns] for row in rows]
    return {"columns": columns, "rows": rows}


@app.get("/api/goldenview/list")
def goldenview_list() -> dict:
    with _connect() as conn:
        _init_goldenview_db(conn)
        rows = conn.execute(
            f"""
            SELECT q.id, q.name, q.description, q.created_at, q.excel_path, q.md_path, q.generated_at, q.commit_at,
                   (SELECT COUNT(1) FROM {GOLDENVIEW_FIELDS_TABLE} f WHERE f.query_id = q.id) AS field_count
            FROM {GOLDENVIEW_QUERIES_TABLE} q
            ORDER BY q.id DESC
            """
        ).fetchall()
    return {
        "items": [dict(row) for row in rows],
        "repo_root": str(GOLDENVIEW_EXPORT_DIR),
        "repo_url": os.getenv("GOLDENVIEW_REPO_URL", "https://github.com/crupp-mfd/M3ChatbotExcels"),
    }


@app.get("/api/goldenview/detail/{query_id}")
def goldenview_detail(query_id: int) -> dict:
    with _connect() as conn:
        _init_goldenview_db(conn)
        query = conn.execute(
            f"""
            SELECT id, name, sql_text, description, created_at, excel_path, md_path, generated_at
            FROM {GOLDENVIEW_QUERIES_TABLE}
            WHERE id = ?
            """,
            (query_id,),
        ).fetchone()
        if not query:
            raise HTTPException(status_code=404, detail="Eintrag nicht gefunden.")
        fields = conn.execute(
            f"""
            SELECT field_name, field_description, connected_fields
            FROM {GOLDENVIEW_FIELDS_TABLE}
            WHERE query_id = ?
            ORDER BY id ASC
            """,
            (query_id,),
        ).fetchall()
    return {
        "query": dict(query),
        "fields": [
            {
                "name": row["field_name"],
                "description": row["field_description"] or "",
                "connected_fields": json.loads(row["connected_fields"] or "[]"),
            }
            for row in fields
        ],
    }


@app.post("/api/goldenview/save")
def goldenview_save(payload: dict = Body(...)) -> dict:
    query_id = payload.get("id")
    name = str(payload.get("name") or "").strip()
    sql_text = str(payload.get("sql") or "").strip()
    description = str(payload.get("description") or "").strip()
    fields = payload.get("fields") or []
    if not sql_text:
        raise HTTPException(status_code=400, detail="SQL fehlt.")
    if not isinstance(fields, list):
        raise HTTPException(status_code=400, detail="Felder müssen eine Liste sein.")
    with _connect() as conn:
        _init_goldenview_db(conn)
        if query_id:
            conn.execute(
                f"""
                UPDATE {GOLDENVIEW_QUERIES_TABLE}
                SET name = ?, sql_text = ?, description = ?
                WHERE id = ?
                """,
                (name or None, sql_text, description or None, int(query_id)),
            )
            conn.execute(
                f"DELETE FROM {GOLDENVIEW_FIELDS_TABLE} WHERE query_id = ?",
                (int(query_id),),
            )
            current_id = int(query_id)
        else:
            cur = conn.execute(
                f"INSERT INTO {GOLDENVIEW_QUERIES_TABLE} (name, sql_text, description) VALUES (?, ?, ?)",
                (name or None, sql_text, description or None),
            )
            current_id = cur.lastrowid
        for field in fields:
            field_name = str(field.get("name") or "").strip()
            if not field_name:
                continue
            field_desc = str(field.get("description") or "").strip()
            connected = field.get("connected_fields") or []
            if not isinstance(connected, list):
                connected = []
            conn.execute(
                f"""
                INSERT INTO {GOLDENVIEW_FIELDS_TABLE}
                (query_id, field_name, field_description, connected_fields)
                VALUES (?, ?, ?, ?)
                """,
                (current_id, field_name, field_desc or None, json.dumps(connected)),
            )
        conn.commit()
    return {"id": current_id, "status": "ok"}


@app.post("/api/goldenview/generate")
def goldenview_generate(payload: dict = Body(...)) -> dict:
    query_id = payload.get("id")
    if not query_id:
        raise HTTPException(status_code=400, detail="ID fehlt.")
    job = _create_job("goldenview_generate", "prd")
    threading.Thread(target=_goldenview_job, args=(int(query_id), job["id"]), daemon=True).start()
    return {"job_id": job["id"], "status": job["status"]}


@app.get("/api/goldenview/jobs/{job_id}")
def goldenview_job_status(job_id: str) -> dict:
    return _job_snapshot(job_id)


@app.get("/api/goldenview/file/download")
def goldenview_file(path: str = Query(...)) -> Response:
    file_path = Path(path)
    try:
        resolved = file_path.resolve()
        base = GOLDENVIEW_EXPORT_DIR.resolve()
        if base not in resolved.parents and base != resolved:
            raise HTTPException(status_code=400, detail="Ungültiger Pfad.")
        if not resolved.exists():
            raise HTTPException(status_code=404, detail="Datei nicht gefunden.")
        return FileResponse(resolved)
    except HTTPException:
        raise
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=400, detail=str(exc)) from exc


@app.post("/api/goldenview/commit")
def goldenview_commit(payload: dict = Body(None)) -> dict:
    repo = GOLDENVIEW_EXPORT_DIR
    if not (repo / ".git").exists():
        raise HTTPException(status_code=400, detail="Repo nicht gefunden. Setze GOLDENVIEW_REPO_PATH.")
    message = (payload or {}).get("message") or "GoldenView update"
    query_id = (payload or {}).get("id")
    try:
        subprocess.run(["git", "-C", str(repo), "add", "."], check=True, capture_output=True, text=True)
        subprocess.run(["git", "-C", str(repo), "commit", "-m", message], check=False, capture_output=True, text=True)
        subprocess.run(["git", "-C", str(repo), "push"], check=True, capture_output=True, text=True)
    except subprocess.CalledProcessError as exc:
        detail = (exc.stderr or exc.stdout or str(exc)).strip()
        raise HTTPException(status_code=500, detail=detail) from exc
    if query_id:
        with _connect() as conn:
            _init_goldenview_db(conn)
            conn.execute(
                f"UPDATE {GOLDENVIEW_QUERIES_TABLE} SET commit_at = datetime('now') WHERE id = ?",
                (int(query_id),),
            )
            conn.commit()
    return {"status": "ok"}


@app.get("/api/goldenview/sync_status")
def goldenview_sync_status() -> dict:
    if not GITHUB_SYNC_TOKEN:
        return {"status": "missing_token"}
    headers = {"Authorization": f"Bearer {GITHUB_SYNC_TOKEN}", "Accept": "application/vnd.github+json"}
    url = f"https://api.github.com/repos/{GITHUB_SYNC_REPO}/actions/workflows/{GITHUB_SYNC_WORKFLOW}/runs"
    try:
        with httpx.Client(timeout=15.0) as client:
            resp = client.get(url, headers=headers, params={"per_page": 1})
        if resp.status_code != 200:
            return {"status": "error", "detail": resp.text}
        data = resp.json()
        runs = data.get("workflow_runs") or []
        if not runs:
            return {"status": "no_runs"}
        run = runs[0]
        return {
            "status": run.get("status"),
            "conclusion": run.get("conclusion"),
            "created_at": run.get("created_at"),
            "updated_at": run.get("updated_at"),
            "html_url": run.get("html_url"),
            "name": run.get("name"),
        }
    except Exception as exc:  # noqa: BLE001
        return {"status": "error", "detail": str(exc)}


@app.post("/api/ask_m3_knowledge")
@app.post("/api/ask_m3_knowledge/")
def ask_m3_knowledge(payload: dict = Body(...), request: Request = None) -> dict:
    if GPT_ACTION_API_KEY:
        api_key = request.headers.get("x-api-key", "").strip() if request else ""
        if api_key != GPT_ACTION_API_KEY:
            raise HTTPException(status_code=401, detail="Unauthorized")
    question = str(payload.get("question") or "").strip()
    if not question:
        raise HTTPException(status_code=400, detail="Frage fehlt.")
    if not OPENAI_API_KEY:
        raise HTTPException(status_code=500, detail="OPENAI_API_KEY fehlt.")
    if not OPENAI_VECTOR_STORE_ID:
        raise HTTPException(status_code=500, detail="OPENAI_VECTOR_STORE_ID fehlt.")
    try:
        client = OpenAI(api_key=OPENAI_API_KEY)
        resp = client.responses.create(
            model="gpt-4.1",
            input=question,
            tools=[
                {
                    "type": "file_search",
                    "vector_store_ids": [OPENAI_VECTOR_STORE_ID],
                    "max_num_results": 5,
                }
            ],
            include=["file_search_call.results"],
        )
        output_text = getattr(resp, "output_text", None)
        if output_text is None:
            output_text = ""
            for item in resp.output:
                if item.get("type") == "message":
                    for part in item.get("content", []):
                        if part.get("type") == "output_text":
                            output_text += part.get("text", "")
        citations = []
        for item in resp.output:
            item_type = getattr(item, "type", None)
            if item_type == "message":
                content = getattr(item, "content", []) or []
                for part in content:
                    annotations = getattr(part, "annotations", []) or []
                    for ann in annotations:
                        if getattr(ann, "type", None) == "file_citation":
                            citations.append(
                                {
                                    "file_id": getattr(ann, "file_id", None),
                                    "filename": getattr(ann, "filename", None),
                                }
                            )
        if not citations:
            return {"answer": "Keine passende Quelle in der Wissensbasis gefunden.", "citations": []}
        return {"answer": output_text, "citations": citations}
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=500, detail=str(exc)) from exc


_MEHR_MODULE = None
MEHR_DEFAULT_YEAR = 2025
MEHR_FABRIC_IMPORT_JOB_TYPE = "mehrkilometer_fabric_import"
MEHR_ALLOWED_DOWNLOAD_KINDS = {
    "source_overview",
    "source_kilometer",
    "output_overview",
    "output_details",
    "special_output_overview",
    "special_output_details",
}


def _mehr_module():
    global _MEHR_MODULE
    if _MEHR_MODULE is None:
        try:
            _MEHR_MODULE = importlib.import_module("app")
        except Exception as exc:  # noqa: BLE001
            raise HTTPException(
                status_code=500,
                detail=f"Mehrkilometer-Modul konnte nicht geladen werden: {exc}",
            ) from exc
    return _MEHR_MODULE


def _mehr_sanitize_year(year: int | None) -> int:
    if year is None:
        return MEHR_DEFAULT_YEAR
    return max(2000, min(3000, int(year)))


def _mehr_set_job_result(job_id: str, values: Dict[str, Any]) -> None:
    if not values:
        return
    with _jobs_lock:
        job = _jobs.get(job_id)
        if not job:
            return
        current = job.get("result")
        if not isinstance(current, dict):
            current = {}
        current.update(values)
        job["result"] = current


def _mehr_get_job_result(job_id: str) -> Dict[str, Any]:
    with _jobs_lock:
        job = _jobs.get(job_id)
        if not job:
            return {}
        result = job.get("result")
        if isinstance(result, dict):
            return dict(result)
    return {}


def _mehr_job_elapsed_seconds(snapshot: Dict[str, Any]) -> int | None:
    started_raw = snapshot.get("started")
    if not started_raw:
        return None
    try:
        started_at = datetime.fromisoformat(str(started_raw))
    except Exception:
        return None
    finished_raw = snapshot.get("finished")
    if finished_raw:
        try:
            finished_at = datetime.fromisoformat(str(finished_raw))
        except Exception:
            finished_at = datetime.utcnow()
    else:
        finished_at = datetime.utcnow()
    elapsed = int((finished_at - started_at).total_seconds())
    return max(0, elapsed)


def _mehr_fabric_progress_callback(job_id: str):
    def _callback(payload: Dict[str, Any]) -> None:
        stage = str(payload.get("stage") or "").strip()
        message = str(payload.get("message") or "").strip()
        rows_total = payload.get("rows_total")
        rows_written = payload.get("rows_written")
        update: Dict[str, Any] = {}
        if stage:
            update["stage"] = stage
        if message:
            update["message"] = message
        if rows_total is not None:
            update["rows_total"] = int(rows_total)
        if rows_written is not None:
            update["rows_written"] = int(rows_written)
        _mehr_set_job_result(job_id, update)
        if message:
            _append_job_log(job_id, message)

    return _callback


def _mehr_run_fabric_import_job(job_id: str, year: int) -> None:
    mod = _mehr_module()
    _mehr_set_job_result(
        job_id,
        {
            "year": year,
            "stage": "starting",
            "message": "Import startet ...",
            "rows_total": None,
            "rows_written": 0,
        },
    )
    try:
        payload = mod._refresh_mehrkilometer_fabric_sqlite(
            year,
            progress=_mehr_fabric_progress_callback(job_id),
        )
        final = _mehr_get_job_result(job_id)
        final.update(payload)
        final["stage"] = "completed"
        final["message"] = "Import abgeschlossen."
        final["rows_total"] = int(payload.get("row_count") or 0)
        final["rows_written"] = int(payload.get("row_count") or 0)
        _finish_job(job_id, "completed", result=final)
    except Exception as exc:  # noqa: BLE001
        final = _mehr_get_job_result(job_id)
        final["year"] = year
        final["stage"] = "failed"
        final["message"] = str(exc)
        _finish_job(job_id, "failed", result=final, error=str(exc))


def _mehr_file_info(mod, file_path: Path | None, kind: str) -> dict:
    if file_path is None:
        return {"exists": False}
    download_url = (
        f"/api/mehrkilometer/download?kind={quote(kind, safe='')}&name={quote(file_path.name, safe='')}"
    )
    try:
        rel_path = file_path.resolve().relative_to(mod.REPO_ROOT.resolve()).as_posix()
        download_url += f"&rel={quote(rel_path, safe='')}"
    except ValueError:
        pass
    return {
        "exists": True,
        "name": file_path.name,
        "download_url": download_url,
    }


def _mehr_extract_year_from_filename(file_path: Path | None) -> int | None:
    if file_path is None:
        return None
    match = re.search(r"_(\d{4})_\d{8}_\d{6}\.xlsx$", file_path.name)
    if not match:
        return None
    try:
        return int(match.group(1))
    except ValueError:
        return None


def _mehr_build_sources_payload(year: int) -> dict:
    mod = _mehr_module()
    source_overview, source_km = mod._discover_source_files(year)
    output_overview = mod._pick_latest_file(
        mod.MEHR_OUTPUT_DIR,
        lambda n: n.endswith(".xlsx") and n.startswith(f"vertragsuebersicht_{year}_"),
    )
    output_details = mod._pick_latest_file(
        mod.MEHR_OUTPUT_DIR,
        lambda n: n.endswith(".xlsx") and n.startswith(f"einzelabrechnungen_detail_{year}_"),
    )
    special_output_overview = mod._pick_latest_file(
        mod.MEHR_OUTPUT_DIR,
        lambda n: n.endswith(".xlsx") and n.startswith(f"special_vertragsuebersicht_{year}_"),
    )
    special_output_details = mod._pick_latest_file(
        mod.MEHR_OUTPUT_DIR,
        lambda n: n.endswith(".xlsx")
        and n.startswith(f"special_einzelabrechnungen_detail_{year}_"),
    )
    special_from_other_year = False
    if special_output_overview is None:
        special_output_overview = mod._pick_latest_file(
            mod.MEHR_OUTPUT_DIR,
            lambda n: n.endswith(".xlsx") and n.startswith("special_vertragsuebersicht_"),
        )
        special_from_other_year = special_output_overview is not None
    if special_output_details is None:
        special_output_details = mod._pick_latest_file(
            mod.MEHR_OUTPUT_DIR,
            lambda n: n.endswith(".xlsx")
            and n.startswith("special_einzelabrechnungen_detail_"),
        )
        special_from_other_year = special_from_other_year or (special_output_details is not None)

    browser_url = None
    try:
        browser_url = f"/{mod.MEHR_OUTPUT_DIR.resolve().relative_to(mod.REPO_ROOT.resolve()).as_posix()}/"
    except ValueError:
        browser_url = None

    payload = {
        "recommended_year": year,
        "sources": {
            "overview": _mehr_file_info(mod, source_overview, "source_overview"),
            "kilometer": _mehr_file_info(mod, source_km, "source_kilometer"),
        },
        "outputs": {
            "overview": _mehr_file_info(mod, output_overview, "output_overview"),
            "details": _mehr_file_info(mod, output_details, "output_details"),
        },
        "special_outputs": {
            "overview": _mehr_file_info(
                mod, special_output_overview, "special_output_overview"
            ),
            "details": _mehr_file_info(mod, special_output_details, "special_output_details"),
        },
        "paths": {
            "output_dir": {
                "path": str(mod.MEHR_OUTPUT_DIR),
                "browser_url": browser_url,
            },
            "source_dirs": [str(path) for path in mod._source_search_dirs()],
        },
    }

    if source_overview is None or source_km is None:
        payload["warning"] = "Eine oder mehrere Quelldateien fehlen."
    elif output_overview is None or output_details is None:
        payload["warning"] = "Für dieses Jahr liegen noch keine lokalen Output-Dateien vor."
    elif special_from_other_year:
        fallback_year = _mehr_extract_year_from_filename(special_output_overview)
        if fallback_year is None:
            fallback_year = _mehr_extract_year_from_filename(special_output_details)
        if fallback_year is not None and fallback_year != year:
            payload["warning"] = (
                f"Für {year} wurden keine Spezialdateien gefunden. "
                f"Es werden die letzten verfügbaren Spezialdateien aus {fallback_year} angezeigt."
            )
    return payload


def _mehr_safe_repo_relative_file(mod, rel_path: str) -> Path | None:
    if not rel_path or rel_path.startswith("/") or rel_path.startswith("\\"):
        return None
    if ".." in rel_path:
        return None
    candidate = (mod.REPO_ROOT / rel_path).resolve()
    try:
        candidate.relative_to(mod.REPO_ROOT.resolve())
    except ValueError:
        return None
    if not candidate.is_file():
        return None
    return candidate


def _mehr_safe_file(base_dir: Path, filename: str) -> Path | None:
    if "/" in filename or "\\" in filename or ".." in filename:
        return None
    candidate = (base_dir / filename).resolve()
    try:
        candidate.relative_to(base_dir.resolve())
    except ValueError:
        return None
    if not candidate.is_file():
        return None
    return candidate


def _mehr_is_allowed_download_file(mod, path: Path, kind: str) -> bool:
    if kind in {"source_overview", "source_kilometer"}:
        return mod._is_under_any_directory(path, mod._source_search_dirs())
    return mod._is_under_directory(path, mod.MEHR_OUTPUT_DIR)


def _mehr_resolve_source_file_for_download(mod, filename: str, kind: str) -> Path | None:
    if "/" in filename or "\\" in filename or ".." in filename:
        return None
    candidates = [path for path in mod._collect_source_excel_files() if path.name == filename]
    if not candidates:
        return None
    year_hint = mod._guess_source_year(filename) or MEHR_DEFAULT_YEAR
    source_kind = "overview" if kind == "source_overview" else "kilometer"
    return mod._select_best_source_file(candidates, source_kind, year_hint)


@app.get("/api/mehrkilometer/sources")
def mehrkilometer_sources(year: int = Query(MEHR_DEFAULT_YEAR)) -> dict:
    return _mehr_build_sources_payload(_mehr_sanitize_year(year))


@app.get("/api/mehrkilometer/download")
def mehrkilometer_download(
    kind: str = Query(...),
    name: str = Query(...),
    rel: str = Query(""),
) -> Response:
    mod = _mehr_module()
    if kind not in MEHR_ALLOWED_DOWNLOAD_KINDS:
        raise HTTPException(status_code=400, detail="Ungültiger Download-Typ.")

    file_path = None
    if rel:
        decoded_rel = unquote(rel)
        file_path = _mehr_safe_repo_relative_file(mod, decoded_rel)
        if file_path is not None and file_path.name != name:
            file_path = None
        if file_path is not None and not _mehr_is_allowed_download_file(mod, file_path, kind):
            file_path = None

    if file_path is None and kind in {"source_overview", "source_kilometer"}:
        file_path = _mehr_resolve_source_file_for_download(mod, name, kind)

    if file_path is None and kind in {
        "output_overview",
        "output_details",
        "special_output_overview",
        "special_output_details",
    }:
        file_path = _mehr_safe_file(mod.MEHR_OUTPUT_DIR, name)

    if file_path is None:
        raise HTTPException(status_code=404, detail="Datei nicht gefunden.")

    return FileResponse(
        file_path,
        filename=file_path.name,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )


@app.get("/api/mehrkilometer/fabric/health")
def mehrkilometer_fabric_health() -> dict:
    mod = _mehr_module()
    try:
        return mod._fabric_health_check()
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=500, detail=str(exc)) from exc


@app.get("/api/mehrkilometer/fabric/status")
def mehrkilometer_fabric_status(year: int = Query(MEHR_DEFAULT_YEAR)) -> dict:
    mod = _mehr_module()
    safe_year = _mehr_sanitize_year(year)
    try:
        return mod._mehrkilometer_fabric_status(safe_year)
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=500, detail=str(exc)) from exc


@app.post("/api/mehrkilometer/fabric/import/start")
def mehrkilometer_fabric_import_start(year: int = Query(MEHR_DEFAULT_YEAR)) -> dict:
    safe_year = _mehr_sanitize_year(year)
    job = _create_job(MEHR_FABRIC_IMPORT_JOB_TYPE, "prd")
    _mehr_set_job_result(
        job["id"],
        {
            "year": safe_year,
            "stage": "queued",
            "message": "Import wurde gestartet.",
            "rows_total": None,
            "rows_written": 0,
        },
    )
    threading.Thread(
        target=_mehr_run_fabric_import_job,
        args=(job["id"], safe_year),
        daemon=True,
    ).start()
    return {"job_id": job["id"], "status": "running", "year": safe_year}


@app.get("/api/mehrkilometer/fabric/import/status")
def mehrkilometer_fabric_import_status(job_id: str = Query(...)) -> dict:
    snapshot = _job_snapshot(job_id)
    if snapshot.get("type") != MEHR_FABRIC_IMPORT_JOB_TYPE:
        raise HTTPException(status_code=400, detail="Ungültiger Job-Typ.")
    result = snapshot.get("result")
    if not isinstance(result, dict):
        result = {}
    return {
        "job_id": job_id,
        "status": snapshot.get("status"),
        "stage": result.get("stage"),
        "message": result.get("message"),
        "year": result.get("year"),
        "rows_written": result.get("rows_written"),
        "rows_total": result.get("rows_total"),
        "elapsed_seconds": _mehr_job_elapsed_seconds(snapshot),
        "started": snapshot.get("started"),
        "finished": snapshot.get("finished"),
        "error": snapshot.get("error"),
        "result": result if snapshot.get("status") == "completed" else None,
    }


@app.post("/api/mehrkilometer/fabric/import")
def mehrkilometer_fabric_import(year: int = Query(MEHR_DEFAULT_YEAR)) -> dict:
    mod = _mehr_module()
    safe_year = _mehr_sanitize_year(year)
    try:
        return mod._refresh_mehrkilometer_fabric_sqlite(safe_year)
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=500, detail=str(exc)) from exc


@app.get("/api/mehrkilometer/fabric/stagli")
def mehrkilometer_fabric_stagli(limit: int = Query(100)) -> dict:
    mod = _mehr_module()
    safe_limit = max(1, min(5000, int(limit)))
    try:
        values = mod._fetch_stagli_scnm(safe_limit)
        return {
            "count": len(values),
            "query": f"SELECT TOP ({safe_limit}) SCNM FROM landing.stagli",
            "rows": values,
        }
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=500, detail=str(exc)) from exc


@app.post("/api/mehrkilometer/create")
def mehrkilometer_create(year: int = Query(MEHR_DEFAULT_YEAR)) -> dict:
    mod = _mehr_module()
    safe_year = _mehr_sanitize_year(year)
    try:
        created_overview, created_details = mod._create_settlement_files(safe_year)
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(status_code=500, detail=f"Abrechnung fehlgeschlagen: {exc}") from exc

    payload = _mehr_build_sources_payload(safe_year)
    payload["outputs"] = {
        "overview": _mehr_file_info(mod, created_overview, "output_overview"),
        "details": _mehr_file_info(mod, created_details, "output_details"),
    }
    return payload


@app.post("/api/mehrkilometer/create-special")
def mehrkilometer_create_special(year: int = Query(MEHR_DEFAULT_YEAR)) -> dict:
    mod = _mehr_module()
    safe_year = _mehr_sanitize_year(year)
    try:
        created_overview, created_details = mod._create_special_settlement_files(safe_year)
    except Exception as exc:  # noqa: BLE001
        raise HTTPException(
            status_code=500,
            detail=f"Spezialabrechnung fehlgeschlagen: {exc}",
        ) from exc

    payload = _mehr_build_sources_payload(safe_year)
    payload["special_outputs"] = {
        "overview": _mehr_file_info(mod, created_overview, "special_output_overview"),
        "details": _mehr_file_info(mod, created_details, "special_output_details"),
    }
    payload["warning"] = "Spezialabrechnung für Grampet, Railcare und Raildox wurde erzeugt."
    return payload


@app.post("/api/mehrkilometer/vertragsexcel/import")
def mehrkilometer_vertragsexcel_import(year: int = Query(MEHR_DEFAULT_YEAR)) -> dict:
    safe_year = _mehr_sanitize_year(year)
    raise HTTPException(
        status_code=501,
        detail=f"Vertragsexcel-SQL für Jahr {safe_year} ist noch nicht hinterlegt.",
    )


@app.get("/apps/christian/AppRSRD", include_in_schema=False)
@app.get("/apps/christian/AppRSRD/", include_in_schema=False)
def rsrd_legacy_entry_redirect() -> RedirectResponse:
    return RedirectResponse(
        url="/apps/christian/AppRSRD/frontend/rsrd2.html",
        status_code=302,
    )


@app.get("/apps/christian/AppMehrkilometer/frontend", include_in_schema=False)
@app.get("/apps/christian/AppMehrkilometer/frontend/", include_in_schema=False)
def mehrkilometer_frontend_redirect() -> RedirectResponse:
    return RedirectResponse(
        url="/apps/christian/AppMehrkilometer/frontend/index.html",
        status_code=302,
    )


@app.get("/api-config.js", include_in_schema=False)
def shared_api_config_js() -> PlainTextResponse:
    return PlainTextResponse(
        "window.__SPAREPART_API_CONFIG__ = { CORE_API_BASE_URL: '' };",
        media_type="application/javascript",
    )


@app.get("/", include_in_schema=False)
def portal_index() -> FileResponse:
    portal_file = PROJECT_ROOT / "index.html"
    if not portal_file.exists():
        raise HTTPException(status_code=404, detail="Portal index.html nicht gefunden.")
    return FileResponse(
        portal_file,
        media_type="text/html",
        headers={"Cache-Control": "no-store, no-cache, must-revalidate, max-age=0"},
    )


# Serve frontend assets
APPS_DIR = PROJECT_ROOT / "apps"
if APPS_DIR.exists():
    app.mount(
        "/apps",
        AuthStaticFiles(directory=APPS_DIR, html=True),
        name="apps-static",
    )

if FRONTEND_DIR.exists():
    app.mount(
        "/",
        AuthStaticFiles(directory=FRONTEND_DIR, html=True),
        name="frontend",
    )
else:
    @app.get("/")
    def placeholder() -> dict:
        return {"message": "Frontend noch nicht angelegt. Lege Dateien in /frontend/ ab."}
